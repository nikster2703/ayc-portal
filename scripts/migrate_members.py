"""
migrate_members.py — Import existing member data from the spreadsheet into SQLite.

Run from inside the ayc-portal directory:
    python3 scripts/migrate_members.py

The script is safe to re-run: it uses INSERT OR IGNORE on member_id so it won't
duplicate records. Contacts are cleared and re-inserted on each run so edits in
the spreadsheet are picked up cleanly.
"""

import os
import sys

import sqlcipher3 as sqlite3  # SQLCipher — transparent AES-256 encryption at rest

try:
    import openpyxl
except ImportError:
    sys.exit('openpyxl not found. Run: pip3 install openpyxl')

try:
    from dotenv import load_dotenv
except ImportError:
    sys.exit('python-dotenv not found. Run: pip3 install python-dotenv')

# ── Paths ──────────────────────────────────────────────────────────────────────

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
APP_DIR      = os.path.dirname(SCRIPT_DIR)
# INSTANCE_DIR separates runtime data from code — used in Docker deployments.
# Falls back to APP_DIR for direct (non-Docker) installs.
INSTANCE_DIR = os.environ.get('INSTANCE_DIR', APP_DIR)
DB_PATH      = os.path.join(INSTANCE_DIR, 'data', 'ayc.db')

# Load .env — try instance dir first, fall back to app dir for direct installs.
load_dotenv(os.path.join(INSTANCE_DIR, '.env'))
load_dotenv(os.path.join(APP_DIR, '.env'))


def _connect_db(path):
    """Open a SQLCipher-encrypted DB connection. Raises if key is missing."""
    key = os.environ.get('DB_ENCRYPTION_KEY')
    if not key:
        sys.exit('ERROR: DB_ENCRYPTION_KEY is not set in .env — cannot open the database.')
    conn = sqlite3.connect(path)
    conn.execute(f"PRAGMA key='{key}'")
    conn.execute('SELECT count(*) FROM sqlite_master')  # verify key immediately
    return conn

# The spreadsheet lives one level up (inside the AYC Member Lookup folder)
XLSX_PATH    = os.path.join(BASE_DIR, '..', 'SYC Member Details-2.xlsx')
XLSX_PATH    = os.path.normpath(XLSX_PATH)

# ── Helpers ────────────────────────────────────────────────────────────────────

def clean(val):
    """Return a stripped string, or empty string for None/NaN."""
    if val is None:
        return ''
    s = str(val).strip()
    return '' if s.lower() in ('nan', 'none', '-') else s

def to_date(val):
    """Convert a date value to YYYY-MM-DD string, or empty string."""
    if val is None:
        return ''
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    s = clean(val)
    return s

def yn_to_int(val):
    """Convert YES/NO/TRUE/FALSE/1/0 to 1 or 0."""
    s = clean(val).upper()
    return 1 if s in ('YES', 'TRUE', '1') else 0

CLUB_SHORT_NAME = os.environ.get('CLUB_SHORT_NAME', 'AYC')

def next_member_id(conn):
    """Find the highest existing member number and return the next one."""
    prefix = CLUB_SHORT_NAME
    prefix_len = len(prefix)
    row = conn.execute(
        f"SELECT member_id FROM members WHERE member_id LIKE '{prefix}%'"
        f" ORDER BY CAST(SUBSTR(member_id, {prefix_len + 1}) AS INTEGER) DESC LIMIT 1"
    ).fetchone()
    if row:
        try:
            return int(row[0][prefix_len:]) + 1
        except (ValueError, AttributeError):
            pass
    return 1

# ── Main ───────────────────────────────────────────────────────────────────────

def migrate():
    if not os.path.exists(XLSX_PATH):
        sys.exit(f'Spreadsheet not found at:\n  {XLSX_PATH}\n'
                 f'Check the path and try again.')

    if not os.path.exists(DB_PATH):
        sys.exit(f'Database not found at:\n  {DB_PATH}\n'
                 f'Run "python3 app.py" once first to create it, '
                 f'then run this script.')

    print(f'Reading spreadsheet: {XLSX_PATH}')
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f'Columns found: {headers}')

    conn = _connect_db(DB_PATH)
    conn.execute('PRAGMA foreign_keys = ON')

    counter    = next_member_id(conn)
    imported   = 0
    skipped    = 0
    updated    = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data   = dict(zip(headers, row))
        first_name = clean(row_data.get('First Name'))
        surname    = clean(row_data.get('Surname'))

        # Skip genuinely blank rows
        if not first_name and not surname:
            skipped += 1
            continue

        member_id = f'{CLUB_SHORT_NAME}{str(counter).zfill(3)}'

        existing = conn.execute(
            'SELECT id FROM members WHERE first_name = ? AND surname = ? AND date_of_birth = ?',
            (first_name, surname, to_date(row_data.get('D.O.B.')))
        ).fetchone()

        if existing:
            # Already in DB — update rather than duplicate
            member_db_id = existing[0]
            conn.execute('''
                UPDATE members SET
                    address            = ?,
                    postcode           = ?,
                    ethnicity_religion = ?,
                    medical_sen        = ?,
                    gp_contact         = ?,
                    unattended_exit    = ?,
                    gdpr_consent       = ?,
                    status             = ?,
                    session            = ?,
                    date_registered    = ?,
                    comments           = ?,
                    updated_at         = datetime('now')
                WHERE id = ?
            ''', (
                clean(row_data.get('Address')),
                clean(row_data.get('PostCode')),
                clean(row_data.get('Ethnicity/Religion')),
                clean(row_data.get('Medical / SEN')),
                clean(row_data.get('GP Contact')),
                yn_to_int(row_data.get('Unattended Exit')),
                yn_to_int(row_data.get('GDPR Consent')),
                clean(row_data.get('Status')) or 'Active',
                clean(row_data.get('Session')),
                to_date(row_data.get('Date Registered')),
                clean(row_data.get('Comments')),
                member_db_id,
            ))
            updated += 1
        else:
            conn.execute('''
                INSERT INTO members
                    (member_id, first_name, surname, date_of_birth,
                     address, postcode, ethnicity_religion,
                     medical_sen, gp_contact, unattended_exit,
                     gdpr_consent, status, session,
                     date_registered, comments)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                member_id,
                first_name,
                surname,
                to_date(row_data.get('D.O.B.')),
                clean(row_data.get('Address')),
                clean(row_data.get('PostCode')),
                clean(row_data.get('Ethnicity/Religion')),
                clean(row_data.get('Medical / SEN')),
                clean(row_data.get('GP Contact')),
                yn_to_int(row_data.get('Unattended Exit')),
                yn_to_int(row_data.get('GDPR Consent')),
                clean(row_data.get('Status')) or 'Active',
                clean(row_data.get('Session')),
                to_date(row_data.get('Date Registered')),
                clean(row_data.get('Comments')),
            ))
            member_db_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            counter += 1
            imported += 1

        # Contacts — delete and re-insert so re-runs stay clean
        conn.execute('DELETE FROM member_contacts WHERE member_id = ?', (member_db_id,))

        c1_name  = clean(row_data.get('Contact Name 1'))
        c1_phone = clean(row_data.get('Contact Number 1'))
        c1_email = clean(row_data.get('Contact Email Address'))
        if c1_name or c1_phone or c1_email:
            conn.execute(
                'INSERT INTO member_contacts (member_id, contact_order, contact_name, contact_phone, contact_email)'
                ' VALUES (?,1,?,?,?)',
                (member_db_id, c1_name, c1_phone, c1_email)
            )

        c2_name  = clean(row_data.get('Contact Name 2'))
        c2_phone = clean(row_data.get('Contact Number 2'))
        c2_email = clean(row_data.get('Contact Email Address.1'))
        if c2_name or c2_phone or c2_email:
            conn.execute(
                'INSERT INTO member_contacts (member_id, contact_order, contact_name, contact_phone, contact_email)'
                ' VALUES (?,2,?,?,?)',
                (member_db_id, c2_name, c2_phone, c2_email)
            )

    conn.commit()
    conn.close()

    print(f'\nMigration complete:')
    print(f'  Imported : {imported} new members')
    print(f'  Updated  : {updated} existing members')
    print(f'  Skipped  : {skipped} blank rows')


if __name__ == '__main__':
    migrate()
