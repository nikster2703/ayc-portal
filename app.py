"""
AYC Portal — Flask Application  v8.0
Phases 1-7: Auth, members, audit, user admin, approvals, register,
            documents, comms, term calendar, staff registrations,
            configurable Roles + Permissions, Member Alert Rules.

Phase roadmap (this file grows into blueprints as phases are added):
  Phase 1 — Auth, members lookup, edit/delete, audit log, user admin    ✓
  Phase 2 — Approvals: review pending registrations                     ✓
  Phase 3 — Digital session register + attendance history + auto-leaver ✓
  Phase 4 — Document repository, email templates, mailshots             ✓
  Phase 5 — Term calendar, staff registrations, user permanent delete   ✓
  Phase 6 — Configurable Roles + Permissions (DB-driven, replaces all   ✓
             hard-coded role checks with permission_required decorator)
  Phase 7 — Member Alert Rules: configurable multi-rule flag engine,    ✓
             nightly auto-check, replaces hardcoded At Risk status
  Phase 8 — Duke of Edinburgh module

To split into blueprints later, each section marked ## BLUEPRINT: <name>
can be extracted to blueprints/<name>.py and registered with app.register_blueprint().
"""

import os
import json
import secrets
import sqlcipher3 as sqlite3  # SQLCipher — transparent AES-256 encryption at rest
import hashlib
import base64
import re
import smtplib
import time
import urllib.request
import urllib.error
import urllib.parse
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from functools import wraps

import bcrypt
from cryptography.fernet import Fernet
from dotenv import load_dotenv
from flask import (Flask, g, jsonify, redirect, render_template,
                   request, session, url_for, send_from_directory,
                   Response, stream_with_context)
from flask_wtf.csrf import CSRFProtect, CSRFError
from werkzeug.utils import secure_filename
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

# ── Instance directory (multi-tenant) ─────────────────────────────────────────
# INSTANCE_DIR is set by the service manager (launchd / systemd) for each club.
# It points to the club's own folder containing .env and data/.
# When running as a single-instance install (no INSTANCE_DIR), BASE_DIR is used.
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
INSTANCE_DIR = os.environ.get('INSTANCE_DIR', BASE_DIR)

# Load .env from the instance directory before anything reads os.environ
load_dotenv(os.path.join(INSTANCE_DIR, '.env'))

# ── App setup ──────────────────────────────────────────────────────────────────

app = Flask(__name__)
_secret_key = os.environ.get('SECRET_KEY')
if not _secret_key:
    import warnings
    warnings.warn(
        'SECRET_KEY is not set in .env — a random key will be used and all sessions '
        'will be invalidated on every restart. Set SECRET_KEY in your .env file.',
        stacklevel=2,
    )
    _secret_key = secrets.token_hex(32)
app.secret_key = _secret_key
app.permanent_session_lifetime = timedelta(hours=8)

# ── CSRF protection ────────────────────────────────────────────────────────────
# CSRFProtect validates the X-CSRFToken header on every non-GET/HEAD/OPTIONS
# request. The token is rendered into base.html via {{ csrf_token() }} and read
# by utils.js on page load. Endpoints that must remain CSRF-exempt are decorated
# with @csrf.exempt (login — pre-auth; public registration — unauthenticated).
app.config['WTF_CSRF_HEADERS']     = ['X-CSRFToken']   # header name used by apiFetch()
app.config['WTF_CSRF_TIME_LIMIT']  = None               # bounded by session lifetime (8 h)
csrf = CSRFProtect(app)

DATABASE   = os.path.join(INSTANCE_DIR, 'data', 'ayc.db')
UPLOAD_DIR = os.path.join(INSTANCE_DIR, 'data', 'documents')
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {'pdf', 'docx', 'doc', 'jpg', 'jpeg', 'png', 'xlsx', 'xls'}
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024  # 20 MB max upload

APP_VERSION = 'v8.9'  # v8.9: Field Builder overhaul — edit fields inline, catalogue preview+edit+delete

# ── Permission catalogue ───────────────────────────────────────────────────────
# Single source of truth for every permission code the app supports.
# Used to seed the DB on first run and to populate the roles editor UI.
ALL_PERMISSIONS = [
    # Members
    ('members.view',        'View Members',            'View member list and full detail cards',               'members'),
    ('members.edit',        'Edit Members',             'Edit member records',                                  'members'),
    ('members.delete',      'Soft Delete Members',      'Mark a member as Leaver (reversible)',                 'members'),
    ('members.hard_delete', 'Permanent Delete Members', 'Permanently and irreversibly delete a member',         'members'),
    # Register / attendance
    ('register.signin',     'Sign In',                  'Sign members in on the session register',              'register'),
    ('register.signout',    'Sign Out',                 'Sign members out on the session register',              'register'),
    ('register.complete',   'Complete Register',        'Lock the register at end of session',                  'register'),
    ('register.reset',      'Reset Register',           'Wipe all sign-in/out data for a session',             'register'),
    # Alert rules (v8.0)
    ('alerts.view',        'View Alerts',               'See member flags and the alert rules list',            'alerts'),
    ('alerts.manage',      'Manage Alert Rules',        'Create and edit alert rules',                         'alerts'),
    ('alerts.run',         'Run Alert Checks',          'Trigger an alert rule evaluation manually',           'alerts'),
    ('alerts.dismiss',     'Dismiss Flags',             'Manually clear a flag from a member record',          'alerts'),
    ('register.print',      'Print Register',           'Print a paper copy of the session register',           'register'),
    ('members.tags',        'Manage Member Tags',        'Add and remove skill/badge tags on member records',    'members'),
    ('register.notes',      'Session Notes',            'Add and manage session incident and general notes',    'register'),
    ('register.export',     'Export Complete Register', 'Open print-ready export with attendance and notes',    'register'),
    ('register.qr_manage',  'Manage QR Code',           'Regenerate the QR quick sign-in code on the register', 'register'),
    # Approvals
    ('approvals.view',      'View Approvals',           'View pending self-registration submissions',           'approvals'),
    ('approvals.approve',   'Approve Registrations',    'Approve a pending registration',                       'approvals'),
    ('approvals.reject',    'Reject Registrations',     'Reject a pending registration',                        'approvals'),
    # Documents
    ('documents.view',      'View Documents',           'Browse the document repository (per-doc rank still applies)', 'documents'),
    ('documents.upload',    'Upload Documents',         'Upload new files to the repository',                   'documents'),
    ('documents.delete',    'Delete Documents',         'Soft-delete documents from the repository',            'documents'),
    # Calendar
    ('calendar.create',     'Create Calendar Sessions', 'Add sessions to the term calendar',                    'calendar'),
    ('calendar.edit',       'Edit Calendar',            'Update session status, notes and term name',           'calendar'),
    ('calendar.delete',     'Delete Calendar Sessions', 'Remove sessions from the term calendar',               'calendar'),
    # Users
    ('users.view',          'View Users',               'View the portal user list',                            'users'),
    ('users.create',        'Create Users',             'Create new portal staff accounts',                     'users'),
    ('users.edit',          'Edit Users',               'Edit existing portal accounts',                        'users'),
    ('users.create.admin',  'Create Admin Users',       'Assign roles that carry admin-level permissions',      'users'),
    ('users.delete',        'Delete Users',             'Permanently delete a portal user account',             'users'),
    # Admin / settings
    ('admin.settings',      'Manage Settings',          'Access and change club settings and roles',            'admin'),
    ('admin.session_types', 'Manage Session Types',     'Create, edit and reorder session types',               'admin'),
    ('admin.maintenance',   'Maintenance Tools',        'Clear audit logs, attendance and registration data',   'admin'),
    # Audit
    ('audit.view',          'View Audit Log',           'View the full system audit log',                       'audit'),
    # Communications
    ('mailshots.send',      'Send Mailshots',           'Send bulk emails to member contacts',                  'communications'),
    ('mailshots.templates', 'Manage Email Templates',   'Create, edit and delete email templates',              'communications'),
    # Notifications (v8.2)
    ('notifications.view',   'View Notifications',       'View personal and system notifications',               'communications'),
    ('notifications.send',   'Send Notifications',       'Send targeted notifications to users, roles or sessions', 'communications'),
    ('notifications.manage', 'Manage Notifications',     'Delete old notifications (admin only)',                'admin'),
    # Display board
    ('activities.manage',   'Manage Activities Board',  'Add and remove activities from the TV display',        'display'),
]

# Default permission sets — exact match to old hard-coded behaviour.
# These seed the roles table on first run; admins can customise from /admin/roles.
DEFAULT_ROLE_PERMISSIONS = {
    'admin': [
        'members.view', 'members.edit', 'members.delete', 'members.hard_delete', 'members.tags',
        'register.signin', 'register.signout', 'register.complete', 'register.reset',
        'register.print', 'register.notes', 'register.export', 'register.qr_manage',
        'approvals.view', 'approvals.approve', 'approvals.reject',
        'documents.view', 'documents.upload', 'documents.delete',
        'calendar.create', 'calendar.edit', 'calendar.delete',
        'users.view', 'users.create', 'users.edit', 'users.create.admin', 'users.delete',
        'admin.settings', 'admin.session_types', 'admin.maintenance',
        'audit.view',
        'mailshots.send', 'mailshots.templates',
        'activities.manage',
        'alerts.view', 'alerts.manage', 'alerts.run', 'alerts.dismiss',
        'notifications.view', 'notifications.send', 'notifications.manage',
    ],
    'editor': [
        'members.view', 'members.edit', 'members.delete', 'members.tags',
        'register.signin', 'register.signout', 'register.complete', 'register.print',
        'register.notes', 'register.export', 'register.qr_manage',
        'approvals.view', 'approvals.approve', 'approvals.reject',
        'documents.view', 'documents.upload', 'documents.delete',
        'calendar.create', 'calendar.edit', 'calendar.delete',
        'users.view', 'users.create', 'users.edit',
        'admin.settings',
        'audit.view',
        'mailshots.send', 'mailshots.templates',
        'activities.manage',
        'alerts.view', 'alerts.manage', 'alerts.run', 'alerts.dismiss',
        'notifications.view', 'notifications.send',
    ],
    'leader': [
        'members.view',
        'register.signin', 'register.signout', 'register.notes',
        'activities.manage',
        'alerts.view',
        'notifications.view',
    ],
    'readonly': [
        'register.signout',
        'documents.view',
        'activities.manage',
        'notifications.view',
    ],
}

# Human-readable labels for the four built-in role slugs.
# Stored as display_name in the roles table; shown everywhere in the UI.
#   readonly → Read Only   |   leader → User   |   editor → Editor   |   admin → Admin
ROLE_DISPLAY_NAMES = {
    'admin':    'Admin',
    'editor':   'Editor',
    'leader':   'User',
    'readonly': 'Read Only',
}

# Role slug constants — use these instead of bare strings so a typo fails loudly.
ROLE_ADMIN    = 'admin'
ROLE_EDITOR   = 'editor'
ROLE_LEADER   = 'leader'
ROLE_READONLY = 'readonly'

# ── Postcode lookup (getaddress.io) ──────────────────────────────────────────
GETADDRESS_KEY = os.environ.get('GETADDRESS_KEY', '')

# ── SMTP config (set in .env) ─────────────────────────────────────────────────
SMTP_HOST = os.environ.get('MAIL_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('MAIL_PORT', 587))
SMTP_USER = os.environ.get('MAIL_USERNAME', '')
SMTP_PASS = os.environ.get('MAIL_PASSWORD', '')
SMTP_FROM = os.environ.get('MAIL_FROM', SMTP_USER)

# ── Club identity (multi-tenant) ──────────────────────────────────────────────
CLUB_NAME       = os.environ.get('CLUB_NAME',       'Ashford Youth Club')
CLUB_SHORT_NAME = os.environ.get('CLUB_SHORT_NAME', 'AYC')

# ── Database helpers ───────────────────────────────────────────────────────────

def _validate_encryption_key(key: str) -> None:
    """Guard against SQLCipher PRAGMA key injection.

    SQLCipher's PRAGMA key cannot be fully parameterized — the key is
    interpolated into the statement as a string literal surrounded by single
    quotes.  The only character that could break out of that literal is a
    single quote itself, so we reject keys that contain one.

    Any other passphrase (including existing alphanumeric or mixed-character
    keys) is accepted unchanged.
    """
    if not key:
        raise RuntimeError(
            'DB_ENCRYPTION_KEY is not set in .env — refusing to start. '
            'Add the key to .env and restart the portal.'
        )
    if "'" in key:
        raise RuntimeError(
            "DB_ENCRYPTION_KEY must not contain a single-quote character (') "
            "as this would break the SQLCipher PRAGMA key statement. "
            'Generate a safe key with: python -c "import secrets; print(secrets.token_hex(32))"'
        )


def _connect_db(path=None):
    """Open a SQLCipher-encrypted DB connection.

    Raises RuntimeError on startup if DB_ENCRYPTION_KEY is missing or invalid —
    the app must never run without the key once the database is encrypted.
    Verifies the key immediately so a wrong key fails fast and clearly.
    """
    if path is None:
        path = DATABASE
    key = os.environ.get('DB_ENCRYPTION_KEY', '')
    _validate_encryption_key(key)
    conn = sqlite3.connect(path)
    conn.execute(f"PRAGMA key='{key}'")
    conn.execute('SELECT count(*) FROM sqlite_master')  # verify key immediately
    return conn


# ── Document encryption helpers ────────────────────────────────────────────────

def _doc_fernet():
    """Return a Fernet instance derived from DB_ENCRYPTION_KEY.

    We derive a 32-byte key by SHA-256 hashing the DB_ENCRYPTION_KEY so that
    a single key in .env covers both database and document encryption.
    Raises RuntimeError if the key is missing — same hard-fail policy as the DB.
    """
    raw_key = os.environ.get('DB_ENCRYPTION_KEY')
    if not raw_key:
        raise RuntimeError(
            'DB_ENCRYPTION_KEY is not set in .env — cannot encrypt/decrypt documents.'
        )
    derived = hashlib.sha256(raw_key.encode()).digest()  # always 32 bytes
    fernet_key = base64.urlsafe_b64encode(derived)
    return Fernet(fernet_key)


def encrypt_file(data: bytes) -> bytes:
    """Encrypt raw file bytes. Returns Fernet token (bytes)."""
    return _doc_fernet().encrypt(data)


def decrypt_file(token: bytes) -> bytes:
    """Decrypt a Fernet token back to the original file bytes."""
    return _doc_fernet().decrypt(token)


def get_db():
    """Return a request-scoped DB connection with SQLCipher encryption."""
    if 'db' not in g:
        g.db = _connect_db()
        g.db.row_factory = sqlite3.Row
        g.db.execute('PRAGMA foreign_keys = ON')
    return g.db

@app.teardown_appcontext
def close_db(error):
    db = g.pop('db', None)
    if db is not None:
        db.close()

@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    """Return a clean JSON 400 when a CSRF token is missing or invalid."""
    return jsonify({'error': 'CSRF token missing or invalid — please refresh the page and try again.'}), 400

def log_action(action, table_name=None, record_id=None, details=None):
    """Write an entry to the audit log. Never raises — logging must not break the app.
    If the DB write fails, the failure is recorded to the application error log so
    that gaps in the audit trail are detectable even when the database is unavailable.
    """
    try:
        db = get_db()
        db.execute(
            'INSERT INTO audit_log (user_id, action, table_name, record_id, details, ip_address)'
            ' VALUES (?,?,?,?,?,?)',
            (
                session.get('user_id'),
                action,
                table_name,
                record_id,
                json.dumps(details) if details else None,
                request.remote_addr,
            )
        )
        db.commit()
    except Exception as _audit_exc:
        # Do not re-raise — audit logging must never break the application.
        # Log to stderr so the failure is visible in server logs / syslog.
        import sys
        print(
            f'AUDIT LOG FAILED — action={action} table={table_name} record={record_id}: {_audit_exc}',
            file=sys.stderr,
        )

def send_notification(sender_id, title, body, notification_type='Info',
                      target_type='all', target_value=None, is_system=0,
                      related_table=None, related_id=None, _db=None):
    """Create a notification record.

    Works inside *and* outside a Flask request context (e.g. from the background
    scheduler).  Pass _db to reuse an existing connection — the caller is then
    responsible for committing.  When _db is omitted a new connection is opened,
    committed, and closed automatically.
    """
    own_conn = _db is None
    db = _db if _db is not None else _connect_db()
    try:
        db.execute(
            'INSERT INTO notifications '
            '(sender_id, title, body, notification_type, target_type, target_value, '
            ' is_system, related_table, related_id) '
            'VALUES (?,?,?,?,?,?,?,?,?)',
            (
                sender_id, title, body, notification_type, target_type,
                json.dumps(target_value) if isinstance(target_value, (list, dict)) else target_value,
                is_system, related_table, related_id,
            )
        )
        if own_conn:
            db.commit()
    finally:
        if own_conn:
            db.close()
    # log_action uses Flask context — it self-suppresses outside a request, so safe to call here
    log_action('notification.sent', 'notifications', None,
               {'title': title, 'type': notification_type, 'target': target_type, 'is_system': is_system})


def init_db():
    """Initialise the database from schema.sql. Safe to run multiple times."""
    os.makedirs(os.path.dirname(DATABASE), exist_ok=True)
    db = _connect_db()
    with open(os.path.join(BASE_DIR, 'schema.sql'), 'r') as f:
        db.executescript(f.read())
    db.commit()
    db.close()
    print(f'Database initialised at {DATABASE}')

@app.cli.command('init-db')
def init_db_command():
    """Flask CLI: flask init-db"""
    init_db()

def ensure_tables():
    """Create any tables added after initial deploy without requiring a full init-db.
    Safe to run on every startup — all operations are idempotent."""
    db = _connect_db()

    # ── Tables ────────────────────────────────────────────────────────────────
    db.executescript('''
        CREATE TABLE IF NOT EXISTS session_activities (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            session_type TEXT    NOT NULL,
            activity     TEXT    NOT NULL,
            added_by     INTEGER REFERENCES users(id),
            created_at   TEXT    DEFAULT (datetime('now')),
            active       INTEGER NOT NULL DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS term_sessions (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            session_date TEXT    NOT NULL,
            session_type TEXT    NOT NULL,
            term_name    TEXT,
            status       TEXT    NOT NULL DEFAULT 'planned',
            notes        TEXT,
            created_by   INTEGER REFERENCES users(id),
            created_at   TEXT    DEFAULT (datetime('now')),
            UNIQUE(session_date, session_type)
        );
        CREATE INDEX IF NOT EXISTS idx_term_sessions_date ON term_sessions(session_date);
        CREATE INDEX IF NOT EXISTS idx_audit_timestamp    ON audit_log(timestamp);
        CREATE TABLE IF NOT EXISTS session_completions (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            session_date       TEXT    NOT NULL,
            session_type       TEXT    NOT NULL,
            completed_by       INTEGER REFERENCES users(id),
            completed_at       TEXT    DEFAULT (datetime('now')),
            auto_signout_count INTEGER DEFAULT 0,
            UNIQUE(session_date, session_type)
        );
        CREATE TABLE IF NOT EXISTS session_types (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT    NOT NULL UNIQUE,
            weekday    INTEGER NOT NULL,
            active     INTEGER NOT NULL DEFAULT 1,
            sort_order INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS settings (
            key        TEXT PRIMARY KEY,
            value      TEXT NOT NULL,
            updated_at TEXT,
            updated_by INTEGER REFERENCES users(id)
        );
        -- v6.0: permissions catalogue and configurable roles
        CREATE TABLE IF NOT EXISTS permissions (
            code        TEXT PRIMARY KEY,
            name        TEXT NOT NULL,
            description TEXT,
            category    TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS roles (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT    UNIQUE NOT NULL,
            is_default  INTEGER DEFAULT 0,
            permissions TEXT    NOT NULL,
            created_at  TEXT    DEFAULT (datetime('now'))
        );
        -- v7.1: member skills & badge tags
        CREATE TABLE IF NOT EXISTS tag_definitions (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT    NOT NULL UNIQUE,
            category   TEXT    NOT NULL DEFAULT 'General',
            icon       TEXT    DEFAULT '🏷',
            colour     TEXT    NOT NULL DEFAULT '#3b82f6',
            active     INTEGER NOT NULL DEFAULT 1,
            sort_order INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS member_tags (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            member_id  INTEGER NOT NULL REFERENCES members(id) ON DELETE CASCADE,
            tag_id     INTEGER NOT NULL REFERENCES tag_definitions(id) ON DELETE CASCADE,
            expires_at TEXT,
            notes      TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            UNIQUE(member_id, tag_id)
        );
        CREATE INDEX IF NOT EXISTS idx_member_tags_member ON member_tags(member_id);
        -- v7.1: session notes & incidents
        CREATE TABLE IF NOT EXISTS session_notes (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            session_date TEXT    NOT NULL,
            session_type TEXT    NOT NULL,
            member_id    INTEGER REFERENCES members(id),
            note_type    TEXT    NOT NULL DEFAULT 'General',
            title        TEXT,
            details      TEXT,
            added_by     INTEGER REFERENCES users(id),
            created_at   TEXT    DEFAULT (datetime('now'))
        );
        CREATE INDEX IF NOT EXISTS idx_session_notes_date_type ON session_notes(session_date, session_type);
        CREATE INDEX IF NOT EXISTS idx_session_notes_member    ON session_notes(member_id);
        -- v8.0: configurable member field system
        CREATE TABLE IF NOT EXISTS member_types (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            name                TEXT    NOT NULL UNIQUE,
            slug                TEXT    NOT NULL UNIQUE,
            icon                TEXT    NOT NULL DEFAULT '👤',
            colour              TEXT    NOT NULL DEFAULT '#1b2d4f',
            description         TEXT,
            public_registration INTEGER NOT NULL DEFAULT 0,
            active              INTEGER NOT NULL DEFAULT 1,
            sort_order          INTEGER NOT NULL DEFAULT 0,
            created_at          TEXT    DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS field_definitions (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            key          TEXT    NOT NULL UNIQUE,
            label        TEXT    NOT NULL,
            field_type   TEXT    NOT NULL DEFAULT 'text',
            options      TEXT,
            help_text    TEXT,
            placeholder  TEXT,
            system_field INTEGER NOT NULL DEFAULT 0,
            column_name  TEXT,
            active       INTEGER NOT NULL DEFAULT 1,
            sort_order   INTEGER NOT NULL DEFAULT 0,
            created_at   TEXT    DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS member_type_fields (
            id                   INTEGER PRIMARY KEY AUTOINCREMENT,
            member_type_id       INTEGER NOT NULL REFERENCES member_types(id) ON DELETE CASCADE,
            field_id             INTEGER NOT NULL REFERENCES field_definitions(id) ON DELETE CASCADE,
            sort_order           INTEGER NOT NULL DEFAULT 0,
            required             INTEGER NOT NULL DEFAULT 0,
            show_on_registration INTEGER NOT NULL DEFAULT 1,
            show_on_list         INTEGER NOT NULL DEFAULT 0,
            show_on_card         INTEGER NOT NULL DEFAULT 0,
            show_on_detail       INTEGER NOT NULL DEFAULT 1,
            show_on_print        INTEGER NOT NULL DEFAULT 1,
            show_on_export       INTEGER NOT NULL DEFAULT 0,
            UNIQUE(member_type_id, field_id)
        );
        CREATE TABLE IF NOT EXISTS member_field_values (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            member_id  INTEGER NOT NULL REFERENCES members(id) ON DELETE CASCADE,
            field_id   INTEGER NOT NULL REFERENCES field_definitions(id) ON DELETE CASCADE,
            value      TEXT,
            updated_at TEXT    DEFAULT (datetime('now')),
            UNIQUE(member_id, field_id)
        );
        CREATE INDEX IF NOT EXISTS idx_mfv_member ON member_field_values(member_id);
        CREATE INDEX IF NOT EXISTS idx_mtf_type   ON member_type_fields(member_type_id);
        -- v8.0: Member Alert Rules — configurable multi-rule flag engine
        CREATE TABLE IF NOT EXISTS alert_rules (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            name                TEXT    NOT NULL,
            rule_type           TEXT    NOT NULL,
            target_field        TEXT,
            condition           TEXT,
            threshold_value     INTEGER,
            threshold_unit      TEXT,
            applies_to_session  TEXT,
            flag_label          TEXT    NOT NULL,
            flag_colour         TEXT    NOT NULL DEFAULT '#f59e0b',
            auto_resolve        INTEGER NOT NULL DEFAULT 1,
            resolve_field       TEXT,
            is_active           INTEGER NOT NULL DEFAULT 1,
            created_at          TEXT    DEFAULT (datetime('now')),
            created_by          INTEGER REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS member_flags (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            member_id    INTEGER NOT NULL REFERENCES members(id) ON DELETE CASCADE,
            rule_id      INTEGER NOT NULL REFERENCES alert_rules(id) ON DELETE CASCADE,
            flagged_at   TEXT    NOT NULL DEFAULT (datetime('now')),
            flagged_by   TEXT    NOT NULL DEFAULT 'auto',
            resolved_at  TEXT,
            resolved_by  TEXT,
            note         TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_member_flags_member ON member_flags(member_id);
        CREATE INDEX IF NOT EXISTS idx_member_flags_rule   ON member_flags(rule_id);
        -- v8.2: Notifications system
        CREATE TABLE IF NOT EXISTS notifications (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            sender_id         INTEGER REFERENCES users(id),
            title             TEXT    NOT NULL,
            body              TEXT    NOT NULL,
            notification_type TEXT    NOT NULL DEFAULT 'Info',
            target_type       TEXT    NOT NULL DEFAULT 'all',
            target_value      TEXT,
            is_system         INTEGER NOT NULL DEFAULT 0,
            related_table     TEXT,
            related_id        INTEGER,
            created_at        TEXT    DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS notification_reads (
            notification_id   INTEGER NOT NULL REFERENCES notifications(id) ON DELETE CASCADE,
            user_id           INTEGER NOT NULL REFERENCES users(id),
            read_at           TEXT    DEFAULT (datetime('now')),
            PRIMARY KEY (notification_id, user_id)
        );
        CREATE INDEX IF NOT EXISTS idx_notifications_created ON notifications(created_at DESC);
        CREATE INDEX IF NOT EXISTS idx_notifications_sender  ON notifications(sender_id);
        CREATE INDEX IF NOT EXISTS idx_notification_reads    ON notification_reads(user_id);
        -- v8.3: QR quick-session tokens (sign-in / sign-out via mobile)
        CREATE TABLE IF NOT EXISTS quick_signin_tokens (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            token          TEXT    NOT NULL UNIQUE,
            session_type   TEXT    NOT NULL,
            session_date   TEXT    NOT NULL,
            created_at     TEXT    NOT NULL DEFAULT (datetime('now')),
            invalidated_at TEXT    NULL
        );
        CREATE INDEX IF NOT EXISTS idx_qst_token   ON quick_signin_tokens(token);
        CREATE INDEX IF NOT EXISTS idx_qst_session ON quick_signin_tokens(session_type, session_date);
    ''')

    # ── ALTER TABLE migrations (idempotent — each wrapped individually) ───────
    alter_stmts = [
        "ALTER TABLE members ADD COLUMN member_type TEXT NOT NULL DEFAULT 'member'",
        "ALTER TABLE members ADD COLUMN staff_role TEXT",
        "ALTER TABLE pending_registrations ADD COLUMN registration_type TEXT NOT NULL DEFAULT 'member'",
        "ALTER TABLE pending_registrations ADD COLUMN applicant_role TEXT",
        "ALTER TABLE pending_registrations ADD COLUMN mobile TEXT",
        "ALTER TABLE pending_registrations ADD COLUMN email TEXT",
        "ALTER TABLE members ADD COLUMN status_note TEXT",
        # v8.0: configurable member field system — registration support
        "ALTER TABLE member_types ADD COLUMN registration_style TEXT NOT NULL DEFAULT 'member'",
        "ALTER TABLE pending_registrations ADD COLUMN custom_fields TEXT",
        "ALTER TABLE pending_registrations ADD COLUMN member_type_slug TEXT NOT NULL DEFAULT 'member'",
        # v6.0: link users to the new roles table
        "ALTER TABLE users ADD COLUMN role_id INTEGER REFERENCES roles(id)",
        # v7.1: role display names
        "ALTER TABLE roles ADD COLUMN display_name TEXT",
        # v7.1: track whether a completed register has been exported
        "ALTER TABLE session_completions ADD COLUMN exported_at TEXT",
        "ALTER TABLE session_completions ADD COLUMN exported_by INTEGER REFERENCES users(id)",
        # v8.1: configurable export columns
        "ALTER TABLE member_type_fields ADD COLUMN show_on_export INTEGER NOT NULL DEFAULT 0",
        # v8.3: track how each attendance record was created ('web' | 'qr-self')
        "ALTER TABLE attendance ADD COLUMN source TEXT NOT NULL DEFAULT 'web'",
    ]
    for stmt in alter_stmts:
        try:
            db.execute(stmt)
        except Exception:
            pass  # Column already exists — safe to ignore

    # v8.3: unique guard on attendance — one row per member per session date+type.
    # De-duplicate first (live DBs may have historical duplicates) then create index.
    try:
        db.execute('''
            DELETE FROM attendance
            WHERE rowid NOT IN (
                SELECT MIN(rowid)
                FROM attendance
                GROUP BY member_id, session_date, session_type
            )
        ''')
        db.execute('''
            CREATE UNIQUE INDEX IF NOT EXISTS idx_attendance_unique
                ON attendance(member_id, session_date, session_type)
        ''')
    except Exception:
        pass  # Index already exists — safe to ignore

    db.commit()
    db.close()

    # ── Seed default settings ─────────────────────────────────────────────────
    sdb = _connect_db()
    sdb.row_factory = sqlite3.Row
    # Seed alerts_last_run timestamp placeholder (updated by the scheduler on each run)
    if not sdb.execute("SELECT key FROM settings WHERE key = 'alerts_last_run'").fetchone():
        sdb.execute("INSERT INTO settings (key, value) VALUES ('alerts_last_run', '')")
    # v8.3: QR quick-session settings
    _qr_defaults = [
        ('quick_signin_enabled',    'true'),
        ('quick_signout_enabled',   'false'),
        ('quick_signin_welcome_msg',    'Welcome, {name}! Great to see you tonight! 🎉'),
        ('quick_signin_already_msg',    "You're already signed in, {name}! See you inside 👋"),
        ('quick_signout_goodbye_msg',   'Goodbye, {name}! See you next time 👋'),
        ('quick_signout_already_msg',   "You're already signed out, {name}. Safe journey home!"),
    ]
    for key, val in _qr_defaults:
        if not sdb.execute('SELECT key FROM settings WHERE key = ?', (key,)).fetchone():
            sdb.execute('INSERT INTO settings (key, value) VALUES (?, ?)', (key, val))
    sdb.commit()
    sdb.close()

    # ── Seed default session types ─────────────────────────────────────────────
    tdb = _connect_db()
    for sort_order, (name, weekday) in enumerate([('Tuesday', 1), ('Thursday', 3)]):
        if not tdb.execute('SELECT id FROM session_types WHERE name = ?', (name,)).fetchone():
            tdb.execute(
                'INSERT INTO session_types (name, weekday, active, sort_order) VALUES (?,?,1,?)',
                (name, weekday, sort_order),
            )
    tdb.commit()
    tdb.close()

    # ── Seed permissions catalogue (v6.0) ──────────────────────────────────────
    pdb = _connect_db()
    for code, name, desc, cat in ALL_PERMISSIONS:
        pdb.execute(
            'INSERT OR IGNORE INTO permissions (code, name, description, category) VALUES (?,?,?,?)',
            (code, name, desc, cat),
        )
    pdb.commit()
    pdb.close()

    # ── Seed default roles (v6.0) + display names (v7.1) ──────────────────────
    rdb = _connect_db()
    rdb.row_factory = sqlite3.Row
    for role_name, perms in DEFAULT_ROLE_PERMISSIONS.items():
        display = ROLE_DISPLAY_NAMES.get(role_name, role_name)
        rdb.execute(
            'INSERT OR IGNORE INTO roles (name, permissions, is_default, display_name) VALUES (?,?,1,?)',
            (role_name, json.dumps(perms), display),
        )
        # Migrate existing rows that pre-date the display_name column
        rdb.execute(
            'UPDATE roles SET display_name = ? WHERE name = ? AND (display_name IS NULL OR display_name = "")',
            (display, role_name),
        )
        # Merge any newly added permissions into existing role records so that
        # deploying new code automatically grants new default permissions without
        # requiring a full DB re-initialisation.
        #
        # ⚠️  INTENTIONAL BEHAVIOUR: this is a union — it only ever adds permissions,
        # never removes them.  If a permission appears in DEFAULT_ROLE_PERMISSIONS but
        # an admin has manually removed it from a role via the Roles UI, this merge
        # will silently restore it on the next app restart.  That is the trade-off
        # chosen to ensure new permissions propagate automatically on deploy.
        # Admins who need to permanently remove a default permission should be aware
        # of this and can re-remove it after each deployment if required.
        existing_row = rdb.execute(
            'SELECT permissions FROM roles WHERE name = ?', (role_name,)
        ).fetchone()
        if existing_row:
            try:
                existing_perms = set(json.loads(existing_row['permissions'] or '[]'))
            except (ValueError, TypeError):
                existing_perms = set()
            new_perms = set(perms)
            if not new_perms.issubset(existing_perms):
                merged = sorted(existing_perms | new_perms, key=lambda p: perms.index(p) if p in perms else 999)
                rdb.execute(
                    'UPDATE roles SET permissions = ? WHERE name = ?',
                    (json.dumps(merged), role_name),
                )
    rdb.commit()

    # ── Migrate existing users → role_id (v6.0) ───────────────────────────────
    # Any user whose role_id is still NULL gets it set from their old role text column.
    users_needing_migration = rdb.execute(
        'SELECT id, role FROM users WHERE role_id IS NULL'
    ).fetchall()
    for user in users_needing_migration:
        role_row = rdb.execute(
            'SELECT id FROM roles WHERE name = ?', (user['role'],)
        ).fetchone()
        if role_row:
            rdb.execute(
                'UPDATE users SET role_id = ? WHERE id = ?',
                (role_row['id'], user['id']),
            )
    rdb.commit()
    rdb.close()

    # ── Seed member types (v8.0) ───────────────────────────────────────────────
    mtdb = _connect_db()
    mtdb.row_factory = sqlite3.Row
    for slug, name, icon, colour, description, public_registration, sort_order in [
        ('member', 'Member', '👦', '#1b2d4f', 'Young people attending club sessions', 1, 0),
        ('staff',  'Staff',  '🧑', '#0f766e', 'Leaders, coaches and volunteers',      0, 1),
    ]:
        mtdb.execute(
            '''INSERT OR IGNORE INTO member_types
               (slug, name, icon, colour, description, public_registration, sort_order)
               VALUES (?,?,?,?,?,?,?)''',
            (slug, name, icon, colour, description, public_registration, sort_order),
        )
    # Ensure registration_style is set correctly for built-in types
    mtdb.execute(
        "UPDATE member_types SET registration_style = 'staff' "
        "WHERE slug = 'staff' AND (registration_style IS NULL OR registration_style = 'member')"
    )
    mtdb.commit()

    # ── Seed system field definitions (v8.0) ───────────────────────────────────
    for key, label, field_type, column_name, placeholder, help_text, sort_order in [
        ('first_name',            'First Name',                      'text',      'first_name',         'e.g. Isabella',                             None,                                                                      1),
        ('surname',               'Surname',                         'text',      'surname',            'e.g. Fitzpatrick',                          None,                                                                      2),
        ('date_of_birth',         'Date of Birth',                   'date',      'date_of_birth',      None,                                        None,                                                                      3),
        ('address',               'Home Address',                    'text',      'address',            'Start typing or use the postcode finder',    None,                                                                      4),
        ('postcode',              'Postcode',                        'postcode',  'postcode',           'e.g. TW15 3EL',                             None,                                                                      5),
        ('ethnicity_religion',    'Ethnicity / Religion',            'text',      'ethnicity_religion', 'e.g. English / Christian',                  None,                                                                      6),
        ('medical_sen',           'Medical Needs, Allergies or SEN', 'textarea',  'medical_sen',        None,                                        'Describe any medical conditions, allergies or special educational needs.', 7),
        ('gp_contact',            'GP / Doctor Surgery Contact',     'text',      'gp_contact',         'e.g. Stanwell Road Surgery — 01784 123456', None,                                                                      8),
        ('unattended_exit',       'Unattended Exit',                 'boolean',   'unattended_exit',    None,                                        'Will make their own way home unaccompanied at the end of the session.',    9),
        ('gdpr_consent',          'Communications Consent',          'boolean',   'gdpr_consent',       None,                                        'Happy to be contacted about upcoming events and club information.',        10),
        ('session',               'Session',                         'text',      'session',            None,                                        'Which session this person attends.',                                       11),
        ('staff_role',            'Staff Role',                      'text',      'staff_role',         None,                                        None,                                                                      12),
        ('comments',              'Internal Notes',                  'textarea',  'comments',           None,                                        'Internal notes — not visible to members or parents.',                      13),
        # Contact fields — previously hardcoded in the registration form
        ('contact1_name',         'Primary Contact — Full Name',     'text',      'contact1_name',      'e.g. Charlotte Day',                        None,                                                                      30),
        ('contact1_phone',        'Primary Contact — Phone',         'phone',     'contact1_phone',     'e.g. 07590 118098',                         None,                                                                      31),
        ('contact1_email',        'Primary Contact — Email',         'email',     'contact1_email',     'e.g. charlotte.day@email.com',              None,                                                                      32),
        ('contact2_name',         'Second Contact — Full Name',      'text',      'contact2_name',      'e.g. James Day',                            None,                                                                      33),
        ('contact2_phone',        'Second Contact — Phone',          'phone',     'contact2_phone',     'e.g. 07700 900000',                         None,                                                                      34),
        ('contact2_email',        'Second Contact — Email',          'email',     'contact2_email',     'e.g. james.day@email.com',                  None,                                                                      35),
        ('mobile',                'Mobile Number',                   'phone',     'mobile',             'e.g. 07700 900000',                         None,                                                                      36),
        ('email',                 'Email Address',                   'email',     'email',              'e.g. you@email.com',                        None,                                                                      37),
        # Signature field — renders guardian confirmation block on registration form
        ('guardian_confirmation', 'Guardian Confirmation',           'signature', None,                 None,                                        'Parent or guardian types their full name to confirm the registration.',   38),
    ]:
        mtdb.execute(
            '''INSERT OR IGNORE INTO field_definitions
               (key, label, field_type, column_name, placeholder, help_text, sort_order, system_field)
               VALUES (?,?,?,?,?,?,?,1)''',
            (key, label, field_type, column_name, placeholder, help_text, sort_order),
        )

    # ── Seed declaration field definitions (v8.1) ──────────────────────────────
    # These are non-system (system_field=0) — they render as Yes/No consent rows
    # on member registration forms.  Use {club} in the label as a placeholder for
    # the club name; the registration template substitutes it at render time.
    for key, label, sort_order in [
        ('consent_attend',
         'I am the parent / guardian of the young person named above and I give '
         'consent for them to attend activities organised by {club}.',
         20),
        ('consent_photos',
         'I agree that photos and videos can be taken of my child to publicise '
         'the group\'s activities.',
         21),
        ('consent_comms',
         'I am happy to be emailed or texted with up-and-coming events or '
         'important information regarding {club}.',
         22),
        ('consent_belongings',
         'I understand that {club} is NOT responsible for personal belongings.',
         23),
        ('consent_medical',
         'I give consent for my child to be taken for medical treatment in the '
         'event of an emergency.',
         24),
    ]:
        mtdb.execute(
            '''INSERT OR IGNORE INTO field_definitions
               (key, label, field_type, column_name, placeholder, help_text, sort_order, system_field)
               VALUES (?,?,'declaration',NULL,NULL,NULL,?,0)''',
            (key, label, sort_order),
        )
    mtdb.commit()

    # ── Seed default member_type_fields (v8.0) ─────────────────────────────────
    _default_fields = {
        'member': [
            # key,                   req, reg, list, card, detail, print, sort
            ('first_name',           1,   1,   1,    1,    1,      1,     1),
            ('surname',              1,   1,   1,    1,    1,      1,     2),
            ('date_of_birth',        1,   1,   0,    0,    1,      1,     3),
            ('address',              1,   1,   0,    0,    1,      1,     4),
            ('postcode',             1,   1,   0,    0,    1,      0,     5),
            ('ethnicity_religion',   0,   1,   0,    0,    1,      0,     6),
            ('medical_sen',          0,   1,   0,    0,    1,      1,     7),
            ('gp_contact',           1,   1,   0,    0,    1,      1,     8),
            ('unattended_exit',      0,   1,   0,    1,    1,      1,     9),
            ('gdpr_consent',         0,   1,   0,    0,    1,      0,     10),
            ('session',              1,   0,   0,    0,    1,      1,     11),
            ('comments',             0,   0,   0,    0,    1,      0,     12),
            # Declaration fields — shown only on registration form
            ('consent_attend',       1,   1,   0,    0,    0,      0,     20),
            ('consent_photos',       1,   1,   0,    0,    0,      0,     21),
            ('consent_comms',        1,   1,   0,    0,    0,      0,     22),
            ('consent_belongings',   1,   1,   0,    0,    0,      0,     23),
            ('consent_medical',      1,   1,   0,    0,    0,      0,     24),
            # Contact fields — now Field Builder-driven
            ('contact1_name',        1,   1,   0,    0,    1,      1,     30),
            ('contact1_phone',       1,   1,   0,    0,    1,      1,     31),
            ('contact1_email',       1,   1,   0,    0,    1,      1,     32),
            ('contact2_name',        0,   1,   0,    0,    1,      1,     33),
            ('contact2_phone',       0,   1,   0,    0,    1,      0,     34),
            ('contact2_email',       0,   1,   0,    0,    1,      0,     35),
            ('guardian_confirmation',0,   1,   0,    0,    0,      0,     38),
        ],
        'staff': [
            # key,          req, reg, list, card, detail, print, sort
            ('first_name',  1,   1,   1,    1,    1,      1,     1),
            ('surname',     1,   1,   1,    1,    1,      1,     2),
            ('mobile',      1,   1,   0,    0,    1,      0,     3),
            ('email',       1,   1,   0,    0,    1,      0,     4),
            ('staff_role',  0,   1,   0,    1,    1,      1,     5),
            ('session',     1,   0,   0,    0,    1,      1,     6),
            ('comments',    0,   0,   0,    0,    1,      0,     7),
        ],
    }
    for type_slug, field_rows in _default_fields.items():
        mt = mtdb.execute('SELECT id FROM member_types WHERE slug = ?', (type_slug,)).fetchone()
        if not mt:
            continue
        mt_id = mt['id']
        # Only seed if this type has no fields assigned yet
        existing_count = mtdb.execute(
            'SELECT COUNT(*) FROM member_type_fields WHERE member_type_id = ?', (mt_id,)
        ).fetchone()[0]
        if existing_count > 0:
            continue
        for field_key, required, show_on_reg, show_on_list, show_on_card, show_on_detail, show_on_print, sort_order in field_rows:
            fd = mtdb.execute('SELECT id FROM field_definitions WHERE key = ?', (field_key,)).fetchone()
            if not fd:
                continue
            mtdb.execute(
                '''INSERT OR IGNORE INTO member_type_fields
                   (member_type_id, field_id, required, show_on_registration, show_on_list,
                    show_on_card, show_on_detail, show_on_print, show_on_export, sort_order)
                   VALUES (?,?,?,?,?,?,?,?,0,?)''',
                (mt_id, fd['id'], required, show_on_reg, show_on_list,
                 show_on_card, show_on_detail, show_on_print, sort_order),
            )
    mtdb.commit()

    # ── Migrate existing deployments: assign declaration fields to member type ──
    # Uses INSERT OR IGNORE so it is safe to run on every startup.
    _mt = mtdb.execute('SELECT id FROM member_types WHERE slug = ?', ('member',)).fetchone()
    if _mt:
        for _fkey, _sort in [
            ('consent_attend',     20),
            ('consent_photos',     21),
            ('consent_comms',      22),
            ('consent_belongings', 23),
            ('consent_medical',    24),
        ]:
            _fd = mtdb.execute(
                'SELECT id FROM field_definitions WHERE key = ?', (_fkey,)
            ).fetchone()
            if _fd:
                mtdb.execute(
                    '''INSERT OR IGNORE INTO member_type_fields
                       (member_type_id, field_id, required, show_on_registration,
                        show_on_list, show_on_card, show_on_detail, show_on_print,
                        show_on_export, sort_order)
                       VALUES (?,?,1,1,0,0,0,0,0,?)''',
                    (_mt['id'], _fd['id'], _sort),
                )
    mtdb.commit()

    # ── Migrate existing deployments: add contact/signature fields to member type ─
    # Previously hardcoded in the registration form; now Field Builder-driven.
    # Uses INSERT OR IGNORE so safe to run on every startup.
    _mt = mtdb.execute('SELECT id FROM member_types WHERE slug = ?', ('member',)).fetchone()
    if _mt:
        for _fkey, _required, _show_print, _sort in [
            ('contact1_name',         1, 1, 30),
            ('contact1_phone',        1, 1, 31),
            ('contact1_email',        1, 1, 32),
            ('contact2_name',         0, 1, 33),
            ('contact2_phone',        0, 0, 34),
            ('contact2_email',        0, 0, 35),
            ('guardian_confirmation', 0, 0, 38),
        ]:
            _fd = mtdb.execute(
                'SELECT id FROM field_definitions WHERE key = ?', (_fkey,)
            ).fetchone()
            if _fd:
                mtdb.execute(
                    '''INSERT OR IGNORE INTO member_type_fields
                       (member_type_id, field_id, required, show_on_registration,
                        show_on_list, show_on_card, show_on_detail, show_on_print,
                        show_on_export, sort_order)
                       VALUES (?,?,?,1,0,0,1,?,0,?)''',
                    (_mt['id'], _fd['id'], _required, _show_print, _sort),
                )

    # ── Migrate existing deployments: add mobile/email to staff type ──────────────
    _st = mtdb.execute('SELECT id FROM member_types WHERE slug = ?', ('staff',)).fetchone()
    if _st:
        for _fkey, _required, _sort in [
            ('mobile', 1, 3),
            ('email',  1, 4),
        ]:
            _fd = mtdb.execute(
                'SELECT id FROM field_definitions WHERE key = ?', (_fkey,)
            ).fetchone()
            if _fd:
                mtdb.execute(
                    '''INSERT OR IGNORE INTO member_type_fields
                       (member_type_id, field_id, required, show_on_registration,
                        show_on_list, show_on_card, show_on_detail, show_on_print,
                        show_on_export, sort_order)
                       VALUES (?,?,?,1,0,0,1,0,0,?)''',
                    (_st['id'], _fd['id'], _required, _sort),
                )

    mtdb.commit()
    mtdb.close()

# Run migration on startup
with app.app_context():
    if os.path.exists(DATABASE):
        ensure_tables()

# ── Password policy ────────────────────────────────────────────────────────────

_HEX_COLOUR_RE = re.compile(r'^#[0-9a-fA-F]{6}$')


def _validate_hex_colour(value: str, default: str) -> tuple[str, str | None]:
    """Return (sanitised_colour, error_message_or_None).

    Accepts any valid 6-digit hex colour (#rrggbb, case-insensitive).
    Returns the default if the value is empty; returns an error if the value
    is non-empty but not a valid hex colour — this prevents DOM-injection via
    a crafted colour string.
    """
    val = (value or '').strip()
    if not val:
        return default, None
    if not _HEX_COLOUR_RE.match(val):
        return default, f'Invalid colour "{val}" — must be a 6-digit hex code (e.g. #3b82f6)'
    return val, None


def validate_password(password):
    """Enforce the portal password policy.
    Returns an error string if the password fails, or None if it passes.
    Policy: 8+ characters, at least one uppercase letter, one number, one special character.
    """
    if len(password) < 8:
        return 'Password must be at least 8 characters'
    if not re.search(r'[A-Z]', password):
        return 'Password must contain at least one uppercase letter'
    if not re.search(r'[0-9]', password):
        return 'Password must contain at least one number'
    if not re.search(r'[^A-Za-z0-9]', password):
        return 'Password must contain at least one special character'
    return None


# ── Auth helpers ───────────────────────────────────────────────────────────────

_SESSION_IDLE_TIMEOUT = 30 * 60  # 30 minutes of inactivity


@app.before_request
def enforce_idle_timeout():
    """Expire sessions that have been idle for more than _SESSION_IDLE_TIMEOUT seconds.

    Every authenticated request refreshes the last_activity timestamp stored
    inside the (server-signed) session cookie. Unauthenticated routes and the
    display/stream endpoint are skipped — they don't use sessions.
    """
    if 'user_id' not in session:
        return
    # Skip the SSE stream — it holds open long connections and sends no real activity
    if request.path == '/api/display/stream':
        return
    now          = time.time()
    last_active  = session.get('last_activity', now)
    if now - last_active > _SESSION_IDLE_TIMEOUT:
        session.clear()
        if request.path.startswith('/api/'):
            return jsonify({'error': 'Session expired due to inactivity. Please log in again.'}), 401
        return redirect(url_for('login_page'))
    session['last_activity'] = now


def login_required(f):
    """Redirect to login (or return 401) if user is not authenticated."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorised'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def permission_required(permission_code):
    """Restrict access to users whose role includes the given permission code."""
    def decorator(f):
        @wraps(f)
        @login_required
        def decorated(*args, **kwargs):
            if permission_code not in session.get('permissions', []):
                if request.path.startswith('/api/'):
                    return jsonify({'error': 'Forbidden'}), 403
                return redirect(url_for('dashboard_page'))
            return f(*args, **kwargs)
        return decorated
    return decorator

def has_permission(permission_code):
    """Return True if the current session user has the given permission.
    Safe to call from templates and helper functions."""
    return permission_code in session.get('permissions', [])

def tpl_ctx():
    """Inject current user info into every protected template."""
    return {
        'current_user':         session.get('username', ''),
        'current_role':         session.get('role', ''),
        'current_role_display': session.get('role_display', session.get('role', '')),
        'current_session':      session.get('session_assigned', ''),
        'app_version':          APP_VERSION,
        'session_types':        get_session_types(),        # [{id, name, weekday}, ...]
        'user_permissions':     session.get('permissions', []),  # list of permission codes
        'club_name':            CLUB_NAME,
        'club_short_name':      CLUB_SHORT_NAME,
    }

# ── Page routes ────────────────────────────────────────────────────────────────

@app.route('/')
def login_page():
    if 'user_id' in session:
        return redirect(url_for('dashboard_page'))
    return render_template('index.html', app_version=APP_VERSION,
                           club_name=CLUB_NAME, club_short_name=CLUB_SHORT_NAME)

@app.route('/dashboard')
@login_required
def dashboard_page():
    db          = get_db()
    reg_types   = db.execute(
        'SELECT slug, name, icon, colour, description, public_registration '
        'FROM member_types WHERE active = 1 ORDER BY sort_order'
    ).fetchall()
    return render_template('dashboard.html', active_page='dashboard',
                           reg_types=[dict(t) for t in reg_types],
                           **tpl_ctx())

@app.route('/members')
@permission_required('members.view')
def members_page():
    return render_template('members.html', active_page='members', **tpl_ctx())

@app.route('/approvals')
@permission_required('approvals.view')
def approvals_page():
    return render_template('approvals.html', active_page='approvals', **tpl_ctx())

@app.route('/register')
@login_required
def register_page():
    # Pre-create QR tokens for all active session types so the register
    # page JS can fetch them immediately without a separate write round-trip.
    _ensure_qr_tokens_for_today()
    return render_template('register.html', active_page='register', **tpl_ctx())


@app.route('/register/print')
@login_required
def print_register_page():
    """
    Render a printable paper register for a given session and date.
    Accessible to editors and admins only (register.print permission).
    Query params: ?session=Tuesday&date=2026-04-18&type=member
    Columns are driven by the show_on_print field configuration for the given type.
    """
    if not has_permission('register.print'):
        return 'Access denied', 403

    session_type = request.args.get('session', '').strip()
    date         = request.args.get('date', '').strip()
    type_slug    = request.args.get('type', 'member').strip() or 'member'

    if not session_type or not date:
        return 'Missing session or date parameter', 400

    # Validate session type
    valid_sessions = get_valid_session_names()
    if session_type not in valid_sessions:
        return 'Invalid session type', 400

    db = get_db()

    # Resolve member type — fall back to first active type if slug not found
    mtype = db.execute(
        'SELECT * FROM member_types WHERE slug = ? AND active = 1', (type_slug,)
    ).fetchone()
    if not mtype:
        mtype = db.execute(
            'SELECT * FROM member_types WHERE active = 1 ORDER BY sort_order LIMIT 1'
        ).fetchone()
    mtype_dict = dict(mtype) if mtype else {}

    # Fetch show_on_print fields for this type (ordered by sort_order)
    print_fields_raw = []
    if mtype:
        pf_rows = db.execute('''
            SELECT  fd.id, fd.key, fd.label, fd.field_type,
                    fd.column_name, fd.system_field
            FROM    member_type_fields mtf
            JOIN    field_definitions fd ON fd.id = mtf.field_id
            WHERE   mtf.member_type_id = ? AND fd.active = 1 AND mtf.show_on_print = 1
            ORDER   BY mtf.sort_order
        ''', (mtype['id'],)).fetchall()
        print_fields_raw = [dict(r) for r in pf_rows]

    # first_name and surname are always rendered as dedicated columns — skip from dynamic list
    SKIP_PRINT_KEYS = {'first_name', 'surname'}
    dynamic_fields  = [f for f in print_fields_raw if f['key'] not in SKIP_PRINT_KEYS]

    # Fetch all active members for this type + session, sorted alphabetically
    members_raw = db.execute('''
        SELECT  m.*
        FROM    members m
        WHERE   m.status      != "Leaver"
          AND   m.member_type = ?
          AND   m.session     = ?
        ORDER   BY m.first_name, m.surname
    ''', (type_slug, session_type)).fetchall()
    members = [dict(r) for r in members_raw]

    # Batch-fetch custom field values when any dynamic field is non-system
    has_custom = any(not f['system_field'] for f in dynamic_fields)
    if members and has_custom:
        member_ids   = [m['id'] for m in members]
        placeholders = ','.join('?' * len(member_ids))
        cfv_rows = db.execute(
            f'SELECT mfv.member_id, fd.key, mfv.value '
            f'FROM member_field_values mfv '
            f'JOIN field_definitions fd ON fd.id = mfv.field_id '
            f'WHERE mfv.member_id IN ({placeholders})',
            member_ids
        ).fetchall()
        custom_map = {}
        for cfv in cfv_rows:
            custom_map.setdefault(cfv['member_id'], {})[cfv['key']] = cfv['value']
        for m in members:
            m['custom_fields'] = custom_map.get(m['id'], {})
    else:
        for m in members:
            m['custom_fields'] = {}

    # Fetch session notes for this date + session type
    notes = db.execute('''
        SELECT  sn.id, sn.note_type, sn.title, sn.details, sn.created_at,
                u.username   AS added_by_name,
                m.first_name AS member_first, m.surname AS member_surname
        FROM    session_notes sn
        LEFT JOIN users   u ON u.id = sn.added_by
        LEFT JOIN members m ON m.id = sn.member_id
        WHERE   sn.session_date = ? AND sn.session_type = ?
        ORDER   BY sn.created_at
    ''', (date, session_type)).fetchall()

    # Format date nicely for display (YYYY-MM-DD → DD/MM/YYYY)
    try:
        from datetime import datetime as _dt
        display_date = _dt.strptime(date, '%Y-%m-%d').strftime('%d/%m/%Y')
    except ValueError:
        display_date = date

    return render_template(
        'print_register.html',
        session_type   = session_type,
        date           = date,
        display_date   = display_date,
        members        = members,
        notes          = [dict(r) for r in notes],
        dynamic_fields = dynamic_fields,
        mtype          = mtype_dict,
        club_name      = CLUB_NAME,
        club_short_name= CLUB_SHORT_NAME,
    )


@app.route('/register/export')
@login_required
def export_register_page():
    """
    Print-ready export of a completed register: full attendance with sign-in/out
    times, duration, and all session notes. Opens in a new tab for browser print/save.
    Requires register.export permission.
    """
    if not has_permission('register.export'):
        return 'Access denied', 403

    session_type = request.args.get('session', '').strip()
    date         = request.args.get('date', '').strip()

    if not session_type or not date:
        return 'Missing session or date parameter', 400

    valid_sessions = get_valid_session_names()
    if session_type not in valid_sessions:
        return 'Invalid session type', 400

    db = get_db()

    # Must be a completed register
    completion = db.execute('''
        SELECT sc.completed_at, sc.auto_signout_count,
               u.username AS completed_by_name
        FROM   session_completions sc
        LEFT JOIN users u ON u.id = sc.completed_by
        WHERE  sc.session_date = ? AND sc.session_type = ?
    ''', (date, session_type)).fetchone()
    if not completion:
        return 'This register has not been completed yet.', 400

    # ── Configurable export fields for the 'member' type ─────────────────────
    member_mt = db.execute(
        "SELECT id FROM member_types WHERE slug = 'member' AND active = 1"
    ).fetchone()
    export_fields = []
    if member_mt:
        ef_rows = db.execute('''
            SELECT  fd.key, fd.label, fd.field_type, fd.column_name, fd.system_field
            FROM    member_type_fields mtf
            JOIN    field_definitions fd ON fd.id = mtf.field_id
            WHERE   mtf.member_type_id = ? AND fd.active = 1 AND mtf.show_on_export = 1
              AND   fd.key NOT IN ('first_name', 'surname')
            ORDER   BY mtf.sort_order
        ''', (member_mt['id'],)).fetchall()
        export_fields = [dict(f) for f in ef_rows]

    # Full member attendance — everyone expected, whether they arrived or not
    members = db.execute('''
        SELECT  m.*,
                a.signed_in_at, a.signed_out_at
        FROM    members m
        LEFT JOIN attendance a
               ON  a.member_id   = m.id
               AND a.session_date = ?
               AND a.session_type = ?
        WHERE   m.status     != 'Leaver'
          AND   m.member_type = 'member'
          AND   m.session     = ?
        ORDER   BY m.surname, m.first_name
    ''', (date, session_type, session_type)).fetchall()

    # Batch-fetch custom field values for all members in this export
    member_ids = [m['id'] for m in members]
    custom_fields_map = {}
    if member_ids and export_fields:
        placeholders = ','.join('?' * len(member_ids))
        cfv_rows = db.execute(
            f'SELECT mfv.member_id, fd.key, mfv.value '
            f'FROM member_field_values mfv '
            f'JOIN field_definitions fd ON fd.id = mfv.field_id '
            f'WHERE mfv.member_id IN ({placeholders})',
            member_ids
        ).fetchall()
        for cfv in cfv_rows:
            custom_fields_map.setdefault(cfv['member_id'], {})[cfv['key']] = cfv['value']

    # Resolve export field values per member (system columns + custom fields)
    members_out = []
    for m in members:
        md = dict(m)
        extra = {}
        for f in export_fields:
            if f['system_field'] and f['column_name']:
                extra[f['key']] = md.get(f['column_name'], '') or ''
            else:
                extra[f['key']] = custom_fields_map.get(md['id'], {}).get(f['key'], '') or ''
        md['export_extra'] = extra
        members_out.append(md)
    members = members_out

    # Staff attendance
    staff = db.execute('''
        SELECT  m.first_name, m.surname, m.staff_role,
                a.signed_in_at, a.signed_out_at
        FROM    members m
        LEFT JOIN attendance a
               ON  a.member_id   = m.id
               AND a.session_date = ?
               AND a.session_type = ?
        WHERE   m.status     != 'Leaver'
          AND   m.member_type = 'staff'
          AND   m.session     = ?
        ORDER   BY m.surname, m.first_name
    ''', (date, session_type, session_type)).fetchall()

    # Session notes
    notes = db.execute('''
        SELECT  sn.note_type, sn.title, sn.details, sn.created_at,
                u.username   AS added_by_name,
                m.first_name AS member_first, m.surname AS member_surname
        FROM    session_notes sn
        LEFT JOIN users   u ON u.id = sn.added_by
        LEFT JOIN members m ON m.id = sn.member_id
        WHERE   sn.session_date = ? AND sn.session_type = ?
        ORDER   BY sn.created_at
    ''', (date, session_type)).fetchall()

    try:
        from datetime import datetime as _dt
        display_date = _dt.strptime(date, '%Y-%m-%d').strftime('%A %d %B %Y')
    except ValueError:
        display_date = date

    # Calculate attended / not arrived counts for the header
    attended    = sum(1 for m in members if m['signed_in_at'])
    not_arrived = sum(1 for m in members if not m['signed_in_at'])

    return render_template(
        'register_export.html',
        session_type    = session_type,
        date            = date,
        display_date    = display_date,
        completion      = dict(completion),
        members         = members,
        staff           = [dict(s) for s in staff],
        notes           = [dict(n) for n in notes],
        export_fields   = export_fields,
        attended        = attended,
        not_arrived     = not_arrived,
        club_name       = CLUB_NAME,
        club_short_name = CLUB_SHORT_NAME,
        app_version     = APP_VERSION,
    )


@app.route('/registration')
def registration_page():
    """Landing page — dynamically lists all publicly-registerable member types."""
    db    = get_db()
    types = db.execute(
        'SELECT * FROM member_types WHERE public_registration = 1 AND active = 1 ORDER BY sort_order'
    ).fetchall()
    return render_template('registration_landing.html',
                           reg_types=[dict(t) for t in types],
                           club_name=CLUB_NAME, club_short_name=CLUB_SHORT_NAME)


@app.route('/registration/<slug>')
def registration_slug_page(slug):
    """Dynamic registration form — rendered by JS from field config, works for any member type."""
    db    = get_db()
    mtype = db.execute(
        'SELECT * FROM member_types WHERE slug = ? AND active = 1', (slug,)
    ).fetchone()
    if not mtype:
        return redirect('/registration')
    return render_template('registration_dynamic.html',
                           type_slug=slug,
                           type_info=dict(mtype),
                           version=APP_VERSION,
                           club_name=CLUB_NAME, club_short_name=CLUB_SHORT_NAME)

@app.route('/documents')
@permission_required('documents.view')
def documents_page():
    return render_template('documents.html', active_page='documents', **tpl_ctx())

@app.route('/communications')
@permission_required('mailshots.send')
def communications_page():
    return render_template('communications.html', active_page='communications', **tpl_ctx())

@app.route('/admin/users')
@permission_required('users.view')
def users_page():
    return render_template('admin/users.html', active_page='users', **tpl_ctx())

@app.route('/admin/audit')
@permission_required('audit.view')
def audit_page():
    return render_template('admin/audit.html', active_page='audit', **tpl_ctx())

@app.route('/admin/settings')
@permission_required('admin.settings')
def settings_page():
    return render_template('admin/settings.html', active_page='settings', **tpl_ctx())

@app.route('/admin/roles')
@permission_required('admin.settings')
def roles_page():
    return render_template('admin/roles.html', active_page='roles', **tpl_ctx())

@app.route('/admin/staff-roles')
@permission_required('admin.settings')
def staff_roles_page():
    return render_template('admin/staff_roles.html', active_page='staff_roles', **tpl_ctx())

@app.route('/admin/tags')
@permission_required('admin.settings')
def tags_page():
    return render_template('admin/tags.html', active_page='tags', **tpl_ctx())

@app.route('/admin/member-types')
@permission_required('admin.settings')
def member_types_page():
    return render_template('admin/member_types.html', active_page='settings', **tpl_ctx())

@app.route('/admin/field-builder/<int:type_id>')
@permission_required('admin.settings')
def field_builder_page(type_id):
    db    = get_db()
    mtype = db.execute('SELECT * FROM member_types WHERE id = ?', (type_id,)).fetchone()
    if not mtype:
        return redirect(url_for('member_types_page'))
    return render_template('admin/field_builder.html',
                           active_page='settings',
                           mtype=dict(mtype),
                           **tpl_ctx())

@app.route('/admin/settings/attendance')
@permission_required('admin.settings')
def attendance_settings_page():
    return render_template('admin/attendance_settings.html', active_page='settings', **tpl_ctx())

@app.route('/admin/settings/session-types')
@permission_required('admin.session_types')
def session_types_page():
    return render_template('admin/session_types.html', active_page='settings', **tpl_ctx())

@app.route('/admin/settings/maintenance')
@permission_required('admin.maintenance')
def maintenance_page():
    return render_template('admin/maintenance.html', active_page='settings', **tpl_ctx())

@app.route('/api/settings')
@permission_required('admin.settings')
def api_settings_get():
    """Return all settings as a key/value dict."""
    db   = get_db()
    rows = db.execute('SELECT key, value FROM settings').fetchall()
    return jsonify({r['key']: r['value'] for r in rows})

@app.route('/api/settings', methods=['POST'])
@permission_required('admin.settings')
def api_settings_save():
    """Save one or more settings. Body: {key: value, ...}
    v8.0: at_risk_threshold_* keys are no longer in use — alert rule thresholds
    are now configured per-rule in the alert_rules table (/admin/alerts).
    This endpoint is kept for any future generic settings that need it."""
    data = request.get_json() or {}
    # Allowlist of writable keys — extend here when new settings are added
    ALLOWED_KEYS = set()   # No generic settings require saving via this endpoint post-v8.0
    db = get_db()
    saved = {}
    for key, value in data.items():
        if key not in ALLOWED_KEYS:
            continue
        db.execute(
            'INSERT INTO settings (key, value, updated_at, updated_by) VALUES (?, ?, datetime("now"), ?)'
            ' ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at, updated_by = excluded.updated_by',
            (key, str(value), session['user_id'])
        )
        saved[key] = value
    db.commit()
    if saved:
        log_action('update_settings', 'settings', None, {'changes': saved})
    return jsonify({'success': True})

# ── Login rate limiter (in-memory, per-IP) ────────────────────────────────────
# Tracks failed login attempts per client IP.  After _LOGIN_MAX_FAILURES
# consecutive failures the IP is locked out for _LOGIN_LOCKOUT_SECONDS.
# State lives only in-process — resets on restart (acceptable for this scale).

_login_attempts: dict = {}   # ip -> {'count': int, 'locked_until': float}
_LOGIN_MAX_FAILURES    = 10
_LOGIN_LOCKOUT_SECONDS = 15 * 60   # 15 minutes


def _check_login_rate_limit(ip: str):
    """Return (allowed: bool, retry_after_seconds: int)."""
    now  = time.time()
    rec  = _login_attempts.get(ip, {'count': 0, 'locked_until': 0})
    if now < rec['locked_until']:
        return False, int(rec['locked_until'] - now)
    return True, 0


def _record_login_failure(ip: str):
    """Increment failure count; lock IP if threshold reached."""
    now = time.time()
    rec = _login_attempts.get(ip, {'count': 0, 'locked_until': 0})
    rec['count'] += 1
    if rec['count'] >= _LOGIN_MAX_FAILURES:
        rec['locked_until'] = now + _LOGIN_LOCKOUT_SECONDS
        rec['count']        = 0   # reset counter so next lockout is fresh
    _login_attempts[ip] = rec


def _clear_login_failures(ip: str):
    """Reset failure state on successful login."""
    _login_attempts.pop(ip, None)


# ── BLUEPRINT: auth ────────────────────────────────────────────────────────────

@app.route('/api/auth/login', methods=['POST'])
@csrf.exempt   # pre-auth — no session or token exists yet
def api_login():
    ip       = request.remote_addr or '0.0.0.0'
    allowed, retry_after = _check_login_rate_limit(ip)
    if not allowed:
        return jsonify({
            'error': f'Too many failed login attempts. Please try again in {retry_after // 60 + 1} minutes.'
        }), 429

    data     = request.get_json() or {}
    username = data.get('username', '').strip().lower()
    password = data.get('password', '')

    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400

    db   = get_db()
    user = db.execute(
        'SELECT * FROM users WHERE lower(username) = ? AND active = 1',
        (username,)
    ).fetchone()

    if user and bcrypt.checkpw(password.encode('utf-8'),
                                user['password_hash'].encode('utf-8')):
        # Load permissions from the roles table via role_id.
        # Fall back to the role text column if role_id is not yet set
        # (safety net for the one-release transition period).
        perms        = []
        role_name    = user['role']
        role_display = ROLE_DISPLAY_NAMES.get(role_name, role_name)
        if user['role_id']:
            role_row = db.execute(
                'SELECT name, permissions, display_name FROM roles WHERE id = ?', (user['role_id'],)
            ).fetchone()
            if role_row:
                role_name    = role_row['name']
                role_display = role_row['display_name'] or ROLE_DISPLAY_NAMES.get(role_name, role_name)
                try:
                    perms = json.loads(role_row['permissions'])
                except (TypeError, ValueError):
                    perms = []
        else:
            # Fallback: look up by role name (covers users not yet migrated)
            role_row = db.execute(
                'SELECT name, permissions, display_name FROM roles WHERE name = ?', (user['role'],)
            ).fetchone()
            if role_row:
                role_display = role_row['display_name'] or ROLE_DISPLAY_NAMES.get(role_name, role_name)
                try:
                    perms = json.loads(role_row['permissions'])
                except (TypeError, ValueError):
                    perms = []

        session['user_id']            = user['id']
        session['username']           = user['username']
        session['role']               = role_name          # slug — kept for _assigned_session() + templates
        session['role_display']       = role_display       # v7.1: human-readable label
        session['permissions']        = perms              # v6.0: full permission list
        session['session_assigned']   = user['session_assigned'] or ''

        db.execute('UPDATE users SET last_login = ? WHERE id = ?',
                   (datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S'), user['id']))
        db.commit()

        _clear_login_failures(ip)
        log_action('login')

        return jsonify({
            'success': True,
            'redirect': '/dashboard',
            'user': {
                'username':         user['username'],
                'role':             role_name,
                'session_assigned': user['session_assigned'],
            }
        })

    _record_login_failure(ip)
    log_action('login_failed', details={'attempted_username': username})
    return jsonify({'error': 'Incorrect username or password'}), 401

@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    log_action('logout')   # must log BEFORE clearing session (user_id is still set)
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/auth/me')
@login_required
def api_me():
    return jsonify({
        'username':         session['username'],
        'role':             session['role'],
        'role_display':     session.get('role_display', session['role']),
        'session_assigned': session.get('session_assigned', ''),
        'permissions':      session.get('permissions', []),
    })

@app.route('/api/auth/change-password', methods=['POST'])
@login_required
def api_change_password():
    data         = request.get_json() or {}
    current_pw   = data.get('current_password', '')
    new_pw       = data.get('new_password', '')

    if not current_pw or not new_pw:
        return jsonify({'error': 'Both current and new passwords are required'}), 400
    pw_error = validate_password(new_pw)
    if pw_error:
        return jsonify({'error': pw_error}), 400

    db   = get_db()
    user = db.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not user or not bcrypt.checkpw(current_pw.encode('utf-8'),
                                      user['password_hash'].encode('utf-8')):
        return jsonify({'error': 'Current password is incorrect'}), 401

    new_hash = bcrypt.hashpw(new_pw.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    db.execute('UPDATE users SET password_hash = ? WHERE id = ?',
               (new_hash, session['user_id']))
    db.commit()
    log_action('change_password')
    return jsonify({'success': True})

# ── BLUEPRINT: members ─────────────────────────────────────────────────────────

@app.route('/api/members')
@permission_required('members.view')
def api_members():
    db = get_db()

    status_filter  = request.args.get('status', 'active')  # active | leaver | flagged | all
    session_filter = request.args.get('session', 'all')     # all | Tuesday | Thursday
    flag_filter    = request.args.get('flag_rule_id')        # optional: filter by specific rule id

    conditions = ['1=1']
    params = []

    if status_filter == 'active':
        conditions.append("m.status = 'Active'")
    elif status_filter == 'leaver':
        conditions.append("m.status = 'Leaver'")

    if session_filter != 'all':
        conditions.append('m.session = ?')
        params.append(session_filter)

    # All non-admin roles are scoped to their assigned session.
    # No session assigned → return empty (safest default).
    scoped = _assigned_session()
    if scoped is not None:           # non-admin
        if not scoped:
            return jsonify([])       # no session set — return empty
        conditions.append('m.session = ?')
        params.append(scoped)

    # Leaders cannot see staff/volunteer records — youth members only.
    if session.get('role') == 'leader':
        conditions.append("m.member_type = 'member'")

    # flagged filter — join to member_flags; optionally restrict to a specific rule
    flag_join = ''
    if status_filter == 'flagged' or flag_filter:
        if flag_filter:
            try:
                rule_id_int = int(flag_filter)
            except (ValueError, TypeError):
                rule_id_int = None
            if rule_id_int:
                flag_join = (
                    f'INNER JOIN member_flags mf ON mf.member_id = m.id '
                    f'AND mf.rule_id = {rule_id_int} AND mf.resolved_at IS NULL'
                )
            else:
                flag_join = 'INNER JOIN member_flags mf ON mf.member_id = m.id AND mf.resolved_at IS NULL'
        else:
            flag_join = 'INNER JOIN member_flags mf ON mf.member_id = m.id AND mf.resolved_at IS NULL'

    where = ' AND '.join(conditions)

    rows = db.execute(f'''
        SELECT  DISTINCT m.*,
                c1.contact_name  AS contact1_name,
                c1.contact_phone AS contact1_phone,
                c1.contact_email AS contact1_email,
                c2.contact_name  AS contact2_name,
                c2.contact_phone AS contact2_phone,
                c2.contact_email AS contact2_email
        FROM    members m
        {flag_join}
        LEFT JOIN member_contacts c1
               ON c1.member_id = m.id AND c1.contact_order = 1
        LEFT JOIN member_contacts c2
               ON c2.member_id = m.id AND c2.contact_order = 2
        WHERE   {where}
        ORDER   BY m.first_name, m.surname
    ''', params).fetchall()

    member_ids = [r['id'] for r in rows]
    custom_fields_map = {}
    flags_map = {}
    if member_ids:
        placeholders = ','.join('?' * len(member_ids))
        cfv_rows = db.execute(
            f'SELECT mfv.member_id, fd.key, mfv.value '
            f'FROM member_field_values mfv '
            f'JOIN field_definitions fd ON fd.id = mfv.field_id '
            f'WHERE mfv.member_id IN ({placeholders})',
            member_ids
        ).fetchall()
        for cfv in cfv_rows:
            custom_fields_map.setdefault(cfv['member_id'], {})[cfv['key']] = cfv['value']

        # Attach active flags so the member list can render badges
        flag_rows = db.execute(
            f'SELECT mf.member_id, mf.id AS flag_id, mf.flagged_at, '
            f'ar.id AS rule_id, ar.flag_label, ar.flag_colour '
            f'FROM member_flags mf '
            f'JOIN alert_rules ar ON ar.id = mf.rule_id '
            f'WHERE mf.member_id IN ({placeholders}) AND mf.resolved_at IS NULL',
            member_ids
        ).fetchall()
        for f in flag_rows:
            flags_map.setdefault(f['member_id'], []).append({
                'flag_id':    f['flag_id'],
                'rule_id':    f['rule_id'],
                'flag_label': f['flag_label'],
                'flag_colour': f['flag_colour'],
                'flagged_at': f['flagged_at'],
            })

    result = []
    for r in rows:
        d = dict(r)
        d['custom_fields'] = custom_fields_map.get(r['id'], {})
        d['flags'] = flags_map.get(r['id'], [])
        result.append(d)
    return jsonify(result)

@app.route('/api/members/<int:member_id>')
@permission_required('members.view')
def api_member_detail(member_id):
    db     = get_db()
    member = db.execute('SELECT * FROM members WHERE id = ?', (member_id,)).fetchone()
    if not member:
        return jsonify({'error': 'Not found'}), 404

    # Enforce session scope for non-admin roles
    scoped = _assigned_session()
    if scoped is not None and (member['session'] or '') != scoped:
        return jsonify({'error': 'Forbidden'}), 403

    contacts = db.execute(
        'SELECT * FROM member_contacts WHERE member_id = ? ORDER BY contact_order',
        (member_id,)
    ).fetchall()

    result             = dict(member)
    result['contacts'] = [dict(c) for c in contacts]
    return jsonify(result)

@app.route('/api/members/<int:member_id>/viewed', methods=['POST'])
@permission_required('members.view')
def api_member_viewed(member_id):
    """Record in the audit log that a member's card was opened and their details viewed."""
    db     = get_db()
    member = db.execute('SELECT first_name, surname, session FROM members WHERE id = ?', (member_id,)).fetchone()
    if not member:
        return jsonify({'error': 'Not found'}), 404

    # Enforce session scope for non-admin roles
    scoped = _assigned_session()
    if scoped is not None and (member['session'] or '') != scoped:
        return jsonify({'error': 'Forbidden'}), 403

    log_action('view_member', 'members', member_id, {
        'member':    f"{member['first_name'] or ''} {member['surname'] or ''}".strip(),
        'viewed_by': session['username'],
    })
    return jsonify({'ok': True})


@app.route('/api/field-config')
@permission_required('members.view')
def api_field_config():
    """Return field configuration for all active member types, keyed by slug."""
    db    = get_db()
    types = db.execute(
        'SELECT * FROM member_types WHERE active = 1 ORDER BY sort_order'
    ).fetchall()

    result = {}
    for mtype in types:
        rows = db.execute('''
            SELECT  fd.id, fd.key, fd.label, fd.field_type,
                    fd.column_name, fd.system_field,
                    fd.placeholder, fd.help_text, fd.options,
                    mtf.required,
                    mtf.show_on_registration, mtf.show_on_list,
                    mtf.show_on_card, mtf.show_on_detail, mtf.show_on_print, mtf.show_on_export,
                    mtf.sort_order
            FROM    member_type_fields mtf
            JOIN    field_definitions fd ON fd.id = mtf.field_id
            WHERE   mtf.member_type_id = ? AND fd.active = 1
            ORDER   BY mtf.sort_order
        ''', (mtype['id'],)).fetchall()

        all_fields = [dict(r) for r in rows]
        result[mtype['slug']] = {
            'type':         dict(mtype),
            'list':         [f for f in all_fields if f['show_on_list']],
            'card':         [f for f in all_fields if f['show_on_card']],
            'detail':       [f for f in all_fields if f['show_on_detail']],
            'print':        [f for f in all_fields if f['show_on_print']],
            'export':       [f for f in all_fields if f['show_on_export']],
            'registration': [f for f in all_fields if f['show_on_registration']],
        }

    return jsonify(result)


@app.route('/api/public/field-config/<slug>')
def api_public_field_config(slug):
    """Public endpoint — no auth required.
    Returns field config (show_on_registration fields only) for a single member type.
    """
    db    = get_db()
    mtype = db.execute(
        'SELECT * FROM member_types WHERE slug = ? AND active = 1', (slug,)
    ).fetchone()
    if not mtype:
        return jsonify({'error': 'Not found'}), 404

    rows = db.execute('''
        SELECT  fd.id, fd.key, fd.label, fd.field_type,
                fd.column_name, fd.system_field,
                fd.placeholder, fd.help_text, fd.options,
                mtf.required, mtf.sort_order
        FROM    member_type_fields mtf
        JOIN    field_definitions fd ON fd.id = mtf.field_id
        WHERE   mtf.member_type_id = ? AND fd.active = 1
          AND   mtf.show_on_registration = 1
        ORDER   BY mtf.sort_order
    ''', (mtype['id'],)).fetchall()

    return jsonify({
        'type':   dict(mtype),
        'fields': [dict(r) for r in rows],
    })


@app.route('/api/public/session-types')
def api_public_session_types():
    """Public endpoint — no auth required. Returns active session types for public forms."""
    return jsonify(get_session_types())


@app.route('/api/members/<int:member_id>', methods=['PUT'])
@permission_required('members.edit')
def api_member_update(member_id):
    """Edit a member record. Logs every change to the audit trail."""
    data = request.get_json() or {}
    db   = get_db()

    before = db.execute('SELECT * FROM members WHERE id = ?', (member_id,)).fetchone()
    if not before:
        return jsonify({'error': 'Not found'}), 404

    # Enforce session scope for non-admin roles
    scoped = _assigned_session()
    if scoped is not None and (before['session'] or '') != scoped:
        return jsonify({'error': 'Forbidden'}), 403

    text_fields = ['first_name', 'surname', 'date_of_birth', 'address', 'postcode',
                   'ethnicity_religion', 'medical_sen', 'gp_contact', 'status',
                   'session', 'comments']
    bool_fields = ['unattended_exit', 'gdpr_consent']

    updates, params = [], []
    changes = {}

    for field in text_fields:
        if field in data:
            updates.append(f'{field} = ?')
            params.append(data[field])
            if str(before[field] or '') != str(data[field] or ''):
                changes[field] = {'from': before[field], 'to': data[field]}

    for field in bool_fields:
        if field in data:
            new_val = 1 if data[field] else 0
            updates.append(f'{field} = ?')
            params.append(new_val)
            if before[field] != new_val:
                changes[field] = {'from': bool(before[field]), 'to': bool(new_val)}

    updates += ['updated_at = datetime("now")', 'updated_by = ?']
    params  += [session['user_id'], member_id]

    db.execute(f"UPDATE members SET {', '.join(updates)} WHERE id = ?", params)

    # Contacts — replace wholesale if provided
    if 'contacts' in data:
        db.execute('DELETE FROM member_contacts WHERE member_id = ?', (member_id,))
        for c in data['contacts']:
            if c.get('contact_name') or c.get('contact_phone') or c.get('contact_email'):
                db.execute(
                    'INSERT INTO member_contacts'
                    ' (member_id, contact_order, contact_name, contact_phone, contact_email)'
                    ' VALUES (?,?,?,?,?)',
                    (member_id, c.get('contact_order', 1),
                     c.get('contact_name', ''), c.get('contact_phone', ''),
                     c.get('contact_email', ''))
                )

    db.commit()
    log_action('edit_member', 'members', member_id, {
        'member': f"{before['first_name'] or ''} {before['surname'] or ''}".strip(),
        'editor': session['username'],
        'changes': changes,
    })
    return jsonify({'success': True})


@app.route('/api/members/<int:member_id>', methods=['DELETE'])
@permission_required('members.delete')
def api_member_delete(member_id):
    """
    Soft-delete: mark member as Leaver. Requires a reason.
    Core Leader and Admin only. Record is kept for audit/history.
    """
    data   = request.get_json() or {}
    reason = data.get('reason', '').strip()
    if not reason:
        return jsonify({'error': 'A reason is required when marking a member as Leaver'}), 400

    db     = get_db()
    member = db.execute('SELECT * FROM members WHERE id = ?', (member_id,)).fetchone()
    if not member:
        return jsonify({'error': 'Not found'}), 404

    # Enforce session scope for non-admin roles
    scoped = _assigned_session()
    if scoped is not None and (member['session'] or '') != scoped:
        return jsonify({'error': 'Forbidden'}), 403

    db.execute(
        "UPDATE members SET status = 'Leaver', status_note = ?, "
        "updated_at = datetime('now'), updated_by = ? WHERE id = ?",
        (reason, session['user_id'], member_id)
    )
    db.commit()
    log_action('soft_delete_member', 'members', member_id,
               {'member':  f"{member['first_name'] or ''} {member['surname'] or ''}".strip(),
                'reason':  reason,
                'by':      session['username']})
    return jsonify({'success': True})


@app.route('/api/members/<int:member_id>/permanent', methods=['DELETE'])
@permission_required('members.hard_delete')
def api_member_permanent_delete(member_id):
    """
    Permanently and irreversibly delete a member and ALL associated data.
    Admin only. Wrapped in a transaction — either everything goes or nothing does.
    Requires confirmation token (member's full name) to prevent accidental calls.
    """
    data  = request.get_json() or {}
    token = data.get('confirm_name', '').strip()

    db     = get_db()
    member = db.execute('SELECT * FROM members WHERE id = ?', (member_id,)).fetchone()
    if not member:
        return jsonify({'error': 'Member not found'}), 404

    expected = f"{member['first_name']} {member['surname']}"
    if token.lower() != expected.lower():
        return jsonify({'error': 'Confirmation name does not match — deletion cancelled'}), 400

    # Gather counts for the audit entry before we delete anything
    att_count = db.execute(
        'SELECT COUNT(*) AS n FROM attendance WHERE member_id = ?', (member_id,)
    ).fetchone()['n']
    dofe_count = db.execute(
        'SELECT COUNT(*) AS n FROM dofe_participants WHERE member_id = ?', (member_id,)
    ).fetchone()['n']
    contact_count = db.execute(
        'SELECT COUNT(*) AS n FROM member_contacts WHERE member_id = ?', (member_id,)
    ).fetchone()['n']

    # Write the audit entry BEFORE deletion (same connection, inside transaction)
    # so the log is committed atomically with the delete.
    try:
        db.execute('BEGIN')

        # 1. Attendance records
        db.execute('DELETE FROM attendance WHERE member_id = ?', (member_id,))
        # 2. DoE participants
        db.execute('DELETE FROM dofe_participants WHERE member_id = ?', (member_id,))
        # 3. Member contacts (CASCADE would handle this but being explicit)
        db.execute('DELETE FROM member_contacts WHERE member_id = ?', (member_id,))
        # 4. Nullify the reviewed_by link on any pending_registrations (no FK, just housekeeping)
        db.execute(
            "UPDATE pending_registrations SET notes = COALESCE(notes || ' ', '') || '[Member record permanently deleted]'"
            " WHERE id IN ("
            "  SELECT id FROM pending_registrations WHERE status = 'approved'"
            "  AND first_name = ? AND surname = ?"
            ")",
            (member['first_name'], member['surname'])
        )
        # 5. Audit log entry — written inside transaction so it commits or rolls back together
        db.execute(
            'INSERT INTO audit_log (user_id, action, table_name, record_id, details, ip_address)'
            ' VALUES (?,?,?,?,?,?)',
            (
                session.get('user_id'),
                'permanent_delete_member',
                'members',
                member_id,
                json.dumps({
                    'member':        expected,
                    'member_id':     member['member_id'],
                    'session':       member['session'],
                    'deleted_by':    session.get('username'),
                    'att_deleted':   att_count,
                    'dofe_deleted':  dofe_count,
                    'contacts_deleted': contact_count,
                }),
                request.remote_addr,
            )
        )
        # 6. The member row itself — last
        db.execute('DELETE FROM members WHERE id = ?', (member_id,))

        db.execute('COMMIT')
    except Exception as e:
        db.execute('ROLLBACK')
        app.logger.error(f'Permanent member delete failed (member_id={member_id}): {e}')
        return jsonify({'error': 'Deletion failed and was rolled back. Check server logs for details.'}), 500

    return jsonify({
        'success':  True,
        'deleted':  expected,
        'summary': {
            'attendance': att_count,
            'contacts':   contact_count,
            'dofe':       dofe_count,
        }
    })


@app.route('/api/dashboard')
@login_required
def api_dashboard():
    """Summary stats for the dashboard home page. Session-scoped for all non-admin roles."""
    db    = get_db()
    today = datetime.now().strftime('%Y-%m-%d')

    scoped = _assigned_session()   # None for admin, session string for everyone else

    if scoped is None:
        # Admin — global counts
        counts = db.execute('''
            SELECT
                SUM(CASE WHEN member_type = "member"                        THEN 1 ELSE 0 END) AS total,
                SUM(CASE WHEN member_type = "member" AND status = "Active"  THEN 1 ELSE 0 END) AS active,
                SUM(CASE WHEN member_type = "member" AND status = "Leaver"  THEN 1 ELSE 0 END) AS leavers,
                SUM(CASE WHEN member_type = "staff"  AND status != "Leaver" THEN 1 ELSE 0 END) AS staff_active
            FROM members
        ''').fetchone()
        # Per-session counts (dynamic)
        session_rows = db.execute('''
            SELECT session,
                   SUM(CASE WHEN member_type = "member" AND status != "Leaver" THEN 1 ELSE 0 END) AS members,
                   SUM(CASE WHEN member_type = "staff"  AND status != "Leaver" THEN 1 ELSE 0 END) AS staff
            FROM members GROUP BY session
        ''').fetchall()
        pending = db.execute(
            'SELECT COUNT(*) AS n FROM pending_registrations WHERE status = "pending"'
        ).fetchone()['n']
        # Alert flags summary — per rule, count of active flags (admin: global)
        alert_rows = db.execute('''
            SELECT ar.id, ar.flag_label, ar.flag_colour,
                   COUNT(mf.id) AS flag_count
            FROM alert_rules ar
            LEFT JOIN member_flags mf ON mf.rule_id = ar.id AND mf.resolved_at IS NULL
            WHERE ar.is_active = 1
            GROUP BY ar.id
            ORDER BY flag_count DESC, ar.name
        ''').fetchall()
    else:
        # Scoped user — counts restricted to their session only
        counts = db.execute('''
            SELECT
                SUM(CASE WHEN member_type = "member"                        THEN 1 ELSE 0 END) AS total,
                SUM(CASE WHEN member_type = "member" AND status = "Active"  THEN 1 ELSE 0 END) AS active,
                SUM(CASE WHEN member_type = "member" AND status = "Leaver"  THEN 1 ELSE 0 END) AS leavers,
                SUM(CASE WHEN member_type = "staff"  AND status != "Leaver" THEN 1 ELSE 0 END) AS staff_active
            FROM members WHERE session = ?
        ''', (scoped,)).fetchone()
        # Per-session counts (scoped — only the user's session)
        session_rows = db.execute('''
            SELECT session,
                   SUM(CASE WHEN member_type = "member" AND status != "Leaver" THEN 1 ELSE 0 END) AS members,
                   SUM(CASE WHEN member_type = "staff"  AND status != "Leaver" THEN 1 ELSE 0 END) AS staff
            FROM members WHERE session = ? GROUP BY session
        ''', (scoped,)).fetchall()
        pending = db.execute(
            'SELECT COUNT(*) AS n FROM pending_registrations WHERE status = "pending"'
            ' AND (assigned_session = ? OR assigned_session IS NULL OR assigned_session = "")',
            (scoped,)
        ).fetchone()['n']
        # Alert flags summary — scoped to members in this session
        alert_rows = db.execute('''
            SELECT ar.id, ar.flag_label, ar.flag_colour,
                   COUNT(mf.id) AS flag_count
            FROM alert_rules ar
            LEFT JOIN member_flags mf ON mf.rule_id = ar.id AND mf.resolved_at IS NULL
            LEFT JOIN members m ON m.id = mf.member_id AND m.session = ?
            WHERE ar.is_active = 1
            GROUP BY ar.id
            ORDER BY flag_count DESC, ar.name
        ''', (scoped,)).fetchall()

    # Build per-session dict: {session_name: {members: N, staff: N}}
    # Filter out rows where session IS NULL (e.g. members imported without a session assigned).
    session_counts = {r['session']: {'members': r['members'], 'staff': r['staff']}
                      for r in session_rows if r['session'] is not None}

    # Today's attendance (scoped where applicable)
    if scoped is None:
        today_att = db.execute('''
            SELECT COUNT(*) AS total_signed_in,
                   SUM(CASE WHEN signed_out_at IS NOT NULL THEN 1 ELSE 0 END) AS signed_out,
                   session_type
            FROM attendance
            WHERE session_date = ?
            GROUP BY session_type
        ''', (today,)).fetchall()
    else:
        today_att = db.execute('''
            SELECT COUNT(*) AS total_signed_in,
                   SUM(CASE WHEN signed_out_at IS NOT NULL THEN 1 ELSE 0 END) AS signed_out,
                   session_type
            FROM attendance
            WHERE session_date = ? AND session_type = ?
            GROUP BY session_type
        ''', (today, scoped)).fetchall()

    # Recent audit log (last 8 entries)
    recent = db.execute('''
        SELECT a.action, a.details, a.timestamp, u.username
        FROM   audit_log a
        LEFT JOIN users u ON u.id = a.user_id
        ORDER  BY a.timestamp DESC
        LIMIT  8
    ''').fetchall()

    alerts_last_run = db.execute(
        "SELECT value FROM settings WHERE key = 'alerts_last_run'"
    ).fetchone()

    return jsonify({
        'members':          dict(counts),
        'session_counts':   session_counts,   # {session_name: {members, staff}}
        'pending_approvals': pending,
        'today_attendance': [dict(r) for r in today_att],
        'recent_activity':  [dict(r) for r in recent],
        'scoped_session':   scoped,
        'alert_flags':      [dict(r) for r in alert_rows],
        'alerts_last_run':  alerts_last_run['value'] if alerts_last_run else '',
    })


@app.route('/api/admin/audit')
@permission_required('audit.view')
def api_audit_log():
    """Return recent audit log entries."""
    limit  = min(int(request.args.get('limit', 200)), 500)
    offset = int(request.args.get('offset', 0))
    db     = get_db()
    rows   = db.execute('''
        SELECT  a.*, u.username
        FROM    audit_log a
        LEFT JOIN users u ON u.id = a.user_id
        ORDER   BY a.timestamp DESC
        LIMIT   ? OFFSET ?
    ''', (limit, offset)).fetchall()
    return jsonify([dict(r) for r in rows])


# ── BLUEPRINT: postcode lookup proxy ─────────────────────────────────────────

@app.route('/api/postcode/<path:postcode>')
def api_postcode_lookup(postcode):
    """Proxy getaddress.io lookups server-side to avoid browser CORS restrictions."""
    if not GETADDRESS_KEY:
        return jsonify({'error': 'Address lookup not configured on this server'}), 503

    clean = urllib.parse.quote(postcode.replace(' ', '').upper())
    url   = f'https://ga.ideal-postcodes.co.uk/find/{clean}?api-key={GETADDRESS_KEY}'

    try:
        with urllib.request.urlopen(url, timeout=8) as resp:
            data = json.loads(resp.read().decode())
        return jsonify(data)
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        try:
            parsed = json.loads(body)
            return jsonify(parsed), e.code
        except Exception:
            return jsonify({'error': body, 'http_status': e.code}), e.code
    except Exception as e:
        app.logger.error(f'Postcode lookup network error ({postcode}): {e}')
        return jsonify({'error': 'Address lookup failed — please enter your address manually.'}), 502



# ── BLUEPRINT: staff roles (public read) ──────────────────────────────────────

@app.route('/api/staff-roles')
def api_staff_roles_public():
    """Return active staff roles ordered by display_order — no auth required (used on public registration form)."""
    db   = get_db()
    rows = db.execute(
        'SELECT id, name, display_order FROM staff_roles WHERE active = 1 ORDER BY display_order, name'
    ).fetchall()
    return jsonify([dict(r) for r in rows])


# ── BLUEPRINT: registration (public) ──────────────────────────────────────────

@app.route('/api/registration', methods=['POST'])
@csrf.exempt   # public unauthenticated endpoint — no session token available
def api_registration():
    """Accept a public self-registration and store it as pending.
    Fully field-driven — no hardcoded member/staff branching.
    Supports both legacy (registration_type) and new dynamic (member_type_slug) payloads.
    """
    data = request.get_json() or {}
    db   = get_db()

    # Determine member type and registration style
    slug  = (data.get('member_type_slug') or data.get('registration_type') or 'member').strip()
    mtype = db.execute(
        'SELECT * FROM member_types WHERE slug = ? AND active = 1', (slug,)
    ).fetchone()

    if mtype:
        style         = mtype['registration_style'] or 'member'
        type_slug_val = mtype['slug']
        rtype         = 'staff' if style == 'staff' else 'member'
    else:
        # Fallback for old payloads
        rtype         = 'staff' if slug == 'staff' else 'member'
        style         = rtype
        type_slug_val = slug

    # Validate staff role if provided
    applicant_role = (data.get('applicant_role') or data.get('staff_role') or '').strip()
    if applicant_role:
        valid_roles = [r['name'] for r in db.execute(
            'SELECT name FROM staff_roles WHERE active = 1'
        ).fetchall()]
        if applicant_role not in valid_roles:
            return jsonify({'error': 'Invalid role'}), 400

    # Validate session if provided
    session_pref = (data.get('assigned_session') or data.get('session') or '').strip()
    valid_sessions = get_valid_session_names()
    if session_pref and session_pref not in valid_sessions:
        return jsonify({'error': 'Invalid session preference'}), 400

    # Serialize any custom fields submitted by the dynamic form
    raw_custom = data.get('custom_fields')
    custom_fields_json = json.dumps(raw_custom) if isinstance(raw_custom, dict) and raw_custom else None

    # Single unified INSERT — all columns nullable, unused ones default to NULL/0
    db.execute('''
        INSERT INTO pending_registrations
            (first_name, surname, date_of_birth, address, postcode,
             ethnicity_religion, medical_sen, gp_contact,
             unattended_exit, gdpr_consent, comms_consent,
             contact1_name, contact1_phone, contact1_email,
             contact2_name, contact2_phone, contact2_email,
             mobile, email,
             declarations, registration_type,
             custom_fields, member_type_slug,
             applicant_role, assigned_session)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (
        data.get('first_name', '').strip(),
        data.get('surname', '').strip(),
        data.get('date_of_birth', ''),
        data.get('address', '').strip(),
        data.get('postcode', '').strip(),
        data.get('ethnicity_religion', '').strip(),
        data.get('medical_sen', '').strip(),
        data.get('gp_contact', '').strip(),
        1 if data.get('unattended_exit') else 0,
        1 if data.get('gdpr_consent') else 0,
        1 if data.get('comms_consent') else 0,
        data.get('contact1_name', '').strip(),
        data.get('contact1_phone', '').strip(),
        data.get('contact1_email', '').strip(),
        data.get('contact2_name', '').strip(),
        data.get('contact2_phone', '').strip(),
        data.get('contact2_email', '').strip(),
        data.get('mobile', '').strip(),
        data.get('email', '').strip(),
        json.dumps(data.get('declarations', {})),
        rtype,
        custom_fields_json,
        type_slug_val,
        applicant_role,
        session_pref,
    ))

    db.commit()
    return jsonify({'success': True})

# ── BLUEPRINT: admin ───────────────────────────────────────────────────────────

# ── Session types API ──────────────────────────────────────────────────────────

@app.route('/api/session-types')
@login_required
def api_session_types_list():
    """Public read endpoint — returns active session types for dropdowns/JS."""
    return jsonify(get_session_types())


@app.route('/api/admin/session-types', methods=['GET'])
@permission_required('admin.session_types')
def api_admin_session_types_get():
    """Return all session types (including inactive)."""
    db   = get_db()
    rows = db.execute(
        'SELECT id, name, weekday, active, sort_order FROM session_types ORDER BY sort_order, name'
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/admin/session-types', methods=['POST'])
@permission_required('admin.session_types')
def api_admin_session_types_create():
    """Create a new session type."""
    data    = request.get_json() or {}
    name    = data.get('name', '').strip()
    weekday = data.get('weekday')

    if not name:
        return jsonify({'error': 'name is required'}), 400
    if weekday is None or not isinstance(weekday, int) or weekday < 0 or weekday > 6:
        return jsonify({'error': 'weekday must be an integer 0–6 (Mon=0)'}), 400

    db = get_db()
    max_order = db.execute('SELECT COALESCE(MAX(sort_order), -1) FROM session_types').fetchone()[0]
    try:
        cur = db.execute(
            'INSERT INTO session_types (name, weekday, active, sort_order) VALUES (?,?,1,?)',
            (name, weekday, max_order + 1)
        )
        db.commit()
        log_action('create_session_type', 'session_types', cur.lastrowid, {'name': name, 'weekday': weekday})
        row = db.execute('SELECT * FROM session_types WHERE id = ?', (cur.lastrowid,)).fetchone()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A session type named "{name}" already exists'}), 409


@app.route('/api/admin/session-types/<int:st_id>', methods=['PUT'])
@permission_required('admin.session_types')
def api_admin_session_types_update(st_id):
    """Update a session type's name, weekday, or active status."""
    data    = request.get_json() or {}
    db      = get_db()
    current = db.execute('SELECT * FROM session_types WHERE id = ?', (st_id,)).fetchone()
    if not current:
        return jsonify({'error': 'Session type not found'}), 404

    name    = data.get('name', current['name']).strip()
    weekday = data.get('weekday', current['weekday'])
    active  = int(data.get('active', current['active']))

    if not name:
        return jsonify({'error': 'name is required'}), 400
    if not isinstance(weekday, int) or weekday < 0 or weekday > 6:
        return jsonify({'error': 'weekday must be an integer 0–6'}), 400

    # Prevent deactivating the last active session type
    if not active:
        active_count = db.execute(
            'SELECT COUNT(*) FROM session_types WHERE active = 1 AND id != ?', (st_id,)
        ).fetchone()[0]
        if active_count == 0:
            return jsonify({'error': 'Cannot deactivate the only active session type'}), 400

    try:
        db.execute(
            'UPDATE session_types SET name = ?, weekday = ?, active = ? WHERE id = ?',
            (name, weekday, active, st_id)
        )
        db.commit()
        log_action('update_session_type', 'session_types', st_id,
                   {'name': name, 'weekday': weekday, 'active': active})
        return jsonify(dict(db.execute('SELECT * FROM session_types WHERE id = ?', (st_id,)).fetchone()))
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A session type named "{name}" already exists'}), 409


@app.route('/api/admin/session-types/<int:st_id>', methods=['DELETE'])
@permission_required('admin.session_types')
def api_admin_session_types_delete(st_id):
    """Delete a session type (only if no members or attendance records use it)."""
    db  = get_db()
    row = db.execute('SELECT * FROM session_types WHERE id = ?', (st_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Session type not found'}), 404

    # Safety checks
    member_count = db.execute(
        'SELECT COUNT(*) FROM members WHERE session = ?', (row['name'],)
    ).fetchone()[0]
    if member_count:
        return jsonify({'error': f'Cannot delete — {member_count} member(s) are assigned to this session'}), 400

    active_count = db.execute(
        'SELECT COUNT(*) FROM session_types WHERE active = 1 AND id != ?', (st_id,)
    ).fetchone()[0]
    if row['active'] and active_count == 0:
        return jsonify({'error': 'Cannot delete the only active session type'}), 400

    db.execute('DELETE FROM session_types WHERE id = ?', (st_id,))
    db.commit()
    log_action('delete_session_type', 'session_types', st_id, {'name': row['name']})
    return jsonify({'success': True})


@app.route('/api/admin/session-types/reorder', methods=['POST'])
@permission_required('admin.session_types')
def api_admin_session_types_reorder():
    """Reorder session types. Body: [{id, sort_order}, ...]"""
    items = request.get_json() or []
    db    = get_db()
    for item in items:
        db.execute(
            'UPDATE session_types SET sort_order = ? WHERE id = ?',
            (item.get('sort_order', 0), item.get('id'))
        )
    db.commit()
    return jsonify({'success': True})


# ── BLUEPRINT: staff roles admin CRUD ─────────────────────────────────────────

@app.route('/api/admin/staff-roles', methods=['GET'])
@permission_required('admin.settings')
def api_admin_staff_roles_get():
    """Return all staff roles including inactive ones (admin view)."""
    db   = get_db()
    rows = db.execute(
        'SELECT id, name, active, display_order FROM staff_roles ORDER BY display_order, name'
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/admin/staff-roles', methods=['POST'])
@permission_required('admin.settings')
def api_admin_staff_roles_create():
    """Create a new staff role."""
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'name is required'}), 400

    db        = get_db()
    max_order = db.execute('SELECT COALESCE(MAX(display_order), -1) FROM staff_roles').fetchone()[0]
    try:
        cur = db.execute(
            'INSERT INTO staff_roles (name, active, display_order) VALUES (?,1,?)',
            (name, max_order + 1)
        )
        db.commit()
        log_action('create_staff_role', 'staff_roles', cur.lastrowid, {'name': name})
        row = db.execute('SELECT * FROM staff_roles WHERE id = ?', (cur.lastrowid,)).fetchone()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A role named "{name}" already exists'}), 409


@app.route('/api/admin/staff-roles/<int:role_id>', methods=['PUT'])
@permission_required('admin.settings')
def api_admin_staff_roles_update(role_id):
    """Update a staff role's name or active status."""
    data    = request.get_json() or {}
    db      = get_db()
    current = db.execute('SELECT * FROM staff_roles WHERE id = ?', (role_id,)).fetchone()
    if not current:
        return jsonify({'error': 'Role not found'}), 404

    name   = data.get('name', current['name']).strip()
    active = int(data.get('active', current['active']))

    if not name:
        return jsonify({'error': 'name is required'}), 400

    # Prevent deactivating the last active role
    if not active:
        active_count = db.execute(
            'SELECT COUNT(*) FROM staff_roles WHERE active = 1 AND id != ?', (role_id,)
        ).fetchone()[0]
        if active_count == 0:
            return jsonify({'error': 'Cannot deactivate the only active staff role'}), 400

    try:
        db.execute(
            'UPDATE staff_roles SET name = ?, active = ? WHERE id = ?',
            (name, active, role_id)
        )
        db.commit()
        log_action('update_staff_role', 'staff_roles', role_id,
                   {'name': name, 'active': active})
        return jsonify(dict(db.execute('SELECT * FROM staff_roles WHERE id = ?', (role_id,)).fetchone()))
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A role named "{name}" already exists'}), 409


@app.route('/api/admin/staff-roles/reorder', methods=['POST'])
@permission_required('admin.settings')
def api_admin_staff_roles_reorder():
    """Reorder staff roles. Body: [{id, display_order}, ...]"""
    items = request.get_json() or []
    db    = get_db()
    for item in items:
        db.execute(
            'UPDATE staff_roles SET display_order = ? WHERE id = ?',
            (item.get('display_order', 0), item.get('id'))
        )
    db.commit()
    return jsonify({'success': True})


@app.route('/api/admin/users')
@permission_required('users.view')
def api_users_list():
    db     = get_db()
    scoped = _assigned_session()
    if scoped is not None:
        # Editors see only non-admin users for their own session
        users = db.execute(
            'SELECT id, username, email, role, session_assigned, active, '
            'created_at, last_login FROM users '
            "WHERE role != 'admin' AND session_assigned = ? ORDER BY username",
            (scoped,)
        ).fetchall()
    else:
        users = db.execute(
            'SELECT id, username, email, role, session_assigned, active, '
            'created_at, last_login FROM users ORDER BY username'
        ).fetchall()
    return jsonify([dict(u) for u in users])

@app.route('/api/admin/users', methods=['POST'])
@permission_required('users.create')
def api_users_create():
    data     = request.get_json() or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')
    role     = data.get('role', 'readonly')
    email    = data.get('email', '').strip()
    sess     = data.get('session_assigned', '')

    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400
    pw_error = validate_password(password)
    if pw_error:
        return jsonify({'error': pw_error}), 400

    db = get_db()

    # Validate role exists in the roles table
    target_role_row = db.execute('SELECT id, permissions FROM roles WHERE name = ?', (role,)).fetchone()
    if not target_role_row:
        return jsonify({'error': 'Invalid role'}), 400

    # If the target role carries admin-level privileges, require users.create.admin
    target_perms = json.loads(target_role_row['permissions'])
    if 'users.create.admin' in target_perms and not has_permission('users.create.admin'):
        return jsonify({'error': 'You do not have permission to assign this role'}), 403

    # Non-admin-level users must have a session assigned
    # (sessions scope data — leaving it blank would expose all data)
    if 'admin.maintenance' not in target_perms and not sess:
        return jsonify({'error': 'A session must be assigned for non-admin users'}), 400
    if sess and sess not in get_valid_session_names():
        return jsonify({'error': 'Invalid session'}), 400

    # Scoped users (Core Leaders) can only create users for their own session
    scoped = _assigned_session()
    if scoped is not None and sess != scoped:
        return jsonify({'error': 'You can only create users for your own session'}), 403

    pw_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    try:
        db.execute(
            'INSERT INTO users (username, email, password_hash, role, role_id, session_assigned)'
            ' VALUES (?,?,?,?,?,?)',
            (username, email, pw_hash, role, target_role_row['id'], sess)
        )
        db.commit()
        log_action('create_user', 'users', None,
                   {'username': username, 'role': role, 'created_by': session.get('username')})
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Username already exists'}), 409

@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@permission_required('users.edit')
def api_users_update(user_id):
    data    = request.get_json() or {}
    db      = get_db()
    updates = []
    params  = []

    # Safety: cannot deactivate your own account
    if user_id == session['user_id'] and data.get('active') is False:
        return jsonify({'error': 'You cannot deactivate your own account'}), 400

    # Read current state now so we can log before/after changes
    before_row = db.execute(
        'SELECT username, email, role, session_assigned, active FROM users WHERE id = ?', (user_id,)
    ).fetchone()

    # If changing role, validate the target role and check admin-level permission
    if 'role' in data:
        target_role_row = db.execute(
            'SELECT id, permissions FROM roles WHERE name = ?', (data['role'],)
        ).fetchone()
        if not target_role_row:
            return jsonify({'error': 'Invalid role'}), 400
        target_perms = json.loads(target_role_row['permissions'])
        if 'users.create.admin' in target_perms and not has_permission('users.create.admin'):
            return jsonify({'error': 'You do not have permission to assign this role'}), 403

    # Editors can only modify users in their own session
    scoped = _assigned_session()
    if scoped is not None:
        target = db.execute('SELECT role, session_assigned FROM users WHERE id = ?', (user_id,)).fetchone()
        if not target or target['role'] == 'admin':
            return jsonify({'error': 'Forbidden'}), 403
        if target['session_assigned'] != scoped:
            return jsonify({'error': 'You can only manage users in your own session'}), 403
        # Cannot re-assign to a different session
        if 'session_assigned' in data and data['session_assigned'] != scoped:
            return jsonify({'error': 'You cannot move a user to a different session'}), 403

    if 'email' in data:
        updates.append('email = ?')
        params.append(data['email'])

    if 'role' in data:
        # role and role_id are updated together; validation already done above
        target_role_row = db.execute('SELECT id FROM roles WHERE name = ?', (data['role'],)).fetchone()
        updates.append('role = ?')
        params.append(data['role'])
        if target_role_row:
            updates.append('role_id = ?')
            params.append(target_role_row['id'])

    if 'session_assigned' in data:
        updates.append('session_assigned = ?')
        params.append(data['session_assigned'])

    if 'active' in data:
        updates.append('active = ?')
        params.append(1 if data['active'] else 0)

    if 'password' in data and data['password']:
        pw_error = validate_password(data['password'])
        if pw_error:
            return jsonify({'error': pw_error}), 400
        pw_hash = bcrypt.hashpw(data['password'].encode('utf-8'),
                                bcrypt.gensalt()).decode('utf-8')
        updates.append('password_hash = ?')
        params.append(pw_hash)

    if updates:
        params.append(user_id)
        db.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ?", params)
        db.commit()

        # Build before/after change log
        if before_row:
            field_changes = {}
            if 'email' in data and data.get('email') != before_row['email']:
                field_changes['email'] = {'from': before_row['email'], 'to': data['email']}
            if 'role' in data and data['role'] != before_row['role']:
                field_changes['role'] = {'from': before_row['role'], 'to': data['role']}
            if 'session_assigned' in data and data.get('session_assigned') != before_row['session_assigned']:
                field_changes['session_assigned'] = {'from': before_row['session_assigned'], 'to': data['session_assigned']}
            if 'active' in data:
                new_active = 1 if data['active'] else 0
                if new_active != before_row['active']:
                    field_changes['active'] = {'from': bool(before_row['active']), 'to': bool(new_active)}
            log_action('update_user', 'users', user_id, {
                'username':       before_row['username'],
                'changes':        field_changes,
                'password_reset': True if ('password' in data and data['password']) else None,
            })

    return jsonify({'success': True})

@app.route('/api/admin/users/<int:user_id>/permanent', methods=['DELETE'])
@permission_required('users.delete')
def api_users_permanent_delete(user_id):
    """Permanently delete a portal user account.

    Requires JSON body: { "confirm_username": "<username>" }
    Safety: admin cannot delete their own account this way.
    All FK references in other tables are nullified; audit_log rows are kept.
    """
    data = request.get_json() or {}
    db   = get_db()

    # Prevent self-deletion
    if user_id == session['user_id']:
        return jsonify({'error': 'You cannot delete your own account'}), 400

    user = db.execute('SELECT id, username FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user:
        return jsonify({'error': 'User not found'}), 404

    # Verify confirmation
    confirm = (data.get('confirm_username') or '').strip().lower()
    if confirm != user['username'].lower():
        return jsonify({'error': 'Username confirmation does not match'}), 400

    username = user['username']

    try:
        db.execute('BEGIN')

        # Nullify all FK references across tables
        db.execute('UPDATE members               SET updated_by = NULL WHERE updated_by = ?', (user_id,))
        db.execute('UPDATE pending_registrations SET reviewed_by = NULL WHERE reviewed_by = ?', (user_id,))
        db.execute('UPDATE documents             SET uploaded_by = NULL WHERE uploaded_by = ?', (user_id,))
        db.execute('UPDATE email_templates       SET created_by  = NULL WHERE created_by  = ?', (user_id,))
        db.execute('UPDATE mailshot_log          SET sent_by     = NULL WHERE sent_by     = ?', (user_id,))
        db.execute('UPDATE attendance            SET recorded_by = NULL WHERE recorded_by = ?', (user_id,))
        db.execute('UPDATE term_sessions         SET created_by  = NULL WHERE created_by  = ?', (user_id,))
        db.execute('UPDATE session_activities    SET added_by    = NULL WHERE added_by    = ?', (user_id,))
        # audit_log rows are retained; user_id is nullified to preserve history
        db.execute('UPDATE audit_log             SET user_id     = NULL WHERE user_id     = ?', (user_id,))

        # Delete the user
        db.execute('DELETE FROM users WHERE id = ?', (user_id,))

        # Audit entry for the deletion (recorded under the acting admin)
        db.execute(
            'INSERT INTO audit_log (user_id, action, table_name, record_id, details, ip_address)'
            ' VALUES (?, ?, ?, ?, ?, ?)',
            (
                session['user_id'], 'delete_user', 'users', user_id,
                json.dumps({'username': username}),
                request.remote_addr,
            )
        )

        db.execute('COMMIT')
    except Exception as exc:
        db.execute('ROLLBACK')
        app.logger.error(f'Permanent user delete failed (username={username}): {exc}')
        return jsonify({'error': 'Deletion failed and was rolled back. Check server logs for details.'}), 500

    return jsonify({'success': True, 'deleted': username})

# ── BLUEPRINT: approvals ──────────────────────────────────────────────────────

def _next_member_id(db):
    """Generate the next sequential member ID.

    Uses a custom format persisted in settings if one was learned during a
    bulk import (e.g. 'M-' prefix with 3-digit padding).  Falls back to
    CLUB_SHORT_NAME + 3-digit zero-padding (e.g. ARA001) if no custom
    format has been set.
    """
    import re as _re

    prefix_row  = db.execute("SELECT value FROM settings WHERE key='member_id_prefix'").fetchone()
    padding_row = db.execute("SELECT value FROM settings WHERE key='member_id_padding'").fetchone()

    if prefix_row is not None and padding_row is not None:
        prefix  = prefix_row['value']
        padding = int(padding_row['value'])
    else:
        prefix  = CLUB_SHORT_NAME
        padding = 3

    # Find the highest existing numeric suffix for this prefix across all members.
    all_ids  = db.execute('SELECT member_id FROM members').fetchall()
    max_num  = 0
    suffix_re = _re.compile(r'(\d+)$')
    for r in all_ids:
        mid = r['member_id'] or ''
        if not mid.startswith(prefix):
            continue
        m = suffix_re.search(mid)
        if m:
            max_num = max(max_num, int(m.group(1)))

    return f'{prefix}{max_num + 1:0{padding}d}'


def _save_id_format_from_import(db, imported_ids):
    """Detect and persist the member ID format from successfully imported IDs.

    Analyses the list, extracts a common non-numeric prefix and the maximum
    zero-padding width, then stores them in settings so _next_member_id picks
    them up for all future auto-generated IDs.

    Does nothing if the IDs are inconsistent (mixed prefixes, unparseable
    values, or only a single ID that could be coincidental).
    """
    import re as _re
    if not imported_ids:
        return

    pattern  = _re.compile(r'^([^0-9]*)(\d+)$')
    prefixes = []
    paddings = []

    for mid in imported_ids:
        m = pattern.match(str(mid).strip())
        if not m:
            return   # Unparseable ID — bail, don't guess
        prefixes.append(m.group(1))
        paddings.append(len(m.group(2)))

    if len(set(prefixes)) != 1:
        return   # Mixed prefixes — inconsistent, bail

    prefix  = prefixes[0]
    padding = max(paddings)

    db.execute(
        "INSERT INTO settings (key,value) VALUES ('member_id_prefix',?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=datetime('now')",
        (prefix,)
    )
    db.execute(
        "INSERT INTO settings (key,value) VALUES ('member_id_padding',?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=datetime('now')",
        (str(padding),)
    )
    db.commit()


@app.route('/api/approvals')
@permission_required('approvals.view')
def api_approvals_list():
    """List pending registrations. ?status=pending|approved|rejected|all
    Editors (Core Leaders) only see registrations for their assigned session."""
    status = request.args.get('status', 'pending')
    db     = get_db()
    scoped = _assigned_session()

    base_query = (
        'SELECT pr.*, u.username AS reviewed_by_name'
        ' FROM pending_registrations pr'
        ' LEFT JOIN users u ON u.id = pr.reviewed_by'
    )

    if scoped is not None:
        # Scoped: show pending (unassigned or matching their session) and
        # their own approved/rejected work.
        if status == 'all':
            rows = db.execute(
                base_query +
                ' WHERE (pr.assigned_session = ? OR pr.assigned_session IS NULL OR pr.assigned_session = "")'
                ' ORDER BY pr.submitted_at DESC',
                (scoped,)
            ).fetchall()
        else:
            rows = db.execute(
                base_query +
                ' WHERE pr.status = ?'
                ' AND (pr.assigned_session = ? OR pr.assigned_session IS NULL OR pr.assigned_session = "")'
                ' ORDER BY pr.submitted_at DESC',
                (status, scoped)
            ).fetchall()
    else:
        # Admin — see everything
        if status == 'all':
            rows = db.execute(
                base_query + ' ORDER BY pr.submitted_at DESC'
            ).fetchall()
        else:
            rows = db.execute(
                base_query + ' WHERE pr.status = ? ORDER BY pr.submitted_at DESC',
                (status,)
            ).fetchall()

    return jsonify([dict(r) for r in rows])


@app.route('/api/approvals/<int:reg_id>/approve', methods=['POST'])
@permission_required('approvals.approve')
def api_approvals_approve(reg_id):
    """Approve a pending registration — handles both member and staff types."""
    data = request.get_json() or {}
    db   = get_db()

    reg = db.execute(
        'SELECT * FROM pending_registrations WHERE id = ? AND status = "pending"',
        (reg_id,)
    ).fetchone()
    if not reg:
        return jsonify({'error': 'Registration not found or already reviewed'}), 404

    assigned_session = data.get('session_assigned', '').strip()
    if not assigned_session:
        return jsonify({'error': 'Session must be assigned when approving'}), 400

    # Editors can only approve into their own session
    scoped = _assigned_session()
    if scoped is not None and assigned_session != scoped:
        return jsonify({'error': 'You can only approve registrations for your own session'}), 403

    rtype      = reg['registration_type'] or 'member'
    mid        = _next_member_id(db)
    portal_user_id = None

    if rtype == 'staff':
        # ── Staff/volunteer approval ──────────────────────────────
        staff_role = data.get('staff_role', reg['applicant_role'] or '').strip()
        if not staff_role:
            return jsonify({'error': 'Staff role is required'}), 400

        db.execute('''
            INSERT INTO members
                (member_id, first_name, surname, date_of_birth, address, postcode,
                 status, session, member_type, staff_role, date_registered)
            VALUES (?,?,?,?,?,?,"Active",?,?,?,date("now"))
        ''', (
            mid,
            reg['first_name'], reg['surname'], reg['date_of_birth'],
            reg['address'], reg['postcode'],
            assigned_session,
            'staff',
            staff_role,
        ))
        member_db_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]

        # Store mobile and email as contact 1 (self-contact for staff)
        full_name = f"{reg['first_name']} {reg['surname']}".strip()
        if reg['mobile'] or reg['email']:
            db.execute(
                'INSERT INTO member_contacts'
                ' (member_id, contact_order, contact_name, contact_phone, contact_email)'
                ' VALUES (?,1,?,?,?)',
                (member_db_id, full_name, reg['mobile'] or '', reg['email'] or '')
            )

        # Optionally create a portal login
        create_login  = data.get('create_login', False)
        portal_role   = data.get('portal_role', '').strip()
        username      = data.get('username', '').strip()
        temp_password = data.get('temp_password', '').strip()

        if create_login:
            if not username or not temp_password:
                return jsonify({'error': 'Username and password are required to create a login'}), 400
            pw_error = validate_password(temp_password)
            if pw_error:
                return jsonify({'error': pw_error}), 400
            # Validate portal_role against the roles table
            role_row = db.execute('SELECT id, permissions FROM roles WHERE name = ?', (portal_role,)).fetchone()
            if not role_row:
                return jsonify({'error': 'Invalid portal role'}), 400
            role_perms = json.loads(role_row['permissions'])
            if 'users.create.admin' in role_perms and not has_permission('users.create.admin'):
                return jsonify({'error': 'You do not have permission to assign this role'}), 403
            existing = db.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone()
            if existing:
                return jsonify({'error': f'Username "{username}" is already taken'}), 409
            pw_hash = bcrypt.hashpw(temp_password.encode(), bcrypt.gensalt()).decode()
            cur = db.execute(
                'INSERT INTO users (username, email, password_hash, role, role_id, session_assigned, active)'
                ' VALUES (?,?,?,?,?,?,1)',
                (username, reg['email'] or '', pw_hash, portal_role, role_row['id'], assigned_session)
            )
            portal_user_id = cur.lastrowid
            log_action('create_user', 'users', portal_user_id, {
                'username': username, 'role': portal_role,
                'created_by': session.get('username'),
                'via': 'staff_approval',
            })

    else:
        # ── Standard member approval ──────────────────────────────
        db.execute('''
            INSERT INTO members
                (member_id, first_name, surname, date_of_birth, address, postcode,
                 ethnicity_religion, medical_sen, gp_contact,
                 unattended_exit, gdpr_consent, status, session,
                 member_type, date_registered)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,"Active",?,"member",date("now"))
        ''', (
            mid,
            reg['first_name'], reg['surname'], reg['date_of_birth'],
            reg['address'], reg['postcode'],
            reg['ethnicity_religion'], reg['medical_sen'], reg['gp_contact'],
            reg['unattended_exit'], reg['gdpr_consent'],
            assigned_session,
        ))
        member_db_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]

        # Insert contacts
        for order, (name, phone, email) in enumerate([
            (reg['contact1_name'], reg['contact1_phone'], reg['contact1_email']),
            (reg['contact2_name'], reg['contact2_phone'], reg['contact2_email']),
        ], start=1):
            if name or phone or email:
                db.execute(
                    'INSERT INTO member_contacts'
                    ' (member_id, contact_order, contact_name, contact_phone, contact_email)'
                    ' VALUES (?,?,?,?,?)',
                    (member_db_id, order, name or '', phone or '', email or '')
                )

    # Write custom fields from registration into member_field_values
    custom_raw = reg['custom_fields'] if 'custom_fields' in reg.keys() else None
    if custom_raw:
        try:
            custom_fields = json.loads(custom_raw)
            for key, val in custom_fields.items():
                if val is None or val == '':
                    continue
                fd = db.execute(
                    'SELECT id FROM field_definitions WHERE key = ?', (key,)
                ).fetchone()
                if fd:
                    db.execute(
                        'INSERT OR REPLACE INTO member_field_values'
                        ' (member_id, field_id, value, updated_at)'
                        ' VALUES (?, ?, ?, datetime("now"))',
                        (member_db_id, fd['id'], str(val))
                    )
        except (json.JSONDecodeError, TypeError):
            pass

    # Mark registration as approved
    db.execute(
        'UPDATE pending_registrations'
        ' SET status = "approved", assigned_session = ?, reviewed_by = ?, reviewed_at = datetime("now")'
        ' WHERE id = ?',
        (assigned_session, session['user_id'], reg_id)
    )
    db.commit()

    log_action('approve_registration', 'pending_registrations', reg_id, {
        'new_member_id':   mid,
        'name':            f"{reg['first_name']} {reg['surname']}",
        'type':            rtype,
        'session':         assigned_session,
        'portal_user_id':  portal_user_id,
        'approved_by':     session['username'],
    })
    return jsonify({
        'success':      True,
        'member_id':    mid,
        'user_created': portal_user_id is not None,
    })


@app.route('/api/approvals/<int:reg_id>/reject', methods=['POST'])
@permission_required('approvals.reject')
def api_approvals_reject(reg_id):
    """Reject a pending registration with optional notes."""
    data = request.get_json() or {}
    db   = get_db()

    reg = db.execute(
        'SELECT * FROM pending_registrations WHERE id = ? AND status = "pending"',
        (reg_id,)
    ).fetchone()
    if not reg:
        return jsonify({'error': 'Registration not found or already reviewed'}), 404

    # Editors can only reject registrations in their session (or unassigned ones)
    scoped = _assigned_session()
    if scoped is not None:
        reg_session = reg['assigned_session'] or ''
        if reg_session and reg_session != scoped:
            return jsonify({'error': 'Access denied for this registration'}), 403

    notes = data.get('notes', '').strip()
    db.execute(
        'UPDATE pending_registrations'
        ' SET status = "rejected", notes = ?, reviewed_by = ?, reviewed_at = datetime("now")'
        ' WHERE id = ?',
        (notes, session['user_id'], reg_id)
    )
    db.commit()

    log_action('reject_registration', 'pending_registrations', reg_id, {
        'name': f"{reg['first_name']} {reg['surname']}",
        'notes': notes,
        'rejected_by': session['username'],
    })
    return jsonify({'success': True})


# ── BLUEPRINT: attendance (Phase 3) ───────────────────────────────────────────

def _fetch_tags_for_members(db, member_ids):
    """Return {member_id: [{id, name, icon, colour, category}, ...]} for a list of member IDs."""
    if not member_ids:
        return {}
    placeholders = ','.join('?' * len(member_ids))
    rows = db.execute(f'''
        SELECT  mt.member_id,
                td.id, td.name, td.icon, td.colour, td.category
        FROM    member_tags mt
        JOIN    tag_definitions td ON td.id = mt.tag_id AND td.active = 1
        WHERE   mt.member_id IN ({placeholders})
        ORDER   BY td.sort_order, td.name
    ''', member_ids).fetchall()
    result = {}
    for r in rows:
        result.setdefault(r['member_id'], []).append({
            'id':       r['id'],
            'name':     r['name'],
            'icon':     r['icon'],
            'colour':   r['colour'],
            'category': r['category'],
        })
    return result


@app.route('/api/attendance/<session_type>/<date>')
@login_required
def api_attendance_get(session_type, date):
    """
    Return all active members for a session on a given date,
    annotated with their sign-in/out times if recorded.
    session_type: Tuesday | Thursday
    date: YYYY-MM-DD
    """
    db = get_db()

    # Session-scope for all non-admin roles
    scoped = _assigned_session()
    if scoped is not None:
        if not scoped:
            return jsonify([])
        if scoped != session_type:
            return jsonify({'error': 'Access denied for this session'}), 403

    rows = db.execute('''
        SELECT  m.*,
                a.id         AS att_id,
                a.signed_in_at,
                a.signed_out_at
        FROM    members m
        LEFT JOIN attendance a
               ON  a.member_id   = m.id
               AND a.session_date = ?
               AND a.session_type = ?
        WHERE   m.status      != "Leaver"
          AND   m.member_type  = "member"
          AND   m.session      = ?
        ORDER   BY m.first_name, m.surname
    ''', (date, session_type, session_type)).fetchall()

    member_ids     = [r['id'] for r in rows]
    tags_by_member = _fetch_tags_for_members(db, member_ids)

    # Batch-fetch custom field values
    custom_map = {}
    if member_ids:
        placeholders = ','.join('?' * len(member_ids))
        cfv_rows = db.execute(
            f'SELECT mfv.member_id, fd.key, mfv.value '
            f'FROM member_field_values mfv '
            f'JOIN field_definitions fd ON fd.id = mfv.field_id '
            f'WHERE mfv.member_id IN ({placeholders})',
            member_ids
        ).fetchall()
        for cfv in cfv_rows:
            custom_map.setdefault(cfv['member_id'], {})[cfv['key']] = cfv['value']

    result = []
    for r in rows:
        d = dict(r)
        d['tags']          = tags_by_member.get(r['id'], [])
        d['custom_fields'] = custom_map.get(r['id'], {})
        result.append(d)
    return jsonify(result)


@app.route('/api/attendance/staff/<session_type>/<date>')
@login_required
def api_attendance_staff_get(session_type, date):
    """
    Return all active staff members for a session on a given date,
    annotated with sign-in/out times. Staff are never subject to alert rule evaluation.
    """
    if session_type not in get_valid_session_names():
        return jsonify({'error': 'Invalid session'}), 400

    scoped = _assigned_session()
    if scoped is not None and scoped != session_type:
        return jsonify({'error': 'Access denied for this session'}), 403

    db   = get_db()
    rows = db.execute('''
        SELECT  m.id, m.first_name, m.surname, m.staff_role,
                a.signed_in_at,
                a.signed_out_at
        FROM    members m
        LEFT JOIN attendance a
               ON  a.member_id    = m.id
               AND a.session_date = ?
               AND a.session_type = ?
        WHERE   m.status      != "Leaver"
          AND   m.member_type  = "staff"
          AND   m.session      = ?
        ORDER   BY m.first_name, m.surname
    ''', (date, session_type, session_type)).fetchall()

    member_ids     = [r['id'] for r in rows]
    tags_by_member = _fetch_tags_for_members(db, member_ids)
    result = []
    for r in rows:
        d = dict(r)
        d['tags'] = tags_by_member.get(r['id'], [])
        result.append(d)
    return jsonify(result)


@app.route('/api/attendance/signin', methods=['POST'])
@permission_required('register.signin')
def api_attendance_signin():
    data        = request.get_json() or {}
    member_id   = data.get('member_id')
    sess_type   = data.get('session_type', '').strip()
    sess_date   = data.get('date', '').strip()

    if not all([member_id, sess_type, sess_date]):
        return jsonify({'error': 'member_id, session_type and date are required'}), 400

    # Enforce session scope for non-admin roles
    scoped = _assigned_session()
    if scoped is not None and scoped != sess_type:
        return jsonify({'error': 'Access denied for this session'}), 403

    # Reject if register is locked
    if _is_register_locked(sess_type, sess_date):
        return jsonify({'error': 'This register has been completed and is now locked.'}), 403

    db = get_db()
    # Upsert — insert if not exists, update signed_in_at if already there
    existing = db.execute(
        'SELECT id FROM attendance WHERE member_id = ? AND session_date = ? AND session_type = ?',
        (member_id, sess_date, sess_type)
    ).fetchone()

    now = datetime.now().strftime('%H:%M')
    if existing:
        db.execute(
            'UPDATE attendance SET signed_in_at = ?, recorded_by = ? WHERE id = ?',
            (now, session['user_id'], existing['id'])
        )
    else:
        db.execute(
            'INSERT INTO attendance (member_id, session_date, session_type, signed_in_at, recorded_by)'
            ' VALUES (?,?,?,?,?)',
            (member_id, sess_date, sess_type, now, session['user_id'])
        )
    # Auto-resolve: clear any active attendance-rule flags for this member now they've attended
    att_flags = db.execute(
        "SELECT mf.id FROM member_flags mf "
        "JOIN alert_rules ar ON ar.id = mf.rule_id "
        "WHERE mf.member_id = ? AND mf.resolved_at IS NULL "
        "AND ar.rule_type = 'attendance' AND ar.auto_resolve = 1",
        (member_id,)
    ).fetchall()
    for flag in att_flags:
        db.execute(
            "UPDATE member_flags SET resolved_at = datetime('now'), resolved_by = 'auto' "
            "WHERE id = ?",
            (flag['id'],)
        )
    if att_flags:
        member_row = db.execute('SELECT first_name, surname FROM members WHERE id = ?',
                                (member_id,)).fetchone()
        log_action('auto_resolve_flags', 'members', member_id, {
            'member': f"{member_row['first_name'] or ''} {member_row['surname'] or ''}".strip(),
            'reason': 'attended session — attendance alert flags auto-resolved',
            'count':  len(att_flags),
        })

    db.commit()
    _touch_attendance()
    return jsonify({'success': True, 'signed_in_at': now})


@app.route('/api/attendance/signout', methods=['POST'])
@permission_required('register.signout')
def api_attendance_signout():
    data        = request.get_json() or {}
    member_id   = data.get('member_id')
    sess_type   = data.get('session_type', '').strip()
    sess_date   = data.get('date', '').strip()
    clear       = data.get('clear', False)   # True = undo sign-out (set to NULL)

    if not all([member_id, sess_type, sess_date]):
        return jsonify({'error': 'member_id, session_type and date are required'}), 400

    # Enforce session scope for non-admin roles
    scoped = _assigned_session()
    if scoped is not None and scoped != sess_type:
        return jsonify({'error': 'Access denied for this session'}), 403

    # Reject if register is locked
    if _is_register_locked(sess_type, sess_date):
        return jsonify({'error': 'This register has been completed and is now locked.'}), 403

    db        = get_db()
    out_value = None if clear else datetime.now().strftime('%H:%M')
    result = db.execute(
        'UPDATE attendance SET signed_out_at = ?, recorded_by = ?'
        ' WHERE member_id = ? AND session_date = ? AND session_type = ?',
        (out_value, session['user_id'], member_id, sess_date, sess_type)
    )
    if result.rowcount == 0:
        return jsonify({'error': 'No sign-in record found — member may not be signed in yet'}), 404
    db.commit()
    _touch_attendance()
    return jsonify({'success': True, 'signed_out_at': out_value})


@app.route('/api/attendance/complete/<session_type>/<date>')
@login_required
def api_attendance_complete_status(session_type, date):
    """Return whether the register for this session+date has been completed."""
    db  = get_db()
    row = db.execute(
        '''SELECT sc.completed_at, sc.auto_signout_count,
                  u.username AS completed_by_name
           FROM session_completions sc
           LEFT JOIN users u ON u.id = sc.completed_by
           WHERE sc.session_date = ? AND sc.session_type = ?''',
        (date, session_type)
    ).fetchone()
    if row:
        return jsonify({
            'completed':          True,
            'completed_by':       row['completed_by_name'],
            'completed_at':       row['completed_at'],
            'auto_signout_count': row['auto_signout_count'],
        })
    return jsonify({'completed': False})


@app.route('/api/attendance/complete', methods=['POST'])
@permission_required('register.complete')
def api_attendance_complete():
    """
    Mark a session register as complete (locked).
    Auto signs out any members still signed in.
    Restricted to admin (any session) and editor (own session only).
    """
    data      = request.get_json() or {}
    sess_type = data.get('session_type', '').strip()
    sess_date = data.get('date', '').strip()

    if not all([sess_type, sess_date]):
        return jsonify({'error': 'session_type and date are required'}), 400

    # Editors are scoped to their assigned session
    scoped = _assigned_session()
    if scoped is not None and scoped != sess_type:
        return jsonify({'error': 'You can only complete your own session register'}), 403

    db = get_db()

    # Prevent double-completion — checked again inside the transaction to avoid race
    existing = db.execute(
        'SELECT id FROM session_completions WHERE session_date = ? AND session_type = ?',
        (sess_date, sess_type)
    ).fetchone()
    if existing:
        return jsonify({'error': 'This register has already been completed'}), 409

    # Auto sign out anyone still signed in
    now      = datetime.now().strftime('%H:%M')
    still_in = db.execute(
        '''SELECT id FROM attendance
           WHERE session_date = ? AND session_type = ?
             AND signed_in_at IS NOT NULL AND signed_out_at IS NULL''',
        (sess_date, sess_type)
    ).fetchall()
    auto_count = len(still_in)

    try:
        db.execute('BEGIN IMMEDIATE')   # exclusive write lock — prevents race condition
        if auto_count:
            db.execute(
                '''UPDATE attendance SET signed_out_at = ?, recorded_by = ?
                   WHERE session_date = ? AND session_type = ?
                     AND signed_in_at IS NOT NULL AND signed_out_at IS NULL''',
                (now, session['user_id'], sess_date, sess_type)
            )
        db.execute(
            '''INSERT INTO session_completions
                   (session_date, session_type, completed_by, completed_at, auto_signout_count)
               VALUES (?, ?, ?, datetime('now'), ?)''',
            (sess_date, sess_type, session['user_id'], auto_count)
        )
        db.execute('COMMIT')
    except sqlite3.IntegrityError:
        db.execute('ROLLBACK')
        return jsonify({'error': 'This register was already completed by another user'}), 409
    except Exception as exc:
        db.execute('ROLLBACK')
        return jsonify({'error': f'Could not complete register: {exc}'}), 500

    _touch_attendance()

    # Invalidate any active QR token for this session so the QR code
    # on the display stops working the moment the register is locked.
    _invalidate_qr_token_for_session(sess_type, sess_date)

    # Fetch summary counts for the response
    totals = db.execute(
        '''SELECT
             COUNT(*) AS total,
             SUM(CASE WHEN signed_out_at IS NOT NULL THEN 1 ELSE 0 END) AS signed_out
           FROM attendance
           WHERE session_date = ? AND session_type = ?''',
        (sess_date, sess_type)
    ).fetchone()

    log_action('register_complete', 'session_completions', None, {
        'session_type':      sess_type,
        'session_date':      sess_date,
        'auto_signout_count': auto_count,
    })

    return jsonify({
        'success':           True,
        'auto_signout_count': auto_count,
        'total_members':     totals['total'] if totals else 0,
        'total_signed_out':  totals['signed_out'] if totals else 0,
    })


@app.route('/api/attendance/reset', methods=['POST'])
@permission_required('register.reset')
def api_attendance_reset():
    """
    Admin-only: completely wipe a session register.
    Deletes all attendance rows and the session_completions record for
    the given session_type + date, leaving the register blank and unlocked.
    """
    data      = request.get_json() or {}
    sess_type = data.get('session_type', '').strip()
    sess_date = data.get('date', '').strip()

    if not all([sess_type, sess_date]):
        return jsonify({'error': 'session_type and date are required'}), 400

    db = get_db()

    # Count rows being deleted so we can return a useful summary
    att_count = db.execute(
        'SELECT COUNT(*) AS n FROM attendance WHERE session_date = ? AND session_type = ?',
        (sess_date, sess_type)
    ).fetchone()['n']

    # Wipe attendance records
    db.execute(
        'DELETE FROM attendance WHERE session_date = ? AND session_type = ?',
        (sess_date, sess_type)
    )

    # Unlock the register (remove completion record if one exists)
    db.execute(
        'DELETE FROM session_completions WHERE session_date = ? AND session_type = ?',
        (sess_date, sess_type)
    )

    db.commit()
    _touch_attendance()

    log_action('register_reset', 'attendance', None, {
        'session_type':       sess_type,
        'session_date':       sess_date,
        'attendance_deleted': att_count,
    })

    return jsonify({
        'success':            True,
        'attendance_deleted': att_count,
    })


def get_setting(key, default=None):
    """Return a value from the settings table, or default if not found."""
    db  = get_db()
    row = db.execute('SELECT value FROM settings WHERE key = ?', (key,)).fetchone()
    return row['value'] if row else default


def _touch_attendance():
    """Stamp the current time into settings so the SSE stream can detect changes."""
    db  = get_db()
    now = datetime.now().isoformat()
    db.execute(
        'INSERT INTO settings (key, value) VALUES (?,?) '
        'ON CONFLICT(key) DO UPDATE SET value=excluded.value',
        ('last_attendance_change', now)
    )
    db.commit()


@app.route('/api/display/stream')
def api_display_stream():
    """SSE endpoint — pushes a 'refresh' event whenever attendance changes.

    The generator has a hard 4-hour lifetime. When it expires it sends a
    'timeout' event and exits cleanly, so the thread is released. The JS client
    treats an EventSource close as an error and the browser reconnects
    automatically, starting a fresh 4-hour window.

    A heartbeat comment line is sent every 30 seconds so proxies don't close
    idle connections before they time out naturally.
    """
    _SSE_MAX_SECONDS  = 4 * 3600  # 4 hours — hard lifetime per connection
    _SSE_HEARTBEAT_S  = 30        # keepalive comment interval in seconds
    _SSE_POLL_S       = 1         # how often to check for attendance changes

    def generate():
        last         = None
        deadline     = time.time() + _SSE_MAX_SECONDS
        last_hb      = time.time()
        while time.time() < deadline:
            if time.time() - last_hb >= _SSE_HEARTBEAT_S:
                yield ': heartbeat\n\n'  # SSE comment — ignored by clients, prevents proxy timeout
                last_hb = time.time()
            current = get_setting('last_attendance_change')
            if current != last:
                last = current
                yield 'data: refresh\n\n'
            time.sleep(_SSE_POLL_S)
        yield 'data: timeout\n\n'  # signal client to reconnect (EventSource onerror fires)

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'}
    )


def get_session_types():
    """Return active session types from the DB, cached per request in Flask g."""
    if 'session_types' not in g:
        db   = get_db()
        rows = db.execute(
            'SELECT id, name, weekday FROM session_types WHERE active = 1 ORDER BY sort_order, name'
        ).fetchall()
        g.session_types = [dict(r) for r in rows]
    return g.session_types


def get_valid_session_names():
    """Return a tuple of valid session name strings (for validation checks)."""
    return tuple(s['name'] for s in get_session_types())


def weekday_to_session_map():
    """Return dict mapping Python weekday int -> session name."""
    return {s['weekday']: s['name'] for s in get_session_types()}


def session_to_weekday_map():
    """Return dict mapping session name -> Python weekday int."""
    return {s['name']: s['weekday'] for s in get_session_types()}


def _is_register_locked(sess_type, sess_date):
    """Return True if the register for this session+date has been completed/locked."""
    db = get_db()
    row = db.execute(
        'SELECT id FROM session_completions WHERE session_date = ? AND session_type = ?',
        (sess_date, sess_type)
    ).fetchone()
    return row is not None


def _assigned_session():
    """
    Return the session this user is scoped to, or None for admin (unscoped).
    All non-admin roles MUST have a session_assigned.
    """
    if session.get('role') == ROLE_ADMIN:
        return None
    return session.get('session_assigned') or ''


# NOTE: /api/attendance/check-at-risk and /api/attendance/mark-at-risk removed in v8.0.
# Attendance-based flagging is now handled by the alert_rules engine (rule_type='attendance').


@app.route('/api/attendance/history/<int:member_id>')
@login_required
def api_attendance_history(member_id):
    """Return last 20 attendance records for a member."""
    db = get_db()
    # Enforce session scope — non-admin users can only view history for members in their session
    scoped = _assigned_session()
    if scoped is not None:
        member = db.execute('SELECT session FROM members WHERE id = ?', (member_id,)).fetchone()
        if not member or (member['session'] or '') != scoped:
            return jsonify({'error': 'Forbidden'}), 403

    rows = db.execute(
        'SELECT session_date, session_type, signed_in_at, signed_out_at'
        ' FROM attendance WHERE member_id = ?'
        ' ORDER BY session_date DESC LIMIT 20',
        (member_id,)
    ).fetchall()
    return jsonify([dict(r) for r in rows])


# ── BLUEPRINT: reception display (public, read-only) ──────────────────────────

@app.route('/display')
def display_page():
    """Full-screen reception TV display — no login required."""
    return render_template('display.html',
                           current_session=session.get('session_assigned', ''),
                           session_types=get_session_types(),
                           club_name=CLUB_NAME, club_short_name=CLUB_SHORT_NAME)


@app.route('/api/display/<session_type>')
def api_display(session_type):
    """
    Return first names of members currently signed IN (not yet signed out) for
    today's session, plus on-duty leaders (first name + role) and active activities.

    Intentionally unauthenticated — this endpoint powers a public TV/reception
    display. To limit exposure, only first names are returned for members;
    full names are never included. Leaders show first name and role only.
    """
    if session_type not in get_valid_session_names():
        return jsonify({'error': 'Invalid session'}), 400

    today = datetime.now().strftime('%Y-%m-%d')
    db    = get_db()

    rows = db.execute('''
        SELECT  m.first_name, m.surname,
                a.signed_in_at
        FROM    attendance a
        JOIN    members m ON m.id = a.member_id
        WHERE   a.session_date  = ?
          AND   a.session_type  = ?
          AND   a.signed_in_at  IS NOT NULL
          AND   a.signed_out_at IS NULL
          AND   m.member_type   = "member"
        ORDER   BY a.signed_in_at ASC
    ''', (today, session_type)).fetchall()

    # Staff who have signed in today and not yet signed out — these are "on duty"
    leader_rows = db.execute('''
        SELECT  m.first_name, m.surname, m.staff_role
        FROM    attendance a
        JOIN    members m ON m.id = a.member_id
        WHERE   a.session_date  = ?
          AND   a.session_type  = ?
          AND   a.signed_in_at  IS NOT NULL
          AND   a.signed_out_at IS NULL
          AND   m.member_type   = "staff"
        ORDER   BY a.signed_in_at ASC
    ''', (today, session_type)).fetchall()

    # Active activities for this session
    activity_rows = db.execute('''
        SELECT id, activity
        FROM   session_activities
        WHERE  session_type = ? AND active = 1
        ORDER  BY created_at ASC
    ''', (session_type,)).fetchall()

    return jsonify({
        'session':    session_type,
        'date':       today,
        # First names only — this endpoint is unauthenticated (public display board)
        'members':    [{'first_name': r['first_name'],
                        'signed_in_at': r['signed_in_at']} for r in rows],
        'leaders':    [{'name': r['first_name'] or '', 'role': r['staff_role'] or ''} for r in leader_rows],
        'activities': [{'id': r['id'], 'activity': r['activity']} for r in activity_rows],
    })


@app.route('/api/activities/<session_type>', methods=['GET'])
@login_required
def api_activities_list(session_type):
    """List active activities for a session."""
    if session_type not in get_valid_session_names():
        return jsonify({'error': 'Invalid session'}), 400
    db   = get_db()
    rows = db.execute('''
        SELECT id, activity, created_at
        FROM   session_activities
        WHERE  session_type = ? AND active = 1
        ORDER  BY created_at ASC
    ''', (session_type,)).fetchall()
    return jsonify([{'id': r['id'], 'activity': r['activity'],
                     'created_at': r['created_at']} for r in rows])


@app.route('/api/activities', methods=['POST'])
@permission_required('activities.manage')
def api_activity_add():
    """Add an activity to the display board."""
    data     = request.get_json() or {}
    sess     = data.get('session_type', '').strip()
    activity = data.get('activity', '').strip()
    if sess not in get_valid_session_names():
        return jsonify({'error': 'Invalid session'}), 400
    if not activity:
        return jsonify({'error': 'Activity text is required'}), 400
    if len(activity) > 120:
        return jsonify({'error': 'Activity text is too long (max 120 characters)'}), 400
    db = get_db()
    cur = db.execute(
        'INSERT INTO session_activities (session_type, activity, added_by) VALUES (?,?,?)',
        (sess, activity, session.get('user_id'))
    )
    db.commit()
    return jsonify({'id': cur.lastrowid, 'activity': activity}), 201


@app.route('/api/activities/<int:activity_id>', methods=['PUT'])
@permission_required('activities.manage')
def api_activity_update(activity_id):
    """Edit an activity's text on the display board."""
    db       = get_db()
    row      = db.execute('SELECT id FROM session_activities WHERE id = ? AND active = 1',
                          (activity_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Activity not found'}), 404
    data     = request.get_json() or {}
    activity = data.get('activity', '').strip()
    if not activity:
        return jsonify({'error': 'Activity text is required'}), 400
    if len(activity) > 120:
        return jsonify({'error': 'Activity text is too long (max 120 characters)'}), 400
    db.execute('UPDATE session_activities SET activity = ? WHERE id = ?', (activity, activity_id))
    db.commit()
    return jsonify({'id': activity_id, 'activity': activity})


@app.route('/api/activities/<int:activity_id>', methods=['DELETE'])
@permission_required('activities.manage')
def api_activity_delete(activity_id):
    """Remove an activity from the display board."""
    db = get_db()
    db.execute('UPDATE session_activities SET active = 0 WHERE id = ?', (activity_id,))
    db.commit()
    return jsonify({'ok': True})


# ── BLUEPRINT: calendar (Phase 5) ────────────────────────────────────────────

VALID_STATUSES = ('planned', 'cancelled', 'special')


@app.route('/calendar')
@login_required
def calendar_page():
    return render_template('calendar.html', **tpl_ctx(), active_page='calendar')


@app.route('/api/calendar', methods=['GET'])
@login_required
def api_calendar_list():
    """Return term sessions, optionally filtered by year/month or term."""
    db    = get_db()
    year  = request.args.get('year')
    month = request.args.get('month')
    term  = request.args.get('term')

    query  = 'SELECT * FROM term_sessions WHERE 1=1'
    params = []

    if year and month:
        prefix = f"{int(year):04d}-{int(month):02d}-"
        query += ' AND session_date LIKE ?'
        params.append(prefix + '%')
    elif term:
        query += ' AND term_name = ?'
        params.append(term)

    query += ' ORDER BY session_date ASC, session_type ASC'
    rows   = db.execute(query, params).fetchall()

    return jsonify([dict(r) for r in rows])


@app.route('/api/calendar/terms', methods=['GET'])
@login_required
def api_calendar_terms():
    """Return distinct term names present in the database."""
    db   = get_db()
    rows = db.execute(
        "SELECT DISTINCT term_name FROM term_sessions WHERE term_name IS NOT NULL ORDER BY term_name"
    ).fetchall()
    return jsonify([r['term_name'] for r in rows])


@app.route('/api/calendar', methods=['POST'])
@permission_required('calendar.create')
def api_calendar_add():
    """Add a single session to the calendar."""
    data         = request.get_json() or {}
    session_date = data.get('session_date', '').strip()
    session_type = data.get('session_type', '').strip()
    status       = data.get('status', 'planned').strip()
    notes        = data.get('notes', '').strip() or None
    term_name    = data.get('term_name', '').strip() or None

    if not session_date:
        return jsonify({'error': 'Session date is required'}), 400
    if session_type not in get_valid_session_names():
        return jsonify({'error': 'Invalid session type'}), 400

    # Editors can only add sessions for their own session type
    scoped = _assigned_session()
    if scoped is not None and session_type != scoped:
        return jsonify({'error': f'You can only add {scoped} sessions'}), 403
    if status not in VALID_STATUSES:
        return jsonify({'error': 'Invalid status'}), 400

    # Validate date and check day-of-week matches session type
    try:
        from datetime import date as dt_date
        d            = dt_date.fromisoformat(session_date)
        wday_map     = session_to_weekday_map()
        expected_day = wday_map.get(session_type)
        if expected_day is None:
            return jsonify({'error': 'Unknown session type'}), 400
        if d.weekday() != expected_day:
            return jsonify({'error': f'{session_date} is not a {session_type}'}), 400
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400

    db = get_db()
    try:
        cur = db.execute(
            '''INSERT INTO term_sessions (session_date, session_type, term_name, status, notes, created_by)
               VALUES (?,?,?,?,?,?)''',
            (session_date, session_type, term_name, status, notes, session.get('user_id'))
        )
        db.commit()
        log_action('add_term_session', 'term_sessions', cur.lastrowid,
                   {'date': session_date, 'type': session_type, 'term': term_name})
        row = db.execute('SELECT * FROM term_sessions WHERE id = ?', (cur.lastrowid,)).fetchone()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A {session_type} session on {session_date} already exists'}), 409


@app.route('/api/calendar/bulk', methods=['POST'])
@permission_required('calendar.create')
def api_calendar_bulk():
    """
    Bulk-generate sessions between two dates.
    Body: { start_date, end_date, days ['Tuesday','Thursday'], term_name, exclude_dates[] }
    """
    from datetime import date as dt_date, timedelta as dt_timedelta

    data          = request.get_json() or {}
    start_str     = data.get('start_date', '').strip()
    end_str       = data.get('end_date', '').strip()
    days          = data.get('days', [])
    term_name     = data.get('term_name', '').strip() or None
    exclude_dates = set(data.get('exclude_dates', []))

    if not start_str or not end_str:
        return jsonify({'error': 'start_date and end_date are required'}), 400
    if not days:
        return jsonify({'error': 'At least one day must be selected'}), 400

    # Editors can only bulk-generate their own session type
    scoped = _assigned_session()
    if scoped is not None:
        if any(d != scoped for d in days):
            return jsonify({'error': f'You can only generate {scoped} sessions'}), 403
        days = [scoped]   # enforce — discard any other values

    try:
        start = dt_date.fromisoformat(start_str)
        end   = dt_date.fromisoformat(end_str)
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400

    if end < start:
        return jsonify({'error': 'end_date must be on or after start_date'}), 400
    if (end - start).days > 365:
        return jsonify({'error': 'Date range cannot exceed 365 days'}), 400

    # Map day names to Python weekday numbers using DB-driven session types
    s2w = session_to_weekday_map()
    # weekday -> session_name for days that were requested
    target_weekdays = {}
    for d in days:
        if d in s2w:
            target_weekdays[s2w[d]] = d

    # Walk the date range
    created, skipped = 0, 0
    db = get_db()
    current = start
    while current <= end:
        if current.weekday() in target_weekdays:
            date_str     = current.isoformat()
            session_type = target_weekdays[current.weekday()]
            if date_str not in exclude_dates:
                try:
                    db.execute(
                        '''INSERT INTO term_sessions
                               (session_date, session_type, term_name, status, created_by)
                           VALUES (?,?,?,?,?)''',
                        (date_str, session_type, term_name, 'planned', session.get('user_id'))
                    )
                    created += 1
                except sqlite3.IntegrityError:
                    skipped += 1  # already exists
        current += dt_timedelta(days=1)

    db.commit()
    log_action('bulk_term_sessions', 'term_sessions', None,
               {'term': term_name, 'created': created, 'skipped': skipped})
    return jsonify({'created': created, 'skipped': skipped})


@app.route('/api/calendar/<int:session_id>', methods=['PUT'])
@permission_required('calendar.edit')
def api_calendar_update(session_id):
    """Update status or notes on a session."""
    data   = request.get_json() or {}
    status = data.get('status')
    notes  = data.get('notes')
    term   = data.get('term_name')
    db     = get_db()

    row = db.execute('SELECT * FROM term_sessions WHERE id = ?', (session_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404

    # Editors can only update sessions for their own session type
    scoped = _assigned_session()
    if scoped is not None and row['session_type'] != scoped:
        return jsonify({'error': 'You can only edit your own session entries'}), 403

    updates, params = [], []
    if status is not None:
        if status not in VALID_STATUSES:
            return jsonify({'error': 'Invalid status'}), 400
        updates.append('status = ?'); params.append(status)
    if notes is not None:
        updates.append('notes = ?'); params.append(notes.strip() or None)
    if term is not None:
        updates.append('term_name = ?'); params.append(term.strip() or None)

    if updates:
        params.append(session_id)
        db.execute(f"UPDATE term_sessions SET {', '.join(updates)} WHERE id = ?", params)
        db.commit()
        log_action('update_term_session', 'term_sessions', session_id,
                   {'status': status, 'notes': notes})

    return jsonify(dict(db.execute('SELECT * FROM term_sessions WHERE id = ?', (session_id,)).fetchone()))


@app.route('/api/calendar/<int:session_id>', methods=['DELETE'])
@permission_required('calendar.delete')
def api_calendar_delete(session_id):
    """Delete a session from the calendar."""
    db  = get_db()
    row = db.execute('SELECT * FROM term_sessions WHERE id = ?', (session_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404

    # Editors can only delete sessions for their own session type
    scoped = _assigned_session()
    if scoped is not None and row['session_type'] != scoped:
        return jsonify({'error': 'You can only delete your own session entries'}), 403

    db.execute('DELETE FROM term_sessions WHERE id = ?', (session_id,))
    db.commit()
    log_action('delete_term_session', 'term_sessions', session_id,
               {'date': row['session_date'], 'type': row['session_type']})
    return jsonify({'ok': True})


@app.route('/api/calendar/upcoming', methods=['GET'])
@login_required
def api_calendar_upcoming():
    """Return the next N planned sessions from today (used by dashboard)."""
    limit = min(int(request.args.get('limit', 6)), 20)
    today = datetime.now().strftime('%Y-%m-%d')
    db    = get_db()
    rows  = db.execute('''
        SELECT * FROM term_sessions
        WHERE  session_date >= ? AND status != 'cancelled'
        ORDER  BY session_date ASC, session_type ASC
        LIMIT  ?
    ''', (today, limit)).fetchall()
    return jsonify([dict(r) for r in rows])


# ── Page routes for Phase 2 & 3 ───────────────────────────────────────────────

# (approvals_page and register_page already defined above — no change needed)

# ── BLUEPRINT: documents ──────────────────────────────────────────────────────

CATEGORY_LABELS = ('policy', 'template', 'form', 'general', 'registers')
# Numeric rank used to gate document access.
# readonly=0, leader=1, editor=2, admin=3
# When uploading a doc, set access_role to the minimum rank required.
ROLE_RANK = {'readonly': 0, 'leader': 1, 'editor': 2, 'admin': 3}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def user_can_access_doc(doc):
    """Return True if the current session role meets the document's access_role."""
    user_rank = ROLE_RANK.get(session.get('role', 'readonly'), 0)
    req_rank  = ROLE_RANK.get(doc['access_role'] or 'readonly', 0)
    return user_rank >= req_rank


@app.route('/api/documents')
@permission_required('documents.view')
def api_documents_list():
    db   = get_db()
    rows = db.execute('''
        SELECT d.*, u.username AS uploaded_by_name
        FROM   documents d
        LEFT JOIN users u ON u.id = d.uploaded_by
        WHERE  d.active = 1
        ORDER  BY d.category, d.title
    ''').fetchall()
    # Filter by access_role
    docs = [dict(r) for r in rows if user_can_access_doc(r)]
    return jsonify(docs)


@app.route('/api/documents', methods=['POST'])
@permission_required('documents.upload')
def api_documents_upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'No file selected'}), 400
    if not allowed_file(f.filename):
        return jsonify({'error': 'File type not allowed'}), 400

    title       = request.form.get('title', '').strip() or f.filename
    category    = request.form.get('category', 'general')
    access_role = request.form.get('access_role', 'readonly')

    if category    not in CATEGORY_LABELS:  category    = 'general'
    if access_role not in ROLE_RANK:        access_role = 'readonly'

    safe_name = secure_filename(f.filename)
    # Prefix with timestamp to avoid collisions
    stored_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{safe_name}"
    encrypted_data = encrypt_file(f.read())
    with open(os.path.join(UPLOAD_DIR, stored_name), 'wb') as fh:
        fh.write(encrypted_data)

    mime = f.mimetype or 'application/octet-stream'
    db   = get_db()
    cur  = db.execute(
        'INSERT INTO documents (title, filename, file_path, mime_type, category, access_role, uploaded_by)'
        ' VALUES (?,?,?,?,?,?,?)',
        (title, safe_name, stored_name, mime, category, access_role, session['user_id'])
    )
    db.commit()
    log_action('upload_document', 'documents', cur.lastrowid, {'title': title, 'category': category})
    return jsonify({'success': True})


@app.route('/api/documents/<int:doc_id>/download')
@login_required
def api_documents_download(doc_id):
    db  = get_db()
    doc = db.execute('SELECT * FROM documents WHERE id = ? AND active = 1', (doc_id,)).fetchone()
    if not doc:
        return jsonify({'error': 'Not found'}), 404
    if not user_can_access_doc(doc):
        return jsonify({'error': 'Forbidden'}), 403
    log_action('download_document', 'documents', doc_id, {'title': doc['title']})
    with open(os.path.join(UPLOAD_DIR, doc['file_path']), 'rb') as fh:
        decrypted = decrypt_file(fh.read())
    return app.response_class(
        decrypted,
        mimetype=doc['mime_type'] or 'application/octet-stream',
        headers={'Content-Disposition': f'attachment; filename="{doc["filename"]}"'},
    )


@app.route('/api/documents/<int:doc_id>/view')
@login_required
def api_documents_view(doc_id):
    """Serve the document inline so the browser can render it directly."""
    db  = get_db()
    doc = db.execute('SELECT * FROM documents WHERE id = ? AND active = 1', (doc_id,)).fetchone()
    if not doc:
        return jsonify({'error': 'Not found'}), 404
    if not user_can_access_doc(doc):
        return jsonify({'error': 'Forbidden'}), 403
    log_action('view_document', 'documents', doc_id, {'title': doc['title']})
    with open(os.path.join(UPLOAD_DIR, doc['file_path']), 'rb') as fh:
        decrypted = decrypt_file(fh.read())
    return app.response_class(
        decrypted,
        mimetype=doc['mime_type'] or 'application/octet-stream',
        headers={'Content-Disposition': f'inline; filename="{doc["filename"]}"'},
    )


@app.route('/api/documents/<int:doc_id>', methods=['DELETE'])
@permission_required('documents.delete')
def api_documents_delete(doc_id):
    db  = get_db()
    doc = db.execute('SELECT * FROM documents WHERE id = ? AND active = 1', (doc_id,)).fetchone()
    if not doc:
        return jsonify({'error': 'Not found'}), 404
    db.execute('UPDATE documents SET active = 0 WHERE id = ?', (doc_id,))
    db.commit()
    # Remove the encrypted file from disk — no point keeping it once deleted from the repo
    file_path = os.path.join(UPLOAD_DIR, doc['file_path'])
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except OSError:
        pass  # File already gone — not a reason to fail the request
    log_action('delete_document', 'documents', doc_id, {'title': doc['title']})
    return jsonify({'success': True})


# ── BLUEPRINT: email templates ────────────────────────────────────────────────

@app.route('/api/email-templates')
@permission_required('mailshots.templates')
def api_email_templates_list():
    db   = get_db()
    rows = db.execute(
        'SELECT et.*, u.username AS created_by_name FROM email_templates et'
        ' LEFT JOIN users u ON u.id = et.created_by ORDER BY et.name'
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/email-templates', methods=['POST'])
@permission_required('mailshots.templates')
def api_email_templates_create():
    data    = request.get_json() or {}
    name    = data.get('name', '').strip()
    subject = data.get('subject', '').strip()
    body    = data.get('body_html', '').strip()
    if not name or not subject or not body:
        return jsonify({'error': 'Name, subject and body are required'}), 400
    db = get_db()
    db.execute(
        'INSERT INTO email_templates (name, subject, body_html, created_by) VALUES (?,?,?,?)',
        (name, subject, body, session['user_id'])
    )
    db.commit()
    log_action('create_email_template', 'email_templates', None, {'name': name})
    return jsonify({'success': True})


@app.route('/api/email-templates/<int:tmpl_id>', methods=['PUT'])
@permission_required('mailshots.templates')
def api_email_templates_update(tmpl_id):
    data    = request.get_json() or {}
    name    = data.get('name', '').strip()
    subject = data.get('subject', '').strip()
    body    = data.get('body_html', '').strip()
    if not name or not subject or not body:
        return jsonify({'error': 'Name, subject and body are required'}), 400
    db = get_db()
    db.execute(
        'UPDATE email_templates SET name=?, subject=?, body_html=?, updated_at=datetime("now")'
        ' WHERE id=?',
        (name, subject, body, tmpl_id)
    )
    db.commit()
    log_action('edit_email_template', 'email_templates', tmpl_id, {'name': name})
    return jsonify({'success': True})


@app.route('/api/email-templates/<int:tmpl_id>', methods=['DELETE'])
@permission_required('mailshots.templates')
def api_email_templates_delete(tmpl_id):
    db   = get_db()
    tmpl = db.execute('SELECT name FROM email_templates WHERE id = ?', (tmpl_id,)).fetchone()
    db.execute('DELETE FROM email_templates WHERE id = ?', (tmpl_id,))
    db.commit()
    log_action('delete_email_template', 'email_templates', tmpl_id,
               {'name': tmpl['name']} if tmpl else None)
    return jsonify({'success': True})


# ── BLUEPRINT: mailshots ───────────────────────────────────────────────────────

def _get_recipients(session_filter, status_filter):
    """
    Return a deduplicated list of (email, member_name) tuples from
    member_contacts (contact_order=1) matching the given filters.
    Editors are automatically scoped to their own session.
    """
    db = get_db()
    conditions = ["m.status != 'Leaver'"]
    params     = []

    if status_filter and status_filter != 'all':
        conditions.append('m.status = ?')
        params.append(status_filter)

    # Editors are scoped to their session; admins can filter freely
    scoped = _assigned_session()
    if scoped is not None:
        conditions.append('m.session = ?')
        params.append(scoped)
    elif session_filter and session_filter != 'all':
        conditions.append('m.session = ?')
        params.append(session_filter)

    where = ' AND '.join(conditions)
    rows  = db.execute(f'''
        SELECT  DISTINCT c.contact_email,
                m.first_name || " " || m.surname AS member_name
        FROM    members m
        JOIN    member_contacts c ON c.member_id = m.id AND c.contact_order = 1
        WHERE   {where}
          AND   c.contact_email IS NOT NULL
          AND   trim(c.contact_email) != ""
        ORDER   BY m.first_name
    ''', params).fetchall()
    return [{'email': r['contact_email'], 'name': r['member_name']} for r in rows]


@app.route('/api/mailshots/preview', methods=['POST'])
@permission_required('mailshots.send')
def api_mailshots_preview():
    """Return how many unique recipient emails a mailshot would reach."""
    data       = request.get_json() or {}
    recipients = _get_recipients(data.get('session_filter'), data.get('status_filter'))
    return jsonify({'count': len(recipients), 'recipients': recipients})


@app.route('/api/mailshots/send', methods=['POST'])
@permission_required('mailshots.send')
def api_mailshots_send():
    """Send a mailshot via Gmail SMTP and log it."""
    data           = request.get_json() or {}
    subject        = data.get('subject', '').strip()
    body           = data.get('body_html', '').strip()
    template_id    = data.get('template_id')
    explicit_recip = data.get('recipients')        # list of {email, name} from frontend checklist
    document_ids   = data.get('document_ids', [])  # list of document IDs to attach

    if not subject or not body:
        return jsonify({'error': 'Subject and body are required'}), 400
    if not SMTP_USER or not SMTP_PASS:
        return jsonify({'error': 'Email not configured — add MAIL_USERNAME and MAIL_PASSWORD to your .env file'}), 503

    # Use explicit selection from the frontend checklist; fall back to filter query
    if explicit_recip and isinstance(explicit_recip, list):
        recipients = [r for r in explicit_recip if r.get('email')]
    else:
        session_filter = data.get('session_filter', 'all')
        recipients = _get_recipients(session_filter, 'Active')

    if not recipients:
        return jsonify({'error': 'No recipients selected'}), 400

    # Resolve and validate attachments from the document repository
    db          = get_db()
    attachments = []   # list of {filename, mime_type, data (bytes)}
    if document_ids:
        for doc_id in document_ids:
            doc = db.execute(
                'SELECT * FROM documents WHERE id = ? AND active = 1', (doc_id,)
            ).fetchone()
            if not doc:
                return jsonify({'error': f'Document ID {doc_id} not found in repository'}), 400
            if not user_can_access_doc(doc):
                return jsonify({'error': f'Access denied to document: {doc["title"]}'}), 403
            file_path = os.path.join(UPLOAD_DIR, doc['file_path'])
            if not os.path.exists(file_path):
                return jsonify({'error': f'File not found on server for: {doc["title"]}'}), 500
            with open(file_path, 'rb') as f:
                attachments.append({
                    'filename':  doc['filename'],
                    'mime_type': doc['mime_type'] or 'application/octet-stream',
                    'data':      decrypt_file(f.read()),
                })
            log_action('attach_to_mailshot', 'documents', doc_id, {'title': doc['title'], 'subject': subject})

    emails_sent = 0
    errors      = []
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as srv:
            srv.ehlo()
            srv.starttls()
            srv.login(SMTP_USER, SMTP_PASS)
            for r in recipients:
                try:
                    # Use 'mixed' when there are attachments, 'alternative' for body-only
                    msg = MIMEMultipart('mixed') if attachments else MIMEMultipart('alternative')
                    msg['Subject'] = subject
                    msg['From']    = SMTP_FROM
                    msg['To']      = r['email']
                    msg.attach(MIMEText(body, 'html', 'utf-8'))

                    for att in attachments:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(att['data'])
                        encoders.encode_base64(part)
                        part.add_header(
                            'Content-Disposition',
                            'attachment',
                            filename=att['filename'],
                        )
                        part.set_type(att['mime_type'])
                        msg.attach(part)

                    srv.sendmail(SMTP_FROM, [r['email']], msg.as_string())
                    emails_sent += 1
                except Exception as e:
                    errors.append({'email': r['email'], 'error': str(e)})
    except smtplib.SMTPAuthenticationError:
        return jsonify({'error': 'Gmail authentication failed — check your App Password in .env'}), 503
    except Exception as e:
        app.logger.error(f'Mailshot SMTP error: {e}')
        return jsonify({'error': 'Failed to connect to the mail server. Check SMTP settings in .env.'}), 503

    # Log mailshot — store document IDs in filter_criteria for audit trail
    log_meta = {
        'recipients':       len(recipients),
        'manual_selection': bool(explicit_recip),
        'document_ids':     list(document_ids) if document_ids else [],
    }
    db.execute(
        'INSERT INTO mailshot_log (template_id, subject, sent_by, recipient_count, filter_criteria, notes)'
        ' VALUES (?,?,?,?,?,?)',
        (
            template_id,
            subject,
            session['user_id'],
            emails_sent,
            json.dumps(log_meta),
            f'{len(errors)} error(s)' if errors else None,
        )
    )
    db.commit()
    log_action('send_mailshot', 'mailshot_log', None, {
        'subject': subject, 'sent': emails_sent, 'errors': len(errors),
        'attachments': len(attachments),
    })
    return jsonify({'success': True, 'sent': emails_sent, 'errors': errors})


@app.route('/api/mailshots')
@permission_required('mailshots.send')
def api_mailshots_history():
    db   = get_db()
    rows = db.execute('''
        SELECT  ml.*, u.username AS sent_by_name,
                et.name AS template_name
        FROM    mailshot_log ml
        LEFT JOIN users u ON u.id = ml.sent_by
        LEFT JOIN email_templates et ON et.id = ml.template_id
        ORDER   BY ml.sent_at DESC
        LIMIT   50
    ''').fetchall()
    return jsonify([dict(r) for r in rows])


# ── BLUEPRINT: maintenance (admin only) ───────────────────────────────────────

@app.route('/api/admin/maintenance/counts')
@permission_required('admin.maintenance')
def api_maintenance_counts():
    """Return record counts for each clearable data category."""
    db = get_db()
    return jsonify({
        'audit_log':      db.execute('SELECT COUNT(*) FROM audit_log').fetchone()[0],
        'attendance':     db.execute('SELECT COUNT(*) FROM attendance').fetchone()[0],
        'mailshot_log':   db.execute('SELECT COUNT(*) FROM mailshot_log').fetchone()[0],
        'registrations':  db.execute('SELECT COUNT(*) FROM pending_registrations').fetchone()[0],
    })


@app.route('/api/admin/maintenance/audit-log', methods=['DELETE'])
@permission_required('admin.maintenance')
def api_maintenance_clear_audit():
    db = get_db()
    n  = db.execute('SELECT COUNT(*) FROM audit_log').fetchone()[0]
    db.execute('DELETE FROM audit_log')
    db.commit()
    # Write a single fresh entry so the log is never completely empty
    log_action('maintenance_clear', 'audit_log', None,
               {'cleared': n, 'by': session['username']})
    return jsonify({'success': True, 'deleted': n})


@app.route('/api/admin/maintenance/attendance', methods=['DELETE'])
@permission_required('admin.maintenance')
def api_maintenance_clear_attendance():
    db = get_db()
    n  = db.execute('SELECT COUNT(*) FROM attendance').fetchone()[0]
    db.execute('DELETE FROM attendance')
    db.commit()
    log_action('maintenance_clear', 'attendance', None,
               {'cleared': n, 'by': session['username']})
    return jsonify({'success': True, 'deleted': n})


@app.route('/api/admin/maintenance/mailshot-log', methods=['DELETE'])
@permission_required('admin.maintenance')
def api_maintenance_clear_mailshots():
    db = get_db()
    n  = db.execute('SELECT COUNT(*) FROM mailshot_log').fetchone()[0]
    db.execute('DELETE FROM mailshot_log')
    db.commit()
    log_action('maintenance_clear', 'mailshot_log', None,
               {'cleared': n, 'by': session['username']})
    return jsonify({'success': True, 'deleted': n})


@app.route('/api/admin/maintenance/registrations', methods=['DELETE'])
@permission_required('admin.maintenance')
def api_maintenance_clear_registrations():
    db = get_db()
    n  = db.execute('SELECT COUNT(*) FROM pending_registrations').fetchone()[0]
    db.execute('DELETE FROM pending_registrations')
    db.commit()
    log_action('maintenance_clear', 'pending_registrations', None,
               {'cleared': n, 'by': session['username']})
    return jsonify({'success': True, 'deleted': n})


@app.route('/api/admin/maintenance/backup')
@permission_required('admin.maintenance')
def api_maintenance_backup():
    """Stream a hot backup of the SQLCipher database as a downloadable file.

    Uses SQLite's online backup API so the backup is consistent even while the
    app is serving requests.  The resulting file is itself an SQLCipher-encrypted
    database — the same DB_ENCRYPTION_KEY is required to open it.
    """
    import io
    import tempfile

    timestamp  = datetime.now().strftime('%Y%m%d_%H%M%S')
    slug       = CLUB_SHORT_NAME.lower().replace(' ', '_')
    filename   = f'{slug}_backup_{timestamp}.db'

    # Write a hot backup to a temporary file, then stream it.
    # We use a temp file rather than an in-memory buffer because SQLite's
    # backup API requires a real file path as the destination.
    with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp:
        tmp_path = tmp.name

    try:
        src  = _connect_db()          # source: live encrypted DB
        dest = sqlite3.connect(tmp_path)
        dest.execute(f"PRAGMA key='{os.environ.get('DB_ENCRYPTION_KEY', '')}'")
        src.backup(dest)              # online backup — consistent snapshot
        dest.close()
        src.close()

        with open(tmp_path, 'rb') as f:
            data = f.read()
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    log_action('maintenance_backup', 'database', None,
               {'filename': filename, 'size_bytes': len(data), 'by': session['username']})

    return Response(
        data,
        mimetype='application/octet-stream',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Content-Length':      str(len(data)),
        },
    )


@app.route('/api/admin/maintenance/restore', methods=['POST'])
@permission_required('admin.maintenance')
def api_maintenance_restore():
    """Restore the database from an uploaded backup file.

    Safety steps (in order):
      1. Validate the upload is a real SQLCipher database openable with the
         current DB_ENCRYPTION_KEY — rejects corrupt files or wrong-key backups.
      2. Collect a record-count summary from the uploaded DB to return to the UI.
      3. Write an automatic pre-restore snapshot to data/backups/ so the
         current data is never lost.
      4. Atomically replace the live database with the uploaded file.
      5. Close the current request-scoped DB handle so Flask opens a fresh
         connection on the next request.
    """
    import shutil
    import tempfile

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded.'}), 400

    upload = request.files['file']
    if not upload.filename:
        return jsonify({'error': 'Empty filename.'}), 400

    # Save the upload to a temp file for validation
    with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp:
        tmp_path = tmp.name
        upload.save(tmp_path)

    try:
        # ── Step 1: validate the uploaded file ────────────────────────────────
        try:
            check = sqlite3.connect(tmp_path)
            check.execute(f"PRAGMA key='{os.environ.get('DB_ENCRYPTION_KEY', '')}'")
            check.execute('SELECT count(*) FROM sqlite_master')
        except Exception:
            return jsonify({
                'error': 'The uploaded file could not be opened. It may be corrupt '
                         'or was created with a different encryption key.'
            }), 400
        finally:
            try:
                check.close()
            except Exception:
                pass

        # ── Step 2: collect summary from uploaded DB ──────────────────────────
        try:
            info_conn = sqlite3.connect(tmp_path)
            info_conn.execute(f"PRAGMA key='{os.environ.get('DB_ENCRYPTION_KEY', '')}'")
            def _count(tbl):
                try:
                    return info_conn.execute(f'SELECT COUNT(*) FROM {tbl}').fetchone()[0]
                except Exception:
                    return None
            summary = {
                'members':     _count('members'),
                'users':       _count('users'),
                'audit_log':   _count('audit_log'),
                'attendance':  _count('attendance'),
            }
            info_conn.close()
        except Exception:
            summary = {}

        # ── Step 3: auto-snapshot current DB before overwriting ───────────────
        backups_dir = os.path.join(INSTANCE_DIR, 'data', 'backups')
        os.makedirs(backups_dir, exist_ok=True)
        snapshot_ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
        snapshot_path = os.path.join(backups_dir, f'pre_restore_{snapshot_ts}.db')
        shutil.copy2(DATABASE, snapshot_path)

        # ── Step 4: atomic swap ───────────────────────────────────────────────
        shutil.copy2(tmp_path, DATABASE)

        # ── Step 5: drop the stale request-scoped connection ─────────────────
        if hasattr(g, 'db'):
            try:
                g.db.close()
            except Exception:
                pass
            g.db = None

    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    log_action('maintenance_restore', 'database', None, {
        'snapshot': snapshot_path,
        'summary':  summary,
        'by':       session['username'],
    })

    return jsonify({
        'success':  True,
        'summary':  summary,
        'snapshot': os.path.basename(snapshot_path),
    })


# ── BLUEPRINT: data import (v8.4) ────────────────────────────────────────────
import csv as _csv_mod
import uuid as _uuid_mod

# Core members-table columns available for import mapping.
# 'email' and 'member_id' are special cases handled separately in api_import_run.
_IMPORT_CORE_FIELDS = [
    {'key': 'member_id',       'label': 'Member ID (existing)', 'field_type': 'text',     'required': False},
    {'key': 'first_name',      'label': 'First Name',           'field_type': 'text',     'required': True},
    {'key': 'surname',         'label': 'Surname',              'field_type': 'text',     'required': True},
    {'key': 'email',           'label': 'Email Address',        'field_type': 'email',    'required': False},
    {'key': 'postcode',        'label': 'Postcode',             'field_type': 'text',     'required': False},
    {'key': 'address',         'label': 'Address',              'field_type': 'text',     'required': False},
    {'key': 'date_of_birth',   'label': 'Date of Birth',        'field_type': 'date',     'required': False},
    {'key': 'date_registered', 'label': 'Date Registered',      'field_type': 'date',     'required': False},
    {'key': 'comments',        'label': 'Notes / Comments',     'field_type': 'textarea', 'required': False},
    {'key': 'status',          'label': 'Status',               'field_type': 'select',   'required': False,
     'options': 'Active,Inactive,Leaver'},
]

def _fmt_cell(v):
    """Normalise a spreadsheet cell value to a plain string."""
    import datetime as _dt
    if v is None:
        return ''
    if isinstance(v, _dt.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, _dt.date):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, _dt.time):
        return ''  # time-only cells are not useful for member data
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else str(v)
    return str(v).strip()

def _read_xlsx_file(path, sheet_name=None):
    """Return (sheet_names, active_sheet, headers, data_rows) from an xlsx or xls file."""
    ext = path.rsplit('.', 1)[-1].lower()

    if ext == 'xls':
        import xlrd
        wb          = xlrd.open_workbook(path)
        sheet_names = wb.sheet_names()
        ws          = wb.sheet_by_name(sheet_name) if (sheet_name and sheet_name in sheet_names) \
                      else wb.sheet_by_index(0)
        active      = ws.name
        rows        = [ws.row_values(i) for i in range(ws.nrows)]
    else:
        import openpyxl
        wb          = openpyxl.load_workbook(path, data_only=True)
        sheet_names = wb.sheetnames
        ws          = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
        active      = ws.title
        rows        = list(ws.iter_rows(values_only=True))

    if not rows:
        return sheet_names, active, [], []

    # Headers — trim trailing empty/None columns
    raw_headers = [str(c).strip() if c is not None else '' for c in rows[0]]
    last_col = len(raw_headers)
    while last_col > 0 and not raw_headers[last_col - 1]:
        last_col -= 1
    headers = raw_headers[:last_col]

    # Data rows — skip entirely blank rows
    data_rows = []
    for row in rows[1:]:
        vals = [_fmt_cell(v) for v in row[:last_col]]
        if any(vals):
            data_rows.append(vals)

    return sheet_names, active, headers, data_rows

def _read_csv_file(path):
    """Return (headers, data_rows) from a CSV file."""
    with open(path, 'r', encoding='utf-8-sig', errors='replace') as fh:
        reader = list(_csv_mod.reader(fh))
    if not reader:
        return [], []
    headers   = [h.strip() for h in reader[0]]
    data_rows = [r for r in reader[1:] if any(c.strip() for c in r)]
    return headers, data_rows


@app.route('/admin/settings/import')
@permission_required('admin.maintenance')
def import_page():
    return render_template('admin/import.html', active_page='settings', **tpl_ctx())


@app.route('/api/admin/import/analyse', methods=['POST'])
@csrf.exempt
@permission_required('admin.maintenance')
def api_import_analyse():
    """Upload a file and return sheet names, column headers, and preview rows.

    Form fields:
        file        — multipart .xlsx/.csv upload
        sheet_name  — (optional) which xlsx sheet to use; defaults to first sheet
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'No file selected'}), 400

    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    if ext not in ('xlsx', 'xls', 'csv'):
        return jsonify({'error': 'Only .xlsx, .xls and .csv files are supported'}), 400

    imports_dir = os.path.join(INSTANCE_DIR, 'data', 'imports')
    os.makedirs(imports_dir, exist_ok=True)
    file_id   = str(_uuid_mod.uuid4())
    save_path = os.path.join(imports_dir, f'{file_id}.{ext}')
    f.save(save_path)

    sheet_name = (request.form.get('sheet_name') or '').strip() or None

    try:
        if ext in ('xlsx', 'xls'):
            sheet_names, active_sheet, headers, data_rows = _read_xlsx_file(save_path, sheet_name)
        else:
            sheet_names, active_sheet = [], ''
            headers, data_rows = _read_csv_file(save_path)
    except Exception as exc:
        try:
            os.remove(save_path)
        except OSError:
            pass
        return jsonify({'error': f'Could not read file: {exc}'}), 400

    return jsonify({
        'file_id':      file_id,
        'file_ext':     ext,
        'sheet_names':  sheet_names,
        'active_sheet': active_sheet,
        'columns':      headers,
        'preview':      data_rows[:5],
        'total_rows':   len(data_rows),
    })


@app.route('/api/admin/import/fields/<int:type_id>')
@permission_required('admin.maintenance')
def api_import_fields(type_id):
    """Return core + custom fields available for column mapping for a member type."""
    db = get_db()
    if not db.execute('SELECT id FROM member_types WHERE id = ?', (type_id,)).fetchone():
        return jsonify({'error': 'Member type not found'}), 404

    custom = db.execute('''
        SELECT  fd.id, fd.key, fd.label, fd.field_type, fd.options
        FROM    member_type_fields  mtf
        JOIN    field_definitions   fd  ON fd.id = mtf.field_id
        WHERE   mtf.member_type_id = ? AND fd.active = 1
        ORDER   BY mtf.sort_order, fd.label
    ''', (type_id,)).fetchall()

    return jsonify({
        'core_fields':   _IMPORT_CORE_FIELDS,
        'custom_fields': [dict(f) for f in custom],
    })


@app.route('/api/admin/import/run', methods=['POST'])
@csrf.exempt
@permission_required('admin.maintenance')
def api_import_run():
    """Execute the import with the provided column-to-field mapping.

    Body (JSON):
        file_id         — UUID returned by /analyse
        file_ext        — extension returned by /analyse
        sheet_name      — xlsx sheet (may be omitted for csv)
        member_type_id  — int
        mapping         — { "col_index": "field_key" | "_skip" }
        skip_duplicates — bool (default true); skips rows where
                          first_name + surname + postcode already exist
    """
    data       = request.get_json() or {}
    file_id    = (data.get('file_id') or '').strip()
    file_ext   = (data.get('file_ext') or '').strip()
    sheet_name = (data.get('sheet_name') or '').strip() or None
    type_id    = data.get('member_type_id')
    mapping    = data.get('mapping', {})
    skip_dupes = data.get('skip_duplicates', True)

    if not all([file_id, file_ext, type_id, mapping]):
        return jsonify({'error': 'Missing required parameters'}), 400

    imports_dir = os.path.join(INSTANCE_DIR, 'data', 'imports')
    save_path   = os.path.join(imports_dir, f'{file_id}.{file_ext}')
    if not os.path.exists(save_path):
        return jsonify({'error': 'Upload not found — please re-upload the file'}), 404

    db = get_db()
    mt = db.execute('SELECT * FROM member_types WHERE id = ?', (type_id,)).fetchone()
    if not mt:
        return jsonify({'error': 'Member type not found'}), 404

    # Index custom fields by key
    custom_fields = {}
    for fd in db.execute('''
        SELECT fd.id, fd.key, fd.field_type
        FROM   member_type_fields mtf
        JOIN   field_definitions  fd ON fd.id = mtf.field_id
        WHERE  mtf.member_type_id = ?
    ''', (type_id,)).fetchall():
        custom_fields[fd['key']] = dict(fd)

    # Load all data rows from the saved file
    try:
        if file_ext in ('xlsx', 'xls'):
            _, _, file_headers, data_rows = _read_xlsx_file(save_path, sheet_name)
        else:
            file_headers, data_rows = _read_csv_file(save_path)
    except Exception as exc:
        return jsonify({'error': f'Could not read file: {exc}'}), 400

    CORE_KEYS = {'first_name', 'surname', 'date_of_birth', 'address',
                 'postcode', 'date_registered', 'comments', 'status'}
    # Special-cased keys handled outside the generic core/custom branches
    SPECIAL_KEYS = {'email', 'member_id'}

    imported          = 0
    skipped           = 0
    errors            = []
    not_imported      = []   # [{row, name, reason}] — used for the CSV report
    imported_ids_used = []   # track member_ids that were explicitly provided

    for row_num, row in enumerate(data_rows, start=2):
        try:
            core        = {}
            custom_vals = {}
            email_val   = None
            provided_id = None   # explicit member_id from the spreadsheet

            for col_str, field_key in mapping.items():
                if field_key == '_skip':
                    continue
                col_idx = int(col_str)
                val     = row[col_idx] if col_idx < len(row) else ''
                if not val:
                    continue

                if field_key == 'email':
                    email_val = val
                elif field_key == 'member_id':
                    provided_id = str(val).strip()
                elif field_key in CORE_KEYS:
                    core[field_key] = val
                elif field_key in custom_fields:
                    custom_vals[field_key] = val

            # Helper: identify the row for the report
            def _row_name():
                parts = [core.get('first_name', ''), core.get('surname', '')]
                return ' '.join(p for p in parts if p) or f'Row {row_num}'

            # Skip rows with no name data at all
            if not core.get('first_name') and not core.get('surname'):
                skipped += 1
                not_imported.append({
                    'row': row_num, 'name': f'Row {row_num} (blank)',
                    'reason': 'Blank row — no name data found',
                })
                continue

            # If an explicit member_id was provided, check it isn't already taken
            if provided_id and db.execute(
                'SELECT id FROM members WHERE member_id = ?', (provided_id,)
            ).fetchone():
                skipped += 1
                not_imported.append({
                    'row': row_num, 'name': _row_name(),
                    'reason': f'Member ID {provided_id} already exists in the portal',
                })
                continue

            # Name+postcode duplicate check (only when not using an explicit ID)
            if skip_dupes and not provided_id and db.execute(
                'SELECT id FROM members WHERE first_name=? AND surname=? AND postcode=?',
                (core.get('first_name', ''), core.get('surname', ''), core.get('postcode', ''))
            ).fetchone():
                skipped += 1
                not_imported.append({
                    'row': row_num, 'name': _row_name(),
                    'reason': 'Duplicate — already exists in the portal (same name + postcode)',
                })
                continue

            # Use the provided ID, or auto-generate one
            member_id = provided_id if provided_id else _next_member_id(db)
            db.execute('''
                INSERT INTO members
                    (member_id, first_name, surname, date_of_birth, address,
                     postcode, date_registered, comments, status, member_type)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            ''', (
                member_id,
                core.get('first_name', ''),
                core.get('surname', ''),
                core.get('date_of_birth') or None,
                core.get('address') or None,
                core.get('postcode') or None,
                core.get('date_registered') or None,
                core.get('comments') or None,
                core.get('status', 'Active'),
                mt['slug'],
            ))
            new_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]

            # Primary contact email
            if email_val:
                db.execute(
                    'INSERT INTO member_contacts (member_id, contact_order, contact_email) VALUES (?,1,?)',
                    (new_id, email_val)
                )

            # Custom field values
            for fkey, fval in custom_vals.items():
                fd = custom_fields[fkey]
                db.execute(
                    'INSERT OR REPLACE INTO member_field_values (member_id, field_id, value) VALUES (?,?,?)',
                    (new_id, fd['id'], str(fval))
                )

            db.commit()
            imported += 1
            if provided_id:
                imported_ids_used.append(member_id)

        except Exception as exc:
            errors.append({'row': row_num, 'error': str(exc)})
            not_imported.append({
                'row': row_num,
                'name': ' '.join(filter(None, [
                    (row[int(k)] if int(k) < len(row) else '') for k, v in mapping.items()
                    if v in ('first_name', 'surname')
                ])) or f'Row {row_num}',
                'reason': f'Error: {exc}',
            })
            try:
                db.rollback()
            except Exception:
                pass

    # Clean up temp file regardless of outcome
    try:
        os.remove(save_path)
    except OSError:
        pass

    # If the import provided explicit member IDs, detect their format and
    # persist it so future auto-generated IDs continue in the same style.
    if imported_ids_used:
        _save_id_format_from_import(db, imported_ids_used)

    # Write a CSV report if any rows were not imported
    report_id = None
    if not_imported:
        report_id = str(_uuid_mod.uuid4())
        report_path = os.path.join(imports_dir, f'{report_id}_report.csv')
        with open(report_path, 'w', newline='', encoding='utf-8') as fh:
            writer = _csv_mod.writer(fh)
            writer.writerow(['Row #', 'Name', 'Reason Not Imported'])
            for rec in not_imported:
                writer.writerow([rec['row'], rec['name'], rec['reason']])

    log_action('import.run', 'members', None, {
        'imported': imported, 'skipped': skipped,
        'errors':   len(errors), 'member_type': mt['slug'],
    })

    return jsonify({
        'imported':  imported,
        'skipped':   skipped,
        'errors':    errors,
        'report_id': report_id,   # present only when some rows were not imported
    })


@app.route('/api/admin/import/report/<report_id>')
@permission_required('admin.maintenance')
def api_import_report(report_id):
    """Download the not-imported rows CSV report generated by a previous import run."""
    # Basic sanity-check on the ID to prevent path traversal
    if not report_id or '/' in report_id or '..' in report_id:
        return jsonify({'error': 'Invalid report ID'}), 400
    imports_dir = os.path.join(INSTANCE_DIR, 'data', 'imports')
    report_path = os.path.join(imports_dir, f'{report_id}_report.csv')
    if not os.path.exists(report_path):
        return jsonify({'error': 'Report not found — it may have expired'}), 404
    from flask import send_file
    return send_file(
        report_path,
        mimetype='text/csv',
        as_attachment=True,
        download_name='import_not_imported.csv',
    )


# ── BLUEPRINT: roles + permissions (v6.0) ─────────────────────────────────────

@app.route('/api/admin/permissions')
@permission_required('admin.settings')
def api_permissions_list():
    """Return all permission codes grouped by category."""
    db   = get_db()
    rows = db.execute(
        'SELECT code, name, description, category FROM permissions ORDER BY category, code'
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/admin/roles')
@permission_required('admin.settings')
def api_roles_list():
    """Return all roles with their permission sets."""
    db   = get_db()
    rows = db.execute(
        'SELECT id, name, display_name, is_default, permissions, created_at FROM roles ORDER BY name'
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        # Fall back to slug if display_name not yet set (pre-migration rows)
        if not d.get('display_name'):
            d['display_name'] = ROLE_DISPLAY_NAMES.get(d['name'], d['name'])
        try:
            d['permissions'] = json.loads(d['permissions'])
        except (TypeError, ValueError):
            d['permissions'] = []
        result.append(d)
    return jsonify(result)


@app.route('/api/admin/roles', methods=['POST'])
@permission_required('admin.settings')
def api_roles_create():
    """Create a new custom role."""
    data         = request.get_json() or {}
    name         = data.get('name', '').strip()
    display_name = data.get('display_name', '').strip() or name
    perms        = data.get('permissions', [])

    if not name:
        return jsonify({'error': 'Role name is required'}), 400
    if not isinstance(perms, list):
        return jsonify({'error': 'permissions must be a list'}), 400

    # Validate all permission codes exist
    db          = get_db()
    valid_codes = {r['code'] for r in db.execute('SELECT code FROM permissions').fetchall()}
    bad = [p for p in perms if p not in valid_codes]
    if bad:
        return jsonify({'error': f'Unknown permission code(s): {", ".join(bad)}'}), 400

    # Only users with users.create.admin can include that permission in a new role
    if 'users.create.admin' in perms and not has_permission('users.create.admin'):
        return jsonify({'error': 'You do not have permission to assign users.create.admin'}), 403

    try:
        cur = db.execute(
            'INSERT INTO roles (name, display_name, permissions, is_default) VALUES (?,?,?,0)',
            (name, display_name, json.dumps(perms))
        )
        db.commit()
        log_action('create_role', 'roles', cur.lastrowid, {'name': name, 'permissions': perms})
        row = db.execute('SELECT id, name, display_name, is_default, permissions, created_at FROM roles WHERE id = ?',
                         (cur.lastrowid,)).fetchone()
        d = dict(row)
        d['permissions'] = json.loads(d['permissions'])
        return jsonify(d), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A role named "{name}" already exists'}), 409


@app.route('/api/admin/roles/<int:role_id>', methods=['PUT'])
@permission_required('admin.settings')
def api_roles_update(role_id):
    """Update a role's name, display name, and/or permission set."""
    data  = request.get_json() or {}
    db    = get_db()
    role  = db.execute('SELECT * FROM roles WHERE id = ?', (role_id,)).fetchone()
    if not role:
        return jsonify({'error': 'Role not found'}), 404

    name         = data.get('name', role['name']).strip()
    display_name = data.get('display_name', role['display_name'] or role['name']).strip() or name
    perms        = data.get('permissions', json.loads(role['permissions']))

    if not name:
        return jsonify({'error': 'Role name is required'}), 400
    if not isinstance(perms, list):
        return jsonify({'error': 'permissions must be a list'}), 400

    # Validate all permission codes exist
    valid_codes = {r['code'] for r in db.execute('SELECT code FROM permissions').fetchall()}
    bad = [p for p in perms if p not in valid_codes]
    if bad:
        return jsonify({'error': f'Unknown permission code(s): {", ".join(bad)}'}), 400

    # Only users with users.create.admin can add that permission to a role
    old_perms = json.loads(role['permissions'])
    if 'users.create.admin' in perms and 'users.create.admin' not in old_perms:
        if not has_permission('users.create.admin'):
            return jsonify({'error': 'You do not have permission to assign users.create.admin'}), 403

    try:
        db.execute(
            'UPDATE roles SET name = ?, display_name = ?, permissions = ? WHERE id = ?',
            (name, display_name, json.dumps(perms), role_id)
        )
        db.commit()
        log_action('update_role', 'roles', role_id, {'name': name, 'display_name': display_name, 'permissions': perms})
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A role named "{name}" already exists'}), 409

    # Refresh session permissions if the user's own role was updated
    # (so changes take effect on their next page load rather than next login)
    row = db.execute('SELECT id, name, display_name, is_default, permissions, created_at FROM roles WHERE id = ?',
                     (role_id,)).fetchone()
    d = dict(row)
    d['permissions'] = json.loads(d['permissions'])
    return jsonify(d)


@app.route('/api/admin/roles/<int:role_id>', methods=['DELETE'])
@permission_required('admin.settings')
def api_roles_delete(role_id):
    """Delete a custom role. Default roles and roles with assigned users cannot be deleted."""
    db   = get_db()
    role = db.execute('SELECT * FROM roles WHERE id = ?', (role_id,)).fetchone()
    if not role:
        return jsonify({'error': 'Role not found'}), 404

    if role['is_default']:
        return jsonify({'error': 'Default roles cannot be deleted'}), 400

    # Prevent deletion if any users are assigned to this role
    user_count = db.execute(
        'SELECT COUNT(*) FROM users WHERE role_id = ?', (role_id,)
    ).fetchone()[0]
    if user_count:
        return jsonify({'error': f'Cannot delete — {user_count} user(s) are assigned to this role'}), 400

    db.execute('DELETE FROM roles WHERE id = ?', (role_id,))
    db.commit()
    log_action('delete_role', 'roles', role_id, {'name': role['name']})
    return jsonify({'success': True})


# ── BLUEPRINT: tag definitions admin CRUD (Phase 7.1) ────────────────────────

@app.route('/api/tags')
@permission_required('members.view')
def api_tags_public():
    """Return all active tag definitions — used by the member edit UI for assignment."""
    db   = get_db()
    rows = db.execute(
        'SELECT id, name, category, icon, colour FROM tag_definitions '
        'WHERE active = 1 ORDER BY sort_order, name'
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/admin/tags')
@permission_required('admin.settings')
def api_tags_list():
    """Return all tag definitions including inactive ones (admin view)."""
    db   = get_db()
    rows = db.execute(
        'SELECT id, name, category, icon, colour, active, sort_order '
        'FROM tag_definitions ORDER BY sort_order, name'
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/admin/tags', methods=['POST'])
@permission_required('admin.settings')
def api_tags_create():
    data              = request.get_json() or {}
    name              = data.get('name', '').strip()
    category          = data.get('category', 'General').strip() or 'General'
    icon              = data.get('icon', '🏷').strip() or '🏷'
    colour, col_err   = _validate_hex_colour(data.get('colour', ''), '#3b82f6')

    if not name:
        return jsonify({'error': 'name is required'}), 400
    if col_err:
        return jsonify({'error': col_err}), 400

    db        = get_db()
    max_order = db.execute('SELECT COALESCE(MAX(sort_order), -1) FROM tag_definitions').fetchone()[0]
    try:
        cur = db.execute(
            'INSERT INTO tag_definitions (name, category, icon, colour, active, sort_order) VALUES (?,?,?,?,1,?)',
            (name, category, icon, colour, max_order + 1)
        )
        db.commit()
        log_action('create_tag', 'tag_definitions', cur.lastrowid, {'name': name})
        return jsonify(dict(db.execute('SELECT * FROM tag_definitions WHERE id = ?', (cur.lastrowid,)).fetchone())), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A tag named "{name}" already exists'}), 409


@app.route('/api/admin/tags/<int:tag_id>', methods=['PUT'])
@permission_required('admin.settings')
def api_tags_update(tag_id):
    db  = get_db()
    row = db.execute('SELECT * FROM tag_definitions WHERE id = ?', (tag_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Tag not found'}), 404

    data              = request.get_json() or {}
    name              = data.get('name',     row['name']).strip()
    category          = data.get('category', row['category']).strip() or 'General'
    icon              = data.get('icon',     row['icon']).strip() or '🏷'
    colour, col_err   = _validate_hex_colour(data.get('colour', row['colour']), row['colour'])
    active            = int(data.get('active', row['active']))

    if not name:
        return jsonify({'error': 'name is required'}), 400
    if col_err:
        return jsonify({'error': col_err}), 400
    try:
        db.execute(
            'UPDATE tag_definitions SET name=?, category=?, icon=?, colour=?, active=? WHERE id=?',
            (name, category, icon, colour, active, tag_id)
        )
        db.commit()
        log_action('update_tag', 'tag_definitions', tag_id, {'name': name, 'active': active})
        return jsonify(dict(db.execute('SELECT * FROM tag_definitions WHERE id = ?', (tag_id,)).fetchone()))
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A tag named "{name}" already exists'}), 409


@app.route('/api/admin/tags/reorder', methods=['POST'])
@permission_required('admin.settings')
def api_tags_reorder():
    items = request.get_json() or []
    db    = get_db()
    for item in items:
        db.execute('UPDATE tag_definitions SET sort_order = ? WHERE id = ?',
                   (item.get('sort_order', 0), item.get('id')))
    db.commit()
    return jsonify({'success': True})


# ── Member tag assignment ──────────────────────────────────────────────────────

@app.route('/api/members/<int:member_id>/tags')
@permission_required('members.view')
def api_member_tags_get(member_id):
    """Return all active tags assigned to a member."""
    db   = get_db()
    rows = db.execute('''
        SELECT  mt.id AS assignment_id, mt.expires_at, mt.notes, mt.created_at,
                td.id AS tag_id, td.name, td.icon, td.colour, td.category
        FROM    member_tags mt
        JOIN    tag_definitions td ON td.id = mt.tag_id
        WHERE   mt.member_id = ?
        ORDER   BY td.sort_order, td.name
    ''', (member_id,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/members/<int:member_id>/tags', methods=['POST'])
@permission_required('members.tags')
def api_member_tags_add(member_id):
    """Assign a tag to a member."""
    data       = request.get_json() or {}
    tag_id     = data.get('tag_id')
    expires_at = data.get('expires_at', None)
    notes      = data.get('notes', '').strip() or None

    if not tag_id:
        return jsonify({'error': 'tag_id is required'}), 400

    db  = get_db()
    tag = db.execute('SELECT id, name FROM tag_definitions WHERE id = ? AND active = 1', (tag_id,)).fetchone()
    if not tag:
        return jsonify({'error': 'Tag not found or inactive'}), 404

    try:
        cur = db.execute(
            'INSERT INTO member_tags (member_id, tag_id, expires_at, notes) VALUES (?,?,?,?)',
            (member_id, tag_id, expires_at, notes)
        )
        db.commit()
        m = db.execute('SELECT first_name, surname FROM members WHERE id = ?', (member_id,)).fetchone()
        log_action('add_member_tag', 'member_tags', cur.lastrowid, {
            'member': f"{m['first_name'] or ''} {m['surname'] or ''}".strip() if m else str(member_id),
            'tag':    tag['name'],
        })
        row = db.execute('''
            SELECT mt.id AS assignment_id, mt.expires_at, mt.notes, mt.created_at,
                   td.id AS tag_id, td.name, td.icon, td.colour, td.category
            FROM   member_tags mt JOIN tag_definitions td ON td.id = mt.tag_id
            WHERE  mt.id = ?
        ''', (cur.lastrowid,)).fetchone()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': 'This tag is already assigned to the member'}), 409


@app.route('/api/members/<int:member_id>/tags/<int:tag_id>', methods=['DELETE'])
@permission_required('members.tags')
def api_member_tags_remove(member_id, tag_id):
    """Remove a tag assignment from a member (tag_id here is the tag_definitions.id)."""
    db  = get_db()
    row = db.execute(
        'SELECT id FROM member_tags WHERE member_id = ? AND tag_id = ?',
        (member_id, tag_id)
    ).fetchone()
    if not row:
        return jsonify({'error': 'Tag assignment not found'}), 404

    m   = db.execute('SELECT first_name, surname FROM members WHERE id = ?', (member_id,)).fetchone()
    tag = db.execute('SELECT name FROM tag_definitions WHERE id = ?', (tag_id,)).fetchone()
    db.execute('DELETE FROM member_tags WHERE id = ?', (row['id'],))
    db.commit()
    log_action('remove_member_tag', 'member_tags', row['id'], {
        'member': f"{m['first_name'] or ''} {m['surname'] or ''}".strip() if m else str(member_id),
        'tag':    tag['name'] if tag else str(tag_id),
    })
    return jsonify({'success': True})


# ── BLUEPRINT: session notes (Phase 7.1) ──────────────────────────────────────

NOTE_TYPES = ('General', 'Medical', 'Safeguarding', 'Behaviour', 'Accident', 'Other')


@app.route('/api/register/notes/<session_type>/<date>')
@permission_required('register.signout')   # minimum permission — everyone on the register page
def api_notes_get(session_type, date):
    """Return all session notes for a given session and date."""
    db    = get_db()
    rows  = db.execute('''
        SELECT  sn.id, sn.note_type, sn.title, sn.details, sn.created_at,
                sn.member_id,
                u.username   AS added_by_name,
                m.first_name AS member_first, m.surname AS member_surname
        FROM    session_notes sn
        LEFT JOIN users   u ON u.id = sn.added_by
        LEFT JOIN members m ON m.id = sn.member_id
        WHERE   sn.session_date = ? AND sn.session_type = ?
        ORDER   BY sn.created_at
    ''', (date, session_type)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/register/notes', methods=['POST'])
@permission_required('register.notes')
def api_notes_create():
    """Add a new session note or incident."""
    data         = request.get_json() or {}
    session_date = data.get('session_date', '').strip()
    session_type = data.get('session_type', '').strip()
    note_type    = data.get('note_type', 'General').strip()
    title        = data.get('title', '').strip()
    details      = data.get('details', '').strip()
    member_id    = data.get('member_id') or None

    if not session_date or not session_type:
        return jsonify({'error': 'session_date and session_type are required'}), 400
    if note_type not in NOTE_TYPES:
        note_type = 'General'
    if not title and not details:
        return jsonify({'error': 'Please provide at least a title or details'}), 400

    # Validate session type
    valid_sessions = get_valid_session_names()
    if session_type not in valid_sessions:
        return jsonify({'error': 'Invalid session type'}), 400

    db = get_db()

    # Validate member_id belongs to this session if provided
    if member_id:
        member = db.execute(
            'SELECT id FROM members WHERE id = ? AND session = ? AND status != "Leaver"',
            (member_id, session_type)
        ).fetchone()
        if not member:
            member_id = None  # silently clear invalid member link

    cur = db.execute(
        '''INSERT INTO session_notes (session_date, session_type, member_id, note_type, title, details, added_by)
           VALUES (?, ?, ?, ?, ?, ?, ?)''',
        (session_date, session_type, member_id, note_type, title, details, session['user_id'])
    )
    db.commit()

    log_action('create_session_note', 'session_notes', cur.lastrowid, {
        'session_date': session_date,
        'session_type': session_type,
        'note_type':    note_type,
    })

    row = db.execute('''
        SELECT  sn.id, sn.note_type, sn.title, sn.details, sn.created_at,
                sn.member_id,
                u.username   AS added_by_name,
                m.first_name AS member_first, m.surname AS member_surname
        FROM    session_notes sn
        LEFT JOIN users   u ON u.id = sn.added_by
        LEFT JOIN members m ON m.id = sn.member_id
        WHERE   sn.id = ?
    ''', (cur.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route('/api/register/notes/<int:note_id>', methods=['DELETE'])
@permission_required('register.notes')
def api_notes_delete(note_id):
    """Delete a session note. Only the original author or an admin may delete."""
    db   = get_db()
    note = db.execute('SELECT * FROM session_notes WHERE id = ?', (note_id,)).fetchone()
    if not note:
        return jsonify({'error': 'Note not found'}), 404

    # Only the author or admin can delete
    is_admin   = session.get('role') == 'admin'
    is_author  = note['added_by'] == session['user_id']
    if not is_admin and not is_author:
        return jsonify({'error': 'You can only delete your own notes'}), 403

    db.execute('DELETE FROM session_notes WHERE id = ?', (note_id,))
    db.commit()
    log_action('delete_session_note', 'session_notes', note_id, {
        'session_date': note['session_date'],
        'session_type': note['session_type'],
    })
    return jsonify({'success': True})


# ── BLUEPRINT: member types CRUD (v8.0) ───────────────────────────────────────

import re as _re

def _slugify(text):
    """Convert text to a URL-safe slug."""
    text = text.lower().strip()
    text = _re.sub(r'[^\w\s-]', '', text)
    text = _re.sub(r'[\s_]+', '-', text)
    text = _re.sub(r'-+', '-', text)
    return text.strip('-')


@app.route('/api/admin/member-types', methods=['GET'])
@permission_required('admin.settings')
def api_admin_member_types_list():
    """Return all member types ordered by sort_order, with field_count."""
    db   = get_db()
    rows = db.execute('''
        SELECT mt.*,
               (SELECT COUNT(*) FROM member_type_fields mtf WHERE mtf.member_type_id = mt.id) AS field_count
        FROM   member_types mt
        ORDER  BY mt.sort_order, mt.name
    ''').fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/admin/member-types', methods=['POST'])
@permission_required('admin.settings')
def api_admin_member_types_create():
    """Create a new member type."""
    data                = request.get_json() or {}
    name                = data.get('name', '').strip()
    slug                = data.get('slug', '').strip() or _slugify(name)
    icon                = data.get('icon', '👤').strip() or '👤'
    colour, col_err     = _validate_hex_colour(data.get('colour', ''), '#1b2d4f')
    description         = data.get('description', '').strip() or None
    public_registration = int(data.get('public_registration', 0))

    if not name:
        return jsonify({'error': 'name is required'}), 400
    if not slug:
        return jsonify({'error': 'slug is required'}), 400
    if col_err:
        return jsonify({'error': col_err}), 400

    db        = get_db()
    max_order = db.execute('SELECT COALESCE(MAX(sort_order), -1) FROM member_types').fetchone()[0]
    try:
        cur = db.execute(
            '''INSERT INTO member_types
               (name, slug, icon, colour, description, public_registration, sort_order)
               VALUES (?,?,?,?,?,?,?)''',
            (name, slug, icon, colour, description, public_registration, max_order + 1),
        )
        db.commit()
        log_action('create_member_type', 'member_types', cur.lastrowid, {'name': name, 'slug': slug})
        row = db.execute(
            '''SELECT mt.*, (SELECT COUNT(*) FROM member_type_fields mtf WHERE mtf.member_type_id = mt.id) AS field_count
               FROM member_types mt WHERE mt.id = ?''', (cur.lastrowid,)
        ).fetchone()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A member type with that name or slug already exists'}), 409


@app.route('/api/admin/member-types/<int:type_id>', methods=['PUT'])
@permission_required('admin.settings')
def api_admin_member_types_update(type_id):
    """Update a member type. Slug cannot be changed after creation."""
    data    = request.get_json() or {}
    db      = get_db()
    current = db.execute('SELECT * FROM member_types WHERE id = ?', (type_id,)).fetchone()
    if not current:
        return jsonify({'error': 'Member type not found'}), 404

    name                = (data.get('name') or current['name']).strip()
    icon                = (data.get('icon') or current['icon'] or '👤').strip() or '👤'
    colour, col_err     = _validate_hex_colour(data.get('colour', current['colour']), current['colour'] or '#1b2d4f')
    description         = data.get('description', current['description'])
    if description is not None:
        description = description.strip() or None
    public_registration = int(data.get('public_registration', current['public_registration']))
    active              = int(data.get('active', current['active']))
    sort_order          = int(data.get('sort_order', current['sort_order']))

    if not name:
        return jsonify({'error': 'name is required'}), 400
    if col_err:
        return jsonify({'error': col_err}), 400

    # Prevent deactivating the last active type
    if not active:
        active_count = db.execute(
            'SELECT COUNT(*) FROM member_types WHERE active = 1 AND id != ?', (type_id,)
        ).fetchone()[0]
        if active_count == 0:
            return jsonify({'error': 'Cannot deactivate the only active member type'}), 400

    try:
        db.execute(
            '''UPDATE member_types
               SET name=?, icon=?, colour=?, description=?, public_registration=?, active=?, sort_order=?
               WHERE id=?''',
            (name, icon, colour, description, public_registration, active, sort_order, type_id),
        )
        db.commit()
        log_action('update_member_type', 'member_types', type_id,
                   {'name': name, 'active': active})
        row = db.execute(
            '''SELECT mt.*, (SELECT COUNT(*) FROM member_type_fields mtf WHERE mtf.member_type_id = mt.id) AS field_count
               FROM member_types mt WHERE mt.id = ?''', (type_id,)
        ).fetchone()
        return jsonify(dict(row))
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A member type named "{name}" already exists'}), 409


@app.route('/api/admin/member-types/<int:type_id>', methods=['DELETE'])
@permission_required('admin.settings')
def api_admin_member_types_delete(type_id):
    """Delete a member type. Blocked if members use this type or it's the last active type."""
    db  = get_db()
    row = db.execute('SELECT * FROM member_types WHERE id = ?', (type_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Member type not found'}), 404

    member_count = db.execute(
        'SELECT COUNT(*) FROM members WHERE member_type = ?', (row['slug'],)
    ).fetchone()[0]
    if member_count:
        return jsonify({'error': f'Cannot delete — {member_count} member(s) use this type'}), 409

    active_count = db.execute(
        'SELECT COUNT(*) FROM member_types WHERE active = 1 AND id != ?', (type_id,)
    ).fetchone()[0]
    if row['active'] and active_count == 0:
        return jsonify({'error': 'Cannot delete the only active member type'}), 400

    db.execute('DELETE FROM member_types WHERE id = ?', (type_id,))
    db.commit()
    log_action('delete_member_type', 'member_types', type_id, {'name': row['name']})
    return jsonify({'success': True})


# ── BLUEPRINT: field definitions CRUD (v8.0) ──────────────────────────────────

@app.route('/api/admin/field-definitions', methods=['GET'])
@permission_required('admin.settings')
def api_admin_field_definitions_list():
    """Return all field definitions ordered by sort_order, with assigned_to count."""
    db   = get_db()
    rows = db.execute('''
        SELECT fd.*,
               (SELECT COUNT(*) FROM member_type_fields mtf WHERE mtf.field_id = fd.id) AS assigned_to
        FROM   field_definitions fd
        ORDER  BY fd.sort_order, fd.label
    ''').fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/admin/field-definitions', methods=['POST'])
@permission_required('admin.settings')
def api_admin_field_definitions_create():
    """Create a new custom field (system_field=0)."""
    data        = request.get_json() or {}
    label       = (data.get('label') or '').strip()
    field_type  = (data.get('field_type') or 'text').strip() or 'text'
    placeholder = (data.get('placeholder') or '').strip() or None
    help_text   = (data.get('help_text') or '').strip() or None
    options     = (data.get('options') or '').strip() or None

    if not label:
        return jsonify({'error': 'label is required'}), 400

    key = _slugify(label).replace('-', '_')
    if not key:
        return jsonify({'error': 'Could not generate a valid key from the label'}), 400

    db        = get_db()
    max_order = db.execute('SELECT COALESCE(MAX(sort_order), -1) FROM field_definitions').fetchone()[0]

    # Ensure key uniqueness by appending a suffix if needed
    base_key = key
    suffix   = 1
    while db.execute('SELECT id FROM field_definitions WHERE key = ?', (key,)).fetchone():
        key = f'{base_key}_{suffix}'
        suffix += 1

    try:
        cur = db.execute(
            '''INSERT INTO field_definitions
               (key, label, field_type, placeholder, help_text, options, system_field, sort_order)
               VALUES (?,?,?,?,?,?,0,?)''',
            (key, label, field_type, placeholder, help_text, options, max_order + 1),
        )
        db.commit()
        log_action('create_field_definition', 'field_definitions', cur.lastrowid,
                   {'key': key, 'label': label, 'field_type': field_type})
        row = db.execute(
            '''SELECT fd.*, (SELECT COUNT(*) FROM member_type_fields mtf WHERE mtf.field_id = fd.id) AS assigned_to
               FROM field_definitions fd WHERE fd.id = ?''', (cur.lastrowid,)
        ).fetchone()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A field with key "{key}" already exists'}), 409


@app.route('/api/admin/field-definitions/<int:field_id>', methods=['PUT'])
@permission_required('admin.settings')
def api_admin_field_definitions_update(field_id):
    """Update a field definition. Key and system_field status cannot be changed."""
    data    = request.get_json() or {}
    db      = get_db()
    current = db.execute('SELECT * FROM field_definitions WHERE id = ?', (field_id,)).fetchone()
    if not current:
        return jsonify({'error': 'Field definition not found'}), 404

    # Block field_type changes for system fields
    if current['system_field'] and 'field_type' in data and data['field_type'] != current['field_type']:
        return jsonify({'error': 'Cannot change the field type of a system field'}), 403

    label       = data.get('label', current['label']).strip()
    field_type  = data.get('field_type', current['field_type']).strip() or current['field_type']
    placeholder = data.get('placeholder', current['placeholder'])
    if placeholder is not None:
        placeholder = placeholder.strip() or None
    help_text   = data.get('help_text', current['help_text'])
    if help_text is not None:
        help_text = help_text.strip() or None
    options     = data.get('options', current['options'])
    if options is not None:
        options = options.strip() or None
    active      = int(data.get('active', current['active']))

    if not label:
        return jsonify({'error': 'label is required'}), 400

    db.execute(
        '''UPDATE field_definitions
           SET label=?, field_type=?, placeholder=?, help_text=?, options=?, active=?
           WHERE id=?''',
        (label, field_type, placeholder, help_text, options, active, field_id),
    )
    db.commit()
    log_action('update_field_definition', 'field_definitions', field_id,
               {'label': label, 'active': active})
    row = db.execute(
        '''SELECT fd.*, (SELECT COUNT(*) FROM member_type_fields mtf WHERE mtf.field_id = fd.id) AS assigned_to
           FROM field_definitions fd WHERE fd.id = ?''', (field_id,)
    ).fetchone()
    return jsonify(dict(row))


@app.route('/api/admin/field-definitions/<int:field_id>', methods=['DELETE'])
@permission_required('admin.settings')
def api_admin_field_definitions_delete(field_id):
    """Delete a custom field. Blocked for system fields or if in use."""
    db  = get_db()
    row = db.execute('SELECT * FROM field_definitions WHERE id = ?', (field_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Field definition not found'}), 404

    if row['system_field']:
        return jsonify({'error': 'System fields cannot be deleted'}), 403

    usage = db.execute(
        'SELECT COUNT(*) FROM member_type_fields WHERE field_id = ?', (field_id,)
    ).fetchone()[0]
    if usage:
        return jsonify({'error': f'Cannot delete — this field is assigned to {usage} member type(s)'}), 409

    db.execute('DELETE FROM field_definitions WHERE id = ?', (field_id,))
    db.commit()
    log_action('delete_field_definition', 'field_definitions', field_id, {'key': row['key']})
    return jsonify({'success': True})


# ── BLUEPRINT: type-field config (v8.0) ───────────────────────────────────────

@app.route('/api/admin/member-types/<int:type_id>/fields', methods=['GET'])
@permission_required('admin.settings')
def api_admin_type_fields_list(type_id):
    """Return all fields assigned to this member type with full field definition data."""
    db = get_db()
    if not db.execute('SELECT id FROM member_types WHERE id = ?', (type_id,)).fetchone():
        return jsonify({'error': 'Member type not found'}), 404
    rows = db.execute('''
        SELECT  mtf.id AS assignment_id, mtf.sort_order, mtf.required,
                mtf.show_on_registration, mtf.show_on_list, mtf.show_on_card,
                mtf.show_on_detail, mtf.show_on_print, mtf.show_on_export,
                fd.id AS field_id, fd.key, fd.label, fd.field_type,
                fd.options, fd.help_text, fd.placeholder,
                fd.system_field, fd.column_name, fd.active
        FROM    member_type_fields mtf
        JOIN    field_definitions fd ON fd.id = mtf.field_id
        WHERE   mtf.member_type_id = ?
        ORDER   BY mtf.sort_order, fd.label
    ''', (type_id,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/admin/member-types/<int:type_id>/fields', methods=['POST'])
@permission_required('admin.settings')
def api_admin_type_fields_assign(type_id):
    """Assign a field to a member type."""
    db = get_db()
    if not db.execute('SELECT id FROM member_types WHERE id = ?', (type_id,)).fetchone():
        return jsonify({'error': 'Member type not found'}), 404

    data     = request.get_json() or {}
    field_id = data.get('field_id')
    if not field_id:
        return jsonify({'error': 'field_id is required'}), 400

    if not db.execute('SELECT id FROM field_definitions WHERE id = ?', (field_id,)).fetchone():
        return jsonify({'error': 'Field definition not found'}), 404

    max_order = db.execute(
        'SELECT COALESCE(MAX(sort_order), 0) FROM member_type_fields WHERE member_type_id = ?', (type_id,)
    ).fetchone()[0]

    try:
        cur = db.execute(
            '''INSERT INTO member_type_fields
               (member_type_id, field_id, sort_order, required,
                show_on_registration, show_on_list, show_on_card, show_on_detail, show_on_print, show_on_export)
               VALUES (?,?,?,0,1,0,0,1,1,0)''',
            (type_id, field_id, max_order + 1),
        )
        db.commit()
        log_action('assign_type_field', 'member_type_fields', cur.lastrowid,
                   {'member_type_id': type_id, 'field_id': field_id})
        row = db.execute('''
            SELECT  mtf.id AS assignment_id, mtf.sort_order, mtf.required,
                    mtf.show_on_registration, mtf.show_on_list, mtf.show_on_card,
                    mtf.show_on_detail, mtf.show_on_print, mtf.show_on_export,
                    fd.id AS field_id, fd.key, fd.label, fd.field_type,
                    fd.options, fd.help_text, fd.placeholder,
                    fd.system_field, fd.column_name, fd.active
            FROM    member_type_fields mtf
            JOIN    field_definitions fd ON fd.id = mtf.field_id
            WHERE   mtf.id = ?
        ''', (cur.lastrowid,)).fetchone()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': 'This field is already assigned to this member type'}), 409


@app.route('/api/admin/member-types/<int:type_id>/fields/<int:field_id>', methods=['PUT'])
@permission_required('admin.settings')
def api_admin_type_fields_update(type_id, field_id):
    """Update display/required settings for a field on a member type."""
    db  = get_db()
    row = db.execute(
        'SELECT * FROM member_type_fields WHERE member_type_id = ? AND field_id = ?',
        (type_id, field_id)
    ).fetchone()
    if not row:
        return jsonify({'error': 'Assignment not found'}), 404

    data = request.get_json() or {}
    updates = {}
    for col in ('required', 'show_on_registration', 'show_on_list',
                'show_on_card', 'show_on_detail', 'show_on_print', 'show_on_export'):
        if col in data:
            updates[col] = int(data[col])

    if not updates:
        return jsonify({'error': 'No updatable fields provided'}), 400

    set_clause = ', '.join(f'{k} = ?' for k in updates)
    values     = list(updates.values()) + [type_id, field_id]
    db.execute(
        f'UPDATE member_type_fields SET {set_clause} WHERE member_type_id = ? AND field_id = ?',
        values,
    )
    db.commit()

    updated = db.execute('''
        SELECT  mtf.id AS assignment_id, mtf.sort_order, mtf.required,
                mtf.show_on_registration, mtf.show_on_list, mtf.show_on_card,
                mtf.show_on_detail, mtf.show_on_print, mtf.show_on_export,
                fd.id AS field_id, fd.key, fd.label, fd.field_type,
                fd.options, fd.help_text, fd.placeholder,
                fd.system_field, fd.column_name, fd.active
        FROM    member_type_fields mtf
        JOIN    field_definitions fd ON fd.id = mtf.field_id
        WHERE   mtf.member_type_id = ? AND mtf.field_id = ?
    ''', (type_id, field_id)).fetchone()
    return jsonify(dict(updated))


@app.route('/api/admin/member-types/<int:type_id>/fields/<int:field_id>', methods=['DELETE'])
@permission_required('admin.settings')
def api_admin_type_fields_remove(type_id, field_id):
    """Remove a field from a member type. first_name and surname cannot be removed."""
    db  = get_db()
    fd  = db.execute('SELECT key FROM field_definitions WHERE id = ?', (field_id,)).fetchone()
    if fd and fd['key'] in ('first_name', 'surname'):
        return jsonify({'error': 'first_name and surname cannot be removed from a member type'}), 403

    row = db.execute(
        'SELECT id FROM member_type_fields WHERE member_type_id = ? AND field_id = ?',
        (type_id, field_id)
    ).fetchone()
    if not row:
        return jsonify({'error': 'Assignment not found'}), 404

    db.execute(
        'DELETE FROM member_type_fields WHERE member_type_id = ? AND field_id = ?',
        (type_id, field_id)
    )
    db.commit()
    log_action('remove_type_field', 'member_type_fields', row['id'],
               {'member_type_id': type_id, 'field_id': field_id})
    return jsonify({'success': True})


@app.route('/api/admin/member-types/<int:type_id>/fields/reorder', methods=['POST'])
@permission_required('admin.settings')
def api_admin_type_fields_reorder(type_id):
    """Reorder fields for a member type. Body: [{field_id, sort_order}, ...]"""
    items = request.get_json() or []
    db    = get_db()
    for item in items:
        db.execute(
            'UPDATE member_type_fields SET sort_order = ? WHERE member_type_id = ? AND field_id = ?',
            (item.get('sort_order', 0), type_id, item.get('field_id')),
        )
    db.commit()
    return jsonify({'success': True})


# ── BLUEPRINT: alerts ──────────────────────────────────────────────────────────
# Member Alert Rules — configurable flag engine (v8.0)
# Rules evaluate members against date, attendance, empty-field, or numeric
# conditions and write rows to member_flags.  A BackgroundScheduler runs the
# full check nightly at 02:00; admins can also trigger manually via the API.

# Allowed colour presets (hex) for flag badges — validated on create/update
ALERT_COLOURS = {'#ef4444', '#f59e0b', '#3b6fde', '#8b5cf6', '#22a06b'}


def _run_alert_rule(db, rule, today_str):
    """Evaluate one active alert rule against all eligible members.

    For each member that meets the condition:
      - Insert into member_flags if no active flag already exists.
    For each member that no longer meets the condition (auto_resolve=1):
      - Set resolved_at on their active flag.

    Returns (raised_count, resolved_count).
    """
    from collections import defaultdict

    rule_id     = rule['id']
    rule_type   = rule['rule_type']
    auto_res    = rule['auto_resolve']
    scoped_sess = rule['applies_to_session']   # None → all sessions

    # ── Fetch eligible members ────────────────────────────────────────────────
    member_cond = "m.status = 'Active' AND m.member_type = 'member'"
    params_base = []
    if scoped_sess:
        member_cond += ' AND m.session = ?'
        params_base.append(scoped_sess)

    members = db.execute(
        f'SELECT m.id, m.member_id, m.first_name, m.surname, m.session '
        f'FROM members m WHERE {member_cond}',
        params_base
    ).fetchall()

    # ── Existing active flags for this rule ────────────────────────────────────
    existing_flags = {
        row['member_id']: row['id']
        for row in db.execute(
            'SELECT member_id, id FROM member_flags '
            'WHERE rule_id = ? AND resolved_at IS NULL',
            (rule_id,)
        ).fetchall()
    }

    should_flag = set()   # member db ids that currently meet the condition

    # ── Attendance rule ────────────────────────────────────────────────────────
    if rule_type == 'attendance':
        threshold = rule['threshold_value'] or 5
        past_sessions_rows = db.execute(
            "SELECT session_date, session_type FROM term_sessions "
            "WHERE session_date <= ? ORDER BY session_date DESC",
            (today_str,)
        ).fetchall()
        sessions_by_type = defaultdict(list)
        for s in past_sessions_rows:
            sessions_by_type[s['session_type']].append(s['session_date'])

        for m in members:
            relevant = sessions_by_type[m['session']][:threshold]
            if len(relevant) < threshold:
                continue
            attended = db.execute(
                'SELECT COUNT(*) AS n FROM attendance WHERE member_id = ? '
                'AND session_date IN ({})'.format(','.join('?' * len(relevant))),
                [m['id']] + relevant
            ).fetchone()['n']
            if attended == 0:
                should_flag.add(m['id'])

    # ── Date field rule ────────────────────────────────────────────────────────
    elif rule_type == 'date_field':
        target      = rule['target_field']
        condition   = rule['condition']     # older_than | before_today
        threshold_d = rule['threshold_value'] or 0

        # Determine if target is a system column or a custom field
        fd = db.execute(
            'SELECT id, column_name, system_field FROM field_definitions WHERE key = ?',
            (target,)
        ).fetchone()

        for m in members:
            if fd and fd['system_field'] and fd['column_name']:
                row = db.execute(
                    f'SELECT {fd["column_name"]} AS val FROM members WHERE id = ?',
                    (m['id'],)
                ).fetchone()
                val = row['val'] if row else None
            else:
                fid = fd['id'] if fd else None
                if not fid:
                    continue
                row = db.execute(
                    'SELECT value FROM member_field_values WHERE member_id = ? AND field_id = ?',
                    (m['id'], fid)
                ).fetchone()
                val = row['value'] if row else None

            if not val:
                continue
            try:
                field_date = datetime.strptime(val[:10], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                continue

            today_date = datetime.strptime(today_str, '%Y-%m-%d').date()
            if condition == 'older_than':
                if (today_date - field_date).days >= threshold_d:
                    should_flag.add(m['id'])
            elif condition == 'before_today':
                if field_date < today_date:
                    should_flag.add(m['id'])

    # ── Empty field rule ───────────────────────────────────────────────────────
    elif rule_type == 'empty_field':
        target = rule['target_field']
        fd = db.execute(
            'SELECT id, column_name, system_field FROM field_definitions WHERE key = ?',
            (target,)
        ).fetchone()

        for m in members:
            if fd and fd['system_field'] and fd['column_name']:
                row = db.execute(
                    f'SELECT {fd["column_name"]} AS val FROM members WHERE id = ?',
                    (m['id'],)
                ).fetchone()
                val = (row['val'] or '').strip() if row else ''
            else:
                fid = fd['id'] if fd else None
                if not fid:
                    continue
                row = db.execute(
                    'SELECT value FROM member_field_values WHERE member_id = ? AND field_id = ?',
                    (m['id'], fid)
                ).fetchone()
                val = (row['value'] or '').strip() if row else ''

            if not val:
                should_flag.add(m['id'])

    # ── Numeric rule ───────────────────────────────────────────────────────────
    elif rule_type == 'numeric':
        target    = rule['target_field']
        condition = rule['condition']    # above | below
        threshold = rule['threshold_value'] or 0
        fd = db.execute(
            'SELECT id, column_name, system_field FROM field_definitions WHERE key = ?',
            (target,)
        ).fetchone()

        for m in members:
            if fd and fd['system_field'] and fd['column_name']:
                row = db.execute(
                    f'SELECT {fd["column_name"]} AS val FROM members WHERE id = ?',
                    (m['id'],)
                ).fetchone()
                raw = row['val'] if row else None
            else:
                fid = fd['id'] if fd else None
                if not fid:
                    continue
                row = db.execute(
                    'SELECT value FROM member_field_values WHERE member_id = ? AND field_id = ?',
                    (m['id'], fid)
                ).fetchone()
                raw = row['value'] if row else None

            try:
                num = float(raw)
            except (TypeError, ValueError):
                continue

            if condition == 'above' and num > threshold:
                should_flag.add(m['id'])
            elif condition == 'below' and num < threshold:
                should_flag.add(m['id'])

    # ── Raise new flags ────────────────────────────────────────────────────────
    raised = 0
    for mid in should_flag:
        if mid not in existing_flags:
            db.execute(
                'INSERT INTO member_flags (member_id, rule_id, flagged_at, flagged_by) '
                'VALUES (?, ?, datetime("now"), "auto")',
                (mid, rule_id)
            )
            raised += 1

    # ── Auto-resolve flags where condition no longer met ───────────────────────
    # Resolves any existing flag whose member no longer meets the rule condition.
    # This covers both members whose data changed (condition cleared) and members
    # who left or moved session (no longer in the eligible set at all).
    resolved = 0
    if auto_res:
        for mid, flag_id in existing_flags.items():
            if mid not in should_flag:
                db.execute(
                    "UPDATE member_flags SET resolved_at = datetime('now'), "
                    "resolved_by = 'auto' WHERE id = ?",
                    (flag_id,)
                )
                resolved += 1

    return raised, resolved


def run_all_alert_rules():
    """Evaluate every active alert rule. Called by the nightly scheduler and the manual API."""
    db = _connect_db()
    db.row_factory = sqlite3.Row
    today = datetime.now().strftime('%Y-%m-%d')
    try:
        rules = db.execute(
            'SELECT * FROM alert_rules WHERE is_active = 1'
        ).fetchall()

        total_raised   = 0
        total_resolved = 0
        for rule in rules:
            try:
                raised, resolved = _run_alert_rule(db, rule, today)
                total_raised   += raised
                total_resolved += resolved
            except Exception as e:
                # Log but don't abort the full run if one rule errors
                print(f'[alerts] Rule {rule["id"]} ({rule["name"]}) error: {e}')

        db.execute(
            "UPDATE settings SET value = ?, updated_at = datetime('now') "
            "WHERE key = 'alerts_last_run'",
            (datetime.now().strftime('%Y-%m-%d %H:%M'),)
        )

        # ── System notification to admins when new flags are raised ───────────────
        if total_raised > 0:
            flag_word = 'flag' if total_raised == 1 else 'flags'
            try:
                send_notification(
                    sender_id=None,
                    title='⚑ Alert Flags Raised',
                    body=(f'{total_raised} new member {flag_word} raised by the automated '
                          f'alert check. Open the Members section to review.'),
                    notification_type='Urgent',
                    target_type='role',
                    target_value='admin',
                    is_system=1,
                    related_table='member_flags',
                    _db=db,
                )
            except Exception as _notif_exc:
                print(f'[alerts] Failed to send system notification: {_notif_exc}')

        db.commit()
        print(f'[alerts] Run complete — {total_raised} raised, {total_resolved} resolved')
        return total_raised, total_resolved
    except Exception as _run_exc:
        print(f'[alerts] Run failed: {_run_exc}')
        raise
    finally:
        db.close()


# ── Notifications API (v8.2) ──────────────────────────────────────────────────

@app.route('/api/notifications')
@permission_required('notifications.view')
def get_notifications():
    """Return notifications relevant to the current user, newest first."""
    user_id = session.get('user_id')
    db = get_db()
    user = db.execute('SELECT role, session_assigned FROM users WHERE id = ?', (user_id,)).fetchone()
    user_role    = user['role']             if user else None
    user_session = user['session_assigned'] if user else None

    notifs = db.execute('''
        SELECT n.*,
               CASE WHEN nr.read_at IS NOT NULL THEN 1 ELSE 0 END AS is_read,
               u.username AS sender_name
        FROM notifications n
        LEFT JOIN notification_reads nr
            ON n.id = nr.notification_id AND nr.user_id = ?
        LEFT JOIN users u ON u.id = n.sender_id
        WHERE n.target_type = 'all'
           OR (n.target_type = 'role'    AND n.target_value = ?)
           OR (n.target_type = 'session' AND n.target_value = ?)
           OR (n.target_type = 'users'   AND EXISTS (
               SELECT 1 FROM json_each(n.target_value)
               WHERE CAST(json_each.value AS INTEGER) = ?
           ))
        ORDER BY n.created_at DESC
        LIMIT 100
    ''', (user_id, user_role, user_session, user_id)).fetchall()

    unread = sum(1 for n in notifs if not n['is_read'])
    return jsonify({
        'notifications': [dict(n) for n in notifs],
        'unread_count':  unread,
    })


@app.route('/api/notifications/unread-count')
@permission_required('notifications.view')
def get_notifications_unread_count():
    """Lightweight endpoint for badge polling — returns just the unread count."""
    user_id = session.get('user_id')
    db = get_db()
    user = db.execute('SELECT role, session_assigned FROM users WHERE id = ?', (user_id,)).fetchone()
    user_role    = user['role']             if user else None
    user_session = user['session_assigned'] if user else None

    count = db.execute('''
        SELECT COUNT(*) AS n
        FROM notifications n
        WHERE (
            n.target_type = 'all'
            OR (n.target_type = 'role'    AND n.target_value = ?)
            OR (n.target_type = 'session' AND n.target_value = ?)
            OR (n.target_type = 'users'   AND EXISTS (
                SELECT 1 FROM json_each(n.target_value)
                WHERE CAST(json_each.value AS INTEGER) = ?
            ))
        )
        AND NOT EXISTS (
            SELECT 1 FROM notification_reads nr
            WHERE nr.notification_id = n.id AND nr.user_id = ?
        )
    ''', (user_role, user_session, user_id, user_id)).fetchone()['n']

    return jsonify({'unread_count': count})


@app.route('/api/notifications/mark-read', methods=['POST'])
@permission_required('notifications.view')
def mark_notifications_read():
    """Mark a specific list of notification IDs as read for the current user."""
    data = request.get_json() or {}
    ids  = data.get('notification_ids', [])
    if not ids:
        return jsonify({'error': 'notification_ids required'}), 400

    user_id = session['user_id']
    db = get_db()
    for nid in ids:
        db.execute(
            'INSERT OR IGNORE INTO notification_reads (notification_id, user_id) VALUES (?, ?)',
            (nid, user_id)
        )
    db.commit()
    return jsonify({'success': True})


@app.route('/api/notifications/mark-all-read', methods=['POST'])
@permission_required('notifications.view')
def mark_all_notifications_read():
    """Mark every unread notification visible to the current user as read."""
    user_id = session['user_id']
    db = get_db()
    user = db.execute('SELECT role, session_assigned FROM users WHERE id = ?', (user_id,)).fetchone()
    user_role    = user['role']             if user else None
    user_session = user['session_assigned'] if user else None

    unread_ids = db.execute('''
        SELECT n.id
        FROM notifications n
        WHERE (
            n.target_type = 'all'
            OR (n.target_type = 'role'    AND n.target_value = ?)
            OR (n.target_type = 'session' AND n.target_value = ?)
            OR (n.target_type = 'users'   AND EXISTS (
                SELECT 1 FROM json_each(n.target_value)
                WHERE CAST(json_each.value AS INTEGER) = ?
            ))
        )
        AND NOT EXISTS (
            SELECT 1 FROM notification_reads nr
            WHERE nr.notification_id = n.id AND nr.user_id = ?
        )
    ''', (user_role, user_session, user_id, user_id)).fetchall()

    for row in unread_ids:
        db.execute(
            'INSERT OR IGNORE INTO notification_reads (notification_id, user_id) VALUES (?, ?)',
            (row['id'], user_id)
        )
    db.commit()
    return jsonify({'success': True, 'marked': len(unread_ids)})


@app.route('/api/notifications/send', methods=['POST'])
@permission_required('notifications.send')
def send_custom_notification():
    """Send a custom notification to a targeted audience."""
    data         = request.get_json() or {}
    title        = (data.get('title') or '').strip()
    body         = (data.get('body')  or '').strip()
    ntype        = data.get('notification_type', 'Info')
    target_type  = data.get('target_type')
    target_value = data.get('target_value')

    if not title or not body or not target_type:
        return jsonify({'error': 'title, body and target_type are required'}), 400
    if ntype not in ('Info', 'Reminder', 'Urgent', 'Announcement'):
        return jsonify({'error': 'Invalid notification_type'}), 400
    if target_type not in ('all', 'role', 'users', 'session'):
        return jsonify({'error': 'Invalid target_type'}), 400

    send_notification(
        sender_id=session['user_id'],
        title=title,
        body=body,
        notification_type=ntype,
        target_type=target_type,
        target_value=target_value,
        is_system=0,
    )
    return jsonify({'success': True})


@app.route('/api/notifications/<int:notification_id>', methods=['DELETE'])
@permission_required('notifications.manage')
def delete_notification(notification_id):
    """Permanently delete a notification (admin / manage permission only)."""
    db = get_db()
    notif = db.execute('SELECT id FROM notifications WHERE id = ?', (notification_id,)).fetchone()
    if not notif:
        return jsonify({'error': 'Notification not found'}), 404
    db.execute('DELETE FROM notifications WHERE id = ?', (notification_id,))
    db.commit()
    log_action('notification.deleted', 'notifications', notification_id, {})
    return jsonify({'success': True})


# ── Alert Rules API ────────────────────────────────────────────────────────────

@app.route('/api/alert-rules')
@permission_required('alerts.view')
def api_alert_rules_list():
    """List all alert rules with active flag counts."""
    db = get_db()
    rows = db.execute('''
        SELECT ar.*,
               COUNT(CASE WHEN mf.resolved_at IS NULL THEN 1 END) AS active_flag_count
        FROM alert_rules ar
        LEFT JOIN member_flags mf ON mf.rule_id = ar.id
        GROUP BY ar.id
        ORDER BY ar.is_active DESC, ar.name
    ''').fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/alert-rules', methods=['POST'])
@permission_required('alerts.manage')
def api_alert_rules_create():
    """Create a new alert rule."""
    data = request.get_json() or {}
    name        = (data.get('name') or '').strip()
    rule_type   = (data.get('rule_type') or '').strip()
    flag_label  = (data.get('flag_label') or '').strip()
    flag_colour = (data.get('flag_colour') or '#f59e0b').strip()

    if not name:
        return jsonify({'error': 'Rule name is required'}), 400
    if rule_type not in ('attendance', 'date_field', 'empty_field', 'numeric'):
        return jsonify({'error': 'Invalid rule_type'}), 400
    if not flag_label:
        return jsonify({'error': 'Flag label is required'}), 400
    if flag_colour not in ALERT_COLOURS:
        return jsonify({'error': f'Colour must be one of: {", ".join(ALERT_COLOURS)}'}), 400

    target_field   = (data.get('target_field') or '').strip() or None
    condition      = (data.get('condition') or '').strip() or None
    threshold_val  = data.get('threshold_value')
    threshold_unit = (data.get('threshold_unit') or '').strip() or None
    applies_sess   = (data.get('applies_to_session') or '').strip() or None
    auto_resolve   = 1 if data.get('auto_resolve', True) else 0
    resolve_field  = (data.get('resolve_field') or '').strip() or None

    db = get_db()
    cur = db.execute(
        'INSERT INTO alert_rules (name, rule_type, target_field, condition, '
        'threshold_value, threshold_unit, applies_to_session, flag_label, '
        'flag_colour, auto_resolve, resolve_field, is_active, created_by) '
        'VALUES (?,?,?,?,?,?,?,?,?,?,?,1,?)',
        (name, rule_type, target_field, condition, threshold_val, threshold_unit,
         applies_sess, flag_label, flag_colour, auto_resolve, resolve_field,
         session['user_id'])
    )
    db.commit()
    log_action('create_alert_rule', 'alert_rules', cur.lastrowid,
               {'name': name, 'rule_type': rule_type, 'flag_label': flag_label})
    return jsonify({'success': True, 'id': cur.lastrowid})


@app.route('/api/alert-rules/<int:rule_id>', methods=['PUT'])
@permission_required('alerts.manage')
def api_alert_rules_update(rule_id):
    """Update an existing alert rule."""
    rule = get_db().execute('SELECT * FROM alert_rules WHERE id = ?', (rule_id,)).fetchone()
    if not rule:
        return jsonify({'error': 'Not found'}), 404

    data = request.get_json() or {}
    flag_colour = (data.get('flag_colour') or rule['flag_colour']).strip()
    if flag_colour not in ALERT_COLOURS:
        return jsonify({'error': f'Colour must be one of: {", ".join(ALERT_COLOURS)}'}), 400

    db = get_db()
    db.execute(
        'UPDATE alert_rules SET name=?, rule_type=?, target_field=?, condition=?, '
        'threshold_value=?, threshold_unit=?, applies_to_session=?, flag_label=?, '
        'flag_colour=?, auto_resolve=?, resolve_field=?, is_active=? WHERE id=?',
        (
            (data.get('name') or rule['name']).strip(),
            (data.get('rule_type') or rule['rule_type']).strip(),
            (data.get('target_field') or rule['target_field'] or None),
            (data.get('condition') or rule['condition'] or None),
            data.get('threshold_value', rule['threshold_value']),
            (data.get('threshold_unit') or rule['threshold_unit'] or None),
            (data.get('applies_to_session') or rule['applies_to_session'] or None),
            (data.get('flag_label') or rule['flag_label']).strip(),
            flag_colour,
            1 if data.get('auto_resolve', bool(rule['auto_resolve'])) else 0,
            (data.get('resolve_field') or rule['resolve_field'] or None),
            1 if data.get('is_active', bool(rule['is_active'])) else 0,
            rule_id,
        )
    )
    db.commit()
    log_action('update_alert_rule', 'alert_rules', rule_id, {'name': data.get('name', rule['name'])})
    return jsonify({'success': True})


@app.route('/api/alert-rules/<int:rule_id>', methods=['DELETE'])
@permission_required('alerts.manage')
def api_alert_rules_delete(rule_id):
    """Deactivate (soft-delete) a rule and resolve all its open flags."""
    db = get_db()
    rule = db.execute('SELECT * FROM alert_rules WHERE id = ?', (rule_id,)).fetchone()
    if not rule:
        return jsonify({'error': 'Not found'}), 404
    db.execute("UPDATE alert_rules SET is_active = 0 WHERE id = ?", (rule_id,))
    db.execute(
        "UPDATE member_flags SET resolved_at = datetime('now'), resolved_by = 'rule_deleted' "
        "WHERE rule_id = ? AND resolved_at IS NULL",
        (rule_id,)
    )
    db.commit()
    log_action('deactivate_alert_rule', 'alert_rules', rule_id, {'name': rule['name']})
    return jsonify({'success': True})


@app.route('/api/alert-rules/<int:rule_id>/permanent-delete', methods=['POST'])
@permission_required('alerts.manage')
def api_alert_rules_permanent_delete(rule_id):
    """Permanently delete a rule and all its flag history. Irreversible."""
    db = get_db()
    rule = db.execute('SELECT * FROM alert_rules WHERE id = ?', (rule_id,)).fetchone()
    if not rule:
        return jsonify({'error': 'Not found'}), 404
    flag_count = db.execute(
        'SELECT COUNT(*) AS n FROM member_flags WHERE rule_id = ?', (rule_id,)
    ).fetchone()['n']
    db.execute('DELETE FROM member_flags WHERE rule_id = ?', (rule_id,))
    db.execute('DELETE FROM alert_rules WHERE id = ?', (rule_id,))
    db.commit()
    log_action('delete_alert_rule', 'alert_rules', rule_id, {
        'name': rule['name'], 'flags_deleted': flag_count
    })
    return jsonify({'success': True})


@app.route('/api/alert-rules/run', methods=['POST'])
@permission_required('alerts.run')
def api_alert_rules_run():
    """Manually trigger a full evaluation of all active alert rules."""
    raised, resolved = run_all_alert_rules()
    log_action('run_alert_rules', 'alert_rules', None,
               {'raised': raised, 'resolved': resolved, 'triggered_by': 'manual'})
    return jsonify({'success': True, 'raised': raised, 'resolved': resolved})


@app.route('/api/alerts/summary')
@permission_required('alerts.view')
def api_alerts_summary():
    """Per-rule active flag counts — used by the dashboard widget."""
    db     = get_db()
    scoped = _assigned_session()

    if scoped is None:
        rows = db.execute('''
            SELECT ar.id, ar.name, ar.flag_label, ar.flag_colour,
                   COUNT(CASE WHEN mf.resolved_at IS NULL THEN 1 END) AS flag_count
            FROM alert_rules ar
            LEFT JOIN member_flags mf ON mf.rule_id = ar.id
            WHERE ar.is_active = 1
            GROUP BY ar.id
            ORDER BY flag_count DESC, ar.name
        ''').fetchall()
    else:
        rows = db.execute('''
            SELECT ar.id, ar.name, ar.flag_label, ar.flag_colour,
                   COUNT(CASE WHEN mf.resolved_at IS NULL AND m.session = ? THEN 1 END) AS flag_count
            FROM alert_rules ar
            LEFT JOIN member_flags mf ON mf.rule_id = ar.id
            LEFT JOIN members m ON m.id = mf.member_id
            WHERE ar.is_active = 1
            GROUP BY ar.id
            ORDER BY flag_count DESC, ar.name
        ''', (scoped,)).fetchall()

    last_run = db.execute(
        "SELECT value FROM settings WHERE key = 'alerts_last_run'"
    ).fetchone()

    return jsonify({
        'rules':    [dict(r) for r in rows],
        'last_run': last_run['value'] if last_run else '',
    })


@app.route('/api/members/<int:member_id>/flags')
@permission_required('members.view')
def api_member_flags(member_id):
    """Return all flags (active and resolved) for a member."""
    db = get_db()
    scoped = _assigned_session()
    member = db.execute('SELECT * FROM members WHERE id = ?', (member_id,)).fetchone()
    if not member:
        return jsonify({'error': 'Not found'}), 404
    if scoped is not None and (member['session'] or '') != scoped:
        return jsonify({'error': 'Forbidden'}), 403

    flags = db.execute('''
        SELECT mf.id, mf.flagged_at, mf.flagged_by, mf.resolved_at, mf.resolved_by, mf.note,
               ar.id AS rule_id, ar.name AS rule_name, ar.flag_label, ar.flag_colour, ar.rule_type
        FROM member_flags mf
        JOIN alert_rules ar ON ar.id = mf.rule_id
        WHERE mf.member_id = ?
        ORDER BY mf.flagged_at DESC
    ''', (member_id,)).fetchall()

    return jsonify([dict(f) for f in flags])


@app.route('/api/members/<int:member_id>/flags/<int:flag_id>/dismiss', methods=['POST'])
@permission_required('alerts.dismiss')
def api_member_flag_dismiss(member_id, flag_id):
    """Manually dismiss a flag from a member, with an optional note."""
    db   = get_db()
    flag = db.execute(
        'SELECT * FROM member_flags WHERE id = ? AND member_id = ?',
        (flag_id, member_id)
    ).fetchone()
    if not flag:
        return jsonify({'error': 'Flag not found'}), 404
    if flag['resolved_at']:
        return jsonify({'error': 'Flag already resolved'}), 400

    data = request.get_json() or {}
    note = (data.get('note') or '').strip() or None

    db.execute(
        "UPDATE member_flags SET resolved_at = datetime('now'), resolved_by = ?, note = ? "
        "WHERE id = ?",
        (str(session['user_id']), note, flag_id)
    )
    db.commit()

    member = db.execute('SELECT first_name, surname FROM members WHERE id = ?',
                        (member_id,)).fetchone()
    rule   = db.execute('SELECT name FROM alert_rules WHERE id = ?',
                        (flag['rule_id'],)).fetchone()
    log_action('dismiss_flag', 'member_flags', flag_id, {
        'member': f"{member['first_name'] or ''} {member['surname'] or ''}".strip(),
        'rule':   rule['name'] if rule else str(flag['rule_id']),
        'note':   note,
    })
    return jsonify({'success': True})


# ── Admin page route ────────────────────────────────────────────────────────────

@app.route('/admin/alerts')
@permission_required('alerts.view')
def admin_alerts_page():
    """Alert Rules admin page — rule builder and manual run trigger."""
    db           = get_db()
    session_types = get_session_types()
    # Field picker: all active field_definitions available for rule targeting
    fields = db.execute(
        'SELECT key, label, field_type, system_field FROM field_definitions '
        'WHERE active = 1 ORDER BY label'
    ).fetchall()
    last_run = get_setting('alerts_last_run', '')
    return render_template(
        'admin/alerts.html',
        field_definitions=[dict(f) for f in fields],
        alerts_last_run=last_run,
        **tpl_ctx(),
    )


# ── BLUEPRINT: QR quick-session (v8.3) ───────────────────────────────────────
#
# Public mobile page + API for self sign-in / sign-out via QR code.
# A session token (quick_signin_tokens table) is the sole gate on this flow.
# All public API endpoints are CSRF-exempt; they validate via the token instead.

# Simple in-memory rate limiter — (ip, endpoint, minute_bucket) → request count.
# No external dependency needed for a youth-club scale deployment.
_rl_store: dict = {}


def _rl_check(endpoint: str, max_per_min: int) -> bool:
    """Return True if the request is within rate limit, False if exceeded."""
    ip     = request.remote_addr or 'unknown'
    bucket = datetime.now().strftime('%Y%m%d%H%M')
    key    = (ip, endpoint, bucket)
    count  = _rl_store.get(key, 0) + 1
    _rl_store[key] = count
    # Prune old buckets (keep store small)
    if len(_rl_store) > 2000:
        cur = datetime.now().strftime('%Y%m%d%H%M')
        for k in list(_rl_store.keys()):
            if k[2] != cur:
                del _rl_store[k]
    return count <= max_per_min


def _get_or_create_qr_token(session_type: str) -> str:
    """Return today's active QR token for session_type, creating one if needed."""
    today = datetime.now().strftime('%Y-%m-%d')
    db    = get_db()
    row   = db.execute(
        '''SELECT token FROM quick_signin_tokens
           WHERE session_type = ? AND session_date = ? AND invalidated_at IS NULL''',
        (session_type, today),
    ).fetchone()
    if row:
        return row['token']
    token = secrets.token_urlsafe(32)
    db.execute(
        'INSERT INTO quick_signin_tokens (token, session_type, session_date) VALUES (?,?,?)',
        (token, session_type, today),
    )
    db.commit()
    return token


def _ensure_qr_tokens_for_today():
    """Pre-create tokens for every active session type (called when register page loads)."""
    for st in get_session_types():
        _get_or_create_qr_token(st['name'])


def _validate_qr_token(token: str):
    """Return the token row if valid for today and not invalidated, else None."""
    today = datetime.now().strftime('%Y-%m-%d')
    db    = get_db()
    row   = db.execute(
        '''SELECT * FROM quick_signin_tokens
           WHERE token = ? AND session_date = ? AND invalidated_at IS NULL''',
        (token, today),
    ).fetchone()
    return row


def _invalidate_qr_token_for_session(session_type: str, session_date: str):
    """Stamp invalidated_at on all active tokens for this session (called on register complete)."""
    db = get_db()
    db.execute(
        '''UPDATE quick_signin_tokens
           SET invalidated_at = datetime('now')
           WHERE session_type = ? AND session_date = ? AND invalidated_at IS NULL''',
        (session_type, session_date),
    )
    db.commit()


# ── Public mobile landing page ────────────────────────────────────────────────

@app.route('/quick-session')
@csrf.exempt
def quick_session_page():
    """Mobile self-sign-in / sign-out page — no login required."""
    return render_template('quick_session.html', club_name=CLUB_NAME)


# ── Authenticated: token management (called by register page JS) ──────────────

@app.route('/api/quick-signin/token/<session_type>')
@login_required
def api_qr_token_get(session_type):
    """Return (or create) today's QR token + full URL for the given session."""
    if session_type not in get_valid_session_names():
        return jsonify({'error': 'Invalid session type'}), 400
    token    = _get_or_create_qr_token(session_type)
    base_url = request.host_url.rstrip('/')
    qr_url   = f'{base_url}/quick-session?t={token}'
    db       = get_db()
    today    = datetime.now().strftime('%Y-%m-%d')
    qr_in    = db.execute(
        "SELECT COUNT(*) FROM attendance WHERE session_date=? AND session_type=? AND source='qr-self' AND signed_in_at IS NOT NULL",
        (today, session_type),
    ).fetchone()[0]
    qr_out   = db.execute(
        "SELECT COUNT(*) FROM attendance WHERE session_date=? AND session_type=? AND source='qr-self' AND signed_out_at IS NOT NULL",
        (today, session_type),
    ).fetchone()[0]
    return jsonify({'token': token, 'qr_url': qr_url, 'qr_signin_count': qr_in, 'qr_signout_count': qr_out})


@app.route('/api/quick-signin/token/<session_type>/regenerate', methods=['POST'])
@permission_required('register.qr_manage')
def api_qr_token_regenerate(session_type):
    """Invalidate existing token and issue a fresh one."""
    if session_type not in get_valid_session_names():
        return jsonify({'error': 'Invalid session type'}), 400
    today = datetime.now().strftime('%Y-%m-%d')
    db    = get_db()
    db.execute(
        '''UPDATE quick_signin_tokens SET invalidated_at = datetime('now')
           WHERE session_type = ? AND session_date = ? AND invalidated_at IS NULL''',
        (session_type, today),
    )
    db.commit()
    token    = _get_or_create_qr_token(session_type)
    base_url = request.host_url.rstrip('/')
    qr_url   = f'{base_url}/quick-session?t={token}'
    log_action('qr_token_regenerated', 'quick_signin_tokens', None, {'session_type': session_type})
    return jsonify({'token': token, 'qr_url': qr_url})


# ── Public display-token endpoint (unauthenticated, for the TV display) ───────

@app.route('/api/quick-signin/display-token/<session_type>')
@csrf.exempt
def api_qr_display_token(session_type):
    """Return the current QR URL for the TV display, or null if none active."""
    if session_type not in get_valid_session_names():
        return jsonify({'qr_url': None})
    today = datetime.now().strftime('%Y-%m-%d')
    db    = get_db()
    row   = db.execute(
        '''SELECT token FROM quick_signin_tokens
           WHERE session_type = ? AND session_date = ? AND invalidated_at IS NULL''',
        (session_type, today),
    ).fetchone()
    if not row:
        return jsonify({'qr_url': None})
    # Also suppress if register is locked
    if _is_register_locked(session_type, today):
        return jsonify({'qr_url': None})
    base_url = request.host_url.rstrip('/')
    return jsonify({'qr_url': f'{base_url}/quick-session?t={row["token"]}'})


# ── Public API: verify / search / signin / signout ────────────────────────────

@app.route('/api/quick-signin/verify')
@csrf.exempt
def api_qr_verify():
    """Validate a QR token and return which modes are enabled."""
    token = request.args.get('t', '').strip()
    if not token:
        return jsonify({'valid': False, 'reason': 'No token provided.'})
    row = _validate_qr_token(token)
    if not row:
        return jsonify({'valid': False, 'reason': 'This sign-in link has expired or is no longer valid.'})
    sess_type = row['session_type']
    sess_date = row['session_date']
    if _is_register_locked(sess_type, sess_date):
        return jsonify({'valid': False, 'reason': 'The session register has been completed and is now locked.'})
    return jsonify({
        'valid':            True,
        'session_type':     sess_type,
        'session_date':     sess_date,
        'signin_enabled':   get_setting('quick_signin_enabled',  'true')  == 'true',
        'signout_enabled':  get_setting('quick_signout_enabled', 'false') == 'true',
    })


@app.route('/api/quick-signin/search')
@csrf.exempt
def api_qr_search():
    """Search members by first name for the QR session page."""
    if not _rl_check('qr_search', 15):
        return jsonify({'error': 'Too many requests — please slow down.'}), 429

    token = request.args.get('t', '').strip()
    q     = request.args.get('q', '').strip()
    mode  = request.args.get('mode', 'signin').strip()   # 'signin' | 'signout'

    if len(q) < 2:
        return jsonify([])
    tok_row = _validate_qr_token(token)
    if not tok_row:
        return jsonify({'error': 'Invalid or expired token.'}), 403

    sess_type = tok_row['session_type']
    sess_date = tok_row['session_date']

    if _is_register_locked(sess_type, sess_date):
        return jsonify({'error': 'Register is locked.'}), 403

    db = get_db()

    if mode == 'signout':
        # Only show members currently signed IN (not yet signed out)
        rows = db.execute(
            '''SELECT m.id, m.first_name, m.surname,
                      1 AS already_signed_in
               FROM members m
               JOIN attendance a ON a.member_id = m.id
                 AND a.session_date = ? AND a.session_type = ?
                 AND a.signed_in_at IS NOT NULL AND a.signed_out_at IS NULL
               WHERE m.status = 'Active'
                 AND LOWER(m.first_name) LIKE LOWER(?)
               ORDER BY m.first_name, m.surname
               LIMIT 20''',
            (sess_date, sess_type, q + '%'),
        ).fetchall()
    else:
        # Sign-in: all active members for this session, flag those already signed in
        rows = db.execute(
            '''SELECT m.id, m.first_name, m.surname,
                      CASE WHEN a.signed_in_at IS NOT NULL THEN 1 ELSE 0 END AS already_signed_in
               FROM members m
               LEFT JOIN attendance a ON a.member_id = m.id
                 AND a.session_date = ? AND a.session_type = ?
               WHERE m.status = 'Active'
                 AND (m.session = ? OR m.member_type = 'staff')
                 AND LOWER(m.first_name) LIKE LOWER(?)
               ORDER BY m.first_name, m.surname
               LIMIT 20''',
            (sess_date, sess_type, sess_type, q + '%'),
        ).fetchall()

    return jsonify([dict(r) for r in rows])


@app.route('/api/quick-signin/signin', methods=['POST'])
@csrf.exempt
def api_qr_signin():
    """Sign a member in via QR token. Bulletproof against duplicates and races."""
    if not _rl_check('qr_signin', 10):
        return jsonify({'error': 'Too many requests — please slow down.'}), 429

    data      = request.get_json() or {}
    token     = data.get('token', '').strip()
    member_id = data.get('member_id')

    if not get_setting('quick_signin_enabled', 'true') == 'true':
        return jsonify({'error': 'QR sign-in is not enabled.'}), 403

    tok_row = _validate_qr_token(token)
    if not tok_row:
        return jsonify({'error': 'Invalid or expired token.'}), 403

    sess_type = tok_row['session_type']
    sess_date = tok_row['session_date']

    if _is_register_locked(sess_type, sess_date):
        return jsonify({'error': 'Register is locked.'}), 403

    db     = get_db()
    member = db.execute(
        "SELECT id, first_name FROM members WHERE id = ? AND status = 'Active'",
        (member_id,),
    ).fetchone()
    if not member:
        return jsonify({'error': 'Member not found.'}), 404

    first_name = member['first_name'] or 'there'
    now        = datetime.now().strftime('%H:%M')

    # Check if already signed in
    existing = db.execute(
        'SELECT id, signed_in_at FROM attendance WHERE member_id=? AND session_date=? AND session_type=?',
        (member_id, sess_date, sess_type),
    ).fetchone()

    if existing and existing['signed_in_at']:
        msg = get_setting('quick_signin_already_msg', "You're already signed in, {name}! See you inside 👋")
        return jsonify({
            'success':          True,
            'already_signed_in': True,
            'first_name':       first_name,
            'welcome_message':  msg.replace('{name}', first_name),
        })

    # INSERT OR IGNORE protects against concurrent double-taps
    # recorded_by stays NULL (no logged-in user); source = 'qr-self' identifies the channel
    db.execute(
        '''INSERT OR IGNORE INTO attendance
               (member_id, session_date, session_type, signed_in_at, recorded_by, source)
           VALUES (?,?,?,?,NULL,'qr-self')''',
        (member_id, sess_date, sess_type, now),
    )
    if db.execute('SELECT changes()').fetchone()[0] == 0:
        # Row already existed (race) — treat as already signed in
        msg = get_setting('quick_signin_already_msg', "You're already signed in, {name}! See you inside 👋")
        db.commit()
        return jsonify({
            'success':          True,
            'already_signed_in': True,
            'first_name':       first_name,
            'welcome_message':  msg.replace('{name}', first_name),
        })

    # Auto-resolve attendance alert flags (mirrors api_attendance_signin logic)
    att_flags = db.execute(
        "SELECT mf.id FROM member_flags mf "
        "JOIN alert_rules ar ON ar.id = mf.rule_id "
        "WHERE mf.member_id = ? AND mf.resolved_at IS NULL "
        "AND ar.rule_type = 'attendance' AND ar.auto_resolve = 1",
        (member_id,),
    ).fetchall()
    for flag in att_flags:
        db.execute(
            "UPDATE member_flags SET resolved_at = datetime('now'), resolved_by = 'auto' WHERE id = ?",
            (flag['id'],),
        )

    db.commit()
    _touch_attendance()

    msg = get_setting('quick_signin_welcome_msg', 'Welcome, {name}! Great to see you tonight! 🎉')
    return jsonify({
        'success':          True,
        'already_signed_in': False,
        'first_name':       first_name,
        'welcome_message':  msg.replace('{name}', first_name),
    })


@app.route('/api/quick-signin/signout', methods=['POST'])
@csrf.exempt
def api_qr_signout():
    """Sign a member out via QR token. Only active when quick_signout_enabled = true."""
    if not _rl_check('qr_signout', 10):
        return jsonify({'error': 'Too many requests — please slow down.'}), 429

    if get_setting('quick_signout_enabled', 'false') != 'true':
        return jsonify({'error': 'QR sign-out is not enabled for this organisation.'}), 403

    data      = request.get_json() or {}
    token     = data.get('token', '').strip()
    member_id = data.get('member_id')

    tok_row = _validate_qr_token(token)
    if not tok_row:
        return jsonify({'error': 'Invalid or expired token.'}), 403

    sess_type = tok_row['session_type']
    sess_date = tok_row['session_date']

    if _is_register_locked(sess_type, sess_date):
        return jsonify({'error': 'Register is locked.'}), 403

    db     = get_db()
    member = db.execute(
        "SELECT id, first_name FROM members WHERE id = ? AND status = 'Active'",
        (member_id,),
    ).fetchone()
    if not member:
        return jsonify({'error': 'Member not found.'}), 404

    first_name = member['first_name'] or 'there'
    now        = datetime.now().strftime('%H:%M')

    existing = db.execute(
        'SELECT id, signed_in_at, signed_out_at FROM attendance WHERE member_id=? AND session_date=? AND session_type=?',
        (member_id, sess_date, sess_type),
    ).fetchone()

    if not existing or not existing['signed_in_at']:
        return jsonify({'error': 'This member is not currently signed in.'}), 400

    if existing['signed_out_at']:
        msg = get_setting('quick_signout_already_msg', "You're already signed out, {name}. Safe journey home!")
        return jsonify({
            'success':           True,
            'already_signed_out': True,
            'first_name':        first_name,
            'goodbye_message':   msg.replace('{name}', first_name),
        })

    # UPDATE … WHERE signed_out_at IS NULL protects against concurrent races
    # source already set to 'qr-self' on the original sign-in row if via QR;
    # update it here too in case the member signed in via iPad but is signing out via QR
    db.execute(
        '''UPDATE attendance SET signed_out_at = ?, source = 'qr-self'
           WHERE id = ? AND signed_out_at IS NULL''',
        (now, existing['id']),
    )
    if db.execute('SELECT changes()').fetchone()[0] == 0:
        msg = get_setting('quick_signout_already_msg', "You're already signed out, {name}. Safe journey home!")
        db.commit()
        return jsonify({
            'success':           True,
            'already_signed_out': True,
            'first_name':        first_name,
            'goodbye_message':   msg.replace('{name}', first_name),
        })

    db.commit()
    _touch_attendance()

    msg = get_setting('quick_signout_goodbye_msg', 'Goodbye, {name}! See you next time 👋')
    return jsonify({
        'success':           True,
        'already_signed_out': False,
        'first_name':        first_name,
        'goodbye_message':   msg.replace('{name}', first_name),
    })


# ── APScheduler — nightly alert check at 02:00 ────────────────────────────────

def _start_scheduler():
    """Start the background scheduler for nightly alert rule evaluation.
    Guarded against double-start in Flask debug mode (reloader spawns two processes).
    """
    # In debug mode with the reloader the child process has WERKZEUG_RUN_MAIN=true;
    # only start the scheduler in the child so it doesn't run twice.
    if os.environ.get('FLASK_DEBUG') == '1':
        if os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
            return

    scheduler = BackgroundScheduler()
    scheduler.add_job(
        run_all_alert_rules,
        CronTrigger(hour=2, minute=0),
        id='nightly_alert_check',
        replace_existing=True,
    )
    scheduler.start()
    print('[alerts] Nightly scheduler started (02:00 daily)')


_start_scheduler()


# ── Run ────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    # Auto-init DB on first run if it doesn't exist
    if not os.path.exists(DATABASE):
        print('First run — initialising database…')
        init_db()
    _debug = os.environ.get('FLASK_DEBUG', '0') == '1'
    # PORT can be set by the service manager (launchd / systemd) for multi-instance
    _port = int(os.environ.get('PORT', 5001))
    app.run(debug=_debug, host='0.0.0.0', port=_port, threaded=True)
