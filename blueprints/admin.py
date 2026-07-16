"""
AYC Portal — Admin blueprint.
Routes:
  /api/settings, /api/admin/branding/*
  /api/admin/users/*
  /api/admin/staff-roles/*
  /api/admin/permissions, /api/admin/roles/*
  /api/admin/smtp-profiles/*
  /api/tags, /api/admin/tags/*
  /api/members/<id>/tags/*
  /api/admin/member-types/*, /api/admin/field-definitions/*
  /api/admin/logs/*, /api/admin/maintenance/*
  /api/admin/import/*
  /api/register/notes/*  (session notes)
"""

import csv as _csv_mod
import glob
import json
import os
import re as _re
import shutil
import sqlcipher3 as sqlite3
import tempfile
import uuid as _uuid_mod
from datetime import datetime

import bcrypt
from flask import Blueprint, Response, current_app, g, jsonify, request, send_file, session

from config import (
    BRANDING_DIR, DATABASE, INSTANCE_DIR, LOG_DIR,
    ROLE_DISPLAY_NAMES,
)
from extensions import csrf
from helpers import (
    get_db, log_action, permission_required, has_permission, club_slug,
    _assigned_session, _connect_db, _validate_hex_colour, _invalidate_brand_cache,
    get_brand_settings, get_session_types, validate_password,
    get_setting, encrypt_file, decrypt_file,
)

bp = Blueprint('admin', __name__)

ALLOWED_LOGO_EXTENSIONS = {'png', 'jpg', 'jpeg', 'svg', 'webp', 'gif'}


# ── Settings ──────────────────────────────────────────────────────────────────

@bp.route('/api/settings')
@permission_required('admin.settings')
def api_settings_get():
    """Return all settings as a key/value dict."""
    db   = get_db()
    rows = db.execute('SELECT key, value FROM settings').fetchall()
    return jsonify({r['key']: r['value'] for r in rows})


@bp.route('/api/settings', methods=['POST'])
@permission_required('admin.settings')
def api_settings_save():
    """Save one or more settings.
    Only keys in ALLOWED_KEYS are persisted — any other key in the POST body
    is silently ignored.  Add keys here when new configurable settings are introduced.
    Alert-rule thresholds live in the alert_rules table (since v8.0) — not here.
    """
    data = request.get_json() or {}
    ALLOWED_KEYS = {
        # QR quick sign-in settings (attendance_settings page)
        'quick_signin_enabled',
        'quick_signout_enabled',
        'quick_signin_welcome_msg',
        'quick_signin_already_msg',
        'quick_signout_goodbye_msg',
        'quick_signout_already_msg',
        # Member ID format
        'member_id_prefix',
        'member_id_padding',
        # Registration settings
        'require_approval',
        # Auto sign-out threshold (minutes after session end)
        'auto_signout_minutes',
    }
    db   = get_db()
    saved = {}
    for key, value in data.items():
        if key not in ALLOWED_KEYS:
            continue
        db.execute(
            'INSERT INTO settings (key, value, updated_at, updated_by) VALUES (?, ?, datetime("now"), ?)'
            ' ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at,'
            ' updated_by = excluded.updated_by',
            (key, str(value), session['user_id'])
        )
        saved[key] = value
    db.commit()
    if saved:
        log_action('update_settings', 'settings', None, {'changes': saved})
    return jsonify({'success': True})


# ── Branding ──────────────────────────────────────────────────────────────────

@bp.route('/api/admin/branding')
@permission_required('admin.branding')
def api_branding_get():
    brand = get_brand_settings()
    return jsonify({
        'accent':     brand.get('brand_accent', '#0096b4'),
        'nav_style':  brand.get('brand_nav_style', 'dark'),
        'club_name':  brand.get('brand_club_name', ''),
        'short_name': brand.get('brand_short_name', ''),
        'has_logo':   bool(brand.get('brand_logo_file')),
    })


@bp.route('/api/admin/branding', methods=['POST'])
@permission_required('admin.branding')
def api_branding_save():
    data    = request.get_json() or {}
    updates = {}

    if 'accent' in data:
        v = str(data['accent']).strip()
        if not _re.match(r'^#[0-9a-fA-F]{6}$', v):
            return jsonify({'error': 'accent must be a 6-digit hex colour e.g. #ff5500'}), 400
        updates['brand_accent'] = v

    if 'nav_style' in data:
        v = str(data['nav_style']).strip()
        if v not in ('dark', 'accent', 'white'):
            return jsonify({'error': 'nav_style must be dark, accent or white'}), 400
        updates['brand_nav_style'] = v

    if 'club_name' in data:
        updates['brand_club_name'] = str(data['club_name']).strip()[:120]

    if 'short_name' in data:
        updates['brand_short_name'] = str(data['short_name']).strip()[:30]

    if not updates:
        return jsonify({'success': True, 'message': 'Nothing to update'})

    db = get_db()
    for key, val in updates.items():
        db.execute(
            'INSERT INTO settings (key, value, updated_at, updated_by) VALUES (?, ?, datetime("now"), ?)'
            ' ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at,'
            ' updated_by=excluded.updated_by',
            (key, val, session['user_id'])
        )
    db.commit()
    _invalidate_brand_cache()
    log_action('update_branding', 'settings', None, updates)
    return jsonify({'success': True})


@bp.route('/api/admin/branding/logo', methods=['POST'])
@permission_required('admin.branding')
def api_branding_logo_upload():
    if 'logo' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['logo']
    if not f.filename:
        return jsonify({'error': 'No file selected'}), 400

    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    if ext not in ALLOWED_LOGO_EXTENSIONS:
        return jsonify({'error': f'File type .{ext} not allowed — use PNG, JPG, SVG or WebP'}), 400

    filename  = f'logo.{ext}'
    save_path = os.path.join(BRANDING_DIR, filename)

    for old in os.listdir(BRANDING_DIR):
        if old.startswith('logo.'):
            try:
                os.remove(os.path.join(BRANDING_DIR, old))
            except OSError:
                pass

    f.save(save_path)
    db = get_db()
    db.execute(
        'INSERT INTO settings (key, value, updated_at, updated_by) VALUES (?, ?, datetime("now"), ?)'
        ' ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at,'
        ' updated_by=excluded.updated_by',
        ('brand_logo_file', filename, session['user_id'])
    )
    db.commit()
    _invalidate_brand_cache()
    log_action('upload_branding_logo', 'settings', None, {'filename': filename})
    return jsonify({'success': True, 'filename': filename})


@bp.route('/api/admin/branding/logo', methods=['DELETE'])
@permission_required('admin.branding')
def api_branding_logo_delete():
    brand    = get_brand_settings()
    filename = brand.get('brand_logo_file', '')
    if filename:
        path = os.path.join(BRANDING_DIR, os.path.basename(filename))
        try:
            if os.path.isfile(path):
                os.remove(path)
        except OSError:
            pass

    db = get_db()
    db.execute(
        'INSERT INTO settings (key, value, updated_at, updated_by) VALUES (?, ?, datetime("now"), ?)'
        ' ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at,'
        ' updated_by=excluded.updated_by',
        ('brand_logo_file', '', session['user_id'])
    )
    db.commit()
    _invalidate_brand_cache()
    log_action('delete_branding_logo', 'settings', None, {})
    return jsonify({'success': True})


# ── Users CRUD ────────────────────────────────────────────────────────────────

@bp.route('/api/admin/users')
@permission_required('users.view')
def api_users_list():
    db     = get_db()
    scoped = _assigned_session()  # None (admin) or list of session names

    # Base query: user rows with their session names aggregated
    base_q = '''
        SELECT u.id, u.username, u.email, u.role, u.active,
               u.created_at, u.last_login,
               GROUP_CONCAT(st.name, ',') AS session_names_csv
        FROM users u
        LEFT JOIN user_sessions us ON us.user_id = u.id
        LEFT JOIN session_types st ON st.id = us.session_type_id AND st.active = 1
        {where}
        GROUP BY u.id
        ORDER BY u.username
    '''

    if scoped is not None:
        if not scoped:
            return jsonify([])
        placeholders = ','.join('?' * len(scoped))
        rows = db.execute(
            base_q.format(where=f"WHERE u.role != 'admin' AND st.name IN ({placeholders})"),
            scoped
        ).fetchall()
    else:
        rows = db.execute(base_q.format(where='')).fetchall()

    result = []
    for u in rows:
        d = dict(u)
        csv = d.pop('session_names_csv', None)
        d['session_names'] = csv.split(',') if csv else []
        result.append(d)
    return jsonify(result)


@bp.route('/api/admin/users', methods=['POST'])
@permission_required('users.create')
def api_users_create():
    data        = request.get_json() or {}
    username    = data.get('username', '').strip()
    password    = data.get('password', '')
    role        = data.get('role', 'readonly')
    email       = data.get('email', '').strip()
    session_ids = data.get('session_ids', [])   # list of session_type IDs

    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400
    pw_error = validate_password(password)
    if pw_error:
        return jsonify({'error': pw_error}), 400

    db              = get_db()
    target_role_row = db.execute('SELECT id, permissions FROM roles WHERE name = ?', (role,)).fetchone()
    if not target_role_row:
        return jsonify({'error': 'Invalid role'}), 400

    target_perms = json.loads(target_role_row['permissions'])
    if 'users.create.admin' in target_perms and not has_permission('users.create.admin'):
        return jsonify({'error': 'You do not have permission to assign this role'}), 403

    is_admin_role = 'admin.maintenance' in target_perms
    if not is_admin_role and not session_ids:
        return jsonify({'error': 'At least one session must be assigned for non-admin users'}), 400

    # Validate that all supplied session IDs exist and are active
    valid_session_map = {s['id']: s['name'] for s in get_session_types()}
    for sid in session_ids:
        if sid not in valid_session_map:
            return jsonify({'error': f'Invalid or inactive session ID: {sid}'}), 400

    # Scoped (non-admin) creators can only assign sessions they themselves have access to
    scoped = _assigned_session()
    if scoped is not None:
        my_names = set(scoped)
        for sid in session_ids:
            if valid_session_map[sid] not in my_names:
                return jsonify({'error': 'You can only assign sessions you have access to'}), 403

    pw_hash    = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    active_id  = session_ids[0] if session_ids else None
    try:
        cur = db.execute(
            'INSERT INTO users (username, email, password_hash, role, role_id, active_session_id)'
            ' VALUES (?,?,?,?,?,?)',
            (username, email, pw_hash, role, target_role_row['id'], active_id)
        )
        new_user_id = cur.lastrowid
        for sid in session_ids:
            db.execute(
                'INSERT OR IGNORE INTO user_sessions (user_id, session_type_id) VALUES (?,?)',
                (new_user_id, sid)
            )
        db.commit()
        log_action('create_user', 'users', new_user_id,
                   {'username': username, 'role': role,
                    'sessions': [valid_session_map[s] for s in session_ids],
                    'created_by': session.get('username')})
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Username already exists'}), 409


@bp.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@permission_required('users.edit')
def api_users_update(user_id):
    data    = request.get_json() or {}
    db      = get_db()
    updates = []
    params  = []

    if user_id == session['user_id'] and data.get('active') is False:
        return jsonify({'error': 'You cannot deactivate your own account'}), 400

    before_row = db.execute(
        'SELECT u.username, u.email, u.role, u.active, '
        'GROUP_CONCAT(st.name, ",") AS session_names_csv '
        'FROM users u '
        'LEFT JOIN user_sessions us ON us.user_id = u.id '
        'LEFT JOIN session_types st ON st.id = us.session_type_id '
        'WHERE u.id = ? GROUP BY u.id',
        (user_id,)
    ).fetchone()

    target_perms = []
    if 'role' in data:
        target_role_row = db.execute(
            'SELECT id, permissions FROM roles WHERE name = ?', (data['role'],)
        ).fetchone()
        if not target_role_row:
            return jsonify({'error': 'Invalid role'}), 400
        target_perms = json.loads(target_role_row['permissions'])
        if 'users.create.admin' in target_perms and not has_permission('users.create.admin'):
            return jsonify({'error': 'You do not have permission to assign this role'}), 403

    # Scoped users can only manage users who share at least one of their sessions
    scoped = _assigned_session()
    if scoped is not None:
        target_sess_rows = db.execute(
            'SELECT st.name FROM user_sessions us '
            'JOIN session_types st ON st.id = us.session_type_id '
            'WHERE us.user_id = ?', (user_id,)
        ).fetchall()
        target_sess_names = {r['name'] for r in target_sess_rows}
        target_role = db.execute('SELECT role FROM users WHERE id = ?', (user_id,)).fetchone()
        if not target_role or target_role['role'] == 'admin':
            return jsonify({'error': 'Forbidden'}), 403
        if not target_sess_names.intersection(set(scoped)):
            return jsonify({'error': 'You can only manage users in your own session'}), 403

    # Validate and apply session_ids change
    valid_session_map = {s['id']: s['name'] for s in get_session_types()}
    if 'session_ids' in data:
        new_ids = data['session_ids']
        # Scoped creators may not assign sessions outside their own access
        if scoped is not None:
            my_names = set(scoped)
            for sid in new_ids:
                if valid_session_map.get(sid) not in my_names:
                    return jsonify({'error': 'You can only assign sessions you have access to'}), 403
        # Enforce non-admin must have at least one session
        eff_role_name = data.get('role', before_row['role'] if before_row else '')
        eff_role_row  = db.execute('SELECT permissions FROM roles WHERE name = ?', (eff_role_name,)).fetchone()
        eff_perms     = json.loads(eff_role_row['permissions']) if eff_role_row else []
        if 'admin.maintenance' not in eff_perms and not new_ids:
            return jsonify({'error': 'At least one session must be assigned for non-admin users'}), 400
        # Replace junction table entries
        db.execute('DELETE FROM user_sessions WHERE user_id = ?', (user_id,))
        for sid in new_ids:
            if sid not in valid_session_map:
                return jsonify({'error': f'Invalid session ID: {sid}'}), 400
            db.execute(
                'INSERT OR IGNORE INTO user_sessions (user_id, session_type_id) VALUES (?,?)',
                (user_id, sid)
            )
        # Reset active_session_id if it's no longer in the new set
        if new_ids:
            updates.append('active_session_id = ?'); params.append(new_ids[0])
        else:
            updates.append('active_session_id = NULL')

    if 'email' in data:
        updates.append('email = ?'); params.append(data['email'])

    if 'role' in data:
        # v12.41: reuse the role row validated at the top of this function rather
        # than re-querying — the second lookup could race a concurrent role
        # deletion and set role without role_id (desync).
        updates.append('role = ?');    params.append(data['role'])
        updates.append('role_id = ?'); params.append(target_role_row['id'])

    if 'active' in data:
        updates.append('active = ?'); params.append(1 if data['active'] else 0)

    if 'password' in data and data['password']:
        pw_error = validate_password(data['password'])
        if pw_error:
            return jsonify({'error': pw_error}), 400
        pw_hash = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        updates.append('password_hash = ?'); params.append(pw_hash)

    if updates:
        params.append(user_id)
        db.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ?", params)

    db.commit()

    if before_row:
        field_changes = {}
        if 'email' in data and data.get('email') != before_row['email']:
            field_changes['email'] = {'from': before_row['email'], 'to': data['email']}
        if 'role' in data and data['role'] != before_row['role']:
            field_changes['role'] = {'from': before_row['role'], 'to': data['role']}
        if 'session_ids' in data:
            before_sess = before_row['session_names_csv'] or ''
            after_sess  = ','.join(valid_session_map.get(s, str(s)) for s in data['session_ids'])
            if before_sess != after_sess:
                field_changes['sessions'] = {'from': before_sess, 'to': after_sess}
        if 'active' in data:
            new_active = 1 if data['active'] else 0
            if new_active != before_row['active']:
                field_changes['active'] = {'from': bool(before_row['active']), 'to': bool(new_active)}
        log_action('update_user', 'users', user_id, {
            'username':       before_row['username'],
            'changes':        field_changes,
            'password_reset': True if ('password' in data and data['password']) else None,
        })

    return jsonify({'success': True})


@bp.route('/api/admin/users/<int:user_id>/permanent', methods=['DELETE'])
@permission_required('users.delete')
def api_users_permanent_delete(user_id):
    """Permanently delete a portal user account. Requires {"confirm_username": "<username>"}."""
    data = request.get_json() or {}
    db   = get_db()

    if user_id == session['user_id']:
        return jsonify({'error': 'You cannot delete your own account'}), 400

    user = db.execute('SELECT id, username FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user:
        return jsonify({'error': 'User not found'}), 404

    confirm = (data.get('confirm_username') or '').strip().lower()
    if confirm != user['username'].lower():
        return jsonify({'error': 'Username confirmation does not match'}), 400

    username = user['username']
    try:
        db.execute('BEGIN')
        # NULL out every nullable FK that references users(id)
        db.execute('UPDATE members               SET updated_by  = NULL WHERE updated_by  = ?', (user_id,))
        db.execute('UPDATE pending_registrations SET reviewed_by = NULL WHERE reviewed_by = ?', (user_id,))
        db.execute('UPDATE documents             SET uploaded_by = NULL WHERE uploaded_by = ?', (user_id,))
        db.execute('UPDATE email_templates       SET created_by  = NULL WHERE created_by  = ?', (user_id,))
        db.execute('UPDATE mailshot_log          SET sent_by     = NULL WHERE sent_by     = ?', (user_id,))
        db.execute('UPDATE attendance            SET recorded_by = NULL WHERE recorded_by = ?', (user_id,))
        db.execute('UPDATE term_sessions         SET created_by  = NULL WHERE created_by  = ?', (user_id,))
        db.execute('UPDATE session_activities    SET added_by    = NULL WHERE added_by    = ?', (user_id,))
        db.execute('UPDATE audit_log             SET user_id     = NULL WHERE user_id     = ?', (user_id,))
        db.execute('UPDATE session_completions   SET completed_by = NULL WHERE completed_by = ?', (user_id,))
        db.execute('UPDATE session_completions   SET exported_by  = NULL WHERE exported_by  = ?', (user_id,))
        db.execute('UPDATE session_notes         SET added_by     = NULL WHERE added_by     = ?', (user_id,))
        db.execute('UPDATE alert_rules           SET created_by   = NULL WHERE created_by   = ?', (user_id,))
        db.execute('UPDATE notifications         SET sender_id    = NULL WHERE sender_id    = ?', (user_id,))
        db.execute('UPDATE settings              SET updated_by   = NULL WHERE updated_by   = ?', (user_id,))
        # notification_reads.user_id is NOT NULL — must DELETE rows, not NULL them
        db.execute('DELETE FROM notification_reads WHERE user_id = ?', (user_id,))
        # user_sessions has ON DELETE CASCADE so it handles itself, but being explicit is safe
        db.execute('DELETE FROM user_sessions WHERE user_id = ?', (user_id,))
        db.execute('DELETE FROM users WHERE id = ?', (user_id,))
        db.execute(
            'INSERT INTO audit_log (user_id, action, table_name, record_id, details, ip_address)'
            ' VALUES (?, ?, ?, ?, ?, ?)',
            (
                session['user_id'], 'delete_user', 'users', user_id,
                json.dumps({'username': username}),
                request.remote_addr,
            )
        )
        db.execute('COMMIT')
    except Exception as exc:
        db.execute('ROLLBACK')
        current_app.logger.error(f'Permanent user delete failed (username={username}): {exc}')
        return jsonify({'error': 'Deletion failed and was rolled back. Check server logs for details.'}), 500

    return jsonify({'success': True, 'deleted': username})


# ── Staff roles admin CRUD ────────────────────────────────────────────────────

@bp.route('/api/admin/staff-roles', methods=['GET'])
@permission_required('admin.settings')
def api_admin_staff_roles_get():
    db   = get_db()
    rows = db.execute(
        'SELECT id, name, active, display_order FROM staff_roles ORDER BY display_order, name'
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route('/api/admin/staff-roles', methods=['POST'])
@permission_required('admin.settings')
def api_admin_staff_roles_create():
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'name is required'}), 400

    db        = get_db()
    max_order = db.execute('SELECT COALESCE(MAX(display_order), -1) FROM staff_roles').fetchone()[0]
    try:
        cur = db.execute(
            'INSERT INTO staff_roles (name, active, display_order) VALUES (?,1,?)',
            (name, max_order + 1)
        )
        db.commit()
        log_action('create_staff_role', 'staff_roles', cur.lastrowid, {'name': name})
        row = db.execute('SELECT * FROM staff_roles WHERE id = ?', (cur.lastrowid,)).fetchone()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A role named "{name}" already exists'}), 409


@bp.route('/api/admin/staff-roles/<int:role_id>', methods=['PUT'])
@permission_required('admin.settings')
def api_admin_staff_roles_update(role_id):
    data    = request.get_json() or {}
    db      = get_db()
    current = db.execute('SELECT * FROM staff_roles WHERE id = ?', (role_id,)).fetchone()
    if not current:
        return jsonify({'error': 'Role not found'}), 404

    name   = data.get('name', current['name']).strip()
    active = int(data.get('active', current['active']))

    if not name:
        return jsonify({'error': 'name is required'}), 400

    if not active:
        active_count = db.execute(
            'SELECT COUNT(*) FROM staff_roles WHERE active = 1 AND id != ?', (role_id,)
        ).fetchone()[0]
        if active_count == 0:
            return jsonify({'error': 'Cannot deactivate the only active staff role'}), 400

    try:
        db.execute('UPDATE staff_roles SET name = ?, active = ? WHERE id = ?', (name, active, role_id))
        db.commit()
        log_action('update_staff_role', 'staff_roles', role_id, {'name': name, 'active': active})
        return jsonify(dict(db.execute('SELECT * FROM staff_roles WHERE id = ?', (role_id,)).fetchone()))
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A role named "{name}" already exists'}), 409


@bp.route('/api/admin/staff-roles/reorder', methods=['POST'])
@permission_required('admin.settings')
def api_admin_staff_roles_reorder():
    items = request.get_json() or []
    db    = get_db()
    for item in items:
        db.execute('UPDATE staff_roles SET display_order = ? WHERE id = ?',
                   (item.get('display_order', 0), item.get('id')))
    db.commit()
    return jsonify({'success': True})


@bp.route('/api/admin/staff-roles/<int:role_id>', methods=['DELETE'])
@permission_required('admin.settings')
def api_admin_staff_roles_delete(role_id):
    db  = get_db()
    row = db.execute('SELECT id, name FROM staff_roles WHERE id = ?', (role_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404

    in_use = db.execute(
        'SELECT COUNT(*) FROM members WHERE staff_role = ?', (row['name'],)
    ).fetchone()[0]
    if in_use:
        return jsonify({'error': f'Cannot delete — {in_use} member(s) have this role assigned. Reassign them first.'}), 409

    db.execute('DELETE FROM staff_roles WHERE id = ?', (role_id,))
    db.commit()
    log_action('delete_staff_role', 'staff_roles', role_id, {'name': row['name']})
    return jsonify({'ok': True})


# ── Permissions + Roles CRUD ──────────────────────────────────────────────────

@bp.route('/api/admin/permissions')
@permission_required('admin.roles')
def api_permissions_list():
    db   = get_db()
    rows = db.execute(
        'SELECT code, name, description, category FROM permissions ORDER BY category, code'
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route('/api/admin/roles')
@permission_required('admin.roles')
def api_roles_list():
    db   = get_db()
    rows = db.execute(
        'SELECT id, name, display_name, is_default, permissions, created_at FROM roles ORDER BY name'
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if not d.get('display_name'):
            d['display_name'] = ROLE_DISPLAY_NAMES.get(d['name'], d['name'])
        try:
            d['permissions'] = json.loads(d['permissions'])
        except (TypeError, ValueError):
            d['permissions'] = []
        result.append(d)
    return jsonify(result)


@bp.route('/api/admin/roles', methods=['POST'])
@permission_required('admin.roles')
def api_roles_create():
    data         = request.get_json() or {}
    name         = data.get('name', '').strip()
    display_name = data.get('display_name', '').strip() or name
    perms        = data.get('permissions', [])

    if not name:
        return jsonify({'error': 'Role name is required'}), 400
    if not isinstance(perms, list):
        return jsonify({'error': 'permissions must be a list'}), 400

    db          = get_db()
    valid_codes = {r['code'] for r in db.execute('SELECT code FROM permissions').fetchall()}
    bad = [p for p in perms if p not in valid_codes]
    if bad:
        return jsonify({'error': f'Unknown permission code(s): {", ".join(bad)}'}), 400

    if 'users.create.admin' in perms and not has_permission('users.create.admin'):
        return jsonify({'error': 'You do not have permission to assign users.create.admin'}), 403

    try:
        cur = db.execute(
            'INSERT INTO roles (name, display_name, permissions, is_default) VALUES (?,?,?,0)',
            (name, display_name, json.dumps(perms))
        )
        db.commit()
        log_action('create_role', 'roles', cur.lastrowid, {'name': name, 'permissions': perms})
        row = db.execute(
            'SELECT id, name, display_name, is_default, permissions, created_at FROM roles WHERE id = ?',
            (cur.lastrowid,)
        ).fetchone()
        d = dict(row); d['permissions'] = json.loads(d['permissions'])
        return jsonify(d), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A role named "{name}" already exists'}), 409


@bp.route('/api/admin/roles/<int:role_id>', methods=['PUT'])
@permission_required('admin.roles')
def api_roles_update(role_id):
    data  = request.get_json() or {}
    db    = get_db()
    role  = db.execute('SELECT * FROM roles WHERE id = ?', (role_id,)).fetchone()
    if not role:
        return jsonify({'error': 'Role not found'}), 404

    name         = data.get('name', role['name']).strip()
    display_name = data.get('display_name', role['display_name'] or role['name']).strip() or name
    perms        = data.get('permissions', json.loads(role['permissions']))

    if not name:
        return jsonify({'error': 'Role name is required'}), 400
    if not isinstance(perms, list):
        return jsonify({'error': 'permissions must be a list'}), 400

    valid_codes = {r['code'] for r in db.execute('SELECT code FROM permissions').fetchall()}
    bad = [p for p in perms if p not in valid_codes]
    if bad:
        return jsonify({'error': f'Unknown permission code(s): {", ".join(bad)}'}), 400

    old_perms = json.loads(role['permissions'])
    if 'users.create.admin' in perms and 'users.create.admin' not in old_perms:
        if not has_permission('users.create.admin'):
            return jsonify({'error': 'You do not have permission to assign users.create.admin'}), 403

    try:
        db.execute(
            'UPDATE roles SET name = ?, display_name = ?, permissions = ? WHERE id = ?',
            (name, display_name, json.dumps(perms), role_id)
        )
        db.commit()
        log_action('update_role', 'roles', role_id,
                   {'name': name, 'display_name': display_name, 'permissions': perms})
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A role named "{name}" already exists'}), 409

    row = db.execute(
        'SELECT id, name, display_name, is_default, permissions, created_at FROM roles WHERE id = ?',
        (role_id,)
    ).fetchone()
    d = dict(row); d['permissions'] = json.loads(d['permissions'])
    return jsonify(d)


@bp.route('/api/admin/roles/<int:role_id>', methods=['DELETE'])
@permission_required('admin.roles')
def api_roles_delete(role_id):
    db   = get_db()
    role = db.execute('SELECT * FROM roles WHERE id = ?', (role_id,)).fetchone()
    if not role:
        return jsonify({'error': 'Role not found'}), 404

    if role['is_default']:
        return jsonify({'error': 'Default roles cannot be deleted'}), 400

    user_count = db.execute('SELECT COUNT(*) FROM users WHERE role_id = ?', (role_id,)).fetchone()[0]
    if user_count:
        return jsonify({'error': f'Cannot delete — {user_count} user(s) are assigned to this role'}), 400

    db.execute('DELETE FROM roles WHERE id = ?', (role_id,))
    db.commit()
    log_action('delete_role', 'roles', role_id, {'name': role['name']})
    return jsonify({'success': True})


# ── SMTP profiles CRUD ────────────────────────────────────────────────────────

@bp.route('/api/admin/smtp-profiles')
@permission_required('admin.smtp_profiles')
def api_smtp_profiles_list():
    db   = get_db()
    rows = db.execute(
        'SELECT id, name, host, port, username, from_address, is_default, created_at, updated_at'
        ' FROM smtp_profiles ORDER BY is_default DESC, name'
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route('/api/admin/smtp-profiles', methods=['POST'])
@permission_required('admin.smtp_profiles')
def api_smtp_profiles_create():
    data         = request.get_json() or {}
    name         = (data.get('name') or '').strip()
    host         = (data.get('host') or '').strip()
    port         = int(data.get('port') or 587)
    username     = (data.get('username') or '').strip()
    password     = (data.get('password') or '').strip()
    from_address = (data.get('from_address') or username).strip()
    is_default   = int(bool(data.get('is_default')))

    if not name:
        return jsonify({'error': 'Name is required'}), 400
    if not host:
        return jsonify({'error': 'Host is required'}), 400
    if not username:
        return jsonify({'error': 'Username is required'}), 400
    if not password:
        return jsonify({'error': 'Password is required'}), 400

    password_enc = encrypt_file(password.encode()).decode()
    db = get_db()

    if is_default:
        db.execute('UPDATE smtp_profiles SET is_default = 0')

    try:
        cur = db.execute(
            'INSERT INTO smtp_profiles (name, host, port, username, password_enc, from_address, is_default, created_by)'
            ' VALUES (?,?,?,?,?,?,?,?)',
            (name, host, port, username, password_enc, from_address, is_default, session['user_id'])
        )
        # If this is the first profile ever, make it default automatically
        if db.execute('SELECT COUNT(*) FROM smtp_profiles').fetchone()[0] == 1:
            db.execute('UPDATE smtp_profiles SET is_default = 1 WHERE id = ?', (cur.lastrowid,))
        db.commit()
        log_action('create_smtp_profile', 'smtp_profiles', cur.lastrowid, {'name': name, 'host': host})
        row = db.execute(
            'SELECT id, name, host, port, username, from_address, is_default, created_at, updated_at'
            ' FROM smtp_profiles WHERE id = ?', (cur.lastrowid,)
        ).fetchone()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A profile named "{name}" already exists'}), 409


@bp.route('/api/admin/smtp-profiles/<int:profile_id>', methods=['PUT'])
@permission_required('admin.smtp_profiles')
def api_smtp_profiles_update(profile_id):
    data    = request.get_json() or {}
    db      = get_db()
    current = db.execute('SELECT * FROM smtp_profiles WHERE id = ?', (profile_id,)).fetchone()
    if not current:
        return jsonify({'error': 'Profile not found'}), 404

    name         = (data.get('name') or current['name']).strip()
    host         = (data.get('host') or current['host']).strip()
    port         = int(data.get('port') or current['port'])
    username     = (data.get('username') or current['username']).strip()
    from_address = (data.get('from_address') or current['from_address']).strip()
    is_default   = int(bool(data.get('is_default', current['is_default'])))

    # Only update password if a new one is supplied
    if data.get('password', '').strip():
        password_enc = encrypt_file(data['password'].strip().encode()).decode()
    else:
        password_enc = current['password_enc']

    if not name or not host or not username:
        return jsonify({'error': 'Name, host and username are required'}), 400

    if is_default:
        db.execute('UPDATE smtp_profiles SET is_default = 0 WHERE id != ?', (profile_id,))

    try:
        db.execute(
            'UPDATE smtp_profiles SET name=?, host=?, port=?, username=?, password_enc=?,'
            ' from_address=?, is_default=?, updated_at=datetime("now") WHERE id=?',
            (name, host, port, username, password_enc, from_address, is_default, profile_id)
        )
        db.commit()
        log_action('update_smtp_profile', 'smtp_profiles', profile_id, {'name': name})
        row = db.execute(
            'SELECT id, name, host, port, username, from_address, is_default, created_at, updated_at'
            ' FROM smtp_profiles WHERE id = ?', (profile_id,)
        ).fetchone()
        return jsonify(dict(row))
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A profile named "{name}" already exists'}), 409


@bp.route('/api/admin/smtp-profiles/<int:profile_id>', methods=['DELETE'])
@permission_required('admin.smtp_profiles')
def api_smtp_profiles_delete(profile_id):
    db      = get_db()
    profile = db.execute('SELECT * FROM smtp_profiles WHERE id = ?', (profile_id,)).fetchone()
    if not profile:
        return jsonify({'error': 'Profile not found'}), 404

    db.execute('DELETE FROM smtp_profiles WHERE id = ?', (profile_id,))
    # If we just deleted the default, promote the first remaining profile
    if profile['is_default']:
        first = db.execute('SELECT id FROM smtp_profiles ORDER BY id LIMIT 1').fetchone()
        if first:
            db.execute('UPDATE smtp_profiles SET is_default = 1 WHERE id = ?', (first['id'],))
    db.commit()
    log_action('delete_smtp_profile', 'smtp_profiles', profile_id, {'name': profile['name']})
    return jsonify({'success': True})


@bp.route('/api/admin/smtp-profiles/<int:profile_id>/set-default', methods=['POST'])
@permission_required('admin.smtp_profiles')
def api_smtp_profiles_set_default(profile_id):
    db = get_db()
    if not db.execute('SELECT id FROM smtp_profiles WHERE id = ?', (profile_id,)).fetchone():
        return jsonify({'error': 'Profile not found'}), 404
    db.execute('UPDATE smtp_profiles SET is_default = 0')
    db.execute('UPDATE smtp_profiles SET is_default = 1 WHERE id = ?', (profile_id,))
    db.commit()
    log_action('set_default_smtp_profile', 'smtp_profiles', profile_id, {})
    return jsonify({'success': True})


@bp.route('/api/admin/smtp-profiles/<int:profile_id>/test', methods=['POST'])
@permission_required('admin.smtp_profiles')
def api_smtp_profiles_test(profile_id):
    """Send a test email to the logged-in user's address using this profile."""
    import smtplib
    from email.mime.text import MIMEText

    db      = get_db()
    profile = db.execute('SELECT * FROM smtp_profiles WHERE id = ?', (profile_id,)).fetchone()
    if not profile:
        return jsonify({'error': 'Profile not found'}), 404

    # Get logged-in user's email address for the test
    user_row = db.execute('SELECT email FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    to_addr  = (user_row['email'] if user_row else '') or ''
    if not to_addr:
        return jsonify({'error': 'Your portal account has no email address set — add one in your user profile first'}), 400

    try:
        password = decrypt_file(profile['password_enc'].encode()).decode()
    except Exception:
        return jsonify({'error': 'Could not decrypt stored password — re-save the profile and try again'}), 500

    msg            = MIMEText('This is a test email from your AYC Portal email sender configuration.', 'plain', 'utf-8')
    msg['Subject'] = 'AYC Portal — Email Sender Test'
    msg['From']    = profile['from_address']
    msg['To']      = to_addr

    try:
        if profile['port'] == 465:
            import ssl as _ssl
            with smtplib.SMTP_SSL(profile['host'], profile['port'], timeout=15,
                                  context=_ssl.create_default_context()) as srv:
                srv.login(profile['username'], password)
                srv.sendmail(profile['from_address'], [to_addr], msg.as_string())
        else:
            with smtplib.SMTP(profile['host'], profile['port'], timeout=15) as srv:
                srv.ehlo(); srv.starttls(); srv.ehlo()
                srv.login(profile['username'], password)
                srv.sendmail(profile['from_address'], [to_addr], msg.as_string())
    except smtplib.SMTPAuthenticationError:
        return jsonify({'error': 'Authentication failed — check the username and password'}), 400
    except Exception as e:
        return jsonify({'error': f'Send failed: {e}'}), 400

    log_action('test_smtp_profile', 'smtp_profiles', profile_id, {'to': to_addr})
    return jsonify({'success': True, 'sent_to': to_addr})


# ── Tag definitions CRUD ──────────────────────────────────────────────────────

@bp.route('/api/tags')
@permission_required('members.view')
def api_tags_public():
    db   = get_db()
    rows = db.execute(
        'SELECT id, name, category, icon, colour FROM tag_definitions '
        'WHERE active = 1 ORDER BY sort_order, name'
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route('/api/admin/tags')
@permission_required('admin.settings')
def api_tags_list():
    db   = get_db()
    rows = db.execute(
        'SELECT id, name, category, icon, colour, active, sort_order '
        'FROM tag_definitions ORDER BY sort_order, name'
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route('/api/admin/tags', methods=['POST'])
@permission_required('admin.settings')
def api_tags_create():
    data            = request.get_json() or {}
    name            = data.get('name', '').strip()
    category        = data.get('category', 'General').strip() or 'General'
    icon            = data.get('icon', '🏷').strip() or '🏷'
    colour, col_err = _validate_hex_colour(data.get('colour', ''), '#3b82f6')

    if not name:
        return jsonify({'error': 'name is required'}), 400
    if col_err:
        return jsonify({'error': col_err}), 400

    db        = get_db()
    max_order = db.execute('SELECT COALESCE(MAX(sort_order), -1) FROM tag_definitions').fetchone()[0]
    try:
        cur = db.execute(
            'INSERT INTO tag_definitions (name, category, icon, colour, active, sort_order) VALUES (?,?,?,?,1,?)',
            (name, category, icon, colour, max_order + 1)
        )
        db.commit()
        log_action('create_tag', 'tag_definitions', cur.lastrowid, {'name': name})
        return jsonify(dict(db.execute('SELECT * FROM tag_definitions WHERE id = ?', (cur.lastrowid,)).fetchone())), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A tag named "{name}" already exists'}), 409


@bp.route('/api/admin/tags/<int:tag_id>', methods=['PUT'])
@permission_required('admin.settings')
def api_tags_update(tag_id):
    db  = get_db()
    row = db.execute('SELECT * FROM tag_definitions WHERE id = ?', (tag_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Tag not found'}), 404

    data            = request.get_json() or {}
    name            = data.get('name',     row['name']).strip()
    category        = data.get('category', row['category']).strip() or 'General'
    icon            = data.get('icon',     row['icon']).strip() or '🏷'
    colour, col_err = _validate_hex_colour(data.get('colour', row['colour']), row['colour'])
    active          = int(data.get('active', row['active']))

    if not name:
        return jsonify({'error': 'name is required'}), 400
    if col_err:
        return jsonify({'error': col_err}), 400
    try:
        db.execute(
            'UPDATE tag_definitions SET name=?, category=?, icon=?, colour=?, active=? WHERE id=?',
            (name, category, icon, colour, active, tag_id)
        )
        db.commit()
        log_action('update_tag', 'tag_definitions', tag_id, {'name': name, 'active': active})
        return jsonify(dict(db.execute('SELECT * FROM tag_definitions WHERE id = ?', (tag_id,)).fetchone()))
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A tag named "{name}" already exists'}), 409


@bp.route('/api/admin/tags/reorder', methods=['POST'])
@permission_required('admin.settings')
def api_tags_reorder():
    items = request.get_json() or []
    db    = get_db()
    for item in items:
        db.execute('UPDATE tag_definitions SET sort_order = ? WHERE id = ?',
                   (item.get('sort_order', 0), item.get('id')))
    db.commit()
    return jsonify({'success': True})


@bp.route('/api/admin/tags/<int:tag_id>', methods=['DELETE'])
@permission_required('admin.settings')
def api_tags_delete(tag_id):
    db  = get_db()
    row = db.execute('SELECT id, name FROM tag_definitions WHERE id = ?', (tag_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Tag not found'}), 404

    assigned = db.execute(
        'SELECT COUNT(*) FROM member_tags WHERE tag_id = ?', (tag_id,)
    ).fetchone()[0]

    db.execute('DELETE FROM tag_definitions WHERE id = ?', (tag_id,))
    db.commit()
    log_action('delete_tag', 'tag_definitions', tag_id, {'name': row['name'], 'removed_from_members': assigned})
    return jsonify({'ok': True, 'removed_from_members': assigned})


# ── Member statuses CRUD ──────────────────────────────────────────────────────

@bp.route('/api/admin/member-statuses', methods=['GET'])
@permission_required('admin.settings')
def api_admin_member_statuses_list():
    db   = get_db()
    rows = db.execute(
        'SELECT * FROM member_statuses ORDER BY sort_order, name'
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route('/api/admin/member-statuses', methods=['POST'])
@permission_required('admin.settings')
def api_admin_member_statuses_create():
    data      = request.get_json() or {}
    name      = data.get('name', '').strip()
    behaviour = data.get('behaviour', '').strip()
    colour, col_err = _validate_hex_colour(data.get('colour', ''), '#64748b')

    if not name:
        return jsonify({'error': 'Name is required'}), 400
    if behaviour not in ('active', 'inactive', 'leaver'):
        return jsonify({'error': 'Behaviour must be active, inactive, or leaver'}), 400
    if col_err:
        return jsonify({'error': col_err}), 400

    db        = get_db()
    max_order = db.execute('SELECT COALESCE(MAX(sort_order), -1) FROM member_statuses').fetchone()[0]
    try:
        cur = db.execute(
            'INSERT INTO member_statuses (name, behaviour, colour, sort_order, is_default, is_protected) '
            'VALUES (?,?,?,?,0,0)',
            (name, behaviour, colour, max_order + 1)
        )
        db.commit()
        log_action('create_member_status', 'member_statuses', cur.lastrowid,
                   {'name': name, 'behaviour': behaviour})
        return jsonify(dict(db.execute(
            'SELECT * FROM member_statuses WHERE id = ?', (cur.lastrowid,)
        ).fetchone())), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A status named "{name}" already exists'}), 409


@bp.route('/api/admin/member-statuses/<int:status_id>', methods=['PUT'])
@permission_required('admin.settings')
def api_admin_member_statuses_update(status_id):
    db  = get_db()
    row = db.execute('SELECT * FROM member_statuses WHERE id = ?', (status_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Status not found'}), 404

    data      = request.get_json() or {}
    name      = data.get('name', row['name']).strip()
    colour, col_err = _validate_hex_colour(data.get('colour', row['colour']), row['colour'])
    sort_order = int(data.get('sort_order', row['sort_order']))

    # Protected statuses cannot be renamed or have their behaviour changed
    if row['is_protected']:
        name = row['name']  # silently keep original name

    behaviour = row['behaviour']  # behaviour is fixed after creation
    if not row['is_protected']:
        _beh = data.get('behaviour', row['behaviour']).strip()
        if _beh in ('active', 'inactive', 'leaver'):
            behaviour = _beh

    if not name:
        return jsonify({'error': 'Name is required'}), 400
    if col_err:
        return jsonify({'error': col_err}), 400

    try:
        db.execute(
            'UPDATE member_statuses SET name=?, behaviour=?, colour=?, sort_order=? WHERE id=?',
            (name, behaviour, colour, sort_order, status_id)
        )
        db.commit()
        log_action('update_member_status', 'member_statuses', status_id,
                   {'name': name, 'behaviour': behaviour, 'colour': colour})
        return jsonify(dict(db.execute(
            'SELECT * FROM member_statuses WHERE id = ?', (status_id,)
        ).fetchone()))
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A status named "{name}" already exists'}), 409


@bp.route('/api/admin/member-statuses/<int:status_id>', methods=['DELETE'])
@permission_required('admin.settings')
def api_admin_member_statuses_delete(status_id):
    db  = get_db()
    row = db.execute('SELECT * FROM member_statuses WHERE id = ?', (status_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Status not found'}), 404
    if row['is_protected']:
        return jsonify({'error': f'"{row["name"]}" is a protected status and cannot be deleted'}), 400

    # Refuse if any members currently have this status
    member_count = db.execute(
        'SELECT COUNT(*) FROM members WHERE status = ?', (row['name'],)
    ).fetchone()[0]
    if member_count:
        return jsonify({'error': f'{member_count} member{"s" if member_count != 1 else ""} '
                                  f'currently have this status — reassign them first'}), 400

    db.execute('DELETE FROM member_statuses WHERE id = ?', (status_id,))
    db.commit()
    log_action('delete_member_status', 'member_statuses', status_id, {'name': row['name']})
    return jsonify({'success': True})


@bp.route('/api/admin/member-statuses/reorder', methods=['POST'])
@permission_required('admin.settings')
def api_admin_member_statuses_reorder():
    items = request.get_json() or []
    db    = get_db()
    for item in items:
        db.execute('UPDATE member_statuses SET sort_order = ? WHERE id = ?',
                   (item.get('sort_order', 0), item.get('id')))
    db.commit()
    return jsonify({'success': True})


# ── Member types CRUD ─────────────────────────────────────────────────────────

def _slugify(text):
    text = text.lower().strip()
    text = _re.sub(r'[^\w\s-]', '', text)
    text = _re.sub(r'[\s_]+', '-', text)
    text = _re.sub(r'-+', '-', text)
    return text.strip('-')


@bp.route('/api/admin/member-types', methods=['GET'])
@permission_required('admin.settings')
def api_admin_member_types_list():
    db   = get_db()
    rows = db.execute('''
        SELECT mt.*,
               (SELECT COUNT(*) FROM member_type_fields mtf WHERE mtf.member_type_id = mt.id) AS field_count
        FROM   member_types mt
        ORDER  BY mt.sort_order, mt.name
    ''').fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route('/api/admin/member-types', methods=['POST'])
@permission_required('admin.settings')
def api_admin_member_types_create():
    data                = request.get_json() or {}
    name                = data.get('name', '').strip()
    slug                = data.get('slug', '').strip() or _slugify(name)
    icon                = data.get('icon', '👤').strip() or '👤'
    colour, col_err     = _validate_hex_colour(data.get('colour', ''), '#1b2d4f')
    description         = data.get('description', '').strip() or None
    public_registration = int(data.get('public_registration', 0))
    registration_style  = 'staff' if data.get('registration_style') == 'staff' else 'member'

    if not name:
        return jsonify({'error': 'name is required'}), 400
    if not slug:
        return jsonify({'error': 'slug is required'}), 400
    if col_err:
        return jsonify({'error': col_err}), 400

    db        = get_db()
    max_order = db.execute('SELECT COALESCE(MAX(sort_order), -1) FROM member_types').fetchone()[0]
    try:
        cur = db.execute(
            '''INSERT INTO member_types
               (name, slug, icon, colour, description, public_registration, registration_style, sort_order)
               VALUES (?,?,?,?,?,?,?,?)''',
            (name, slug, icon, colour, description, public_registration, registration_style, max_order + 1),
        )
        db.commit()
        log_action('create_member_type', 'member_types', cur.lastrowid, {'name': name, 'slug': slug, 'registration_style': registration_style})
        row = db.execute(
            '''SELECT mt.*, (SELECT COUNT(*) FROM member_type_fields mtf WHERE mtf.member_type_id = mt.id) AS field_count
               FROM member_types mt WHERE mt.id = ?''', (cur.lastrowid,)
        ).fetchone()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': 'A member type with that name or slug already exists'}), 409


@bp.route('/api/admin/member-types/<int:type_id>', methods=['PUT'])
@permission_required('admin.settings')
def api_admin_member_types_update(type_id):
    data    = request.get_json() or {}
    db      = get_db()
    current = db.execute('SELECT * FROM member_types WHERE id = ?', (type_id,)).fetchone()
    if not current:
        return jsonify({'error': 'Member type not found'}), 404

    name                = (data.get('name') or current['name']).strip()
    icon                = (data.get('icon') or current['icon'] or '👤').strip() or '👤'
    colour, col_err     = _validate_hex_colour(data.get('colour', current['colour']), current['colour'] or '#1b2d4f')
    description         = data.get('description', current['description'])
    if description is not None:
        description = description.strip() or None
    public_registration = int(data.get('public_registration', current['public_registration']))
    active              = int(data.get('active', current['active']))
    sort_order          = int(data.get('sort_order', current['sort_order']))
    registration_style  = 'staff' if data.get('registration_style', current['registration_style']) == 'staff' else 'member'

    if not name:
        return jsonify({'error': 'name is required'}), 400
    if col_err:
        return jsonify({'error': col_err}), 400

    if not active:
        active_count = db.execute(
            'SELECT COUNT(*) FROM member_types WHERE active = 1 AND id != ?', (type_id,)
        ).fetchone()[0]
        if active_count == 0:
            return jsonify({'error': 'Cannot deactivate the only active member type'}), 400

    try:
        db.execute(
            '''UPDATE member_types
               SET name=?, icon=?, colour=?, description=?, public_registration=?, active=?, registration_style=?, sort_order=?
               WHERE id=?''',
            (name, icon, colour, description, public_registration, active, registration_style, sort_order, type_id),
        )
        db.commit()
        log_action('update_member_type', 'member_types', type_id, {'name': name, 'active': active, 'registration_style': registration_style})
        row = db.execute(
            '''SELECT mt.*, (SELECT COUNT(*) FROM member_type_fields mtf WHERE mtf.member_type_id = mt.id) AS field_count
               FROM member_types mt WHERE mt.id = ?''', (type_id,)
        ).fetchone()
        return jsonify(dict(row))
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A member type named "{name}" already exists'}), 409


@bp.route('/api/admin/member-types/<int:type_id>', methods=['DELETE'])
@permission_required('admin.settings')
def api_admin_member_types_delete(type_id):
    db  = get_db()
    row = db.execute('SELECT * FROM member_types WHERE id = ?', (type_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Member type not found'}), 404

    member_count = db.execute(
        'SELECT COUNT(*) FROM members WHERE member_type = ?', (row['slug'],)
    ).fetchone()[0]
    if member_count:
        return jsonify({'error': f'Cannot delete — {member_count} member(s) use this type'}), 409

    active_count = db.execute(
        'SELECT COUNT(*) FROM member_types WHERE active = 1 AND id != ?', (type_id,)
    ).fetchone()[0]
    if row['active'] and active_count == 0:
        return jsonify({'error': 'Cannot delete the only active member type'}), 400

    db.execute('DELETE FROM member_types WHERE id = ?', (type_id,))
    db.commit()
    log_action('delete_member_type', 'member_types', type_id, {'name': row['name']})
    return jsonify({'success': True})


# ── Field definitions CRUD ────────────────────────────────────────────────────

@bp.route('/api/admin/field-definitions', methods=['GET'])
@permission_required('admin.settings')
def api_admin_field_definitions_list():
    db   = get_db()
    rows = db.execute('''
        SELECT fd.*,
               (SELECT COUNT(*) FROM member_type_fields mtf WHERE mtf.field_id = fd.id) AS assigned_to
        FROM   field_definitions fd
        ORDER  BY fd.sort_order, fd.label
    ''').fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route('/api/admin/field-definitions', methods=['POST'])
@permission_required('admin.settings')
def api_admin_field_definitions_create():
    data        = request.get_json() or {}
    label       = (data.get('label') or '').strip()
    field_type  = (data.get('field_type') or 'text').strip() or 'text'
    placeholder = (data.get('placeholder') or '').strip() or None
    help_text   = (data.get('help_text') or '').strip() or None
    options     = (data.get('options') or '').strip() or None

    if not label:
        return jsonify({'error': 'label is required'}), 400

    key = _slugify(label).replace('-', '_')
    if not key:
        return jsonify({'error': 'Could not generate a valid key from the label'}), 400

    db        = get_db()
    max_order = db.execute('SELECT COALESCE(MAX(sort_order), -1) FROM field_definitions').fetchone()[0]

    # Ensure key uniqueness
    base_key = key
    suffix   = 1
    while db.execute('SELECT id FROM field_definitions WHERE key = ?', (key,)).fetchone():
        key = f'{base_key}_{suffix}'; suffix += 1

    try:
        cur = db.execute(
            '''INSERT INTO field_definitions
               (key, label, field_type, placeholder, help_text, options, system_field, sort_order)
               VALUES (?,?,?,?,?,?,0,?)''',
            (key, label, field_type, placeholder, help_text, options, max_order + 1),
        )
        db.commit()
        log_action('create_field_definition', 'field_definitions', cur.lastrowid,
                   {'key': key, 'label': label, 'field_type': field_type})
        row = db.execute(
            '''SELECT fd.*, (SELECT COUNT(*) FROM member_type_fields mtf WHERE mtf.field_id = fd.id) AS assigned_to
               FROM field_definitions fd WHERE fd.id = ?''', (cur.lastrowid,)
        ).fetchone()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': f'A field with key "{key}" already exists'}), 409


@bp.route('/api/admin/field-definitions/<int:field_id>', methods=['PUT'])
@permission_required('admin.settings')
def api_admin_field_definitions_update(field_id):
    data    = request.get_json() or {}
    db      = get_db()
    current = db.execute('SELECT * FROM field_definitions WHERE id = ?', (field_id,)).fetchone()
    if not current:
        return jsonify({'error': 'Field definition not found'}), 404

    if current['system_field'] and 'field_type' in data and data['field_type'] != current['field_type']:
        return jsonify({'error': 'Cannot change the field type of a system field'}), 403

    label       = data.get('label', current['label']).strip()
    field_type  = data.get('field_type', current['field_type']).strip() or current['field_type']
    placeholder = data.get('placeholder', current['placeholder'])
    if placeholder is not None:
        placeholder = placeholder.strip() or None
    help_text   = data.get('help_text', current['help_text'])
    if help_text is not None:
        help_text = help_text.strip() or None
    options     = data.get('options', current['options'])
    if options is not None:
        options = options.strip() or None
    active      = int(data.get('active', current['active']))
    use_lookup  = int(data['use_lookup']) if 'use_lookup' in data else int(current['use_lookup'])

    if not label:
        return jsonify({'error': 'label is required'}), 400

    db.execute(
        '''UPDATE field_definitions
           SET label=?, field_type=?, placeholder=?, help_text=?, options=?, active=?, use_lookup=?
           WHERE id=?''',
        (label, field_type, placeholder, help_text, options, active, use_lookup, field_id),
    )
    db.commit()
    log_action('update_field_definition', 'field_definitions', field_id,
               {'label': label, 'active': active})
    row = db.execute(
        '''SELECT fd.*, (SELECT COUNT(*) FROM member_type_fields mtf WHERE mtf.field_id = fd.id) AS assigned_to
           FROM field_definitions fd WHERE fd.id = ?''', (field_id,)
    ).fetchone()
    return jsonify(dict(row))


@bp.route('/api/admin/field-definitions/<int:field_id>', methods=['DELETE'])
@permission_required('admin.settings')
def api_admin_field_definitions_delete(field_id):
    db  = get_db()
    row = db.execute('SELECT * FROM field_definitions WHERE id = ?', (field_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Field definition not found'}), 404

    if row['system_field']:
        return jsonify({'error': 'System fields cannot be deleted'}), 403

    usage = db.execute(
        'SELECT COUNT(*) FROM member_type_fields WHERE field_id = ?', (field_id,)
    ).fetchone()[0]
    if usage:
        return jsonify({'error': f'Cannot delete — this field is assigned to {usage} member type(s)'}), 409

    db.execute('DELETE FROM field_definitions WHERE id = ?', (field_id,))
    db.commit()
    log_action('delete_field_definition', 'field_definitions', field_id, {'key': row['key']})
    return jsonify({'success': True})


# ── Type-field config ─────────────────────────────────────────────────────────

@bp.route('/api/admin/member-types/<int:type_id>/fields', methods=['GET'])
@permission_required('admin.settings')
def api_admin_type_fields_list(type_id):
    db = get_db()
    if not db.execute('SELECT id FROM member_types WHERE id = ?', (type_id,)).fetchone():
        return jsonify({'error': 'Member type not found'}), 404
    rows = db.execute('''
        SELECT  mtf.id AS assignment_id, mtf.sort_order, mtf.required,
                mtf.show_on_registration, mtf.show_on_list, mtf.show_on_attendance,
                mtf.show_on_card, mtf.show_on_print, mtf.show_on_export,
                fd.id AS field_id, fd.key, fd.label, fd.field_type,
                fd.options, fd.help_text, fd.placeholder,
                fd.system_field, fd.column_name, fd.active
        FROM    member_type_fields mtf
        JOIN    field_definitions fd ON fd.id = mtf.field_id
        WHERE   mtf.member_type_id = ?
        ORDER   BY mtf.sort_order, fd.label
    ''', (type_id,)).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route('/api/admin/member-types/<int:type_id>/fields', methods=['POST'])
@permission_required('admin.settings')
def api_admin_type_fields_assign(type_id):
    db = get_db()
    if not db.execute('SELECT id FROM member_types WHERE id = ?', (type_id,)).fetchone():
        return jsonify({'error': 'Member type not found'}), 404

    data     = request.get_json() or {}
    field_id = data.get('field_id')
    if not field_id:
        return jsonify({'error': 'field_id is required'}), 400
    if not db.execute('SELECT id FROM field_definitions WHERE id = ?', (field_id,)).fetchone():
        return jsonify({'error': 'Field definition not found'}), 404

    max_order = db.execute(
        'SELECT COALESCE(MAX(sort_order), 0) FROM member_type_fields WHERE member_type_id = ?', (type_id,)
    ).fetchone()[0]

    try:
        cur = db.execute(
            '''INSERT INTO member_type_fields
               (member_type_id, field_id, sort_order, required,
                show_on_registration, show_on_list, show_on_attendance,
                show_on_card, show_on_print, show_on_export)
               VALUES (?,?,?,0,1,0,0,0,1,0)''',
            (type_id, field_id, max_order + 1),
        )
        db.commit()
        log_action('assign_type_field', 'member_type_fields', cur.lastrowid,
                   {'member_type_id': type_id, 'field_id': field_id})
        row = db.execute('''
            SELECT  mtf.id AS assignment_id, mtf.sort_order, mtf.required,
                    mtf.show_on_registration, mtf.show_on_list, mtf.show_on_attendance,
                    mtf.show_on_card, mtf.show_on_print, mtf.show_on_export,
                    fd.id AS field_id, fd.key, fd.label, fd.field_type,
                    fd.options, fd.help_text, fd.placeholder,
                    fd.system_field, fd.column_name, fd.active
            FROM    member_type_fields mtf
            JOIN    field_definitions fd ON fd.id = mtf.field_id
            WHERE   mtf.id = ?
        ''', (cur.lastrowid,)).fetchone()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': 'This field is already assigned to this member type'}), 409


@bp.route('/api/admin/member-types/<int:type_id>/fields/<int:field_id>', methods=['PUT'])
@permission_required('admin.settings')
def api_admin_type_fields_update(type_id, field_id):
    db  = get_db()
    row = db.execute(
        'SELECT * FROM member_type_fields WHERE member_type_id = ? AND field_id = ?',
        (type_id, field_id)
    ).fetchone()
    if not row:
        return jsonify({'error': 'Assignment not found'}), 404

    data    = request.get_json() or {}
    updates = {}
    for col in ('required', 'show_on_registration', 'show_on_list',
                'show_on_attendance', 'show_on_card', 'show_on_print', 'show_on_export'):
        if col in data:
            updates[col] = int(data[col])

    if not updates:
        return jsonify({'error': 'No updatable fields provided'}), 400

    set_clause = ', '.join(f'{k} = ?' for k in updates)
    values     = list(updates.values()) + [type_id, field_id]
    db.execute(
        f'UPDATE member_type_fields SET {set_clause} WHERE member_type_id = ? AND field_id = ?', values
    )
    db.commit()

    updated = db.execute('''
        SELECT  mtf.id AS assignment_id, mtf.sort_order, mtf.required,
                mtf.show_on_registration, mtf.show_on_list, mtf.show_on_attendance,
                mtf.show_on_card, mtf.show_on_print, mtf.show_on_export,
                fd.id AS field_id, fd.key, fd.label, fd.field_type,
                fd.options, fd.help_text, fd.placeholder,
                fd.system_field, fd.column_name, fd.active
        FROM    member_type_fields mtf
        JOIN    field_definitions fd ON fd.id = mtf.field_id
        WHERE   mtf.member_type_id = ? AND mtf.field_id = ?
    ''', (type_id, field_id)).fetchone()
    return jsonify(dict(updated))


@bp.route('/api/admin/member-types/<int:type_id>/fields/<int:field_id>', methods=['DELETE'])
@permission_required('admin.settings')
def api_admin_type_fields_remove(type_id, field_id):
    db = get_db()
    fd = db.execute('SELECT key FROM field_definitions WHERE id = ?', (field_id,)).fetchone()
    if fd and fd['key'] in ('first_name', 'surname'):
        return jsonify({'error': 'first_name and surname cannot be removed from a member type'}), 403

    row = db.execute(
        'SELECT id FROM member_type_fields WHERE member_type_id = ? AND field_id = ?',
        (type_id, field_id)
    ).fetchone()
    if not row:
        return jsonify({'error': 'Assignment not found'}), 404

    db.execute('DELETE FROM member_type_fields WHERE member_type_id = ? AND field_id = ?',
               (type_id, field_id))
    db.commit()
    log_action('remove_type_field', 'member_type_fields', row['id'],
               {'member_type_id': type_id, 'field_id': field_id})
    return jsonify({'success': True})


@bp.route('/api/admin/member-types/<int:type_id>/fields/reorder', methods=['POST'])
@permission_required('admin.settings')
def api_admin_type_fields_reorder(type_id):
    items = request.get_json() or []
    db    = get_db()
    for item in items:
        db.execute(
            'UPDATE member_type_fields SET sort_order = ? WHERE member_type_id = ? AND field_id = ?',
            (item.get('sort_order', 0), type_id, item.get('field_id')),
        )
    db.commit()
    return jsonify({'success': True})


# ── System log viewer ─────────────────────────────────────────────────────────

def _safe_log_path(filename: str):
    safe = os.path.basename(filename)
    full = os.path.join(LOG_DIR, safe)
    if not os.path.abspath(full).startswith(os.path.abspath(LOG_DIR) + os.sep):
        return None
    return full


@bp.route('/api/admin/logs')
@permission_required('admin.maintenance')
def api_logs_list():
    files = []
    for path in sorted(glob.glob(os.path.join(LOG_DIR, 'app.log*')),
                       key=os.path.getmtime, reverse=True):
        st = os.stat(path)
        files.append({
            'filename': os.path.basename(path),
            'size':     st.st_size,
            'modified': datetime.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
        })
    return jsonify(files)


@bp.route('/api/admin/logs/tail')
@permission_required('admin.maintenance')
def api_logs_tail():
    filename = request.args.get('file', 'app.log')
    n_lines  = min(int(request.args.get('lines', 500)), 5000)
    path     = _safe_log_path(filename)
    if not path:
        return jsonify({'error': 'Invalid filename'}), 400
    if not os.path.exists(path):
        return jsonify({'lines': [], 'truncated': False})

    with open(path, 'rb') as f:
        f.seek(0, 2)
        size  = f.tell()
        chunk = min(size, 512 * 1024)
        f.seek(max(0, size - chunk))
        raw = f.read().decode('utf-8', errors='replace')

    all_lines = raw.splitlines()
    tail      = all_lines[-n_lines:]
    return jsonify({'lines': tail, 'truncated': len(all_lines) > n_lines,
                    'total_in_chunk': len(all_lines)})


@bp.route('/api/admin/logs/<path:filename>/download')
@permission_required('admin.maintenance')
def api_logs_download(filename):
    path = _safe_log_path(filename)
    if not path:
        return jsonify({'error': 'Invalid filename'}), 400
    if not os.path.exists(path):
        return jsonify({'error': 'File not found'}), 404
    return send_file(path, as_attachment=True,
                     download_name=os.path.basename(path),
                     mimetype='text/plain')


@bp.route('/api/admin/logs/clear', methods=['POST'])
@permission_required('admin.maintenance')
def api_logs_clear():
    data     = request.get_json() or {}
    filename = data.get('file', 'app.log')
    path     = _safe_log_path(filename)
    if not path:
        return jsonify({'error': 'Invalid filename'}), 400
    if not os.path.exists(path):
        return jsonify({'error': 'File not found'}), 404
    # Truncate — open in write mode to clear contents
    open(path, 'w').close()
    log_action('clear_log_file', None, None, {'file': filename, 'by': session.get('username')})
    return jsonify({'ok': True})


# ── Maintenance ───────────────────────────────────────────────────────────────

@bp.route('/api/admin/maintenance/counts')
@permission_required('admin.maintenance')
def api_maintenance_counts():
    db = get_db()
    return jsonify({
        'audit_log':       db.execute('SELECT COUNT(*) FROM audit_log').fetchone()[0],
        'attendance':      db.execute('SELECT COUNT(*) FROM attendance').fetchone()[0],
        'mailshot_log':    db.execute('SELECT COUNT(*) FROM mailshot_log').fetchone()[0],
        'registrations':   db.execute('SELECT COUNT(*) FROM pending_registrations').fetchone()[0],
        'members':         db.execute('SELECT COUNT(*) FROM members').fetchone()[0],
        'member_contacts': db.execute('SELECT COUNT(*) FROM member_contacts').fetchone()[0],
        'dofe':            db.execute('SELECT COUNT(*) FROM dofe_participants').fetchone()[0],
    })


@bp.route('/api/admin/maintenance/audit-log', methods=['DELETE'])
@permission_required('admin.maintenance')
def api_maintenance_clear_audit():
    db = get_db()
    n  = db.execute('SELECT COUNT(*) FROM audit_log').fetchone()[0]
    db.execute('DELETE FROM audit_log')
    db.commit()
    log_action('maintenance_clear', 'audit_log', None, {'cleared': n, 'by': session['username']})
    return jsonify({'success': True, 'deleted': n})


@bp.route('/api/admin/maintenance/attendance', methods=['DELETE'])
@permission_required('admin.maintenance')
def api_maintenance_clear_attendance():
    db = get_db()
    n  = db.execute('SELECT COUNT(*) FROM attendance').fetchone()[0]
    db.execute('DELETE FROM attendance')
    db.commit()
    log_action('maintenance_clear', 'attendance', None, {'cleared': n, 'by': session['username']})
    return jsonify({'success': True, 'deleted': n})


@bp.route('/api/admin/maintenance/mailshot-log', methods=['DELETE'])
@permission_required('admin.maintenance')
def api_maintenance_clear_mailshots():
    db = get_db()
    n  = db.execute('SELECT COUNT(*) FROM mailshot_log').fetchone()[0]
    db.execute('DELETE FROM mailshot_log')
    db.commit()
    log_action('maintenance_clear', 'mailshot_log', None, {'cleared': n, 'by': session['username']})
    return jsonify({'success': True, 'deleted': n})


@bp.route('/api/admin/maintenance/registrations', methods=['DELETE'])
@permission_required('admin.maintenance')
def api_maintenance_clear_registrations():
    db = get_db()
    n  = db.execute('SELECT COUNT(*) FROM pending_registrations').fetchone()[0]
    db.execute('DELETE FROM pending_registrations')
    db.commit()
    log_action('maintenance_clear', 'pending_registrations', None, {'cleared': n, 'by': session['username']})
    return jsonify({'success': True, 'deleted': n})


@bp.route('/api/admin/maintenance/members', methods=['DELETE'])
@permission_required('admin.maintenance')
def api_maintenance_clear_members():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Only the admin role can clear all members'}), 403

    body   = request.get_json(silent=True) or {}
    phrase = body.get('confirm', '')
    if phrase != 'DELETE ALL MEMBERS':
        return jsonify({'error': 'Confirmation phrase incorrect'}), 400

    db            = get_db()
    n_members     = db.execute('SELECT COUNT(*) FROM members').fetchone()[0]
    n_attendance  = db.execute('SELECT COUNT(*) FROM attendance').fetchone()[0]
    n_dofe        = db.execute('SELECT COUNT(*) FROM dofe_participants').fetchone()[0]

    # v12.59: every table hanging off members must be cleared BEFORE the members
    # rows, or the DELETE aborts on a foreign-key constraint. session_notes.member_id
    # in particular has no ON DELETE action (and most child FKs are CASCADE only when
    # the pragma is on). We discover member-child tables dynamically (so any future
    # member-linked table is covered) and union a known list in case one lacks a
    # formal FK (e.g. member_contacts). session_notes is special-cased: only the
    # member-linked rows are removed, so general/club session notes are preserved.
    existing = {r['name'] for r in db.execute(
        "SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
    child_tables = set()
    for tname in existing:
        try:
            for fk in db.execute(f'PRAGMA foreign_key_list("{tname}")').fetchall():
                if fk['table'] == 'members':
                    child_tables.add(tname)
                    break
        except Exception:
            pass
    child_tables |= {
        'attendance', 'dofe_participants', 'member_sessions', 'member_contacts',
        'member_tags', 'member_field_values', 'member_flags', 'member_payments',
        'session_notes',
    }

    n_notes = db.execute(
        'SELECT COUNT(*) FROM session_notes WHERE member_id IS NOT NULL'
    ).fetchone()[0] if 'session_notes' in existing else 0

    # Detach member-linked notes only, then full-wipe the strictly per-member tables.
    if 'session_notes' in existing:
        db.execute('DELETE FROM session_notes WHERE member_id IS NOT NULL')
    for t in sorted((child_tables & existing) - {'session_notes'}):
        db.execute(f'DELETE FROM "{t}"')
    db.execute('DELETE FROM members')
    db.commit()

    log_action('maintenance_clear_members', 'members', None, {
        'members_deleted':    n_members,
        'attendance_deleted': n_attendance,
        'dofe_deleted':       n_dofe,
        'member_notes_deleted': n_notes,
        'child_tables_cleared': sorted((child_tables & existing) - {'session_notes'}),
        'by':                 session['username'],
    })
    return jsonify({
        'success':            True,
        'members_deleted':    n_members,
        'attendance_deleted': n_attendance,
        'dofe_deleted':       n_dofe,
    })


@bp.route('/api/admin/maintenance/backup')
@permission_required('admin.maintenance')
def api_maintenance_backup():
    """Stream a hot backup of the SQLCipher database."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename  = f'{club_slug()}_backup_{timestamp}.db'

    with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp:
        tmp_path = tmp.name

    try:
        src  = _connect_db()
        import sqlcipher3 as _sqlite3_cipher
        dest = _sqlite3_cipher.connect(tmp_path)
        dest.execute(f"PRAGMA key='{os.environ.get('DB_ENCRYPTION_KEY', '')}'")
        src.backup(dest)
        dest.close()
        src.close()

        with open(tmp_path, 'rb') as f:
            data = f.read()
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    log_action('maintenance_backup', 'database', None,
               {'filename': filename, 'size_bytes': len(data), 'by': session['username']})

    return Response(
        data,
        mimetype='application/octet-stream',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Content-Length':      str(len(data)),
        },
    )


@bp.route('/api/admin/maintenance/restore', methods=['POST'])
@permission_required('admin.maintenance')
def api_maintenance_restore():
    """Restore the database from an uploaded backup file."""
    import sqlcipher3 as _sqlite3_cipher

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded.'}), 400
    upload = request.files['file']
    if not upload.filename:
        return jsonify({'error': 'Empty filename.'}), 400

    with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp:
        tmp_path = tmp.name
        upload.save(tmp_path)

    try:
        # Step 1: validate
        try:
            check = _sqlite3_cipher.connect(tmp_path)
            check.execute(f"PRAGMA key='{os.environ.get('DB_ENCRYPTION_KEY', '')}'")
            check.execute('SELECT count(*) FROM sqlite_master')
        except Exception:
            return jsonify({
                'error': 'The uploaded file could not be opened. It may be corrupt '
                         'or was created with a different encryption key.'
            }), 400
        finally:
            try: check.close()
            except Exception: pass

        # Step 2: collect summary
        try:
            info_conn = _sqlite3_cipher.connect(tmp_path)
            info_conn.execute(f"PRAGMA key='{os.environ.get('DB_ENCRYPTION_KEY', '')}'")
            def _count(tbl):
                try:   return info_conn.execute(f'SELECT COUNT(*) FROM {tbl}').fetchone()[0]
                except: return None
            summary = {
                'members':    _count('members'),
                'users':      _count('users'),
                'audit_log':  _count('audit_log'),
                'attendance': _count('attendance'),
            }
            info_conn.close()
        except Exception:
            summary = {}

        # Step 3: drop this request's connection BEFORE touching the file
        # (v12.41: was previously done after the swap)
        if hasattr(g, 'db'):
            try: g.db.close()
            except Exception: pass
            g.db = None

        # Step 4: auto-snapshot — checkpoint the WAL into the main file first so
        # the snapshot contains all committed data, then copy.
        try:
            _ck = _connect_db()
            _ck.execute('PRAGMA wal_checkpoint(TRUNCATE)')
            _ck.close()
        except Exception:
            pass  # best effort — snapshot still taken below
        backups_dir   = os.path.join(INSTANCE_DIR, 'data', 'backups')
        os.makedirs(backups_dir, exist_ok=True)
        snapshot_ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
        snapshot_path = os.path.join(backups_dir, f'pre_restore_{snapshot_ts}.db')
        shutil.copy2(DATABASE, snapshot_path)

        # Step 5: remove stale WAL/SHM sidecars, then swap. Leftover -wal/-shm
        # files from the OLD database would be replayed against the restored
        # file on next open and corrupt it (v12.41).
        for _suffix in ('-wal', '-shm'):
            try:
                os.remove(DATABASE + _suffix)
            except OSError:
                pass
        shutil.copy2(tmp_path, DATABASE)

        # Step 6: run migrations on restored DB
        from db import ensure_tables
        with current_app.app_context():
            ensure_tables()

    finally:
        try: os.remove(tmp_path)
        except OSError: pass

    log_action('maintenance_restore', 'database', None, {
        'snapshot': snapshot_path,
        'summary':  summary,
        'by':       session['username'],
    })
    return jsonify({
        'success':  True,
        'summary':  summary,
        'snapshot': os.path.basename(snapshot_path),
    })


# ── Data import ───────────────────────────────────────────────────────────────

_IMPORT_CORE_FIELDS = [
    {'key': 'member_id',          'label': 'Member ID (existing)',      'field_type': 'text',     'required': False},
    {'key': 'first_name',         'label': 'First Name',                'field_type': 'text',     'required': True},
    {'key': 'surname',            'label': 'Surname',                   'field_type': 'text',     'required': True},
    {'key': 'date_of_birth',      'label': 'Date of Birth',             'field_type': 'date',     'required': False},
    {'key': 'status',             'label': 'Status',                    'field_type': 'select',   'required': False,
     'options': 'Active,Inactive,Leaver'},
    {'key': 'session',            'label': 'Session',                   'field_type': 'text',     'required': False},
    {'key': 'date_registered',    'label': 'Date Registered',           'field_type': 'date',     'required': False},
    {'key': 'staff_role',         'label': 'Staff Role',                'field_type': 'text',     'required': False},
    {'key': 'address',            'label': 'Address',                   'field_type': 'text',     'required': False},
    {'key': 'postcode',           'label': 'Postcode',                  'field_type': 'text',     'required': False},
    {'key': 'ethnicity_religion', 'label': 'Ethnicity / Religion',      'field_type': 'text',     'required': False},
    {'key': 'medical_sen',        'label': 'Medical / SEN',             'field_type': 'text',     'required': False},
    {'key': 'gp_contact',         'label': 'GP Contact',                'field_type': 'text',     'required': False},
    {'key': 'unattended_exit',    'label': 'Unattended Exit (Yes/No)',  'field_type': 'text',     'required': False},
    {'key': 'gdpr_consent',       'label': 'GDPR Consent (Yes/No)',     'field_type': 'text',     'required': False},
    {'key': 'comments',           'label': 'Notes / Comments',          'field_type': 'textarea', 'required': False},
    {'key': 'contact1_name',      'label': 'Contact 1 — Full Name',     'field_type': 'text',     'required': False},
    {'key': 'contact1_phone',     'label': 'Contact 1 — Phone',         'field_type': 'text',     'required': False},
    {'key': 'contact1_email',     'label': 'Contact 1 — Email',         'field_type': 'email',    'required': False},
    {'key': 'contact2_name',      'label': 'Contact 2 — Full Name',     'field_type': 'text',     'required': False},
    {'key': 'contact2_phone',     'label': 'Contact 2 — Phone',         'field_type': 'text',     'required': False},
    {'key': 'contact2_email',     'label': 'Contact 2 — Email',         'field_type': 'email',    'required': False},
    # Payment fields — create a member_payments record on import
    {'key': '_payment_paid',      'label': 'Payment — Mark as Paid (Yes/No)',  'field_type': 'text', 'required': False,
     'hint': 'A payment record will be created for any row where this is Yes / True / 1. Uses the current membership period and Membership payment type.'},
]


def _bool_val(v):
    if isinstance(v, (int, float)):
        return 1 if v else 0
    return 1 if str(v).strip().lower() in ('yes', 'true', '1', 'y') else 0


def _fmt_cell(v):
    import datetime as _dt
    if v is None:
        return ''
    if isinstance(v, _dt.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, _dt.date):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, _dt.time):
        return ''
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else str(v)
    return str(v).strip()


def _read_xlsx_file(path, sheet_name=None):
    ext = path.rsplit('.', 1)[-1].lower()
    if ext == 'xls':
        import xlrd
        wb          = xlrd.open_workbook(path)
        sheet_names = wb.sheet_names()
        ws          = wb.sheet_by_name(sheet_name) if (sheet_name and sheet_name in sheet_names) \
                      else wb.sheet_by_index(0)
        active      = ws.name
        rows        = [ws.row_values(i) for i in range(ws.nrows)]
    else:
        import openpyxl
        wb          = openpyxl.load_workbook(path, data_only=True)
        sheet_names = wb.sheetnames
        ws          = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
        active      = ws.title
        rows        = list(ws.iter_rows(values_only=True))

    if not rows:
        return sheet_names, active, [], []

    raw_headers = [str(c).strip() if c is not None else '' for c in rows[0]]
    last_col    = len(raw_headers)
    while last_col > 0 and not raw_headers[last_col - 1]:
        last_col -= 1
    headers   = raw_headers[:last_col]
    data_rows = []
    for row in rows[1:]:
        vals = [_fmt_cell(v) for v in row[:last_col]]
        if any(vals):
            data_rows.append(vals)
    return sheet_names, active, headers, data_rows


def _read_csv_file(path):
    with open(path, 'r', encoding='utf-8-sig', errors='replace') as fh:
        reader = list(_csv_mod.reader(fh))
    if not reader:
        return [], []
    headers   = [h.strip() for h in reader[0]]
    data_rows = [r for r in reader[1:] if any(c.strip() for c in r)]
    return headers, data_rows


@bp.route('/api/admin/import/analyse', methods=['POST'])
@csrf.exempt
@permission_required('admin.maintenance')
def api_import_analyse():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'No file selected'}), 400

    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    if ext not in ('xlsx', 'xls', 'csv'):
        return jsonify({'error': 'Only .xlsx, .xls and .csv files are supported'}), 400

    imports_dir = os.path.join(INSTANCE_DIR, 'data', 'imports')
    os.makedirs(imports_dir, exist_ok=True)
    file_id   = str(_uuid_mod.uuid4())
    save_path = os.path.join(imports_dir, f'{file_id}.{ext}')
    f.save(save_path)

    sheet_name = (request.form.get('sheet_name') or '').strip() or None
    try:
        if ext in ('xlsx', 'xls'):
            sheet_names, active_sheet, headers, data_rows = _read_xlsx_file(save_path, sheet_name)
        else:
            sheet_names, active_sheet = [], ''
            headers, data_rows = _read_csv_file(save_path)
    except Exception as exc:
        try: os.remove(save_path)
        except OSError: pass
        return jsonify({'error': f'Could not read file: {exc}'}), 400

    # ── Detect AYC export format ─────────────────────────────────────────────
    _AYC_SHEETS = {'Members', 'Attendance History', 'Payment History'}
    ayc_export  = bool(ext in ('xlsx', 'xls') and _AYC_SHEETS & set(sheet_names))
    ayc_sheet_info = {}
    if ayc_export:
        for _sn in _AYC_SHEETS:
            ayc_sheet_info[_sn] = _sn in sheet_names

    return jsonify({
        'file_id':       file_id,
        'file_ext':      ext,
        'sheet_names':   sheet_names,
        'active_sheet':  active_sheet,
        'columns':       headers,
        'preview':       data_rows[:5],
        'total_rows':    len(data_rows),
        'ayc_export':    ayc_export,
        'ayc_sheets':    ayc_sheet_info,
    })


@bp.route('/api/admin/import/fields/<int:type_id>')
@permission_required('admin.maintenance')
def api_import_fields(type_id):
    db = get_db()
    if not db.execute('SELECT id FROM member_types WHERE id = ?', (type_id,)).fetchone():
        return jsonify({'error': 'Member type not found'}), 404
    custom = db.execute('''
        SELECT  fd.id, fd.key, fd.label, fd.field_type, fd.options
        FROM    member_type_fields  mtf
        JOIN    field_definitions   fd  ON fd.id = mtf.field_id
        WHERE   mtf.member_type_id = ? AND fd.active = 1
        ORDER   BY mtf.sort_order, fd.label
    ''', (type_id,)).fetchall()
    core    = [f for f in _IMPORT_CORE_FIELDS if not f['key'].startswith('_payment')]
    payment = [f for f in _IMPORT_CORE_FIELDS if f['key'].startswith('_payment')]
    return jsonify({
        'core_fields':    core,
        'payment_fields': payment,
        'custom_fields':  [dict(f) for f in custom],
    })


@bp.route('/api/admin/import/run', methods=['POST'])
@csrf.exempt
@permission_required('admin.maintenance')
def api_import_run():
    from helpers import _next_member_id, _save_id_format_from_import

    data       = request.get_json() or {}
    file_id    = (data.get('file_id') or '').strip()
    file_ext   = (data.get('file_ext') or '').strip()
    sheet_name = (data.get('sheet_name') or '').strip() or None
    type_id    = data.get('member_type_id')
    mapping    = data.get('mapping', {})
    skip_dupes = data.get('skip_duplicates', True)
    # v12.58: multi-session register mode. When on, rows that repeat the same
    # member (per the configurable match_fields) are folded into ONE member and
    # their sessions unioned, instead of creating a duplicate member per row.
    multi_session = bool(data.get('multi_session', False))

    # v12.58: configurable duplicate identity. The chosen fields define "the same
    # person" for BOTH the skip-duplicate check and the multi-session merge.
    # Whitelisted against real member columns to keep the dynamic WHERE injection-safe.
    _MATCH_FIELD_WHITELIST = {
        'first_name', 'surname', 'date_of_birth', 'postcode', 'mobile', 'email',
    }
    match_fields = [f for f in (data.get('match_fields') or [])
                    if f in _MATCH_FIELD_WHITELIST]
    if not match_fields:   # back-compat default = legacy behaviour
        match_fields = ['first_name', 'surname', 'postcode']

    if not all([file_id, file_ext, type_id, mapping]):
        return jsonify({'error': 'Missing required parameters'}), 400

    imports_dir = os.path.join(INSTANCE_DIR, 'data', 'imports')
    save_path   = os.path.join(imports_dir, f'{file_id}.{file_ext}')
    if not os.path.exists(save_path):
        return jsonify({'error': 'Upload not found — please re-upload the file'}), 404

    db = get_db()
    mt = db.execute('SELECT * FROM member_types WHERE id = ?', (type_id,)).fetchone()
    if not mt:
        return jsonify({'error': 'Member type not found'}), 404

    custom_fields = {}
    for fd in db.execute('''
        SELECT fd.id, fd.key, fd.field_type, fd.column_name, fd.system_field
        FROM   member_type_fields mtf
        JOIN   field_definitions  fd ON fd.id = mtf.field_id
        WHERE  mtf.member_type_id = ?
    ''', (type_id,)).fetchall():
        custom_fields[fd['key']] = dict(fd)

    try:
        if file_ext in ('xlsx', 'xls'):
            _, _, file_headers, data_rows = _read_xlsx_file(save_path, sheet_name)
        else:
            file_headers, data_rows = _read_csv_file(save_path)
    except Exception as exc:
        return jsonify({'error': f'Could not read file: {exc}'}), 400

    CORE_KEYS = {
        'first_name', 'surname', 'date_of_birth', 'address', 'postcode',
        'session', 'ethnicity_religion', 'medical_sen', 'gp_contact',
        'unattended_exit', 'gdpr_consent', 'staff_role',
        'date_registered', 'comments', 'status',
    }
    # email, mobile, member_id and the contact_* fields are handled explicitly in
    # the mapping loop below — they bypass the custom_vals / EAV path.
    PAYMENT_KEYS = {'_payment_paid'}

    # Pre-fetch payment type + current period if any payment column is mapped
    _has_payment_col  = any(v in PAYMENT_KEYS for v in mapping.values())
    _membership_type  = None
    _import_period    = ''
    if _has_payment_col:
        _import_period   = get_setting('current_membership_period', '')
        _membership_type = db.execute(
            "SELECT id FROM payment_types WHERE is_membership = 1 LIMIT 1"
        ).fetchone()
        if not _membership_type:
            _membership_type = db.execute(
                'SELECT id FROM payment_types ORDER BY sort_order, id LIMIT 1'
            ).fetchone()

    imported          = 0
    skipped           = 0
    merged            = 0           # v12.58: extra rows folded into an existing member
    errors            = []
    warnings          = []          # v12.58: non-fatal merge notes (e.g. phone clash)
    not_imported      = []
    imported_ids_used = []
    run_member_ids    = {}          # v12.58: name_key -> member DB id seen this run

    for row_num, row in enumerate(data_rows, start=2):
        try:
            core        = {}
            custom_vals = {}
            contacts    = {}
            email_val   = None
            mobile_val  = None
            provided_id = None

            for col_str, field_key in mapping.items():
                if field_key == '_skip':
                    continue
                col_idx = int(col_str)
                val     = row[col_idx] if col_idx < len(row) else ''
                if not val:
                    continue

                if field_key == 'email':
                    email_val = str(val).strip()
                elif field_key == 'mobile':
                    mobile_val = str(val).strip()
                elif field_key == 'member_id':
                    provided_id = str(val).strip()
                elif field_key in (
                    'contact1_name', 'contact1_phone', 'contact1_email',
                    'contact2_name', 'contact2_phone', 'contact2_email',
                ):
                    contacts[field_key] = str(val).strip()
                elif field_key in PAYMENT_KEYS:
                    core[field_key] = val  # store raw value; handled after INSERT
                elif field_key in CORE_KEYS:
                    if field_key in ('unattended_exit', 'gdpr_consent'):
                        core[field_key] = _bool_val(val)
                    elif field_key == 'status':
                        _s = str(val).strip()
                        if _s:
                            # Build a case-insensitive map from the member_statuses table
                            _valid = {r['name'].lower(): r['name'] for r in db.execute(
                                'SELECT name FROM member_statuses'
                            ).fetchall()}
                            _matched = _valid.get(_s.lower())
                            if _matched:
                                core[field_key] = _matched
                            # Unknown status values are silently ignored; the default ('Active')
                            # is applied at INSERT time. Import warnings are surfaced via not_imported.
                    else:
                        core[field_key] = str(val).strip()
                elif field_key in custom_fields:
                    custom_vals[field_key] = val

            def _row_name():
                parts = [core.get('first_name', ''), core.get('surname', '')]
                return ' '.join(p for p in parts if p) or f'Row {row_num}'

            if not core.get('first_name') and not core.get('surname'):
                skipped += 1
                not_imported.append({'row': row_num, 'name': f'Row {row_num} (blank)',
                                     'reason': 'Blank row — no name data found'})
                continue

            # v12.58: build this row's identity from the configurable match fields.
            _match_vals = {}
            for f in match_fields:
                if f == 'mobile':
                    _match_vals[f] = (mobile_val or '')
                elif f == 'email':
                    _match_vals[f] = (email_val or '')
                else:
                    _match_vals[f] = (core.get(f) or '')
            match_key = tuple((_match_vals.get(f) or '').strip().lower()
                              for f in match_fields)
            _has_identity = any(v for v in match_key)

            # v12.58: multi-session — fold repeat members into one, unioning sessions.
            if multi_session and _has_identity:
                target_id    = run_member_ids.get(match_key)
                _status_warn = None
                # Match an existing portal member too, but only when the user has
                # asked to treat matches as the same person (skip-duplicates on).
                if target_id is None and skip_dupes:
                    _existing = _find_existing_member(db, match_fields, _match_vals)
                    if _existing is not None:
                        target_id = _existing['id']
                        # v12.66 (audit fix #3): merging onto an archived/left record
                        # "succeeds" but the member stays invisible on every register —
                        # surface it for review instead of folding the row in silently.
                        if _existing['status_behaviour'] != 'active':
                            _status_warn = (f"combined into an existing member whose status is "
                                            f"'{_existing['status']}' — they will not appear on any "
                                            f"register until their status is changed")
                if target_id is not None:
                    _w = _merge_member_row(db, target_id, core.get('session') or '',
                                           mobile_val, email_val, contacts)
                    # v12.67 (audit fix #4): a merged row can carry the paid marker —
                    # previously it was dropped (the payment block below is only
                    # reached by first-occurrence rows). Record a whole-club
                    # membership payment on the target member unless they already
                    # have a non-voided one for the current period.
                    if _bool_val(core.get('_payment_paid', 0)) and _membership_type and _import_period:
                        _has_pay = db.execute(
                            'SELECT 1 FROM member_payments mp '
                            'JOIN payment_types pt ON pt.id = mp.payment_type_id '
                            'WHERE mp.member_id = ? AND mp.period = ? '
                            '  AND pt.is_membership = 1 AND mp.voided_at IS NULL LIMIT 1',
                            (target_id, _import_period)
                        ).fetchone()
                        if not _has_pay:
                            db.execute(
                                'INSERT INTO member_payments '
                                '(member_id, payment_type_id, period, recorded_by) VALUES (?,?,?,?)',
                                (target_id, _membership_type['id'], _import_period, None)
                            )
                    db.commit()
                    run_member_ids[match_key] = target_id
                    merged += 1
                    _w = '; '.join(x for x in (_w, _status_warn) if x)
                    if _w:
                        warnings.append({'row': row_num, 'name': _row_name(), 'warning': _w})
                    continue

            if provided_id and db.execute(
                'SELECT id FROM members WHERE member_id = ?', (provided_id,)
            ).fetchone():
                skipped += 1
                not_imported.append({'row': row_num, 'name': _row_name(),
                                     'reason': f'Member ID {provided_id} already exists in the portal'})
                continue

            # v12.58: duplicate check runs against the configurable match fields
            # (case-insensitive, NULL treated as blank). In multi-session mode an
            # existing match was already merged above, so this only fires for
            # single-session imports. v12.66: staff records are excluded from the
            # match; a non-active-status match names the status so the results
            # screen shows why the "duplicate" isn't on any register.
            if skip_dupes and not provided_id and not multi_session:
                _existing = _find_existing_member(db, match_fields, _match_vals)
                if _existing is not None:
                    skipped += 1
                    _flds = ' + '.join(f.replace('_', ' ').title() for f in match_fields)
                    _note = ('' if _existing['status_behaviour'] == 'active' else
                             f" — note: the existing record's status is '{_existing['status']}'")
                    not_imported.append({'row': row_num, 'name': _row_name(),
                                         'reason': f'Duplicate — already exists in the portal (matched on {_flds}){_note}'})
                    continue

            member_id = provided_id if provided_id else _next_member_id(db)
            db.execute('''
                INSERT INTO members
                    (member_id, first_name, surname, date_of_birth, address,
                     postcode, session, ethnicity_religion, medical_sen,
                     gp_contact, unattended_exit, gdpr_consent, staff_role,
                     date_registered, comments, status, member_type,
                     email, mobile)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                member_id,
                core.get('first_name', ''), core.get('surname', ''),
                core.get('date_of_birth') or None, core.get('address') or None,
                core.get('postcode') or None, core.get('session') or None,
                core.get('ethnicity_religion') or None, core.get('medical_sen') or None,
                core.get('gp_contact') or None,
                core.get('unattended_exit', 0), core.get('gdpr_consent', 0),
                core.get('staff_role') or None, core.get('date_registered') or None,
                core.get('comments') or None, core.get('status', 'Active'),
                mt['slug'],
                email_val or None, mobile_val or None,
            ))
            new_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]

            for order, prefix in ((1, 'contact1'), (2, 'contact2')):
                name  = contacts.get(f'{prefix}_name',  '').strip()
                phone = contacts.get(f'{prefix}_phone', '').strip()
                email = contacts.get(f'{prefix}_email', '').strip()
                if name or phone or email:
                    db.execute(
                        '''INSERT INTO member_contacts
                               (member_id, contact_order, contact_name, contact_phone, contact_email)
                           VALUES (?,?,?,?,?)''',
                        (new_id, order, name or None, phone or None, email or None)
                    )

            # Columns on members that can be written via UPDATE (beyond the INSERT set)
            _MEMBER_COLUMNS = {
                'first_name', 'surname', 'date_of_birth', 'address', 'postcode',
                'ethnicity_religion', 'medical_sen', 'gp_contact', 'session',
                'unattended_exit', 'gdpr_consent', 'staff_role', 'status',
                'date_registered', 'comments', 'mobile', 'email',
            }
            for fkey, fval in custom_vals.items():
                fd = custom_fields[fkey]
                col = fd.get('column_name')
                if col and col in _MEMBER_COLUMNS:
                    # System field that maps to a named column — write directly
                    db.execute(f'UPDATE members SET {col} = ? WHERE id = ?',
                               (str(fval), new_id))
                else:
                    # Custom/EAV field — write to member_field_values
                    db.execute(
                        'INSERT OR REPLACE INTO member_field_values (member_id, field_id, value) VALUES (?,?,?)',
                        (new_id, fd['id'], str(fval))
                    )

            # v12.55: 'session' may be a delimited list ("Tuesday; Thursday").
            # Parse, validate and write member_sessions + the echo column.
            _raw_sess = core.get('session')
            if _raw_sess is None:
                _cur = db.execute('SELECT session FROM members WHERE id = ?', (new_id,)).fetchone()
                _raw_sess = _cur['session'] if _cur else ''
            _valid_sess, _bad_sess = _apply_member_sessions(db, new_id, _raw_sess)
            if _bad_sess:
                errors.append({'row': row_num,
                               'error': 'unknown session(s) ignored: ' + ', '.join(_bad_sess)})

            # ── Create payment record if marked as paid ──────────────────────
            if _bool_val(core.get('_payment_paid', 0)) and _membership_type and _import_period:
                db.execute(
                    'INSERT INTO member_payments '
                    '(member_id, payment_type_id, period, recorded_by) VALUES (?,?,?,?)',
                    (new_id, _membership_type['id'], _import_period, None)
                )

            db.commit()
            imported += 1
            run_member_ids[match_key] = new_id   # v12.58: later repeats merge here
            if provided_id:
                imported_ids_used.append(member_id)

        except Exception as exc:
            errors.append({'row': row_num, 'error': str(exc)})
            not_imported.append({
                'row': row_num,
                'name': ' '.join(filter(None, [
                    (row[int(k)] if int(k) < len(row) else '') for k, v in mapping.items()
                    if v in ('first_name', 'surname')
                ])) or f'Row {row_num}',
                'reason': f'Error: {exc}',
            })
            try: db.rollback()
            except Exception: pass

    try: os.remove(save_path)
    except OSError: pass

    if imported_ids_used:
        _save_id_format_from_import(db, imported_ids_used)

    report_id = None
    if not_imported:
        report_id   = str(_uuid_mod.uuid4())
        report_path = os.path.join(imports_dir, f'{report_id}_report.csv')
        with open(report_path, 'w', newline='', encoding='utf-8') as fh:
            writer = _csv_mod.writer(fh)
            writer.writerow(['Row #', 'Name', 'Reason Not Imported'])
            for rec in not_imported:
                writer.writerow([rec['row'], rec['name'], rec['reason']])

    log_action('import.run', 'members', None, {
        'imported': imported, 'skipped': skipped, 'merged': merged,
        'errors':   len(errors), 'member_type': mt['slug'],
        'multi_session': multi_session, 'match_fields': match_fields,
    })
    return jsonify({
        'imported':  imported,
        'skipped':   skipped,
        'merged':    merged,
        'warnings':  warnings,
        'errors':    errors,
        'report_id': report_id,
    })


@bp.route('/api/admin/maintenance/resync-sessions', methods=['POST'])
@permission_required('admin.maintenance')
def api_resync_sessions():
    """
    Diagnostic: reports how many active members have a session value that
    matches (or doesn't match) the current session types.
    From v10.17 onwards, renaming a session type cascades automatically.
    To fix existing mismatches, simply rename the session type to its correct
    name via Admin → Session Types — the cascade will update all member records.
    """
    db          = get_db()
    type_rows   = db.execute('SELECT name FROM session_types WHERE active = 1').fetchall()
    valid_names = [r['name'] for r in type_rows]
    if valid_names:
        ph      = ','.join('?' * len(valid_names))
        matched = db.execute(
            f'SELECT COUNT(*) FROM members WHERE session IN ({ph}) AND status != "Leaver"',
            valid_names).fetchone()[0]
        total   = db.execute(
            'SELECT COUNT(*) FROM members WHERE status != "Leaver"').fetchone()[0]
        unset   = total - matched
        samples = db.execute(
            f'SELECT DISTINCT session FROM members '
            f'WHERE (session IS NULL OR session = "" OR session NOT IN ({ph})) '
            f'AND status != "Leaver" LIMIT 10',
            valid_names).fetchall()
        sample_vals = [r['session'] for r in samples]
    else:
        matched, total, unset, sample_vals = 0, 0, 0, []

    log_action('resync_sessions_check', 'session_types', None,
               {'valid': valid_names, 'matched': matched, 'unset': unset})
    return jsonify({
        'session_types':   valid_names,
        'members_matched': matched,
        'members_total':   total,
        'members_unset':   unset,
        'unset_values':    sample_vals,
    })


@bp.route('/api/admin/import/report/<report_id>')
@permission_required('admin.maintenance')
def api_import_report(report_id):
    if not report_id or '/' in report_id or '..' in report_id:
        return jsonify({'error': 'Invalid report ID'}), 400
    imports_dir = os.path.join(INSTANCE_DIR, 'data', 'imports')
    report_path = os.path.join(imports_dir, f'{report_id}_report.csv')
    if not os.path.exists(report_path):
        return jsonify({'error': 'Report not found — it may have expired'}), 404
    return send_file(
        report_path,
        mimetype='text/csv',
        as_attachment=True,
        download_name='import_not_imported.csv',
    )


# ── Multi-session import helper (v12.55, Phase D) ──────────────────────────────

def _apply_member_sessions(db, member_db_id, raw_session_value):
    """Parse a delimited session list ("Tuesday; Thursday" — separators ; , / |),
    validate each name case-insensitively against session_types, rewrite the
    member's member_sessions rows and keep the members.session echo in sync
    (first valid session in sort order; empty when none).
    Returns (valid_names, invalid_tokens)."""
    types    = db.execute('SELECT id, name FROM session_types ORDER BY sort_order, name').fetchall()
    by_lower = {t['name'].lower(): t for t in types}
    order    = [t['name'] for t in types]
    valid, invalid, seen = [], [], set()
    for tok in _re.split(r'[;,/|]', str(raw_session_value or '')):
        tok = tok.strip()
        if not tok:
            continue
        t = by_lower.get(tok.lower())
        if t is None:
            invalid.append(tok)
        elif t['name'] not in seen:
            valid.append(t['name']); seen.add(t['name'])
    valid.sort(key=order.index)
    db.execute('DELETE FROM member_sessions WHERE member_id = ?', (member_db_id,))
    for n in valid:
        db.execute('INSERT OR IGNORE INTO member_sessions (member_id, session_type_id) VALUES (?,?)',
                   (member_db_id, by_lower[n.lower()]['id']))
    db.execute('UPDATE members SET session = ? WHERE id = ?',
               (valid[0] if valid else '', member_db_id))
    return valid, invalid


def _find_existing_member(db, match_fields, vals):
    """v12.58: find a member already in the portal matching on the configurable
    identity fields. Comparison is case-insensitive and treats NULL as blank, so
    an empty postcode/DOB still matches a blank. match_fields is pre-whitelisted
    to real column names by the caller. A match key with every value blank never
    matches (avoids collapsing all no-data rows).

    v12.66 (audit fix #3): returns the matched row (id, status, status_behaviour)
    instead of a bare id, and:
    - STAFF records are never matched — a register row whose match fields
      collide with a staff member must not merge sessions/contacts into (or be
      skipped against) the staff record;
    - among non-staff matches, a member whose status behaviour is 'active' is
      preferred, so a returning member matches their live record ahead of an
      archived duplicate; callers warn when the only match is non-active
      (merged rows would otherwise stay invisible on every register while the
      import reports success).
    Returns the row, or None."""
    if not any((vals.get(f) or '').strip() for f in match_fields):
        return None
    clauses, params = [], []
    for f in match_fields:
        clauses.append(f"lower(COALESCE(m.{f},'')) = ?")
        params.append((vals.get(f) or '').strip().lower())
    return db.execute(f'''
        SELECT m.id, m.status, COALESCE(ms.behaviour, '') AS status_behaviour
        FROM   members m
        LEFT JOIN member_types    mt ON mt.slug = m.member_type
        LEFT JOIN member_statuses ms ON ms.name = m.status
        WHERE  {' AND '.join(clauses)}
          AND  COALESCE(mt.registration_style, '') != 'staff'
        ORDER  BY CASE WHEN ms.behaviour = 'active' THEN 0 ELSE 1 END, m.id
        LIMIT  1''', params).fetchone()


def _merge_member_row(db, target_id, raw_session, mobile_val, email_val, contacts):
    """v12.58 multi-session import: fold a repeat register row into an existing
    member. Unions the row's session onto the member, backfills a blank
    mobile/email/contact from the extra row, and flags (does not overwrite) a
    non-empty phone that disagrees. Returns a warning string, or None."""
    warn = []

    # Union the extra session onto whatever the member already has.
    existing = [r['name'] for r in db.execute(
        '''SELECT st.name FROM member_sessions ms
           JOIN session_types st ON st.id = ms.session_type_id
           WHERE ms.member_id = ?''', (target_id,)).fetchall()]
    combined = ';'.join(existing + ([raw_session] if raw_session else []))
    _apply_member_sessions(db, target_id, combined)

    # Backfill blank member-level fields; flag a genuine phone clash.
    row = db.execute('SELECT mobile, email FROM members WHERE id = ?',
                     (target_id,)).fetchone()
    if mobile_val:
        cur = (row['mobile'] or '').strip()
        if not cur:
            db.execute('UPDATE members SET mobile = ? WHERE id = ?', (mobile_val, target_id))
        elif cur != mobile_val.strip():
            warn.append(f'phone {mobile_val} differs from existing {cur}')
    if email_val and not (row['email'] or '').strip():
        db.execute('UPDATE members SET email = ? WHERE id = ?', (email_val, target_id))

    # Backfill an emergency contact slot only when the member has none there.
    for order, prefix in ((1, 'contact1'), (2, 'contact2')):
        name  = contacts.get(f'{prefix}_name',  '').strip()
        phone = contacts.get(f'{prefix}_phone', '').strip()
        email = contacts.get(f'{prefix}_email', '').strip()
        if not (name or phone or email):
            continue
        if not db.execute(
            'SELECT 1 FROM member_contacts WHERE member_id = ? AND contact_order = ?',
            (target_id, order)).fetchone():
            db.execute(
                '''INSERT INTO member_contacts
                       (member_id, contact_order, contact_name, contact_phone, contact_email)
                   VALUES (?,?,?,?,?)''',
                (target_id, order, name or None, phone or None, email or None))

    return '; '.join(warn) if warn else None


# ── AYC round-trip import ──────────────────────────────────────────────────────

# Maps export column headers (normalised to lower) → import field keys.
# None = skip column; '__member_type' = store slug directly on members row.
_AYC_COL_MAP = {
    'member id':       'member_id',
    'first name':      'first_name',
    'surname':         'surname',
    'date of birth':   'date_of_birth',
    'date registered': 'date_registered',
    'session':         'session',
    'member type':     '__member_type',
    'staff role':      'staff_role',
    'status':          'status',
    'status note':     None,
    'mobile':          'mobile',
    'email':           'email',
    'unattended exit': 'unattended_exit',
    'contact 1 name':  'contact1_name',
    'contact 1 phone': 'contact1_phone',
    'contact 1 email': 'contact1_email',
    'contact 2 name':  'contact2_name',
    'contact 2 phone': 'contact2_phone',
    'contact 2 email': 'contact2_email',
    # summary / computed columns → skip
    'total sessions':  None,
    'last attended':   None,
    'payment count':   None,
    'total paid':      None,
    'last payment':    None,
    'active flags':    None,
}


def _ayc_import_members(db, save_path, file_ext):
    """Import the Members sheet from an AYC export XLSX. Returns result dict."""
    from helpers import _next_member_id, _save_id_format_from_import

    try:
        _, _, headers, data_rows = _read_xlsx_file(save_path, 'Members')
    except Exception as exc:
        return {'error': str(exc), 'imported': 0, 'skipped': 0}

    if not headers:
        return {'imported': 0, 'skipped': 0, 'note': 'Members sheet is empty'}

    # Normalised header → col index
    hmap = {h.lower().strip(): i for i, h in enumerate(headers)}

    # Pre-load ALL custom field definitions for label matching
    all_fd = db.execute(
        'SELECT id, label, key, field_type, column_name FROM field_definitions WHERE active = 1'
    ).fetchall()
    label_to_fd = {fd['label'].lower().strip(): dict(fd) for fd in all_fd}

    # Pre-load member type slug → id map
    mt_rows    = db.execute('SELECT id, slug FROM member_types').fetchall()
    slug_to_id = {r['slug']: r['id'] for r in mt_rows}

    # Default member-type slug for rows that don't specify one (or specify an
    # unknown slug). Use the club's primary active non-staff type rather than a
    # hardcoded 'member' — on clubs whose type was renamed (e.g. 'ara-members')
    # the literal 'member' matches nothing, leaving imported members orphaned
    # (invisible to every mt.slug = m.member_type join) with no custom fields.
    default_mt = db.execute('''
        SELECT slug FROM member_types
        WHERE active = 1 AND registration_style != 'staff'
        ORDER BY sort_order LIMIT 1
    ''').fetchone()
    default_member_type_slug = default_mt['slug'] if default_mt else 'member'

    # Valid statuses (case-insensitive)
    valid_statuses = {r['name'].lower(): r['name'] for r in db.execute(
        'SELECT name FROM member_statuses'
    ).fetchall()}

    CORE_KEYS = {
        'first_name', 'surname', 'date_of_birth', 'address', 'postcode',
        'session', 'ethnicity_religion', 'medical_sen', 'gp_contact',
        'unattended_exit', 'gdpr_consent', 'staff_role', 'date_registered',
        'comments', 'status',
    }
    _MEMBER_COLUMNS = {
        'first_name', 'surname', 'date_of_birth', 'address', 'postcode',
        'ethnicity_religion', 'medical_sen', 'gp_contact', 'session',
        'unattended_exit', 'gdpr_consent', 'staff_role', 'status',
        'date_registered', 'comments', 'mobile', 'email',
    }

    imported = 0
    skipped  = 0
    imported_ids_used = []

    for row in data_rows:
        def _cell(col_label):
            i = hmap.get(col_label)
            return str(row[i]).strip() if i is not None and i < len(row) and row[i] not in (None, '') else ''

        first_name = _cell('first name')
        surname    = _cell('surname')
        if not first_name and not surname:
            skipped += 1
            continue

        provided_id  = _cell('member id')
        member_type  = _cell('member type')   # slug
        # Fall back to the club's primary type when blank or unrecognised, so we
        # never create members with an orphaned member_type that no type matches.
        if member_type not in slug_to_id:
            member_type = default_member_type_slug

        if provided_id and db.execute(
            'SELECT id FROM members WHERE member_id = ?', (provided_id,)
        ).fetchone():
            skipped += 1
            continue

        member_id = provided_id if provided_id else _next_member_id(db)

        # Core fields
        core = {}
        for label_norm, field_key in _AYC_COL_MAP.items():
            if field_key is None or field_key in ('member_id', '__member_type'):
                continue
            i = hmap.get(label_norm)
            if i is None or i >= len(row):
                continue
            val = str(row[i]).strip() if row[i] not in (None, '') else ''
            if not val:
                continue
            if field_key in ('unattended_exit', 'gdpr_consent'):
                core[field_key] = _bool_val(val)
            elif field_key == 'status':
                matched = valid_statuses.get(val.lower())
                if matched:
                    core['status'] = matched
            elif field_key in CORE_KEYS or field_key in ('email', 'mobile'):
                core[field_key] = val

        email_val  = core.pop('email',  None)
        mobile_val = core.pop('mobile', None)

        # Contacts
        contacts = {}
        for prefix in ('contact1', 'contact2'):
            for suffix in ('name', 'phone', 'email'):
                lbl = f'{prefix[:-1]} {prefix[-1]} {suffix}'  # e.g. "contact 1 name"
                i   = hmap.get(lbl)
                if i is not None and i < len(row) and row[i] not in (None, ''):
                    contacts[f'{prefix}_{suffix}'] = str(row[i]).strip()

        try:
            db.execute('''
                INSERT INTO members
                    (member_id, first_name, surname, date_of_birth, address,
                     postcode, session, ethnicity_religion, medical_sen,
                     gp_contact, unattended_exit, gdpr_consent, staff_role,
                     date_registered, comments, status, member_type,
                     email, mobile)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                member_id,
                first_name, surname,
                core.get('date_of_birth') or None, core.get('address') or None,
                core.get('postcode') or None, core.get('session') or None,
                core.get('ethnicity_religion') or None, core.get('medical_sen') or None,
                core.get('gp_contact') or None,
                core.get('unattended_exit', 0), core.get('gdpr_consent', 0),
                core.get('staff_role') or None, core.get('date_registered') or None,
                core.get('comments') or None, core.get('status', 'Active'),
                member_type,
                email_val or None, mobile_val or None,
            ))
            new_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]

            for order, prefix in ((1, 'contact1'), (2, 'contact2')):
                name  = contacts.get(f'{prefix}_name',  '')
                phone = contacts.get(f'{prefix}_phone', '')
                email = contacts.get(f'{prefix}_email', '')
                if name or phone or email:
                    db.execute(
                        'INSERT INTO member_contacts '
                        '(member_id, contact_order, contact_name, contact_phone, contact_email) '
                        'VALUES (?,?,?,?,?)',
                        (new_id, order, name or None, phone or None, email or None)
                    )

            # Custom fields — match remaining columns by label
            type_id = slug_to_id.get(member_type)
            if type_id:
                type_fd_ids = {r['field_id'] for r in db.execute(
                    'SELECT field_id FROM member_type_fields WHERE member_type_id = ?', (type_id,)
                ).fetchall()}
                for label_norm, fd in label_to_fd.items():
                    if fd['id'] not in type_fd_ids:
                        continue
                    i = hmap.get(label_norm)
                    if i is None or i >= len(row):
                        continue
                    val = str(row[i]).strip() if row[i] not in (None, '') else ''
                    if not val:
                        continue
                    col = fd.get('column_name')
                    if col and col in _MEMBER_COLUMNS:
                        db.execute(f'UPDATE members SET {col} = ? WHERE id = ?', (val, new_id))
                    else:
                        db.execute(
                            'INSERT OR REPLACE INTO member_field_values (member_id, field_id, value) VALUES (?,?,?)',
                            (new_id, fd['id'], val)
                        )

            # v12.55: session column may be a delimited list — write junction rows
            _raw_sess = core.get('session')
            if _raw_sess is None:
                _cur = db.execute('SELECT session FROM members WHERE id = ?', (new_id,)).fetchone()
                _raw_sess = _cur['session'] if _cur else ''
            _apply_member_sessions(db, new_id, _raw_sess)

            db.commit()
            imported += 1
            if provided_id:
                imported_ids_used.append(member_id)

        except Exception:
            skipped += 1
            try: db.rollback()
            except Exception: pass

    if imported_ids_used:
        _save_id_format_from_import(db, imported_ids_used)

    return {'imported': imported, 'skipped': skipped}


def _ayc_import_attendance(db, save_path, file_ext):
    """Import the Attendance History sheet from an AYC export XLSX."""
    try:
        _, _, headers, data_rows = _read_xlsx_file(save_path, 'Attendance History')
    except Exception as exc:
        return {'error': str(exc), 'imported': 0, 'skipped': 0}

    if not headers:
        return {'imported': 0, 'skipped': 0, 'note': 'Attendance History sheet is empty'}

    hmap = {h.lower().strip(): i for i, h in enumerate(headers)}
    # Build member_id string → numeric id cache
    mid_cache = {}

    def _resolve_member(member_ref):
        if member_ref in mid_cache:
            return mid_cache[member_ref]
        row = db.execute('SELECT id FROM members WHERE member_id = ?', (member_ref,)).fetchone()
        mid_cache[member_ref] = row['id'] if row else None
        return mid_cache[member_ref]

    imported = 0
    skipped  = 0

    for row in data_rows:
        def _cell(label):
            i = hmap.get(label)
            return str(row[i]).strip() if i is not None and i < len(row) and row[i] not in (None, '') else ''

        member_ref   = _cell('member id')
        session_date = _cell('date')
        session_type = _cell('session')

        if not member_ref or not session_date:
            skipped += 1
            continue

        numeric_id = _resolve_member(member_ref)
        if not numeric_id:
            skipped += 1
            continue

        # Skip duplicate
        if db.execute(
            'SELECT 1 FROM attendance WHERE member_id=? AND session_date=? AND session_type=?',
            (numeric_id, session_date, session_type)
        ).fetchone():
            skipped += 1
            continue

        signed_in  = _cell('signed in')  or None
        signed_out = _cell('signed out') or None

        try:
            db.execute(
                'INSERT INTO attendance (member_id, session_date, session_type, signed_in_at, signed_out_at) '
                'VALUES (?,?,?,?,?)',
                (numeric_id, session_date, session_type, signed_in, signed_out)
            )
            db.commit()
            imported += 1
        except Exception:
            skipped += 1
            try: db.rollback()
            except Exception: pass

    return {'imported': imported, 'skipped': skipped}


def _ayc_import_payments(db, save_path, file_ext):
    """Import the Payment History sheet from an AYC export XLSX."""
    try:
        _, _, headers, data_rows = _read_xlsx_file(save_path, 'Payment History')
    except Exception as exc:
        return {'error': str(exc), 'imported': 0, 'skipped': 0}

    if not headers:
        return {'imported': 0, 'skipped': 0, 'note': 'Payment History sheet is empty'}

    hmap = {h.lower().strip(): i for i, h in enumerate(headers)}

    # Pre-load lookup caches
    mid_cache = {}
    pt_cache  = {}
    pm_cache  = {}

    def _resolve_member(ref):
        if ref not in mid_cache:
            r = db.execute('SELECT id FROM members WHERE member_id = ?', (ref,)).fetchone()
            mid_cache[ref] = r['id'] if r else None
        return mid_cache[ref]

    def _resolve_pt(name):
        if name not in pt_cache:
            r = db.execute('SELECT id FROM payment_types WHERE name = ?', (name,)).fetchone()
            if not r:
                r = db.execute('SELECT id FROM payment_types ORDER BY sort_order, id LIMIT 1').fetchone()
            pt_cache[name] = r['id'] if r else None
        return pt_cache[name]

    def _resolve_pm(name):
        if not name:
            return None
        if name not in pm_cache:
            r = db.execute('SELECT id FROM payment_methods WHERE name = ?', (name,)).fetchone()
            pm_cache[name] = r['id'] if r else None
        return pm_cache[name]

    imported = 0
    skipped  = 0

    for row in data_rows:
        def _cell(label):
            i = hmap.get(label)
            return str(row[i]).strip() if i is not None and i < len(row) and row[i] not in (None, '') else ''

        member_ref   = _cell('member id')
        period       = _cell('period')
        payment_date = _cell('payment date') or None
        pt_name      = _cell('payment type')
        pm_name      = _cell('method')
        amount_str   = _cell('amount')
        notes        = _cell('notes') or None

        if not member_ref or not period:
            skipped += 1
            continue

        numeric_id = _resolve_member(member_ref)
        if not numeric_id:
            skipped += 1
            continue

        pt_id = _resolve_pt(pt_name)
        if not pt_id:
            skipped += 1
            continue

        pm_id = _resolve_pm(pm_name)

        try:
            amount = float(amount_str) if amount_str else None
        except ValueError:
            amount = None

        # Dedup: same member + period + payment_type + date
        if db.execute(
            'SELECT 1 FROM member_payments WHERE member_id=? AND period=? AND payment_type_id=? AND payment_date IS ?',
            (numeric_id, period, pt_id, payment_date)
        ).fetchone():
            skipped += 1
            continue

        try:
            db.execute(
                'INSERT INTO member_payments (member_id, payment_type_id, period, payment_date, amount, method_id, notes) '
                'VALUES (?,?,?,?,?,?,?)',
                (numeric_id, pt_id, period, payment_date, amount, pm_id, notes)
            )
            db.commit()
            imported += 1
        except Exception:
            skipped += 1
            try: db.rollback()
            except Exception: pass

    return {'imported': imported, 'skipped': skipped}


@bp.route('/api/admin/import/run-ayc', methods=['POST'])
@csrf.exempt
@permission_required('admin.maintenance')
def api_import_run_ayc():
    """Round-trip import from an AYC export XLSX — additive, skips existing records."""
    data    = request.get_json() or {}
    file_id = (data.get('file_id')  or '').strip()
    file_ext = (data.get('file_ext') or '').strip()
    do_members    = data.get('import_members',    True)
    do_attendance = data.get('import_attendance', True)
    do_payments   = data.get('import_payments',   True)

    if not file_id or not file_ext:
        return jsonify({'error': 'Missing file_id or file_ext'}), 400

    imports_dir = os.path.join(INSTANCE_DIR, 'data', 'imports')
    save_path   = os.path.join(imports_dir, f'{file_id}.{file_ext}')
    if not os.path.exists(save_path):
        return jsonify({'error': 'Upload not found — please re-upload the file'}), 404

    db      = get_db()
    results = {}
    if do_members:
        results['members']    = _ayc_import_members(db, save_path, file_ext)
    if do_attendance:
        results['attendance'] = _ayc_import_attendance(db, save_path, file_ext)
    if do_payments:
        results['payments']   = _ayc_import_payments(db, save_path, file_ext)

    try: os.remove(save_path)
    except OSError: pass

    log_action('import.ayc_restore', 'members', None, results)
    return jsonify({'results': results})


# ── External history import (attendance + payments from third-party files) ──────

@bp.route('/api/admin/import/history/analyse', methods=['POST'])
@csrf.exempt
@permission_required('admin.maintenance')
def api_import_history_analyse():
    """Upload and analyse an external attendance or payment history file."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'No file selected'}), 400

    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    if ext not in ('xlsx', 'xls', 'csv'):
        return jsonify({'error': 'Only .xlsx, .xls and .csv files are supported'}), 400

    imports_dir = os.path.join(INSTANCE_DIR, 'data', 'imports')
    os.makedirs(imports_dir, exist_ok=True)
    file_id   = str(_uuid_mod.uuid4())
    save_path = os.path.join(imports_dir, f'{file_id}.{ext}')
    f.save(save_path)

    sheet_name = (request.form.get('sheet_name') or '').strip() or None
    try:
        if ext in ('xlsx', 'xls'):
            sheet_names, active_sheet, headers, data_rows = _read_xlsx_file(save_path, sheet_name)
        else:
            sheet_names, active_sheet = [], ''
            headers, data_rows = _read_csv_file(save_path)
    except Exception as exc:
        try: os.remove(save_path)
        except OSError: pass
        return jsonify({'error': f'Could not read file: {exc}'}), 400

    # Detect wide attendance format: count columns that look like dates
    import re as _re2
    _date_pat = _re2.compile(
        r'^\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}$'   # 01/09/2024
        r'|\d{4}[/\-\.]\d{2}[/\-\.]\d{2}$'           # 2024-09-01
        r'|\d{1,2}\s+\w+\s+\d{4}$'                   # 1 Sep 2024
    )
    date_col_count = sum(1 for h in headers if _date_pat.match(h.strip()))
    likely_wide    = date_col_count >= 3 and date_col_count > len(headers) // 2

    return jsonify({
        'file_id':       file_id,
        'file_ext':      ext,
        'sheet_names':   sheet_names,
        'active_sheet':  active_sheet,
        'columns':       headers,
        'preview':       data_rows[:5],
        'total_rows':    len(data_rows),
        'likely_wide':   likely_wide,
        'date_col_count': date_col_count,
    })


@bp.route('/api/admin/import/history/run', methods=['POST'])
@csrf.exempt
@permission_required('admin.maintenance')
def api_import_history_run():
    """Import external attendance or payment history into the portal."""
    data     = request.get_json() or {}
    file_id  = (data.get('file_id')  or '').strip()
    file_ext = (data.get('file_ext') or '').strip()
    sheet_name = (data.get('sheet_name') or '').strip() or None
    import_type = (data.get('import_type') or '').strip()  # 'attendance' or 'payments'
    match_by    = data.get('match_by', 'member_id')         # 'member_id' or 'name'
    mapping     = data.get('mapping', {})                   # colIndex → role

    if not file_id or not file_ext or import_type not in ('attendance', 'payments'):
        return jsonify({'error': 'Missing required parameters'}), 400

    imports_dir = os.path.join(INSTANCE_DIR, 'data', 'imports')
    save_path   = os.path.join(imports_dir, f'{file_id}.{file_ext}')
    if not os.path.exists(save_path):
        return jsonify({'error': 'Upload not found — please re-upload the file'}), 404

    try:
        if file_ext in ('xlsx', 'xls'):
            _, _, headers, data_rows = _read_xlsx_file(save_path, sheet_name)
        else:
            headers, data_rows = _read_csv_file(save_path)
    except Exception as exc:
        return jsonify({'error': f'Could not read file: {exc}'}), 400

    db = get_db()

    # Member resolution helper
    _mid_cache = {}
    def _resolve(row):
        if match_by == 'member_id':
            col_i = next((int(k) for k, v in mapping.items() if v == 'member_id'), None)
            ref   = str(row[col_i]).strip() if col_i is not None and col_i < len(row) else ''
            if not ref: return None
            if ref not in _mid_cache:
                r = db.execute('SELECT id FROM members WHERE member_id = ?', (ref,)).fetchone()
                _mid_cache[ref] = r['id'] if r else None
            return _mid_cache[ref]
        else:  # name
            fn_i = next((int(k) for k, v in mapping.items() if v == 'first_name'), None)
            sn_i = next((int(k) for k, v in mapping.items() if v == 'surname'), None)
            fn   = str(row[fn_i]).strip() if fn_i is not None and fn_i < len(row) else ''
            sn   = str(row[sn_i]).strip() if sn_i is not None and sn_i < len(row) else ''
            key  = f'{fn}|{sn}'
            if key not in _mid_cache:
                r = db.execute('SELECT id FROM members WHERE first_name=? AND surname=?', (fn, sn)).fetchone()
                _mid_cache[key] = r['id'] if r else None
            return _mid_cache[key]

    def _col(row, role):
        i = next((int(k) for k, v in mapping.items() if v == role), None)
        return str(row[i]).strip() if i is not None and i < len(row) and row[i] not in (None, '') else ''

    imported = 0
    skipped  = 0
    errors   = []

    if import_type == 'attendance':
        att_format   = data.get('att_format', 'long')   # 'long' or 'wide'
        default_type = (data.get('default_session_type') or '').strip()

        if att_format == 'wide':
            # Wide: member col + date columns (member is resolved via _resolve())
            date_col_indices = [int(k) for k, v in mapping.items() if v == 'date_column']
            date_labels      = [headers[i] for i in date_col_indices if i < len(headers)]

            for row in data_rows:
                numeric_id = _resolve(row)
                if not numeric_id:
                    skipped += 1
                    continue
                for col_i, date_label in zip(date_col_indices, date_labels):
                    val = str(row[col_i]).strip() if col_i < len(row) and row[col_i] not in (None, '') else ''
                    if not val:
                        continue  # blank = absent
                    if db.execute(
                        'SELECT 1 FROM attendance WHERE member_id=? AND session_date=? AND session_type=?',
                        (numeric_id, date_label, default_type)
                    ).fetchone():
                        skipped += 1
                        continue
                    try:
                        db.execute(
                            'INSERT INTO attendance (member_id, session_date, session_type) VALUES (?,?,?)',
                            (numeric_id, date_label, default_type)
                        )
                        db.commit()
                        imported += 1
                    except Exception as exc:
                        errors.append(str(exc))
                        skipped += 1
                        try: db.rollback()
                        except Exception: pass
        else:
            # Long format
            for row in data_rows:
                numeric_id   = _resolve(row)
                session_date = _col(row, 'date')
                session_type = _col(row, 'session_type') or default_type
                if not numeric_id or not session_date:
                    skipped += 1
                    continue
                if db.execute(
                    'SELECT 1 FROM attendance WHERE member_id=? AND session_date=? AND session_type=?',
                    (numeric_id, session_date, session_type)
                ).fetchone():
                    skipped += 1
                    continue
                try:
                    db.execute(
                        'INSERT INTO attendance (member_id, session_date, session_type) VALUES (?,?,?)',
                        (numeric_id, session_date, session_type)
                    )
                    db.commit()
                    imported += 1
                except Exception as exc:
                    errors.append(str(exc))
                    skipped += 1
                    try: db.rollback()
                    except Exception: pass

    else:  # payments
        pt_cache = {}
        pm_cache = {}

        def _pt(name):
            if name not in pt_cache:
                r = db.execute('SELECT id FROM payment_types WHERE name = ?', (name,)).fetchone()
                if not r:
                    r = db.execute('SELECT id FROM payment_types ORDER BY sort_order, id LIMIT 1').fetchone()
                pt_cache[name] = r['id'] if r else None
            return pt_cache[name]

        def _pm(name):
            if not name: return None
            if name not in pm_cache:
                r = db.execute('SELECT id FROM payment_methods WHERE name = ?', (name,)).fetchone()
                pm_cache[name] = r['id'] if r else None
            return pm_cache[name]

        default_period = (data.get('default_period') or get_setting('current_membership_period', '')).strip()
        default_pt     = (data.get('default_payment_type') or '').strip()

        for row in data_rows:
            numeric_id   = _resolve(row)
            period       = _col(row, 'period') or default_period
            payment_date = _col(row, 'date') or None
            pt_name      = _col(row, 'payment_type') or default_pt
            pm_name      = _col(row, 'method')
            amount_str   = _col(row, 'amount')
            notes        = _col(row, 'notes') or None

            if not numeric_id or not period:
                skipped += 1
                continue

            pt_id = _pt(pt_name)
            if not pt_id:
                skipped += 1
                continue

            try:
                amount = float(amount_str) if amount_str else None
            except ValueError:
                amount = None

            if db.execute(
                'SELECT 1 FROM member_payments WHERE member_id=? AND period=? AND payment_type_id=? AND payment_date IS ?',
                (numeric_id, period, pt_id, payment_date)
            ).fetchone():
                skipped += 1
                continue

            try:
                db.execute(
                    'INSERT INTO member_payments (member_id, payment_type_id, period, payment_date, amount, method_id, notes) '
                    'VALUES (?,?,?,?,?,?,?)',
                    (numeric_id, pt_id, period, payment_date, amount, _pm(pm_name), notes)
                )
                db.commit()
                imported += 1
            except Exception as exc:
                errors.append(str(exc))
                skipped += 1
                try: db.rollback()
                except Exception: pass

    try: os.remove(save_path)
    except OSError: pass

    log_action(f'import.history.{import_type}', 'members', None,
               {'imported': imported, 'skipped': skipped})
    return jsonify({'imported': imported, 'skipped': skipped,
                    'errors': errors[:20]})


# ── Member Export ──────────────────────────────────────────────────────────────

@bp.route('/api/admin/export/preview')
@permission_required('admin.maintenance')
def api_export_preview():
    """Return a count of members matching the current filter options."""
    db              = get_db()
    status_filter   = request.args.get('status_filter', 'active')
    session_filter  = request.args.get('session', 'all')
    type_filter     = request.args.get('member_type', 'all')

    where, params = _export_where(status_filter, session_filter, type_filter)
    row = db.execute(f'SELECT COUNT(*) AS n FROM members m WHERE {where}', params).fetchone()
    return jsonify({'count': row['n']})


@bp.route('/api/admin/export/members.csv')
@permission_required('admin.maintenance')
def api_export_members_csv():
    """Stream a full member data CSV export."""
    import io
    db              = get_db()
    status_filter   = request.args.get('status_filter', 'active')
    session_filter  = request.args.get('session', 'all')
    type_filter     = request.args.get('member_type', 'all')
    sort_by         = request.args.get('sort', 'member_id')
    _sections_raw   = request.args.get('sections', 'core,contacts,custom,attendance')
    sections        = set(s.strip() for s in _sections_raw.split(',') if s.strip())
    if not sections:
        sections = {'core', 'contacts', 'custom', 'attendance'}

    inc_core       = 'core'       in sections
    inc_contacts   = 'contacts'   in sections
    inc_custom     = 'custom'     in sections
    inc_attendance = 'attendance' in sections
    inc_payments   = 'payments'   in sections
    inc_flags      = 'flags'      in sections

    where, params = _export_where(status_filter, session_filter, type_filter)

    # ── Fetch members with contacts ──────────────────────────────────────────
    members = db.execute(f'''
        SELECT  m.*,
                (SELECT group_concat(name, '; ') FROM (
                    SELECT st_s.name FROM member_sessions ms_s
                    JOIN session_types st_s ON st_s.id = ms_s.session_type_id
                    WHERE ms_s.member_id = m.id ORDER BY st_s.sort_order, st_s.name
                )) AS session_list,
                c1.contact_name  AS contact1_name,
                c1.contact_phone AS contact1_phone,
                c1.contact_email AS contact1_email,
                c2.contact_name  AS contact2_name,
                c2.contact_phone AS contact2_phone,
                c2.contact_email AS contact2_email
        FROM    members m
        LEFT JOIN member_contacts c1 ON c1.member_id = m.id AND c1.contact_order = 1
        LEFT JOIN member_contacts c2 ON c2.member_id = m.id AND c2.contact_order = 2
        WHERE   {where}
        ORDER   BY {_export_order(sort_by)}
    ''', params).fetchall()
    members = [dict(r) for r in members]
    for _m in members:   # v12.55: Session column = every assigned session, '; '-joined
        _m['session'] = _m.pop('session_list', None) or _m.get('session') or ''

    if not members:
        return Response('No members match the selected filters.\n',
                        mimetype='text/csv',
                        headers={'Content-Disposition': 'attachment; filename="members_export.csv"'})

    member_ids = [m['id'] for m in members]
    placeholders = ','.join('?' * len(member_ids))

    # ── Custom field values ──────────────────────────────────────────────────
    custom_map = {}
    if inc_custom:
        cfv_rows = db.execute(
            f'SELECT mfv.member_id, fd.key, mfv.value '
            f'FROM member_field_values mfv '
            f'JOIN field_definitions fd ON fd.id = mfv.field_id '
            f'WHERE mfv.member_id IN ({placeholders})',
            member_ids
        ).fetchall()
        for r in cfv_rows:
            custom_map.setdefault(r['member_id'], {})[r['key']] = r['value']

    # ── Attendance summary ───────────────────────────────────────────────────
    att_map = {}
    if inc_attendance:
        att_rows = db.execute(
            f'SELECT member_id, COUNT(*) AS total_sessions, MAX(session_date) AS last_attended '
            f'FROM attendance WHERE member_id IN ({placeholders}) '
            f'GROUP BY member_id',
            member_ids
        ).fetchall()
        att_map = {r['member_id']: dict(r) for r in att_rows}

    # ── Payment summary ──────────────────────────────────────────────────────
    payments_map   = {}
    current_period = ''
    if inc_payments:
        current_period = get_setting('current_membership_period', '')
        pay_rows = db.execute(
            f'''SELECT mp.member_id,
                       COUNT(CASE WHEN mp.voided_at IS NULL THEN 1 END) AS payment_count,
                       SUM(CASE WHEN mp.voided_at IS NULL THEN COALESCE(mp.amount, 0) ELSE 0 END) AS total_paid,
                       MAX(CASE WHEN mp.voided_at IS NULL THEN mp.payment_date ELSE NULL END) AS last_payment_date,
                       MAX(CASE WHEN mp.voided_at IS NULL AND mp.period = ? THEN 1 ELSE 0 END) AS paid_this_period
                FROM member_payments mp
                WHERE mp.member_id IN ({placeholders})
                GROUP BY mp.member_id''',
            [current_period] + member_ids
        ).fetchall()
        payments_map = {r['member_id']: dict(r) for r in pay_rows}

    # ── Active flags ─────────────────────────────────────────────────────────
    flags_map = {}
    if inc_flags:
        flag_rows = db.execute(
            f'''SELECT mf.member_id, GROUP_CONCAT(ar.name, ', ') AS flag_names
                FROM member_flags mf
                JOIN alert_rules ar ON ar.id = mf.rule_id
                WHERE mf.member_id IN ({placeholders}) AND mf.resolved_at IS NULL
                GROUP BY mf.member_id''',
            member_ids
        ).fetchall()
        flags_map = {r['member_id']: r['flag_names'] for r in flag_rows}

    # ── Collect active custom field definitions for relevant member types ────
    custom_fields = []
    if inc_custom:
        type_slugs = list({m['member_type'] for m in members})
        type_ids_rows = db.execute(
            f'SELECT id, slug FROM member_types WHERE slug IN ({",".join("?" * len(type_slugs))})',
            type_slugs
        ).fetchall()
        type_id_map = {r['slug']: r['id'] for r in type_ids_rows}
        custom_field_defs = db.execute(
            f'''SELECT DISTINCT fd.key, fd.label, fd.field_type
                FROM member_type_fields mtf
                JOIN field_definitions fd ON fd.id = mtf.field_id
                WHERE mtf.member_type_id IN ({",".join("?" * len(type_id_map))})
                  AND fd.active = 1 AND fd.system_field = 0
                ORDER BY mtf.sort_order''',
            list(type_id_map.values())
        ).fetchall()
        seen_keys = set()
        for r in custom_field_defs:
            if r['key'] not in seen_keys:
                seen_keys.add(r['key'])
                custom_fields.append(dict(r))

    # ── Build CSV ────────────────────────────────────────────────────────────
    core_cols = [
        ('member_id',       'Member ID'),
        ('first_name',      'First Name'),
        ('surname',         'Surname'),
        ('date_of_birth',   'Date of Birth'),
        ('date_registered', 'Date Registered'),
        ('session',         'Session'),
        ('member_type',     'Member Type'),
        ('staff_role',      'Staff Role'),
        ('status',          'Status'),
        ('status_note',     'Status Note'),
        ('mobile',          'Mobile'),
        ('email',           'Email'),
        ('unattended_exit', 'Unattended Exit'),
    ]
    contact_cols = [
        ('contact1_name',  'Contact 1 Name'),
        ('contact1_phone', 'Contact 1 Phone'),
        ('contact1_email', 'Contact 1 Email'),
        ('contact2_name',  'Contact 2 Name'),
        ('contact2_phone', 'Contact 2 Phone'),
        ('contact2_email', 'Contact 2 Email'),
    ]
    attendance_cols = [
        ('total_sessions', 'Total Sessions'),
        ('last_attended',  'Last Attended'),
    ]
    payment_cols = [
        ('payment_count',    'Payment Count'),
        ('total_paid',       'Total Paid'),
        ('last_payment_date','Last Payment Date'),
        ('paid_this_period', f'Paid This Period ({current_period})' if current_period else 'Paid This Period'),
    ]

    header = (
        ([label for _, label in core_cols]       if inc_core       else []) +
        ([label for _, label in contact_cols]     if inc_contacts   else []) +
        ([label for _, label in attendance_cols]  if inc_attendance else []) +
        ([label for _, label in payment_cols]     if inc_payments   else []) +
        (['Active Flags']                         if inc_flags      else []) +
        ([f['label'] for f in custom_fields]      if inc_custom     else [])
    )

    output = io.StringIO()
    writer = _csv_mod.writer(output)
    writer.writerow(header)

    for m in members:
        att  = att_map.get(m['id'], {})
        cfvs = custom_map.get(m['id'], {})
        pay  = payments_map.get(m['id'], {})

        paid_flag = 'Yes' if pay.get('paid_this_period') else ('No' if pay else '')
        row = (
            ([m.get(key, '') or '' for key, _ in core_cols]       if inc_core       else []) +
            ([m.get(key, '') or '' for key, _ in contact_cols]     if inc_contacts   else []) +
            ([att.get('total_sessions', 0), att.get('last_attended', '') or ''] if inc_attendance else []) +
            ([pay.get('payment_count', 0), pay.get('total_paid', ''), pay.get('last_payment_date', '') or '', paid_flag] if inc_payments else []) +
            ([flags_map.get(m['id'], '')]                         if inc_flags      else []) +
            ([cfvs.get(f['key'], '') or '' for f in custom_fields] if inc_custom     else [])
        )
        writer.writerow(row)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename  = f'{club_slug()}_members_{timestamp}.csv'

    log_action('export_members', 'members', None, {
        'count': len(members), 'status_filter': status_filter,
        'session_filter': session_filter, 'type_filter': type_filter,
    })

    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@bp.route('/api/admin/export/members.xlsx')
@permission_required('admin.maintenance')
def api_export_members_xlsx():
    """XLSX export: Sheet 1 = Members, Sheet 2 = Attendance History, Sheet 3 = Payment History."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    db             = get_db()
    status_filter  = request.args.get('status_filter', 'active')
    session_filter = request.args.get('session', 'all')
    type_filter    = request.args.get('member_type', 'all')
    sort_by        = request.args.get('sort', 'member_id')
    _sections_raw  = request.args.get('sections', 'core,contacts,custom,attendance,payments,flags')
    sections       = set(s.strip() for s in _sections_raw.split(',') if s.strip())
    if not sections:
        sections = {'core', 'contacts', 'custom', 'attendance', 'payments', 'flags'}

    inc_core       = 'core'       in sections
    inc_contacts   = 'contacts'   in sections
    inc_custom     = 'custom'     in sections
    inc_attendance = 'attendance' in sections
    inc_payments   = 'payments'   in sections
    inc_flags      = 'flags'      in sections

    where, params = _export_where(status_filter, session_filter, type_filter)

    # ── Fetch members ────────────────────────────────────────────────────────
    members = db.execute(f'''
        SELECT  m.*,
                (SELECT group_concat(name, '; ') FROM (
                    SELECT st_s.name FROM member_sessions ms_s
                    JOIN session_types st_s ON st_s.id = ms_s.session_type_id
                    WHERE ms_s.member_id = m.id ORDER BY st_s.sort_order, st_s.name
                )) AS session_list,
                c1.contact_name  AS contact1_name,
                c1.contact_phone AS contact1_phone,
                c1.contact_email AS contact1_email,
                c2.contact_name  AS contact2_name,
                c2.contact_phone AS contact2_phone,
                c2.contact_email AS contact2_email
        FROM    members m
        LEFT JOIN member_contacts c1 ON c1.member_id = m.id AND c1.contact_order = 1
        LEFT JOIN member_contacts c2 ON c2.member_id = m.id AND c2.contact_order = 2
        WHERE   {where}
        ORDER   BY {_export_order(sort_by)}
    ''', params).fetchall()
    members = [dict(r) for r in members]
    for _m in members:   # v12.55: Session column = every assigned session, '; '-joined
        _m['session'] = _m.pop('session_list', None) or _m.get('session') or ''

    if not members:
        return jsonify({'error': 'No members match the selected filters'}), 404

    member_ids   = [m['id'] for m in members]
    placeholders = ','.join('?' * len(member_ids))

    # ── Attendance summary for Sheet 1 ───────────────────────────────────────
    att_map = {}
    if inc_attendance:
        att_rows = db.execute(
            f'SELECT member_id, COUNT(*) AS total_sessions, MAX(session_date) AS last_attended '
            f'FROM attendance WHERE member_id IN ({placeholders}) GROUP BY member_id',
            member_ids
        ).fetchall()
        att_map = {r['member_id']: dict(r) for r in att_rows}

    # ── Payment summary for Sheet 1 ──────────────────────────────────────────
    payments_map   = {}
    current_period = ''
    if inc_payments:
        current_period = get_setting('current_membership_period', '')
        pay_rows = db.execute(
            f'''SELECT mp.member_id,
                       COUNT(CASE WHEN mp.voided_at IS NULL THEN 1 END) AS payment_count,
                       SUM(CASE WHEN mp.voided_at IS NULL THEN COALESCE(mp.amount, 0) ELSE 0 END) AS total_paid,
                       MAX(CASE WHEN mp.voided_at IS NULL THEN mp.payment_date ELSE NULL END) AS last_payment_date,
                       MAX(CASE WHEN mp.voided_at IS NULL AND mp.period = ? THEN 1 ELSE 0 END) AS paid_this_period
                FROM member_payments mp WHERE mp.member_id IN ({placeholders})
                GROUP BY mp.member_id''',
            [current_period] + member_ids
        ).fetchall()
        payments_map = {r['member_id']: dict(r) for r in pay_rows}

    # ── Active flags for Sheet 1 ─────────────────────────────────────────────
    flags_map = {}
    if inc_flags:
        flag_rows = db.execute(
            f'''SELECT mf.member_id, GROUP_CONCAT(ar.name, ', ') AS flag_names
                FROM member_flags mf JOIN alert_rules ar ON ar.id = mf.rule_id
                WHERE mf.member_id IN ({placeholders}) AND mf.resolved_at IS NULL
                GROUP BY mf.member_id''',
            member_ids
        ).fetchall()
        flags_map = {r['member_id']: r['flag_names'] for r in flag_rows}

    # ── Custom fields for Sheet 1 ────────────────────────────────────────────
    custom_map    = {}
    custom_fields = []
    if inc_custom:
        cfv_rows = db.execute(
            f'SELECT mfv.member_id, fd.key, mfv.value '
            f'FROM member_field_values mfv '
            f'JOIN field_definitions fd ON fd.id = mfv.field_id '
            f'WHERE mfv.member_id IN ({placeholders})',
            member_ids
        ).fetchall()
        for r in cfv_rows:
            custom_map.setdefault(r['member_id'], {})[r['key']] = r['value']
        type_slugs    = list({m['member_type'] for m in members})
        type_ids_rows = db.execute(
            f'SELECT id, slug FROM member_types WHERE slug IN ({",".join("?" * len(type_slugs))})',
            type_slugs
        ).fetchall()
        type_id_map = {r['slug']: r['id'] for r in type_ids_rows}
        cfd_rows = db.execute(
            f'''SELECT DISTINCT fd.key, fd.label, fd.field_type
                FROM member_type_fields mtf JOIN field_definitions fd ON fd.id = mtf.field_id
                WHERE mtf.member_type_id IN ({",".join("?" * len(type_id_map))})
                  AND fd.active = 1 AND fd.system_field = 0
                ORDER BY mtf.sort_order''',
            list(type_id_map.values())
        ).fetchall()
        seen_keys = set()
        for r in cfd_rows:
            if r['key'] not in seen_keys:
                seen_keys.add(r['key'])
                custom_fields.append(dict(r))

    # ── Full attendance history for Sheet 2 ──────────────────────────────────
    att_history = db.execute(
        f'''SELECT a.member_id, m.first_name, m.surname, m.member_id AS member_ref,
                   a.session_date, a.session_type, a.signed_in_at, a.signed_out_at
            FROM attendance a
            JOIN members m ON m.id = a.member_id
            WHERE a.member_id IN ({placeholders})
            ORDER BY a.session_date, a.session_type, m.surname, m.first_name''',
        member_ids
    ).fetchall()

    # ── Full payment history for Sheet 3 ─────────────────────────────────────
    pay_history = db.execute(
        f'''SELECT mp.member_id, m.first_name, m.surname, m.member_id AS member_ref,
                   mp.period, mp.payment_date, pt.name AS payment_type,
                   pm.name AS payment_method, mp.amount, mp.notes
            FROM member_payments mp
            JOIN members m  ON m.id  = mp.member_id
            JOIN payment_types pt ON pt.id = mp.payment_type_id
            LEFT JOIN payment_methods pm ON pm.id = mp.method_id
            WHERE mp.member_id IN ({placeholders}) AND mp.voided_at IS NULL
            ORDER BY mp.payment_date DESC, m.surname, m.first_name''',
        member_ids
    ).fetchall()

    # ── Build workbook ────────────────────────────────────────────────────────
    HEADER_FILL = PatternFill('solid', start_color='1e3a5f')
    HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    BODY_FONT   = Font(name='Arial', size=10)

    def _style_sheet(ws, col_widths):
        ws.freeze_panes = 'A2'
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for cell in ws[1]:
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY_FONT

    wb = Workbook()

    # ── Sheet 1: Members ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Members'

    core_cols = [
        ('member_id',       'Member ID',        14),
        ('first_name',      'First Name',        16),
        ('surname',         'Surname',           16),
        ('date_of_birth',   'Date of Birth',     14),
        ('date_registered', 'Date Registered',   16),
        ('session',         'Session',           16),
        ('member_type',     'Member Type',       16),
        ('staff_role',      'Staff Role',        16),
        ('status',          'Status',            12),
        ('status_note',     'Status Note',       24),
        ('mobile',          'Mobile',            14),
        ('email',           'Email',             26),
        ('unattended_exit', 'Unattended Exit',   15),
    ]
    contact_cols = [
        ('contact1_name',  'Contact 1 Name',  22),
        ('contact1_phone', 'Contact 1 Phone', 16),
        ('contact1_email', 'Contact 1 Email', 26),
        ('contact2_name',  'Contact 2 Name',  22),
        ('contact2_phone', 'Contact 2 Phone', 16),
        ('contact2_email', 'Contact 2 Email', 26),
    ]
    att_cols = [
        ('total_sessions', 'Total Sessions', 14),
        ('last_attended',  'Last Attended',  14),
    ]
    pay_col_label = f'Paid This Period ({current_period})' if current_period else 'Paid This Period'
    payment_summary_cols = [
        ('payment_count',     'Payment Count',   13),
        ('total_paid',        'Total Paid',       12),
        ('last_payment_date', 'Last Payment',     14),
        ('paid_this_period',  pay_col_label,      22),
    ]

    headers_s1 = (
        [(k, lbl, w) for k, lbl, w in core_cols       if inc_core]      +
        [(k, lbl, w) for k, lbl, w in contact_cols     if inc_contacts]  +
        [(k, lbl, w) for k, lbl, w in att_cols         if inc_attendance]+
        [(k, lbl, w) for k, lbl, w in payment_summary_cols if inc_payments] +
        ([('_flags', 'Active Flags', 30)]               if inc_flags     else []) +
        [('_cf_' + f['key'], f['label'], 18) for f in custom_fields]
    )

    ws1.append([lbl for _, lbl, _ in headers_s1])

    for m in members:
        att  = att_map.get(m['id'], {})
        pay  = payments_map.get(m['id'], {})
        cfvs = custom_map.get(m['id'], {})
        paid_flag = 'Yes' if pay.get('paid_this_period') else ('No' if pay else '')

        row_vals = []
        for key, _, _ in headers_s1:
            if key.startswith('_cf_'):
                row_vals.append(cfvs.get(key[4:], '') or '')
            elif key == '_flags':
                row_vals.append(flags_map.get(m['id'], '') or '')
            elif key in ('total_sessions', 'last_attended'):
                row_vals.append(att.get(key, '') or '')
            elif key == 'paid_this_period':
                row_vals.append(paid_flag)
            elif key in ('payment_count', 'total_paid', 'last_payment_date'):
                row_vals.append(pay.get(key, '') or '')
            else:
                row_vals.append(m.get(key, '') or '')
        ws1.append(row_vals)

    _style_sheet(ws1, [w for _, _, w in headers_s1])

    # ── Sheet 2: Attendance History ──────────────────────────────────────────
    ws2 = wb.create_sheet('Attendance History')
    att_h_headers = ['Member ID', 'First Name', 'Surname', 'Date', 'Session', 'Signed In', 'Signed Out']
    ws2.append(att_h_headers)
    for r in att_history:
        ws2.append([
            r['member_ref'], r['first_name'], r['surname'],
            r['session_date'], r['session_type'],
            r['signed_in_at'] or '', r['signed_out_at'] or '',
        ])
    _style_sheet(ws2, [14, 16, 16, 14, 16, 18, 18])

    # ── Sheet 3: Payment History ─────────────────────────────────────────────
    ws3 = wb.create_sheet('Payment History')
    pay_h_headers = ['Member ID', 'First Name', 'Surname', 'Period', 'Payment Date',
                     'Payment Type', 'Method', 'Amount', 'Notes']
    ws3.append(pay_h_headers)
    for r in pay_history:
        ws3.append([
            r['member_ref'], r['first_name'], r['surname'],
            r['period'], r['payment_date'] or '',
            r['payment_type'], r['payment_method'] or '',
            r['amount'] if r['amount'] is not None else '',
            r['notes'] or '',
        ])
    _style_sheet(ws3, [14, 16, 16, 16, 14, 20, 16, 10, 30])

    # ── Stream response ──────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename  = f'{club_slug()}_members_{timestamp}.xlsx'

    log_action('export_members_xlsx', 'members', None, {
        'count': len(members), 'status_filter': status_filter,
        'session_filter': session_filter, 'type_filter': type_filter,
    })

    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def _export_order(sort_by):
    """Return a safe ORDER BY expression for member export queries."""
    return {
        'surname':    'm.surname, m.first_name',
        'first_name': 'm.first_name, m.surname',
    }.get(sort_by, 'm.member_id')


def _export_where(status_filter, session_filter, type_filter):
    """Build the WHERE clause and params for member export queries."""
    conditions = []
    params     = []

    if status_filter == 'active':
        conditions.append(
            'EXISTS (SELECT 1 FROM member_statuses ms '
            'WHERE ms.name = m.status AND ms.behaviour = \'active\')'
        )
    # 'all' = no status filter

    if session_filter and session_filter != 'all':
        # v12.55: match any assigned session via the member_sessions junction
        conditions.append('EXISTS (SELECT 1 FROM member_sessions ms_x '
                          'JOIN session_types st_x ON st_x.id = ms_x.session_type_id '
                          'WHERE ms_x.member_id = m.id AND st_x.name = ?)')
        params.append(session_filter)

    if type_filter and type_filter != 'all':
        conditions.append('m.member_type = ?')
        params.append(type_filter)

    where = ' AND '.join(conditions) if conditions else '1=1'
    return where, params
