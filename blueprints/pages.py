"""
AYC Portal — HTML page routes blueprint.
All routes that render templates — main portal sections and admin pages.
"""

import os
from datetime import datetime

from flask import (
    Blueprint, Response, current_app, redirect, render_template,
    request, send_from_directory, session,
)

from config import APP_VERSION, BRANDING_DIR, CLUB_NAME, CLUB_SHORT_NAME
from helpers import (
    get_db, login_required, permission_required,
    tpl_ctx, get_brand_settings, get_session_types, get_valid_session_names,
    _ensure_qr_tokens_for_today,
)

bp = Blueprint('pages', __name__)


@bp.route('/dashboard')
@login_required
def dashboard_page():
    db        = get_db()
    reg_types = db.execute(
        'SELECT slug, name, icon, colour, description, public_registration '
        'FROM member_types WHERE active = 1 ORDER BY sort_order'
    ).fetchall()
    return render_template('dashboard.html', active_page='dashboard',
                           reg_types=[dict(t) for t in reg_types],
                           **tpl_ctx())


@bp.route('/members')
@permission_required('members.view')
def members_page():
    return render_template('members.html', active_page='members', **tpl_ctx())


@bp.route('/approvals')
@permission_required('approvals.view')
def approvals_page():
    return render_template('approvals.html', active_page='approvals', **tpl_ctx())


@bp.route('/register')
@login_required
def register_page():
    _ensure_qr_tokens_for_today()
    return render_template('register.html', active_page='register', **tpl_ctx())


@bp.route('/register/print')
@permission_required('register.print')
def print_register_page():
    session_type = request.args.get('session', '').strip()
    date         = request.args.get('date', '').strip()
    type_slug    = request.args.get('type', 'member').strip() or 'member'

    if not session_type or not date:
        return 'Missing session or date parameter', 400

    valid_sessions = get_valid_session_names()
    if session_type not in valid_sessions:
        return 'Invalid session type', 400

    db = get_db()

    mtype = db.execute(
        'SELECT * FROM member_types WHERE slug = ? AND active = 1', (type_slug,)
    ).fetchone()
    if not mtype:
        mtype = db.execute(
            'SELECT * FROM member_types WHERE active = 1 ORDER BY sort_order LIMIT 1'
        ).fetchone()
    mtype_dict = dict(mtype) if mtype else {}

    print_fields_raw = []
    if mtype:
        pf_rows = db.execute('''
            SELECT  fd.id, fd.key, fd.label, fd.field_type,
                    fd.column_name, fd.system_field
            FROM    member_type_fields mtf
            JOIN    field_definitions fd ON fd.id = mtf.field_id
            WHERE   mtf.member_type_id = ? AND fd.active = 1 AND mtf.show_on_print = 1
            ORDER   BY mtf.sort_order
        ''', (mtype['id'],)).fetchall()
        print_fields_raw = [dict(r) for r in pf_rows]

    SKIP_PRINT_KEYS = {'first_name', 'surname'}
    dynamic_fields  = [f for f in print_fields_raw if f['key'] not in SKIP_PRINT_KEYS]

    # Use the resolved member type's slug, not the raw URL default ('member').
    # The print button passes no ?type=, so type_slug falls back to 'member';
    # if the club's member type uses a different slug (e.g. 'ara-members'),
    # mtype is resolved via the fallback above and we must filter members by
    # that same slug — otherwise the query matches zero rows.
    query_slug = mtype['slug'] if mtype else type_slug

    members_raw = db.execute('''
        SELECT  m.*
        FROM    members m
        WHERE   EXISTS (SELECT 1 FROM member_statuses ms WHERE ms.name = m.status AND ms.behaviour = 'active')
          AND   m.member_type = ?
          AND   m.session     = ?
        ORDER   BY m.first_name, m.surname
    ''', (query_slug, session_type)).fetchall()
    members = [dict(r) for r in members_raw]

    has_custom = any(not f['system_field'] for f in dynamic_fields)
    if members and has_custom:
        member_ids   = [m['id'] for m in members]
        placeholders = ','.join('?' * len(member_ids))
        cfv_rows = db.execute(
            f'SELECT mfv.member_id, fd.key, mfv.value '
            f'FROM member_field_values mfv '
            f'JOIN field_definitions fd ON fd.id = mfv.field_id '
            f'WHERE mfv.member_id IN ({placeholders})',
            member_ids
        ).fetchall()
        custom_map = {}
        for cfv in cfv_rows:
            custom_map.setdefault(cfv['member_id'], {})[cfv['key']] = cfv['value']
        for m in members:
            m['custom_fields'] = custom_map.get(m['id'], {})
    else:
        for m in members:
            m['custom_fields'] = {}

    notes = db.execute('''
        SELECT  sn.id, sn.note_type, sn.title, sn.details, sn.created_at,
                u.username   AS added_by_name,
                m.first_name AS member_first, m.surname AS member_surname
        FROM    session_notes sn
        LEFT JOIN users   u ON u.id = sn.added_by
        LEFT JOIN members m ON m.id = sn.member_id
        WHERE   sn.session_date = ? AND sn.session_type = ?
        ORDER   BY sn.created_at
    ''', (date, session_type)).fetchall()

    try:
        display_date = datetime.strptime(date, '%Y-%m-%d').strftime('%d/%m/%Y')
    except ValueError:
        display_date = date

    brand = get_brand_settings()
    return render_template(
        'print_register.html',
        session_type   = session_type,
        date           = date,
        display_date   = display_date,
        members        = members,
        notes          = [dict(r) for r in notes],
        dynamic_fields = dynamic_fields,
        mtype          = mtype_dict,
        club_name      = brand.get('brand_club_name') or CLUB_NAME,
        club_short_name= brand.get('brand_short_name') or CLUB_SHORT_NAME,
        brand          = brand,
    )


@bp.route('/register/print.xlsx')
@permission_required('register.print')
def print_register_xlsx():
    """Excel version of the printable session register.

    Mirrors the print register layout — logo header, the member type's
    show_on_print fields, and blank Tick on Arrival / Tick When Leaving columns —
    so it can be saved or printed from Excel. Replaces the in-page Print button,
    which mis-rendered on some browsers.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    session_type = request.args.get('session', '').strip()
    date         = request.args.get('date', '').strip()
    type_slug    = request.args.get('type', 'member').strip() or 'member'

    if not session_type or not date:
        return 'Missing session or date parameter', 400
    if session_type not in get_valid_session_names():
        return 'Invalid session type', 400

    db = get_db()

    # Resolve member type (fall back to the club's first active type) and use its
    # slug for the member query — never the raw 'member' default. See print page.
    mtype = db.execute(
        'SELECT * FROM member_types WHERE slug = ? AND active = 1', (type_slug,)
    ).fetchone()
    if not mtype:
        mtype = db.execute(
            'SELECT * FROM member_types WHERE active = 1 ORDER BY sort_order LIMIT 1'
        ).fetchone()
    query_slug = mtype['slug'] if mtype else type_slug
    mtype_name = (mtype['name'] if mtype else '') or ''

    print_fields = []
    if mtype:
        print_fields = [dict(r) for r in db.execute('''
            SELECT  fd.key, fd.label, fd.field_type, fd.column_name, fd.system_field
            FROM    member_type_fields mtf
            JOIN    field_definitions fd ON fd.id = mtf.field_id
            WHERE   mtf.member_type_id = ? AND fd.active = 1 AND mtf.show_on_print = 1
              AND   fd.key NOT IN ('first_name', 'surname')
            ORDER   BY mtf.sort_order
        ''', (mtype['id'],)).fetchall()]

    members = [dict(r) for r in db.execute('''
        SELECT  m.*
        FROM    members m
        WHERE   EXISTS (SELECT 1 FROM member_statuses ms WHERE ms.name = m.status AND ms.behaviour = 'active')
          AND   m.member_type = ?
          AND   m.session     = ?
        ORDER   BY m.first_name, m.surname
    ''', (query_slug, session_type)).fetchall()]

    has_custom = any(not f['system_field'] for f in print_fields)
    if members and has_custom:
        ids = [m['id'] for m in members]
        ph  = ','.join('?' * len(ids))
        cmap = {}
        for r in db.execute(
            f'SELECT mfv.member_id, fd.key, mfv.value FROM member_field_values mfv '
            f'JOIN field_definitions fd ON fd.id = mfv.field_id '
            f'WHERE mfv.member_id IN ({ph})', ids
        ).fetchall():
            cmap.setdefault(r['member_id'], {})[r['key']] = r['value']
        for m in members:
            m['custom_fields'] = cmap.get(m['id'], {})
    else:
        for m in members:
            m['custom_fields'] = {}

    def _field_value(field, member):
        if field['system_field']:
            val = member.get(field['column_name'])
        else:
            val = (member.get('custom_fields') or {}).get(field['key'])
        if val is None or val == '':
            return ''
        if field['field_type'] == 'boolean':
            return 'YES' if val in (1, '1', 'true', 'yes', True) else 'NO'
        return val

    try:
        display_date = datetime.strptime(date, '%Y-%m-%d').strftime('%d/%m/%Y')
    except ValueError:
        display_date = date

    brand  = get_brand_settings()
    accent = (brand.get('_palette', {}).get('accent') or '#1b2d4f').lstrip('#').upper()
    short  = brand.get('brand_short_name') or CLUB_SHORT_NAME

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'Register'

    headers  = (['#', 'First Name', 'Surname']
                + [f['label'] for f in print_fields]
                + ['Tick on Arrival', 'Tick When Leaving'])
    ncols    = len(headers)
    HEADER_ROW = 6   # rows 1-2 logo, 3 title, 4 subtitle, 5 spacer

    HEADER_FILL = PatternFill('solid', fgColor=accent)
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
    TITLE_FONT  = Font(bold=True, color=accent, size=15)
    SUB_FONT    = Font(color='475569', size=10)
    _thin       = Side(style='thin', color='CBD5E1')
    BORDER      = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    title = f'Session Register — {session_type}'
    if mtype_name:
        title += f' · {mtype_name}s'
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    tc = ws.cell(row=3, column=1, value=title)
    tc.font = TITLE_FONT
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)
    sc = ws.cell(row=4, column=1,
                 value=f'{display_date}  ·  {len(members)} '
                       f'{(mtype_name or "member").lower()}{"s" if len(members) != 1 else ""}')
    sc.font = SUB_FONT
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[5].height = 6

    # Logo (optional — embedding requires Pillow; degrade gracefully if absent)
    logo_path = None
    if brand.get('brand_logo_file'):
        cand = os.path.join(BRANDING_DIR, os.path.basename(brand['brand_logo_file']))
        if os.path.exists(cand):
            logo_path = cand
    if not logo_path and current_app.static_folder:
        cand = os.path.join(current_app.static_folder, 'images', 'logo.png')
        if os.path.exists(cand):
            logo_path = cand
    if logo_path:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo_path)
            target_h = 44
            if img.height:
                img.width  = int(img.width * (target_h / float(img.height)))
                img.height = target_h
            ws.add_image(img, 'A1')
        except Exception as _e:  # Pillow missing or unreadable image
            current_app.logger.warning('Register XLSX logo embed skipped: %s', _e)

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=HEADER_ROW, column=ci, value=h)
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDER
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    row = HEADER_ROW + 1
    for idx, m in enumerate(members, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=m.get('first_name') or '')
        ws.cell(row=row, column=3, value=m.get('surname') or '')
        for fi, f in enumerate(print_fields):
            ws.cell(row=row, column=4 + fi, value=_field_value(f, m))
        for ci in range(1, ncols + 1):
            ws.cell(row=row, column=ci).border = BORDER
        row += 1
    # Blank rows for walk-ins / late arrivals, mirroring the print template.
    for i in range(5):
        ws.cell(row=row, column=1, value=len(members) + i + 1)
        for ci in range(1, ncols + 1):
            ws.cell(row=row, column=ci).border = BORDER
        row += 1

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    for fi in range(len(print_fields)):
        ws.column_dimensions[get_column_letter(4 + fi)].width = 18
    ws.column_dimensions[get_column_letter(ncols - 1)].width = 16
    ws.column_dimensions[get_column_letter(ncols)].width = 16

    ws.freeze_panes = f'A{HEADER_ROW + 1}'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 9  # A4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fn = f"{short.lower().replace(' ', '_')}_register_{session_type}_{date}.xlsx".replace(' ', '_')
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fn}"'},
    )


@bp.route('/register/export')
@permission_required('register.export')
def export_register_page():
    session_type = request.args.get('session', '').strip()
    date         = request.args.get('date', '').strip()

    if not session_type or not date:
        return 'Missing session or date parameter', 400

    valid_sessions = get_valid_session_names()
    if session_type not in valid_sessions:
        return 'Invalid session type', 400

    db = get_db()

    completion = db.execute('''
        SELECT sc.completed_at, sc.auto_signout_count,
               u.username AS completed_by_name
        FROM   session_completions sc
        LEFT JOIN users u ON u.id = sc.completed_by
        WHERE  sc.session_date = ? AND sc.session_type = ?
    ''', (date, session_type)).fetchone()
    if not completion:
        return 'This register has not been completed yet.', 400

    member_mt = db.execute(
        '''SELECT mt.id FROM member_types mt
           JOIN members m ON m.member_type = mt.slug
           WHERE m.session = ? AND mt.registration_style != "staff" AND mt.active = 1
           LIMIT 1''',
        (session_type,)
    ).fetchone()
    export_fields = []
    if member_mt:
        ef_rows = db.execute('''
            SELECT  fd.key, fd.label, fd.field_type, fd.column_name, fd.system_field
            FROM    member_type_fields mtf
            JOIN    field_definitions fd ON fd.id = mtf.field_id
            WHERE   mtf.member_type_id = ? AND fd.active = 1 AND mtf.show_on_export = 1
              AND   fd.key NOT IN ('first_name', 'surname')
            ORDER   BY mtf.sort_order
        ''', (member_mt['id'],)).fetchall()
        export_fields = [dict(f) for f in ef_rows]

    members = db.execute('''
        SELECT  m.*,
                a.signed_in_at, a.signed_out_at
        FROM    members m
        JOIN    member_types mt ON mt.slug = m.member_type
        LEFT JOIN attendance a
               ON  a.member_id   = m.id
               AND a.session_date = ?
               AND a.session_type = ?
        WHERE   EXISTS (SELECT 1 FROM member_statuses ms WHERE ms.name = m.status AND ms.behaviour = 'active')
          AND   mt.registration_style != 'staff'
          AND   m.session     = ?
        ORDER   BY m.surname, m.first_name
    ''', (date, session_type, session_type)).fetchall()

    member_ids = [m['id'] for m in members]
    custom_fields_map = {}
    if member_ids and export_fields:
        placeholders = ','.join('?' * len(member_ids))
        cfv_rows = db.execute(
            f'SELECT mfv.member_id, fd.key, mfv.value '
            f'FROM member_field_values mfv '
            f'JOIN field_definitions fd ON fd.id = mfv.field_id '
            f'WHERE mfv.member_id IN ({placeholders})',
            member_ids
        ).fetchall()
        for cfv in cfv_rows:
            custom_fields_map.setdefault(cfv['member_id'], {})[cfv['key']] = cfv['value']

    members_out = []
    for m in members:
        md    = dict(m)
        extra = {}
        for f in export_fields:
            if f['system_field'] and f['column_name']:
                extra[f['key']] = md.get(f['column_name'], '') or ''
            else:
                extra[f['key']] = custom_fields_map.get(md['id'], {}).get(f['key'], '') or ''
        md['export_extra'] = extra
        members_out.append(md)
    members = members_out

    staff = db.execute('''
        SELECT  m.first_name, m.surname, m.staff_role,
                a.signed_in_at, a.signed_out_at
        FROM    members m
        LEFT JOIN attendance a
               ON  a.member_id   = m.id
               AND a.session_date = ?
               AND a.session_type = ?
        JOIN    member_types mt ON mt.slug = m.member_type
        WHERE   EXISTS (SELECT 1 FROM member_statuses ms WHERE ms.name = m.status AND ms.behaviour = 'active')
          AND   mt.registration_style = 'staff'
          AND   m.session     = ?
        ORDER   BY m.surname, m.first_name
    ''', (date, session_type, session_type)).fetchall()

    notes = db.execute('''
        SELECT  sn.note_type, sn.title, sn.details, sn.created_at,
                u.username   AS added_by_name,
                m.first_name AS member_first, m.surname AS member_surname
        FROM    session_notes sn
        LEFT JOIN users   u ON u.id = sn.added_by
        LEFT JOIN members m ON m.id = sn.member_id
        WHERE   sn.session_date = ? AND sn.session_type = ?
        ORDER   BY sn.created_at
    ''', (date, session_type)).fetchall()

    try:
        display_date = datetime.strptime(date, '%Y-%m-%d').strftime('%A %d %B %Y')
    except ValueError:
        display_date = date

    attended    = sum(1 for m in members if m['signed_in_at'])
    not_arrived = sum(1 for m in members if not m['signed_in_at'])

    brand = get_brand_settings()
    return render_template(
        'register_export.html',
        session_type    = session_type,
        date            = date,
        display_date    = display_date,
        completion      = dict(completion),
        members         = members,
        staff           = [dict(s) for s in staff],
        notes           = [dict(n) for n in notes],
        export_fields   = export_fields,
        attended        = attended,
        not_arrived     = not_arrived,
        club_name       = brand.get('brand_club_name') or CLUB_NAME,
        club_short_name = brand.get('brand_short_name') or CLUB_SHORT_NAME,
        brand           = brand,
        app_version     = APP_VERSION,
    )


@bp.route('/registration')
def registration_page():
    db    = get_db()
    brand = get_brand_settings()
    types = db.execute(
        'SELECT * FROM member_types WHERE public_registration = 1 AND active = 1 ORDER BY sort_order'
    ).fetchall()
    return render_template('registration_landing.html',
                           reg_types=[dict(t) for t in types],
                           club_name=brand.get('brand_club_name') or CLUB_NAME,
                           club_short_name=brand.get('brand_short_name') or CLUB_SHORT_NAME,
                           brand=brand)


@bp.route('/registration/<slug>')
def registration_slug_page(slug):
    db    = get_db()
    brand = get_brand_settings()
    mtype = db.execute(
        'SELECT * FROM member_types WHERE slug = ? AND active = 1', (slug,)
    ).fetchone()
    if not mtype:
        return redirect('/registration')
    return render_template('registration_dynamic.html',
                           type_slug=slug,
                           type_info=dict(mtype),
                           version=APP_VERSION,
                           club_name=brand.get('brand_club_name') or CLUB_NAME,
                           club_short_name=brand.get('brand_short_name') or CLUB_SHORT_NAME,
                           brand=brand)


@bp.route('/documents')
@permission_required('documents.view')
def documents_page():
    return render_template('documents.html', active_page='documents', **tpl_ctx())


@bp.route('/communications')
@permission_required('mailshots.send')
def communications_page():
    return render_template('communications.html', active_page='communications', **tpl_ctx())


@bp.route('/calendar')
@login_required
def calendar_page():
    return render_template('calendar.html', active_page='calendar', **tpl_ctx())


@bp.route('/display')
def display_page():
    brand = get_brand_settings()
    club  = brand.get('brand_club_name') or CLUB_NAME
    short = brand.get('brand_short_name') or CLUB_SHORT_NAME
    return render_template('display.html',
                           current_session=session.get('active_session', ''),
                           session_types=get_session_types(),
                           club_name=club, club_short_name=short,
                           brand=brand)


@bp.route('/quick-session')
def quick_session_page():
    brand = get_brand_settings()
    return render_template('quick_session.html',
                           club_name=brand.get('brand_club_name') or CLUB_NAME,
                           club_short_name=brand.get('brand_short_name') or CLUB_SHORT_NAME,
                           brand=brand)


# ── Branding — public logo endpoint ──────────────────────────────────────────

@bp.route('/branding/logo')
def branding_logo():
    """Serve the organisation logo — publicly accessible, no auth needed."""
    brand    = get_brand_settings()
    filename = brand.get('brand_logo_file', '')
    if not filename:
        return '', 404
    safe = os.path.join(BRANDING_DIR, os.path.basename(filename))
    if not os.path.isfile(safe):
        return '', 404
    return send_from_directory(BRANDING_DIR, os.path.basename(filename))


# ── Admin HTML page routes ────────────────────────────────────────────────────

@bp.route('/admin/users')
@permission_required('users.view')
def users_page():
    return render_template('admin/users.html', active_page='users', **tpl_ctx())


@bp.route('/admin/audit')
@permission_required('audit.view')
def audit_page():
    return render_template('admin/audit.html', active_page='audit', **tpl_ctx())


@bp.route('/admin/logs')
@permission_required('admin.maintenance')
def system_logs_page():
    return render_template('admin/system_logs.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/branding')
@permission_required('admin.branding')
def branding_page():
    return render_template('admin/branding.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/settings')
@permission_required('admin.settings')
def settings_page():
    return render_template('admin/settings.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/roles')
@permission_required('admin.roles')
def roles_page():
    return render_template('admin/roles.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/smtp-profiles')
@permission_required('admin.smtp_profiles')
def smtp_profiles_page():
    return render_template('admin/smtp_profiles.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/staff-roles')
@permission_required('admin.settings')
def staff_roles_page():
    return render_template('admin/staff_roles.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/tags')
@permission_required('admin.settings')
def tags_page():
    return render_template('admin/tags.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/member-statuses')
@permission_required('admin.settings')
def member_statuses_page():
    return render_template('admin/member_statuses.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/payments')
@permission_required('payments.manage')
def payments_admin_page():
    return render_template('admin/payments.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/document-categories')
@permission_required('admin.settings')
def document_categories_page():
    return render_template('admin/document_categories.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/document-fields/<int:cat_id>')
@permission_required('admin.settings')
def document_fields_page(cat_id):
    db  = get_db()
    cat = db.execute('SELECT id FROM document_categories WHERE id = ?', (cat_id,)).fetchone()
    if not cat:
        return redirect('/admin/document-categories')
    return render_template('admin/document_fields.html',
                           cat_id=cat_id, active_page='settings', **tpl_ctx())


@bp.route('/admin/member-types')
@permission_required('admin.settings')
def member_types_page():
    return render_template('admin/member_types.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/field-builder/<int:type_id>')
@permission_required('admin.settings')
def field_builder_page(type_id):
    db    = get_db()
    mtype = db.execute('SELECT * FROM member_types WHERE id = ?', (type_id,)).fetchone()
    if not mtype:
        return redirect('/admin/member-types')
    return render_template('admin/field_builder.html',
                           active_page='settings',
                           mtype=dict(mtype),
                           **tpl_ctx())


@bp.route('/admin/settings/attendance')
@permission_required('admin.settings')
def attendance_settings_page():
    return render_template('admin/attendance_settings.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/settings/session-types')
@permission_required('admin.session_types')
def session_types_page():
    return render_template('admin/session_types.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/settings/maintenance')
@permission_required('admin.maintenance')
def maintenance_page():
    return render_template('admin/maintenance.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/settings/import')
@permission_required('admin.maintenance')
def import_page():
    return render_template('admin/import.html', active_page='settings', **tpl_ctx())


@bp.route('/admin/settings/import-history')
@permission_required('admin.maintenance')
def import_history_page():
    from helpers import get_db
    db           = get_db()
    session_types = db.execute('SELECT name FROM session_types ORDER BY sort_order, name').fetchall()
    payment_types = db.execute('SELECT name FROM payment_types ORDER BY sort_order, name').fetchall()
    payment_methods = db.execute('SELECT name FROM payment_methods ORDER BY sort_order, name').fetchall()
    return render_template(
        'admin/import_history.html',
        active_page='settings',
        session_types=session_types,
        payment_types=payment_types,
        payment_methods=payment_methods,
        **tpl_ctx()
    )


@bp.route('/admin/settings/export')
@permission_required('admin.maintenance')
def export_page():
    return render_template('admin/export.html', active_page='settings', **tpl_ctx())
